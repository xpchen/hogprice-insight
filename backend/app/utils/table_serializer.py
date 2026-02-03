"""表格JSON序列化工具 - 将Excel sheet转换为JSON格式，支持合并单元格信息"""
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import json


def serialize_worksheet_to_json(
    worksheet: Worksheet,
    max_rows: Optional[int] = None,
    sparse: bool = False
) -> Dict[str, Any]:
    """
    将openpyxl Worksheet转换为JSON格式
    
    Args:
        worksheet: openpyxl Worksheet对象
        max_rows: 最大行数（None表示全部）
        sparse: 是否使用稀疏格式（只存非空单元格）
    
    Returns:
        包含table_json和merged_cells_json的字典
    """
    # 获取实际使用的行数和列数
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    if max_rows:
        max_row = min(max_row, max_rows)
    
    # 提取合并单元格信息
    merged_cells = []
    for merged_range in worksheet.merged_cells.ranges:
        merged_cells.append({
            "min_row": merged_range.min_row,
            "max_row": merged_range.max_row,
            "min_col": merged_range.min_col,
            "max_col": merged_range.max_col
        })
    
    # 提取表格数据
    if sparse:
        # 稀疏格式：只存非空单元格（进一步优化：限制列数）
        table_data = []
        max_cols = min(max_col, 100)  # 限制最大列数，避免存储过多列
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_cols, values_only=False), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    # 限制字符串长度（避免存储过长的文本）
                    cell_value = _serialize_cell_value(cell)
                    if isinstance(cell_value, str) and len(cell_value) > 200:
                        cell_value = cell_value[:197] + "..."
                    row_data.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": cell_value,
                        "data_type": _get_cell_type(cell)
                    })
            if row_data:
                table_data.append(row_data)
    else:
        # 密集格式：二维数组
        table_data = []
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            row_data = []
            for cell_value in row:
                row_data.append(_serialize_cell_value_simple(cell_value))
            table_data.append(row_data)
    
    return {
        "table_json": table_data,
        "merged_cells_json": merged_cells,
        "row_count": max_row,
        "col_count": max_col
    }


def _serialize_cell_value(cell: Cell) -> Any:
    """序列化单元格值（保留类型信息）"""
    value = cell.value
    
    if value is None:
        return None
    
    # 处理日期时间
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    
    # 处理数字
    if isinstance(value, (int, float)):
        return value
    
    # 处理字符串
    if isinstance(value, str):
        return value
    
    # 其他类型转为字符串
    return str(value)


def _serialize_cell_value_simple(cell_value: Any) -> Any:
    """简单序列化单元格值（用于密集格式）"""
    if cell_value is None:
        return None
    
    # 处理日期时间
    if hasattr(cell_value, 'isoformat'):
        return cell_value.isoformat()
    
    return cell_value


def _get_cell_type(cell: Cell) -> str:
    """获取单元格数据类型"""
    if cell.data_type == 'n':
        return 'number'
    elif cell.data_type == 's':
        return 'string'
    elif cell.data_type == 'd':
        return 'date'
    elif cell.data_type == 'b':
        return 'boolean'
    elif cell.data_type == 'f':
        return 'formula'
    else:
        return 'unknown'


def extract_header_signature(worksheet: Worksheet, header_rows: int = 3) -> str:
    """
    提取表头签名（用于sheet识别和匹配）
    
    Args:
        worksheet: openpyxl Worksheet对象
        header_rows: 表头行数（默认3行）
    
    Returns:
        表头签名字符串（JSON格式）
    """
    max_col = min(worksheet.max_column, 20)  # 只取前20列
    header_data = []
    
    for row_idx in range(1, min(header_rows + 1, worksheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                # 只保留前50个字符
                value_str = str(value)[:50]
                row_data.append(value_str)
            else:
                row_data.append("")
        header_data.append(row_data)
    
    # 生成签名（使用JSON序列化后取hash）
    signature_json = json.dumps(header_data, ensure_ascii=False, sort_keys=True)
    
    # 如果太长，截断
    if len(signature_json) > 500:
        signature_json = signature_json[:500]
    
    return signature_json


def deserialize_table_json(table_json: List[List[Any]]) -> List[List[Any]]:
    """
    反序列化表格JSON（用于前端展示）
    
    Args:
        table_json: 表格JSON数据
    
    Returns:
        二维数组
    """
    return table_json


def apply_merged_cells_to_table(
    table_data: List[List[Any]],
    merged_cells: List[Dict[str, int]]
) -> List[List[Any]]:
    """
    将合并单元格信息应用到表格数据（用于前端展示）
    
    Args:
        table_data: 表格数据（二维数组）
        merged_cells: 合并单元格信息列表
    
    Returns:
        应用合并单元格后的表格数据
    """
    # 创建副本
    result = [row[:] for row in table_data]
    
    # 应用合并单元格（将合并区域的值填充到所有单元格）
    for merged in merged_cells:
        min_row = merged["min_row"] - 1  # 转换为0-based索引
        max_row = merged["max_row"] - 1
        min_col = merged["min_col"] - 1
        max_col = merged["max_col"] - 1
        
        # 获取主单元格的值（通常是左上角）
        if min_row < len(result) and min_col < len(result[min_row]):
            master_value = result[min_row][min_col]
            
            # 填充到合并区域的所有单元格
            for row_idx in range(min_row, min(max_row + 1, len(result))):
                for col_idx in range(min_col, min(max_col + 1, len(result[row_idx]) if row_idx < len(result) else 0)):
                    if row_idx < len(result) and col_idx < len(result[row_idx]):
                        result[row_idx][col_idx] = master_value
    
    return result
