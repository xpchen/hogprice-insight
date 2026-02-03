"""宽表解析通用工具（openpyxl + MultiIndex + melt）"""
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import re


def parse_wide_table_with_dates(
    workbook_or_path,
    sheet_name: str,
    date_col_start: Optional[int] = None,
    header_rows: int = 1,
    region_col: int = 0
) -> pd.DataFrame:
    """
    通用宽表解析器：将"省份×日期列"的宽表转换为长表
    
    Args:
        workbook_or_path: openpyxl workbook对象或文件路径/BytesIO
        sheet_name: Sheet名称
        date_col_start: 日期列起始位置（从0开始），如果为None则自动识别
        header_rows: 表头行数（默认1行）
        region_col: 区域列索引（默认第0列）
    
    Returns:
        DataFrame with columns: region, trade_date, value, field (可选)
    """
    # 加载workbook
    if isinstance(workbook_or_path, BytesIO):
        wb = load_workbook(workbook_or_path, data_only=True)
    elif isinstance(workbook_or_path, str):
        wb = load_workbook(workbook_or_path, data_only=True)
    else:
        wb = workbook_or_path
    
    ws = wb[sheet_name]
    
    # 读取数据到DataFrame（不设header，保留原始结构）
    data_rows = []
    for row in ws.iter_rows(min_row=header_rows + 1, values_only=True):
        data_rows.append(row)
    
    if not data_rows:
        return pd.DataFrame(columns=['region', 'trade_date', 'value'])
    
    df = pd.DataFrame(data_rows)
    
    # 读取表头（处理合并单元格）
    header_data = []
    for row_idx in range(1, header_rows + 1):
        row_values = []
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            # 获取合并单元格的值
            cell_value = cell.value
            if cell_value is None:
                # 检查是否在合并单元格范围内
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # 获取合并单元格的左上角值
                        top_left = ws[merged_range.min_col][merged_range.min_row]
                        cell_value = top_left.value
                        break
            row_values.append(cell_value)
        header_data.append(row_values)
    
    # 识别日期列
    if date_col_start is None:
        date_col_start = _detect_date_columns(header_data, region_col)
    
    if date_col_start is None:
        raise ValueError(f"无法识别日期列起始位置，请手动指定 date_col_start")
    
    # 提取区域列（第一列）
    regions = df.iloc[:, region_col].astype(str).str.strip()
    
    # 提取日期列和对应的值
    date_columns = []
    for col_idx in range(date_col_start, df.shape[1]):
        # 从表头中提取日期
        date_str = None
        for header_row in header_data:
            if col_idx < len(header_row) and header_row[col_idx]:
                date_str = str(header_row[col_idx]).strip()
                break
        
        if date_str:
            # 尝试解析日期
            trade_date = _parse_date(date_str)
            if trade_date:
                date_columns.append((col_idx, trade_date))
    
    # Melt成长表
    result_rows = []
    for row_idx in range(len(df)):
        region = regions.iloc[row_idx]
        if pd.isna(region) or region == '' or region == 'nan':
            continue
        
        for col_idx, trade_date in date_columns:
            value = df.iloc[row_idx, col_idx]
            if pd.notna(value) and value != '':
                result_rows.append({
                    'region': region,
                    'trade_date': trade_date,
                    'value': value
                })
    
    result_df = pd.DataFrame(result_rows)
    
    return result_df


def _detect_date_columns(header_data: List[List], region_col: int = 0) -> Optional[int]:
    """
    自动检测日期列起始位置
    
    Args:
        header_data: 表头数据（多行）
        region_col: 区域列索引
    
    Returns:
        日期列起始位置，如果无法识别则返回None
    """
    # 从区域列之后开始查找
    for col_idx in range(region_col + 1, len(header_data[0]) if header_data else 0):
        # 检查该列的表头是否包含日期格式
        for header_row in header_data:
            if col_idx < len(header_row) and header_row[col_idx]:
                cell_value = str(header_row[col_idx]).strip()
                if _parse_date(cell_value):
                    return col_idx
    
    return None


def _parse_date(date_str: str) -> Optional[datetime]:
    """
    解析日期字符串（支持多种格式）
    
    Args:
        date_str: 日期字符串
    
    Returns:
        datetime对象，如果无法解析则返回None
    """
    if not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # 常见日期格式
    date_formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%Y.%m.%d',
        '%Y年%m月%d日',
        '%m/%d/%Y',
        '%d/%m/%Y',
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    # 尝试pandas解析
    try:
        return pd.to_datetime(date_str).to_pydatetime()
    except:
        return None


def parse_multirow_header_wide_table(
    workbook_or_path,
    sheet_name: str,
    header_start_row: Optional[int] = None,
    region_col: int = 0
) -> pd.DataFrame:
    """
    解析多行表头的宽表（用于周度数据）
    
    Args:
        workbook_or_path: openpyxl workbook对象或文件路径/BytesIO
        sheet_name: Sheet名称
        header_start_row: 表头起始行（从1开始），如果为None则自动识别（查找"开始日期"/"结束日期"）
        region_col: 区域列索引（默认第0列）
    
    Returns:
        DataFrame with columns: week_start, week_end, region, field, value
    """
    # 加载workbook
    if isinstance(workbook_or_path, BytesIO):
        wb = load_workbook(workbook_or_path, data_only=True)
    elif isinstance(workbook_or_path, str):
        wb = load_workbook(workbook_or_path, data_only=True)
    else:
        wb = workbook_or_path
    
    ws = wb[sheet_name]
    
    # 自动识别表头起始行
    if header_start_row is None:
        header_start_row = _find_header_start_row(ws)
    
    if header_start_row is None:
        raise ValueError("无法识别表头起始行，请手动指定 header_start_row")
    
    # 读取表头（处理合并单元格，forward-fill）
    # 先读取包含"起始日期"/"结束日期"的那一行（header_start_row）
    header_rows = []
    max_col = ws.max_column
    
    # 读取包含日期列名的那一行
    row_values = []
    prev_value = None
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_start_row, column=col_idx)
        cell_value = cell.value
        
        # 处理合并单元格
        if cell_value is None:
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws[merged_range.min_col][merged_range.min_row]
                    cell_value = top_left.value
                    break
        
        # Forward-fill：如果当前单元格为空，使用前一个值
        if cell_value is None:
            cell_value = prev_value
        else:
            prev_value = cell_value
        
        row_values.append(cell_value)
    header_rows.append(row_values)
    
    # 识别"开始日期"和"结束日期"列
    week_start_col = None
    week_end_col = None
    
    for col_idx, header_value in enumerate(header_rows[0]):
        if header_value:
            header_str = str(header_value).strip()
            # 支持"开始日期"和"起始日期"
            if '开始日期' in header_str or '起始日期' in header_str or 'start' in header_str.lower():
                week_start_col = col_idx
            elif '结束日期' in header_str or 'end' in header_str.lower():
                week_end_col = col_idx
    
    if week_start_col is None or week_end_col is None:
        raise ValueError("无法识别开始日期/结束日期列")
    
    # 读取数据行（数据从表头行的下一行开始）
    data_start_row = header_start_row + 1
    result_rows = []
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        week_start_val = ws.cell(row=row_idx, column=week_start_col + 1).value
        week_end_val = ws.cell(row=row_idx, column=week_end_col + 1).value
        
        if not week_start_val or not week_end_val:
            continue
        
        week_start = _parse_date(str(week_start_val))
        week_end = _parse_date(str(week_end_val))
        
        if not week_start or not week_end:
            continue
        
        # 读取该行的其他列（指标值）
        region = ws.cell(row=row_idx, column=region_col + 1).value
        if not region:
            continue
        
        # 遍历其他列（指标列）
        for col_idx in range(max(week_start_col, week_end_col) + 1, max_col + 1):
            field_name = None
            # 从表头中提取字段名（可能需要组合多行表头）
            field_parts = []
            for header_row in header_rows:
                if col_idx <= len(header_row) and header_row[col_idx - 1]:
                    field_parts.append(str(header_row[col_idx - 1]).strip())
            if field_parts:
                field_name = ' / '.join([p for p in field_parts if p])
            
            value = ws.cell(row=row_idx, column=col_idx).value
            if pd.notna(value) and value != '':
                result_rows.append({
                    'week_start': week_start.date(),
                    'week_end': week_end.date(),
                    'region': str(region).strip(),
                    'field': field_name,
                    'value': value
                })
    
    return pd.DataFrame(result_rows)


def _find_header_start_row(ws) -> Optional[int]:
    """
    查找表头起始行（查找包含"开始日期"/"结束日期"的行）
    
    Args:
        ws: openpyxl worksheet对象
    
    Returns:
        表头起始行号（从1开始），如果找不到则返回None
    """
    for row_idx in range(1, min(10, ws.max_row + 1)):  # 最多检查前10行
        for col_idx in range(1, min(20, ws.max_column + 1)):  # 最多检查前20列
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # 支持"开始日期"和"起始日期"
                if '开始日期' in cell_str or '起始日期' in cell_str or '结束日期' in cell_str or 'start' in cell_str.lower() or 'end' in cell_str.lower():
                    return row_idx
    
    return None
