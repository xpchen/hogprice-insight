"""合并单元格处理工具 - 合并单元格值向下填充"""
from typing import List, Any, Dict, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# 缓存合并单元格映射（避免重复计算）
_merged_cell_cache: Dict[int, Dict[Tuple[int, int], Any]] = {}


def forward_fill_merged_cells(worksheet: Worksheet) -> Worksheet:
    """
    合并单元格值向下填充（处理多行表头中的合并单元格）
    
    Args:
        worksheet: openpyxl Worksheet对象
    
    Returns:
        处理后的Worksheet对象（原地修改）
    """
    # 获取所有合并单元格
    merged_ranges = list(worksheet.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        
        # 获取主单元格的值（通常是左上角）
        master_cell = worksheet.cell(row=min_row, column=min_col)
        master_value = master_cell.value
        
        # 如果主单元格有值，填充到合并区域的所有单元格
        if master_value is not None:
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is None:
                        cell.value = master_value
    
    return worksheet


def get_merged_cell_value(worksheet: Worksheet, row: int, col: int) -> Any:
    """
    获取单元格值（如果是合并单元格，返回主单元格的值）
    
    优化：使用缓存避免重复遍历合并单元格
    
    Args:
        worksheet: openpyxl Worksheet对象
        row: 行号（从1开始）
        col: 列号（从1开始）
    
    Returns:
        单元格值
    """
    # 使用worksheet的id作为缓存key
    cache_key = id(worksheet)
    
    # 如果缓存不存在，构建合并单元格映射
    if cache_key not in _merged_cell_cache:
        merged_map: Dict[Tuple[int, int], Any] = {}
        for merged_range in worksheet.merged_cells.ranges:
            master_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            master_value = master_cell.value
            # 为合并区域内的所有单元格创建映射
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(r, c)] = master_value
        _merged_cell_cache[cache_key] = merged_map
    
    merged_map = _merged_cell_cache[cache_key]
    
    # 检查是否在合并单元格映射中
    if (row, col) in merged_map:
        return merged_map[(row, col)]
    
    # 不在合并单元格中，直接获取单元格值
    cell = worksheet.cell(row=row, column=col)
    return cell.value


def clear_merged_cell_cache():
    """清空合并单元格缓存（用于测试或内存管理）"""
    global _merged_cell_cache
    _merged_cell_cache.clear()
