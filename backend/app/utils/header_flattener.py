"""多行表头扁平化工具 - 将多行表头转换为单行"""
from typing import List, Dict, Any, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from app.utils.merged_cell_handler import get_merged_cell_value


def flatten_multirow_header(
    worksheet: Worksheet,
    header_rows: int = 2,
    start_row: int = 1
) -> List[str]:
    """
    将多行表头扁平化为单行
    
    处理日期跨列+子列结构，生成"日期|子列|字段"格式的列名
    
    Args:
        worksheet: openpyxl Worksheet对象
        header_rows: 表头行数
        start_row: 表头起始行（从1开始）
    
    Returns:
        扁平化后的列名列表
    """
    max_col = worksheet.max_column
    flattened_headers = []
    
    # 读取所有表头行
    header_data = []
    for row_idx in range(start_row, start_row + header_rows):
        row_data = []
        for col_idx in range(1, max_col + 1):
            value = get_merged_cell_value(worksheet, row_idx, col_idx)
            row_data.append(str(value) if value is not None else "")
        header_data.append(row_data)
    
    # 扁平化：将多行合并为单行
    for col_idx in range(max_col):
        parts = []
        for row_idx in range(len(header_data)):
            if col_idx < len(header_data[row_idx]):
                value = header_data[row_idx][col_idx].strip()
                if value and value not in parts:  # 避免重复
                    parts.append(value)
        
        # 合并为"|"分隔的字符串
        if parts:
            flattened_header = "|".join(parts)
        else:
            flattened_header = f"Column_{col_idx + 1}"
        
        flattened_headers.append(flattened_header)
    
    return flattened_headers


def extract_date_grouped_subcols(
    worksheet: Worksheet,
    header_rows: int = 2,
    date_row_idx: int = 1,
    subcol_row_idx: int = 2
) -> List[Dict[str, Any]]:
    """
    提取日期分组列结构（日期跨列 + 子列）
    
    返回结构：
    [
        {
            "date": "2026-01-01",
            "subcols": ["规模场", "小散", "均价", "涨跌"],
            "col_start": 2,
            "col_end": 5
        },
        ...
    ]
    
    性能优化：批量读取表头行数据，避免逐单元格访问
    
    Args:
        worksheet: openpyxl Worksheet对象
        header_rows: 表头行数
        date_row_idx: 日期所在行（从1开始）
        subcol_row_idx: 子列所在行（从1开始）
    
    Returns:
        日期分组列结构列表
    """
    max_col = worksheet.max_column
    date_groups = []
    current_date = None
    current_subcols = []
    col_start = None
    
    # 性能优化：批量读取表头行数据（完全避免逐单元格访问）
    print(f"       [header] 批量读取表头行 (行 {date_row_idx}, {subcol_row_idx}, 共 {max_col} 列)...", flush=True)
    # 批量读取所有表头行
    header_rows_data = {}
    for row_idx in range(min(date_row_idx, subcol_row_idx), max(date_row_idx, subcol_row_idx) + 1):
        row_iter = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx,
                                            min_col=1, max_col=max_col, values_only=True))
        if row_iter:
            header_rows_data[row_idx] = list(row_iter[0])
        else:
            header_rows_data[row_idx] = []
    
    print(f"       [header] 表头行读取完成，开始构建合并单元格映射...", flush=True)
    # 从批量读取的数据构建合并单元格映射（避免再次调用worksheet.cell）
    merged_map: Dict[Tuple[int, int], Any] = {}
    merged_ranges_total = len(worksheet.merged_cells.ranges)
    merged_ranges_processed = 0
    for merged_range in worksheet.merged_cells.ranges:
        merged_ranges_processed += 1
        if merged_ranges_processed % 1000 == 0:
            print(f"       [header] 已处理 {merged_ranges_processed}/{merged_ranges_total} 个合并单元格范围...", flush=True)
        # 只处理涉及表头行的合并单元格
        if (merged_range.min_row <= max(date_row_idx, subcol_row_idx) and 
            merged_range.max_row >= min(date_row_idx, subcol_row_idx)):
            # 从批量读取的数据中获取主单元格值（不调用worksheet.cell）
            master_row_idx = merged_range.min_row
            master_col_idx = merged_range.min_col
            if master_row_idx in header_rows_data:
                master_row_data = header_rows_data[master_row_idx]
                if master_col_idx - 1 < len(master_row_data):
                    master_value = master_row_data[master_col_idx - 1]
                    # 为合并区域内的所有单元格创建映射
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                            merged_map[(r, c)] = master_value
    
    # 获取日期行和子列行数据
    date_row_data = header_rows_data.get(date_row_idx, [])
    subcol_row_data = header_rows_data.get(subcol_row_idx, [])
    
    print(f"       [header] 合并单元格映射完成，开始处理 {max_col} 列...", flush=True)
    
    # 构建日期列的合并单元格主单元格映射（用于判断是否是新的日期组）
    date_merged_master_map = {}  # {(row, col) -> master_col}
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= date_row_idx <= merged_range.max_row:
            master_col = merged_range.min_col
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                date_merged_master_map[(date_row_idx, c)] = master_col
    
    # 处理合并单元格：如果单元格在合并映射中，使用映射值
    prev_date_value = None
    for col_idx in range(1, max_col + 1):
        if col_idx % 1000 == 0:
            print(f"       [header] 已处理 {col_idx}/{max_col} 列...", flush=True)
        col_idx_0 = col_idx - 1  # 数组索引从0开始
        
        # 获取日期值
        if (date_row_idx, col_idx) in merged_map:
            date_value = merged_map[(date_row_idx, col_idx)]
        elif col_idx_0 < len(date_row_data):
            date_value = date_row_data[col_idx_0]
        else:
            date_value = None
        
        # 获取子列值
        if (subcol_row_idx, col_idx) in merged_map:
            subcol_value = merged_map[(subcol_row_idx, col_idx)]
        elif col_idx_0 < len(subcol_row_data):
            subcol_value = subcol_row_data[col_idx_0]
        else:
            subcol_value = None
        
        # 检查是否是新的日期组
        # 只有当日期值发生变化，且当前列是合并单元格的主列（或非合并单元格）时，才创建新组
        is_new_date_group = False
        if date_value and str(date_value).strip():
            date_str = str(date_value).strip()
            
            # 检查是否是合并单元格的主列
            if (date_row_idx, col_idx) in date_merged_master_map:
                # 是合并单元格，只有当是主列时才创建新组
                master_col = date_merged_master_map[(date_row_idx, col_idx)]
                if col_idx == master_col:
                    is_new_date_group = True
            else:
                # 不是合并单元格，直接检查日期是否变化
                if prev_date_value != date_str:
                    is_new_date_group = True
        
        if is_new_date_group:
            # 保存上一个日期组
            if current_date and col_start:
                date_groups.append({
                    "date": str(current_date),
                    "subcols": current_subcols.copy(),
                    "col_start": col_start,
                    "col_end": col_idx - 1
                })
            
            # 开始新的日期组
            current_date = date_value
            current_subcols = []
            col_start = col_idx
            prev_date_value = str(date_value).strip() if date_value else None
        
        # 添加子列（无论是否是新日期组，都要添加子列）
        if subcol_value and str(subcol_value).strip():
            current_subcols.append(str(subcol_value).strip())
    
    # 保存最后一个日期组
    if current_date and col_start:
        # 计算最后一个日期组的实际结束列（基于子列数量）
        # 如果子列数量已知，使用子列数量；否则使用max_col
        if current_subcols:
            # 最后一个日期组的结束列应该是 col_start + len(subcols) - 1
            # 但需要确保不超过max_col
            expected_col_end = col_start + len(current_subcols) - 1
            actual_col_end = min(expected_col_end, max_col)
        else:
            actual_col_end = max_col
        
        date_groups.append({
            "date": str(current_date),
            "subcols": current_subcols.copy(),
            "col_start": col_start,
            "col_end": actual_col_end
        })
    
    print(f"       [header] 完成，共找到 {len(date_groups)} 个日期组", flush=True)
    return date_groups
