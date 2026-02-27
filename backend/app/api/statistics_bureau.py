"""
E4. 统计局数据汇总 API
包含：
1. 表1：统计局季度数据汇总（B-Y列）
2. 图1：统计局生猪出栏量&屠宰量（季度出栏量J列、定点屠宰量P列、规模化率）
3. 图2：猪肉进口（钢联猪肉进口 sheet 或 涌益进口肉 sheet）
"""
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
from datetime import date, datetime, timedelta
from pydantic import BaseModel
from pathlib import Path
import json
import math
import re

from app.core.database import get_db
from app.models.raw_sheet import RawSheet
from app.models.raw_table import RawTable
from app.models.raw_file import RawFile
from app.models.dim_metric import DimMetric
from app.models.fact_observation import FactObservation
from sqlalchemy import func, or_

router = APIRouter(prefix="/api/v1/statistics-bureau", tags=["statistics-bureau"])


class QuarterlyDataRow(BaseModel):
    """季度数据行"""
    period: str  # 季度，如"2024Q1"
    data: Dict[str, Optional[float]]  # B-Y列的数据，键为列名或列索引


class QuarterlyDataResponse(BaseModel):
    """季度数据汇总响应"""
    headers: List[str]  # B-Y列的列名
    data: List[QuarterlyDataRow]


class QuarterlyDataRawResponse(BaseModel):
    """季度数据汇总（按 Excel 原样，多级表头 + 合并单元格）"""
    header_row_0: List[str]
    header_row_1: List[str]
    rows: List[List]
    column_count: int
    merged_cells_json: List[Dict]


class OutputSlaughterPoint(BaseModel):
    """出栏量&屠宰量数据点"""
    period: str  # 季度，如"2024Q1"
    output_volume: Optional[float] = None  # 季度出栏量（J列）
    slaughter_volume: Optional[float] = None  # 定点屠宰量（P列）
    scale_rate: Optional[float] = None  # 规模化率（屠宰量/出栏量）


class OutputSlaughterResponse(BaseModel):
    """出栏量&屠宰量响应"""
    data: List[OutputSlaughterPoint]
    latest_period: Optional[str] = None


class ImportMeatCountryPoint(BaseModel):
    """国别进口量"""
    country: str
    value: Optional[float] = None


class ImportMeatPoint(BaseModel):
    """猪肉进口数据点"""
    month: str  # 月份 YYYY-MM
    total: Optional[float] = None  # 进口总量（万吨）
    top_countries: List[ImportMeatCountryPoint] = []  # 进口量前2的国家


class ImportMeatResponse(BaseModel):
    """猪肉进口响应"""
    data: List[ImportMeatPoint]
    latest_month: Optional[str] = None


def _normalize_table_to_dense(table_data: Any) -> Optional[List[List]]:
    """
    将 raw_table 的 table_json 转为密集二维数组。
    unified_ingestor 使用 sparse=True 保存，每行为 [{row,col,value},...] 格式。
    """
    if not table_data or not isinstance(table_data, list):
        return None
    if len(table_data) == 0:
        return []
    first = table_data[0]
    if isinstance(first, list):
        if len(first) > 0 and isinstance(first[0], dict):
            max_col = 0
            for row_list in table_data:
                if isinstance(row_list, list):
                    for item in row_list:
                        if isinstance(item, dict):
                            max_col = max(max_col, item.get('col', 0))
            rows = []
            for row_list in table_data:
                if not isinstance(row_list, list):
                    rows.append([None] * max_col)
                    continue
                row_array = [None] * max_col
                for item in row_list:
                    if isinstance(item, dict):
                        col_idx = item.get('col', 1) - 1
                        if 0 <= col_idx < max_col:
                            row_array[col_idx] = item.get('value')
                rows.append(row_array)
            return rows
        return table_data
    if isinstance(first, dict):
        max_row = max([item.get('row', 0) for item in table_data if isinstance(item, dict)], default=0)
        max_col = max([item.get('col', 0) for item in table_data if isinstance(item, dict)], default=0)
        rows = [[None] * max_col for _ in range(max_row)]
        for item in table_data:
            if isinstance(item, dict):
                r, c = item.get('row', 1) - 1, item.get('col', 1) - 1
                if 0 <= r < len(rows) and 0 <= c < len(rows[0]):
                    rows[r][c] = item.get('value')
        return rows
    return table_data


def _get_raw_table_data(db: Session, sheet_name: str, filename_pattern: str = None) -> Optional[List[List]]:
    """获取raw_table数据。filename_pattern 用于限定来源文件（如 钢联自动更新模板）"""
    result = _get_raw_table_with_meta(db, sheet_name, filename_pattern)
    raw = result[0] if result else None
    return _normalize_table_to_dense(raw) if raw else None


def _get_raw_table_data_by_like(db: Session, sheet_pattern: str, filename_pattern: str = None) -> Optional[tuple]:
    """
    按 sheet 名称模糊匹配获取 raw_table 数据。
    返回 (table_data, merged) 或 None。用于兼容带前缀的 sheet 名（如 04.猪肉进口）。
    """
    query = db.query(RawSheet).join(RawFile).filter(RawSheet.sheet_name.like(sheet_pattern))
    if filename_pattern:
        query = query.filter(RawFile.filename.like(f'%{filename_pattern}%'))
    sheet = query.order_by(RawFile.id.desc()).first()
    if not sheet:
        return None
    raw_table = db.query(RawTable).filter(RawTable.raw_sheet_id == sheet.id).first()
    if not raw_table:
        return None
    table_data = raw_table.table_json
    if isinstance(table_data, str):
        table_data = json.loads(table_data)
    return (_normalize_table_to_dense(table_data), [])


def _get_raw_table_with_meta(db: Session, sheet_name: str, filename_pattern: str = None) -> Optional[tuple]:
    """获取raw_table数据及merged_cells_json，返回 (table_data, merged_cells_json)"""
    query = db.query(RawSheet).join(RawFile).filter(RawSheet.sheet_name == sheet_name)
    if filename_pattern:
        query = query.filter(RawFile.filename.like(f'%{filename_pattern}%'))
    sheet = query.order_by(RawFile.id.desc()).first()
    if not sheet:
        return None
    raw_table = db.query(RawTable).filter(
        RawTable.raw_sheet_id == sheet.id
    ).first()
    if not raw_table:
        return None
    table_data = raw_table.table_json
    if isinstance(table_data, str):
        table_data = json.loads(table_data)
    merged = raw_table.merged_cells_json
    if isinstance(merged, str):
        merged = json.loads(merged) if merged else []
    return (table_data, merged or [])


def _parse_excel_date(excel_date: any) -> Optional[date]:
    """解析Excel日期"""
    if isinstance(excel_date, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=int(excel_date))).date()
        except:
            pass
    elif isinstance(excel_date, str):
        try:
            if 'T' in excel_date:
                return datetime.fromisoformat(excel_date.replace('Z', '+00:00')).date()
            return datetime.strptime(excel_date, '%Y-%m-%d').date()
        except:
            pass
    elif isinstance(excel_date, date):
        return excel_date
    elif isinstance(excel_date, datetime):
        return excel_date.date()
    return None


def _parse_quarter(period_str: str) -> Optional[str]:
    """解析季度字符串，返回YYYYQ格式"""
    if not period_str:
        return None
    
    period_str = str(period_str).strip()
    
    # 尝试解析"2024Q1"格式
    if 'Q' in period_str.upper():
        parts = period_str.upper().split('Q')
        if len(parts) == 2:
            try:
                year = int(parts[0])
                quarter = int(parts[1])
                if 1 <= quarter <= 4:
                    return f"{year}Q{quarter}"
            except:
                pass
    
    # 尝试解析"2024年1季度"格式
    if '年' in period_str and '季度' in period_str:
        parts = period_str.split('年')
        if len(parts) == 2:
            try:
                year = int(parts[0])
                quarter_str = parts[1].replace('季度', '').strip()
                quarter = int(quarter_str)
                if 1 <= quarter <= 4:
                    return f"{year}Q{quarter}"
            except:
                pass
    
    return None


def _format_statistics_cell(val) -> str:
    """格式化统计局表格单元格"""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return ""
        return str(int(val)) if val == int(val) else f"{val:.2f}"
    if hasattr(val, "isoformat"):
        return val.isoformat()[:10]
    return str(val).strip()


@router.get("/quarterly-data", response_model=QuarterlyDataRawResponse)
def get_quarterly_data(
    db: Session = Depends(get_db)
):
    """
    获取统计局季度数据汇总（表1）
    按 Excel 原样：2 行表头、合并单元格、数据行；B-Y 列。
    数据来源：03.统计局季度数据 sheet
    """
    result = _get_raw_table_with_meta(db, "03.统计局季度数据")
    if not result or len(result[0]) < 2:
        return QuarterlyDataRawResponse(
            header_row_0=[],
            header_row_1=[],
            rows=[],
            column_count=0,
            merged_cells_json=[],
        )
    table_data, merged_cells = result
    max_col = min(
        max(len(table_data[0]), len(table_data[1]), *(len(r) for r in table_data[2:]), 1),
        25,  # B-Y 共 24 列 + B 列季度，限 25 列
    )

    def pad_row(row, length):
        r = list(row) if row else []
        return [r[i] if i < len(r) else "" for i in range(length)]

    row0 = pad_row(table_data[0], max_col)
    row1 = pad_row(table_data[1], max_col)
    header_row_0 = [_format_statistics_cell(v) or "" for v in row0]
    header_row_1 = [_format_statistics_cell(v) or "" for v in row1]

    rows = []
    for row_idx in range(2, len(table_data)):
        row = pad_row(table_data[row_idx], max_col)
        if len(row) < 2 or not row[1]:  # B 列为季度，不能为空
            continue
        parsed_date = _parse_excel_date(row[1])
        if parsed_date:
            year = parsed_date.year
            quarter = (parsed_date.month - 1) // 3 + 1
            row[1] = f"{year}Q{quarter}"
        else:
            q = _parse_quarter(str(row[1]))
            if q:
                row[1] = q
        rows.append([_format_statistics_cell(v) for v in row])

    return QuarterlyDataRawResponse(
        header_row_0=header_row_0,
        header_row_1=header_row_1,
        rows=rows,
        column_count=max_col,
        merged_cells_json=merged_cells,
    )


@router.get("/output-slaughter", response_model=OutputSlaughterResponse)
def get_output_slaughter(
    db: Session = Depends(get_db)
):
    """
    获取统计局生猪出栏量&屠宰量（图1）
    数据来源：03.统计局季度数据 sheet
    - J列（索引9）：季度出栏量
    - P列（索引15）：定点屠宰量
    - 规模化率 = 屠宰量 / 出栏量
    
    数据结构：
    - 第1行（索引0）：主表头
    - 第2行（索引1）：子表头
    - 第3行（索引2）开始：数据行（B列是季度日期，J列是出栏量，P列是屠宰量）
    """
    table_data = _get_raw_table_data(db, "03.统计局季度数据")
    
    if not table_data or len(table_data) < 3:
        return OutputSlaughterResponse(data=[], latest_period=None)
    
    data_points = []
    latest_period = None
    
    # 从第3行开始解析数据（索引2）
    for row_idx, row in enumerate(table_data[2:], start=3):
        if len(row) < 16:
            continue
        
        # B列（索引1）是季度日期
        period_str = row[1] if len(row) > 1 and row[1] else None
        if not period_str:
            continue
        
        # 解析季度日期
        period = None
        parsed_date = _parse_excel_date(period_str)
        if parsed_date:
            # 转换为季度格式：YYYYQ格式
            year = parsed_date.year
            quarter = (parsed_date.month - 1) // 3 + 1
            period = f"{year}Q{quarter}"
        else:
            # 尝试直接解析季度字符串
            period = _parse_quarter(str(period_str))
            if not period:
                period = str(period_str)
        
        # J列（索引9）：季度出栏量
        output_val = row[9] if len(row) > 9 else None
        output_volume = None
        if output_val is not None and output_val != "":
            try:
                output_volume = float(output_val)
                if math.isnan(output_volume) or math.isinf(output_volume):
                    output_volume = None
            except (ValueError, TypeError):
                pass
        
        # P列（索引15）：定点屠宰量
        slaughter_val = row[15] if len(row) > 15 else None
        slaughter_volume = None
        if slaughter_val is not None and slaughter_val != "":
            try:
                slaughter_volume = float(slaughter_val)
                if math.isnan(slaughter_volume) or math.isinf(slaughter_volume):
                    slaughter_volume = None
            except (ValueError, TypeError):
                pass
        
        # 计算规模化率
        scale_rate = None
        if output_volume is not None and slaughter_volume is not None and output_volume > 0:
            scale_rate = slaughter_volume / output_volume
        
        data_points.append(OutputSlaughterPoint(
            period=period,
            output_volume=output_volume,
            slaughter_volume=slaughter_volume,
            scale_rate=scale_rate
        ))
        
        if period and (not latest_period or period > latest_period):
            latest_period = period
    
    # 按季度排序
    data_points.sort(key=lambda x: x.period)
    
    return OutputSlaughterResponse(data=data_points, latest_period=latest_period)


def _parse_pork_import_from_sheet(table_data: List[List]) -> tuple:
    """
    解析猪肉进口数据。支持格式：
    1. 猪肉进口 sheet（含钢联自动更新模板）：
       - 钢联格式：第1行大标题，第2行表头（月份、总量、国别列），第3行单位，第4行更新时间，第5行起数据
       - 通用格式：第1行即表头
    2. 进口肉 sheet：矩阵格式，行=月度(1-2月,3月...)，列=年份(2016年...)
    返回 (data_points, latest_month)
    """
    if not table_data or len(table_data) < 2:
        return [], None

    exclude_keywords = ('日期', '月份', '月', '总量', '合计', '环比', '同比', '比例', '占比', '累计', '猪肉', '猪杂碎', '猪总量')

    # 尝试格式1：猪肉进口（月份+总量+国别列），支持表头在 row 0/1/2
    # 钢联格式：第2行表头含「进口数量合计」「进口数量：XX→中国」；第1列=日期
    for header_row_idx in range(min(3, len(table_data))):
        header = table_data[header_row_idx]
        header_strs = [str(h).strip() if h is not None else '' for h in header]
        country_cols = []
        month_col = -1
        total_col = -1
        for i, h in enumerate(header_strs):
            if not h:
                continue
            # 月份列：指标名称/日期/月份，数据行首列为日期
            if any(kw in h for kw in ('日期', '月份', '指标名称')) or (h.endswith('月') and len(h) <= 4):
                if month_col < 0:
                    month_col = i
            elif '总量' in h or '合计' in h:
                total_col = i
            elif '→中国' in h or '->中国' in h:
                # 钢联格式：猪肉及猪杂碎：进口数量：阿根廷→中国（月）→ 提取国名
                m = re.search(r'[：:]\s*([^：:→]+)\s*[→>\-]', h)
                if m:
                    cname = m.group(1).strip()
                    if len(cname) >= 2 and cname not in ('全球', '进口数量'):
                        country_cols.append((i, cname))
            elif len(h) >= 2 and len(h) <= 8 and not any(kw in h for kw in exclude_keywords):
                country_cols.append((i, h))

        if month_col >= 0 and (total_col >= 0 or country_cols):
            # 数据从表头下一行开始；钢联有单位行、更新行时需跳过（非数字行）
            data_start = header_row_idx + 1
            data_points = []
            latest_month = None
            for row in table_data[data_start:]:
                if len(row) <= month_col:
                    continue
                month_val = row[month_col]
                if month_val is None or str(month_val).strip() == '':
                    continue
                month_str = _parse_month_to_yyyy_mm(month_val)
                if not month_str:
                    continue
                total_val = None
                country_vals = []
                if total_col >= 0 and len(row) > total_col:
                    try:
                        total_val = float(row[total_col]) if row[total_col] not in (None, '') else None
                    except (ValueError, TypeError):
                        pass
                for ci, cname in country_cols:
                    if len(row) > ci and row[ci] not in (None, ''):
                        try:
                            v = float(row[ci])
                            if v > 0:
                                country_vals.append((cname, v))
                        except (ValueError, TypeError):
                            pass
                if total_val is None and country_vals:
                    total_val = sum(v for _, v in country_vals)
                if total_val is None and not country_vals:
                    continue
                top2 = sorted(country_vals, key=lambda x: -x[1])[:2]
                data_points.append(ImportMeatPoint(
                    month=month_str,
                    total=round(total_val, 2) if total_val else None,
                    top_countries=[ImportMeatCountryPoint(country=c, value=round(v, 2)) for c, v in top2]
                ))
                if month_str and (not latest_month or month_str > latest_month):
                    latest_month = month_str
            if data_points:
                data_points.sort(key=lambda x: x.month)
                return data_points, latest_month

    # 尝试格式2：进口肉（矩阵：行=月度，列=年份）
    def _is_year_cell(s: str) -> bool:
        if not s:
            return False
        if '年' in s and len(s) <= 6:
            try:
                y = int(s.replace('年', ''))
                return 2010 <= y <= 2030
            except ValueError:
                pass
        if s.isdigit() and len(s) == 4:
            try:
                return 2010 <= int(s) <= 2030
            except ValueError:
                pass
        if s.replace('.', '').isdigit() and len(s) <= 6:  # 2016.0 from Excel
            try:
                y = int(float(s))
                return 2010 <= y <= 2030
            except ValueError:
                pass
        return False

    year_row_idx = -1
    for ri, row in enumerate(table_data[:5]):
        row_str = [str(c).strip() if c else '' for c in row]
        if any(_is_year_cell(s) for s in row_str):
            year_row_idx = ri
            break
    if year_row_idx >= 0:
        year_row = table_data[year_row_idx]
        year_cols = []
        for i, c in enumerate(year_row):
            s = str(c).strip() if c else ''
            if '年' in s and len(s) <= 6:
                try:
                    y = int(s.replace('年', ''))
                    if 2010 <= y <= 2030:
                        year_cols.append((i, y))
                except ValueError:
                    pass
            elif s.isdigit() and len(s) == 4:
                try:
                    y = int(s)
                    if 2010 <= y <= 2030:
                        year_cols.append((i, y))
                except ValueError:
                    pass
            elif s.replace('.', '').isdigit():
                try:
                    y = int(float(s))
                    if 2010 <= y <= 2030 and (i, y) not in [(ci, yy) for ci, yy in year_cols]:
                        year_cols.append((i, y))
                except ValueError:
                    pass
        month_map = {'1-2月': 1, '1月': 1, '2月': 2, '3月': 3, '4月': 4, '5月': 5, '6月': 6,
                     '7月': 7, '8月': 8, '9月': 9, '10月': 10, '11月': 11, '12月': 12}
        data_points = []
        latest_month = None
        for row in table_data[year_row_idx + 1:]:
            if len(row) == 0:
                continue
            month_key = str(row[0]).strip() if row else ''
            month_num = month_map.get(month_key)
            if month_num is None:
                for k, v in month_map.items():
                    if k in month_key:
                        month_num = v
                        break
            if month_num is None:
                continue
            for col_i, year in year_cols:
                if len(row) <= col_i:
                    continue
                try:
                    val = float(row[col_i]) if row[col_i] not in (None, '') else None
                except (ValueError, TypeError):
                    val = None
                if val is None or val <= 0:
                    continue
                month_str = f"{year}-{month_num:02d}"
                data_points.append(ImportMeatPoint(
                    month=month_str,
                    total=round(val, 2),
                    top_countries=[]
                ))
                if not latest_month or month_str > latest_month:
                    latest_month = month_str
        if data_points:
            data_points.sort(key=lambda x: x.month)
            return data_points, latest_month

    return [], None


def _parse_pork_import_from_quarterly_stats(db: Session) -> tuple:
    """
    从 03.统计局季度数据 Z列（猪肉进口）提取季度数据。
    转为月度格式（用季度首月代表，如 2024Q1 -> 2024-01）供图2显示。
    返回 (data_points, latest_month)
    """
    table_data = _get_raw_table_data(db, "03.统计局季度数据")
    if not table_data or len(table_data) < 3:
        return [], None
    # Z列索引：A=0,B=1,...,Z=25
    PORK_IMPORT_COL_IDX = 25
    data_points = []
    latest_month = None
    for row in table_data[2:]:
        if len(row) <= PORK_IMPORT_COL_IDX:
            continue
        period_str = row[1] if len(row) > 1 and row[1] else None
        if not period_str:
            continue
        year, quarter = None, None
        parsed = _parse_excel_date(period_str)
        if parsed:
            year = parsed.year
            quarter = (parsed.month - 1) // 3 + 1
        else:
            q = _parse_quarter(str(period_str))
            if q:
                parts = q.upper().split("Q")
                if len(parts) == 2:
                    try:
                        year = int(parts[0])
                        quarter = int(parts[1])
                    except ValueError:
                        continue
        if year is None or quarter is None:
            continue
        month_num = (quarter - 1) * 3 + 1
        month_str = f"{year}-{month_num:02d}"
        try:
            val = row[PORK_IMPORT_COL_IDX]
            total = float(val) if val not in (None, '') else None
        except (ValueError, TypeError):
            total = None
        if total is None or (isinstance(total, float) and (math.isnan(total) or math.isinf(total) or total <= 0)):
            continue
        data_points.append(ImportMeatPoint(
            month=month_str,
            total=round(total, 2),
            top_countries=[]
        ))
        if not latest_month or month_str > latest_month:
            latest_month = month_str
    if data_points:
        data_points.sort(key=lambda x: x.month)
    return data_points, latest_month


# 本地 Excel 回退：当数据库无数据时，从 docs 下的钢联模板读取
_PORK_IMPORT_FALLBACK_PATHS = [
    Path(__file__).resolve().parents[2] / "docs" / "1、价格：钢联自动更新模板(3).xlsx",
    Path(__file__).resolve().parents[2] / "docs" / "1、价格：钢联自动更新模板.xlsx",
]


def _load_pork_import_from_local_excel() -> Optional[List[List]]:
    """当数据库无猪肉进口数据时，从本地钢联 Excel 读取"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    workspace = Path(__file__).resolve().parents[2]
    docs_dir = workspace / "docs"
    if not docs_dir.exists():
        return None
    # 优先使用指定文件，其次匹配 钢联自动更新模板*.xlsx
    for p in _PORK_IMPORT_FALLBACK_PATHS:
        if p.exists():
            try:
                wb = load_workbook(p, data_only=True)
                if "猪肉进口" not in wb.sheetnames:
                    continue
                ws = wb["猪肉进口"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    return [list(r) for r in rows]
            except Exception:
                continue
    for f in docs_dir.glob("*钢联自动更新模板*.xlsx"):
        try:
            wb = load_workbook(f, data_only=True)
            if "猪肉进口" in wb.sheetnames:
                ws = wb["猪肉进口"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    return [list(r) for r in rows]
        except Exception:
            continue
    return None


def _parse_month_to_yyyy_mm(val) -> Optional[str]:
    """将月份值解析为 YYYY-MM"""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if len(s) == 7 and s[4] == '-':  # 已是 YYYY-MM
        return s
    try:
        from datetime import datetime
        dt = datetime.strptime(s, '%Y-%m-%d')
        return dt.strftime('%Y-%m')
    except Exception:
        pass
    m = re.search(r'(\d{4})年?\s*(\d{1,2})月?', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.search(r'(\d{4})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None


@router.get("/import-meat", response_model=ImportMeatResponse)
def get_import_meat(
    db: Session = Depends(get_db)
):
    """
    获取猪肉进口数据（图2）
    数据来源：钢联自动更新模板 sheet「猪肉进口」或涌益「进口肉」
    优先从钢联自动更新模板读取猪肉进口 sheet
    返回：每月进口总量 + 进口量前2的国家（如有分国别数据）
    """
    # 1. 优先从钢联自动更新模板读取猪肉进口（精确匹配）
    table_data = _get_raw_table_data(db, "猪肉进口", filename_pattern="钢联自动更新模板")
    if table_data:
        data_points, latest_month = _parse_pork_import_from_sheet(table_data)
        if data_points:
            return ImportMeatResponse(data=data_points, latest_month=latest_month)
    # 2. 精确匹配：猪肉进口、进口肉（不限文件，作为后备）
    for sheet_name in ("猪肉进口", "进口肉"):
        table_data = _get_raw_table_data(db, sheet_name)
        if table_data:
            data_points, latest_month = _parse_pork_import_from_sheet(table_data)
            if data_points:
                return ImportMeatResponse(data=data_points, latest_month=latest_month)
    # 2b. 模糊匹配：兼容带前缀/后缀的 sheet 名（如 04.猪肉进口、猪肉进口汇总）
    for pattern in ("%猪肉%进口%", "%进口肉%"):
        result = _get_raw_table_data_by_like(db, pattern)
        if result:
            table_data = result[0]
            data_points, latest_month = _parse_pork_import_from_sheet(table_data)
            if data_points:
                return ImportMeatResponse(data=data_points, latest_month=latest_month)
    # 3. 数据库无数据时，从本地钢联 Excel 回退（docs/1、价格：钢联自动更新模板(3).xlsx 等）
    table_data = _load_pork_import_from_local_excel()
    if table_data:
        data_points, latest_month = _parse_pork_import_from_sheet(table_data)
        if data_points:
            return ImportMeatResponse(data=data_points, latest_month=latest_month)
    return ImportMeatResponse(data=[], latest_month=None)
