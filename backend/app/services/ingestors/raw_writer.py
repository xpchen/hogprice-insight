"""RawWriter服务 - 保存原始文件、sheet和表格数据到Raw层"""
from typing import Optional, Dict, Any, List
from datetime import date
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
import hashlib

from app.models.raw_file import RawFile
from app.models.raw_sheet import RawSheet
from app.models.raw_table import RawTable
from app.utils.table_serializer import serialize_worksheet_to_json, extract_header_signature


def calculate_file_hash(file_content: bytes) -> str:
    """计算文件SHA256哈希值"""
    return hashlib.sha256(file_content).hexdigest()


def save_raw_file(
    db: Session,
    batch_id: int,
    filename: str,
    file_content: bytes,
    report_date: Optional[date] = None,
    date_range_start: Optional[date] = None,
    date_range_end: Optional[date] = None,
    parser_version: Optional[str] = None,
    storage_path: Optional[str] = None
) -> RawFile:
    """
    保存文件元信息到raw_file
    
    Args:
        db: 数据库会话
        batch_id: 导入批次ID
        filename: 文件名
        file_content: 文件内容（bytes）
        report_date: 报告日期
        date_range_start: 数据日期范围开始
        date_range_end: 数据日期范围结束
        parser_version: 解析器版本
        storage_path: 存储路径
    
    Returns:
        RawFile对象
    """
    file_hash = calculate_file_hash(file_content)
    
    raw_file = RawFile(
        batch_id=batch_id,
        filename=filename,
        file_hash=file_hash,
        report_date=report_date,
        date_range_start=date_range_start,
        date_range_end=date_range_end,
        parser_version=parser_version,
        storage_path=storage_path
    )
    
    db.add(raw_file)
    db.flush()
    
    return raw_file


def save_raw_sheet(
    db: Session,
    raw_file_id: int,
    sheet_name: str,
    row_count: Optional[int] = None,
    col_count: Optional[int] = None,
    header_signature: Optional[str] = None,
    parse_status: str = "pending"
) -> RawSheet:
    """
    保存sheet元信息到raw_sheet
    
    Args:
        db: 数据库会话
        raw_file_id: raw_file ID
        sheet_name: sheet名称
        row_count: 行数
        col_count: 列数
        header_signature: 表头签名
        parse_status: 解析状态（pending/parsed/failed/skipped）
    
    Returns:
        RawSheet对象
    """
    raw_sheet = RawSheet(
        raw_file_id=raw_file_id,
        sheet_name=sheet_name,
        row_count=row_count,
        col_count=col_count,
        header_signature=header_signature,
        parse_status=parse_status
    )
    
    db.add(raw_sheet)
    db.flush()
    
    return raw_sheet


def save_raw_table(
    db: Session,
    raw_sheet_id: int,
    worksheet: Worksheet,
    sparse: bool = False,
    max_rows: Optional[int] = None,
    skip_if_too_large: bool = True,
    max_size_mb: float = 10.0
) -> Optional[RawTable]:
    """
    将Excel sheet转换为JSON格式保存到raw_table
    
    Args:
        db: 数据库会话
        raw_sheet_id: raw_sheet ID
        worksheet: openpyxl Worksheet对象
        sparse: 是否使用稀疏格式
        max_rows: 最大行数（None表示全部）
        skip_if_too_large: 如果预估大小超过阈值，是否跳过存储
        max_size_mb: 最大允许大小（MB），超过则跳过存储
    
    Returns:
        RawTable对象，如果跳过则返回None
    """
    # 检查sheet大小
    row_count = worksheet.max_row
    col_count = worksheet.max_column
    
    # 对于超大sheet，使用更激进的限制
    if max_rows is None:
        # 根据sheet大小动态决定存储行数
        if row_count > 1000 or col_count > 1000:
            # 超大sheet：只存前10行
            max_rows = 10
        elif row_count > 500 or col_count > 500:
            # 大sheet：只存前20行
            max_rows = 20
        elif row_count > 100 or col_count > 100:
            # 中等sheet：只存前50行
            max_rows = 50
        else:
            # 小sheet：存储全部
            max_rows = None
    
    # 估算JSON大小（快速检查前100行）
    if skip_if_too_large:
        sample_rows = min(100, row_count)
        non_empty_count = 0
        for row in worksheet.iter_rows(min_row=1, max_row=sample_rows, values_only=False):
            for cell in row:
                if cell.value is not None:
                    non_empty_count += 1
        
        # 估算：稀疏格式每个非空单元格约50字节
        estimated_size_mb = (non_empty_count * 50 * (row_count / sample_rows)) / 1024 / 1024
        
        if estimated_size_mb > max_size_mb:
            # 超过阈值，跳过存储（只存储元信息）
            return None
    
    # 序列化worksheet（限制行数以优化存储）
    serialized = serialize_worksheet_to_json(worksheet, sparse=sparse, max_rows=max_rows)
    
    raw_table = RawTable(
        raw_sheet_id=raw_sheet_id,
        table_json=serialized["table_json"],
        merged_cells_json=serialized["merged_cells_json"]
    )
    
    db.add(raw_table)
    db.flush()
    
    return raw_table


def save_all_sheets_from_workbook(
    db: Session,
    raw_file_id: int,
    workbook_or_path: Any,
    sparse: bool = False,
    max_rows: Optional[int] = None
) -> List[RawSheet]:
    """
    从workbook保存所有sheet到raw_sheet和raw_table
    
    Args:
        db: 数据库会话
        raw_file_id: raw_file ID
        workbook_or_path: openpyxl Workbook对象或文件路径/BytesIO
        sparse: 是否使用稀疏格式
    
    Returns:
        RawSheet对象列表
    """
    # 加载workbook
    if isinstance(workbook_or_path, BytesIO):
        wb = load_workbook(workbook_or_path, data_only=True)
    elif isinstance(workbook_or_path, str):
        wb = load_workbook(workbook_or_path, data_only=True)
    else:
        wb = workbook_or_path
    
    raw_sheets = []
    
    for sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        
        # 提取表头签名
        header_sig = extract_header_signature(worksheet)
        
        # 保存sheet元信息
        raw_sheet = save_raw_sheet(
            db=db,
            raw_file_id=raw_file_id,
            sheet_name=sheet_name,
            row_count=worksheet.max_row,
            col_count=worksheet.max_column,
            header_signature=header_sig,
            parse_status="pending"
        )
        
        # 保存表格数据（限制行数以优化存储，超大sheet可能跳过）
        raw_table = save_raw_table(
            db=db,
            raw_sheet_id=raw_sheet.id,
            worksheet=worksheet,
            sparse=sparse,
            max_rows=max_rows,
            skip_if_too_large=True,  # 超大sheet跳过存储
            max_size_mb=10.0  # 超过10MB跳过
        )
        # 如果跳过存储，raw_table为None，但raw_sheet元信息已保存
        
        raw_sheets.append(raw_sheet)
    
    return raw_sheets


def update_sheet_parse_status(
    db: Session,
    raw_sheet_id: int,
    parse_status: str,
    parser_type: Optional[str] = None,
    error_count: Optional[int] = None,
    observation_count: Optional[int] = None
) -> RawSheet:
    """
    更新sheet解析状态
    
    Args:
        db: 数据库会话
        raw_sheet_id: raw_sheet ID
        parse_status: 解析状态
        parser_type: 解析器类型
        error_count: 错误数量
        observation_count: observation数量
    
    Returns:
        更新后的RawSheet对象
    """
    raw_sheet = db.query(RawSheet).filter(RawSheet.id == raw_sheet_id).first()
    if not raw_sheet:
        raise ValueError(f"RawSheet not found: {raw_sheet_id}")
    
    raw_sheet.parse_status = parse_status
    if parser_type:
        raw_sheet.parser_type = parser_type
    if error_count is not None:
        raw_sheet.error_count = error_count
    if observation_count is not None:
        raw_sheet.observation_count = observation_count
    
    db.flush()
    
    return raw_sheet
