"""P3解析器：WIDE_DATE_GROUPED_SUBCOLS - 日期跨列 + 子列（规模场/小散/均价/涨跌）"""
from typing import List, Dict, Any, Optional
from datetime import date
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
import sys

from .base_parser import BaseParser, ObservationDict
from app.utils.dt_parse import parse_date, parse_period_start_end
from app.utils.value_cleaner import clean_numeric_value_enhanced
from app.utils.header_flattener import extract_date_grouped_subcols
from app.utils.merged_cell_handler import get_merged_cell_value, clear_merged_cell_cache


class P3WideDateGroupedSubcolsParser(BaseParser):
    """P3解析器：处理日期跨列 + 子列的宽表格式"""
    
    def parse(
        self,
        sheet_data: Any,
        sheet_config: Dict[str, Any],
        profile_defaults: Dict[str, Any],
        source_code: str,
        batch_id: int
    ) -> List[ObservationDict]:
        """
        解析日期跨列 + 子列的宽表格式
        
        示例：出栏价、散户标肥价差、市场主流标猪肥猪价格
        
        性能优化：批量读取数据而不是逐单元格读取
        """
        observations = []
        
        # 需要Worksheet对象来处理合并单元格
        if isinstance(sheet_data, Worksheet):
            worksheet = sheet_data
        else:
            return observations  # 需要Worksheet对象
        
        # 获取配置
        header_config = sheet_config.get("header", {})
        # 兼容两种配置格式
        header_row = header_config.get("header_row") or sheet_config.get("header_row", 1)
        date_row_idx = header_config.get("date_row", header_row)
        subheader_row_idx = header_config.get("subheader_row", header_row)
        data_start_row = header_config.get("data_start_row", header_row + 1)
        
        # 检查是否是周期起止格式（有start_col和end_col配置）
        start_col = sheet_config.get("start_col") or sheet_config.get("start_date_col")
        end_col = sheet_config.get("end_col") or sheet_config.get("end_date_col")
        row_dim_col = sheet_config.get("row_dim_col")
        group_size = sheet_config.get("group_size", 2)
        subheaders = sheet_config.get("subheaders", [])
        metric_templates = sheet_config.get("metric_templates", [])
        
        # 如果没有metric_templates，使用metric_template
        if not metric_templates and sheet_config.get("metric_template"):
            metric_templates = [sheet_config.get("metric_template")]
        
        row_dims = sheet_config.get("row_dims", [])
        sheet_name = sheet_config.get("sheet_name", "")
        
        # 如果是周期起止格式，使用不同的解析逻辑
        if start_col and end_col:
            print(f"     └─ 检测到周期起止格式 (start_col={start_col}, end_col={end_col})", flush=True)
            return self._parse_period_start_end_grouped_subcols(
                worksheet, sheet_config, profile_defaults, source_code, batch_id,
                header_row, start_col, end_col, row_dim_col, group_size, subheaders, metric_templates
            )
        
        # 原有的日期跨列格式
        # 提取日期分组列结构
        print(f"     └─ 提取日期分组列结构...", flush=True)
        date_groups = extract_date_grouped_subcols(
            worksheet,
            header_rows=2,
            date_row_idx=date_row_idx,
            subcol_row_idx=subheader_row_idx
        )
        
        print(f"     └─ 找到 {len(date_groups)} 个日期组", flush=True)
        
        if not date_groups:
            return observations
        
        # 性能优化：批量读取所有数据到二维数组
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        print(f"     └─ [步骤1] 批量读取数据 ({data_start_row} 到 {max_row}, 共 {max_row - data_start_row + 1} 行, {max_col} 列)...", flush=True)
        
        # 使用iter_rows批量读取，比逐单元格读取快100倍以上
        data_matrix = []
        row_count = 0
        for row in worksheet.iter_rows(min_row=data_start_row, max_row=max_row, 
                                       min_col=1, max_col=max_col, values_only=True):
            data_matrix.append(list(row))
            row_count += 1
            if row_count % 500 == 0:
                print(f"     └─ [步骤1] 已读取 {row_count} 行...", flush=True)
        
        print(f"     └─ [步骤1] 数据读取完成，共 {len(data_matrix)} 行", flush=True)
        
        # 处理合并单元格：预先构建行维度列的合并单元格映射
        # 对于行维度列（如省份列），需要处理合并单元格
        print(f"     └─ [步骤2] 构建合并单元格映射（行维度列）...", flush=True)
        row_dim_merged_map = {}
        
        # 先批量读取所有行维度列的数据（避免在循环中调用worksheet.cell）
        dim_cols = [dim_config.get("col", 1) for dim_config in row_dims]
        if dim_cols:
            # 批量读取所有行维度列的数据
            dim_cols_data = {}
            for dim_col in dim_cols:
                col_data = []
                for row in worksheet.iter_rows(min_row=1, max_row=max_row,
                                               min_col=dim_col, max_col=dim_col, values_only=True):
                    col_data.append(row[0] if row else None)
                dim_cols_data[dim_col] = col_data
            
            # 从批量数据构建合并单元格映射
            merged_ranges_count = 0
            for merged_range in worksheet.merged_cells.ranges:
                merged_ranges_count += 1
                if merged_ranges_count % 100 == 0:
                    print(f"     └─ [步骤2] 已处理 {merged_ranges_count} 个合并单元格范围...", flush=True)
                
                for dim_col in dim_cols:
                    if (merged_range.min_col <= dim_col <= merged_range.max_col):
                        # 从批量数据中获取主单元格值
                        master_row_idx = merged_range.min_row - 1  # 转换为0-based索引
                        if master_row_idx < len(dim_cols_data[dim_col]):
                            master_value = dim_cols_data[dim_col][master_row_idx]
                            # 为合并区域内的所有行创建映射
                            for r in range(merged_range.min_row, merged_range.max_row + 1):
                                if r >= data_start_row:
                                    row_idx_in_matrix = r - data_start_row
                                    if row_idx_in_matrix >= 0:
                                        row_dim_merged_map[(row_idx_in_matrix, dim_col)] = master_value
        
        print(f"     └─ [步骤2] 合并单元格映射完成，共 {len(row_dim_merged_map)} 个映射", flush=True)
        
        total_rows = len(data_matrix)
        print(f"     └─ [步骤3] 开始处理 {total_rows} 行数据，{len(date_groups)} 个日期组...", flush=True)
        
        obs_count = 0
        for row_idx_in_matrix, row_data in enumerate(data_matrix):
            row_idx = data_start_row + row_idx_in_matrix
            
            # 每处理500行输出一次进度
            if (row_idx_in_matrix + 1) % 500 == 0 or (row_idx_in_matrix + 1) == total_rows:
                progress = (row_idx_in_matrix + 1) * 100 // total_rows
                print(f"     └─ [步骤3] 处理进度: {row_idx_in_matrix + 1}/{total_rows} 行 ({progress}%), 已生成 {obs_count} 条观测值...", flush=True)
            
            # 提取行维度（省份、区域等）
            row_dim_values = {}
            for dim_config in row_dims:
                dim_col = dim_config.get("col", 1)
                dim_key = dim_config.get("key", "")
                dim_type = dim_config.get("type", "TEXT")
                
                # 检查合并单元格映射
                if (row_idx_in_matrix, dim_col) in row_dim_merged_map:
                    cell_value = row_dim_merged_map[(row_idx_in_matrix, dim_col)]
                else:
                    # 直接从矩阵读取（列索引从0开始，所以dim_col-1）
                    if dim_col - 1 < len(row_data):
                        cell_value = row_data[dim_col - 1]
                    else:
                        cell_value = None
                
                if cell_value is not None:
                    row_dim_values[dim_key] = str(cell_value).strip()
            
            # 处理每个日期组
            for date_group in date_groups:
                date_str = date_group.get("date", "")
                subcols = date_group.get("subcols", [])
                col_start = date_group.get("col_start", 1)
                col_end = date_group.get("col_end", max_col)
                
                # 解析日期
                obs_date = None
                try:
                    parsed_date = parse_date(date_str)
                    if parsed_date:
                        obs_date = parsed_date.date()
                        if obs_date.year < 2000 or obs_date.year > 2100:
                            continue
                except:
                    continue
                
                if obs_date is None:
                    continue
                
                # 处理每个子列
                for subcol_idx, subcol_name in enumerate(subcols):
                    col_idx = col_start + subcol_idx
                    if col_idx > col_end or col_idx > max_col:
                        break
                    
                    # 直接从矩阵读取（列索引从0开始）
                    if col_idx - 1 < len(row_data):
                        cell_value = row_data[col_idx - 1]
                    else:
                        continue
                    
                    numeric_value, raw_value = clean_numeric_value_enhanced(cell_value)
                    
                    if numeric_value is None and raw_value is None:
                        continue
                    
                    # 从子列名提取tags
                    from app.utils.dimension_extractor import extract_tags_from_text
                    subcol_tags = extract_tags_from_text(subcol_name)
                    
                    # 合并tags
                    template_tags = metric_template.get("tags", {})
                    tags = self._merge_tags(template_tags, row_dim_values, subcol_tags)
                    
                    # 生成metric_key（基于子列名）
                    metric_key = metric_template.get("metric_key", "")
                    if not metric_key:
                        metric_key = f"{sheet_name}_{subcol_name}".upper().replace(" ", "_")
                    
                    # 生成dedup_key
                    geo_code = row_dim_values.get("province") or "NATION"
                    dedup_key = self._generate_dedup_key(
                        source_code=source_code,
                        sheet_name=sheet_name,
                        metric_key=metric_key,
                        geo_key=geo_code,
                        obs_date=obs_date,
                        period_end=None,
                        tags=tags
                    )
                    
                    observation = {
                        "source_code": source_code,
                        "sheet_name": sheet_name,
                        "metric_key": metric_key,
                        "metric_name": metric_template.get("metric_name", subcol_name),
                        "geo_key": geo_code,
                        "obs_date": obs_date,
                        "period_type": profile_defaults.get("period_type", "day"),
                        "period_start": None,
                        "period_end": None,
                        "value": numeric_value,
                        "raw_value": raw_value,
                        "geo_code": geo_code,
                        "tags": tags,
                        "unit": metric_template.get("unit"),
                        "dedup_key": dedup_key,
                        "batch_id": batch_id,
                        "row_dim": {
                            "province": geo_code
                        },
                        "subheader": subcol_name  # 用于column_mapper提取
                    }
                    
                    observations.append(observation)
                    obs_count += 1
        
        print(f"     └─ [步骤3] 完成处理，共生成 {len(observations)} 条观测值", flush=True)
        
        # 清空缓存释放内存
        clear_merged_cell_cache()
        
        return observations
    
    def _parse_period_start_end_grouped_subcols(
        self,
        worksheet: Worksheet,
        sheet_config: Dict[str, Any],
        profile_defaults: Dict[str, Any],
        source_code: str,
        batch_id: int,
        header_row: int,
        start_col: str,
        end_col: str,
        row_dim_col: Optional[str],
        group_size: int,
        subheaders: List[str],
        metric_templates: List[Dict[str, Any]]
    ) -> List[ObservationDict]:
        """解析周期起止 + 省份分组子列格式"""
        observations = []
        
        import pandas as pd
        
        # 转换为DataFrame
        data = list(worksheet.values)
        df = pd.DataFrame(data)
        
        if df.empty:
            return observations
        
        # 设置表头
        header_row_idx = header_row - 1  # 转换为0-based
        if header_row_idx < len(df):
            df.columns = df.iloc[header_row_idx]
            df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        
        # 查找列索引
        start_col_idx = None
        end_col_idx = None
        row_dim_col_idx = None
        
        for idx, col in enumerate(df.columns):
            col_str = str(col).strip()
            if col_str == start_col:
                start_col_idx = idx
            elif col_str == end_col:
                end_col_idx = idx
            elif row_dim_col and col_str == row_dim_col:
                row_dim_col_idx = idx
        
        if start_col_idx is None or end_col_idx is None:
            print(f"     └─ ⚠️  未找到开始日期列({start_col})或结束日期列({end_col})", flush=True)
            return observations
        
        # 识别省份分组列（从row_dim_col列之后开始，每group_size列为一组）
        # 列结构：开始日期 | 结束日期 | 指标 | 河南 | 较上周 | 湖南 | 较上周 | ...
        province_start_idx = end_col_idx + 1
        if row_dim_col_idx is not None:
            # 如果有row_dim_col，跳过它
            province_start_idx = row_dim_col_idx + 1
        
        province_groups = []
        for idx in range(province_start_idx, len(df.columns), group_size):
            if idx + group_size - 1 < len(df.columns):
                province_name = str(df.columns[idx]).strip()
                # 排除"较上周"等非省份名
                if province_name and len(province_name) <= 10 and province_name not in ["较上周", "值", "较上期"]:
                    province_groups.append({
                        "province": province_name,
                        "start_col_idx": idx,
                        "end_col_idx": idx + group_size - 1
                    })
        
        print(f"     └─ 找到 {len(province_groups)} 个省份组", flush=True)
        
        if not province_groups:
            return observations
        
        # 处理每一行数据
        for row_idx, row in df.iterrows():
            # 解析周期日期
            start_val = row.iloc[start_col_idx] if start_col_idx < len(row) else None
            end_val = row.iloc[end_col_idx] if end_col_idx < len(row) else None
            
            period_start, period_end = parse_period_start_end(start_val, end_val)
            
            if period_end is None:
                continue
            
            # 提取行维度值
            row_dim_value = None
            if row_dim_col_idx is not None and row_dim_col_idx < len(row):
                row_dim_value = str(row.iloc[row_dim_col_idx]).strip() if pd.notna(row.iloc[row_dim_col_idx]) else None
            
            # 处理每个省份组
            for province_group in province_groups:
                province_name = province_group["province"]
                start_col_idx_group = province_group["start_col_idx"]
                
                # 处理每个子列
                for subcol_idx, subheader in enumerate(subheaders):
                    if subcol_idx >= group_size:
                        break
                    
                    col_idx = start_col_idx_group + subcol_idx
                    if col_idx >= len(row):
                        continue
                    
                    value_val = row.iloc[col_idx]
                    numeric_value, raw_value = clean_numeric_value_enhanced(value_val)
                    
                    if numeric_value is None and raw_value is None:
                        continue
                    
                    # 找到对应的metric_template
                    metric_template = None
                    for mt in metric_templates:
                        if mt.get("subheader_index") == subcol_idx:
                            metric_template = mt
                            break
                    
                    if not metric_template:
                        continue
                    
                    # 构建tags
                    tags = {
                        "province": province_name,
                        "indicator": row_dim_value
                    } if row_dim_value else {"province": province_name}
                    
                    # 生成dedup_key
                    dedup_key = self._generate_dedup_key(
                        source_code=source_code,
                        sheet_name=sheet_config.get("sheet_name", ""),
                        metric_key=metric_template.get("metric_key", ""),
                        geo_key=province_name,
                        obs_date=None,
                        period_end=period_end.date() if period_end else None,
                        tags=tags
                    )
                    
                    observation = {
                        "metric_key": metric_template.get("metric_key", ""),
                        "metric_name": metric_template.get("metric_name", ""),
                        "obs_date": period_end.date() if period_end else None,
                        "period_type": profile_defaults.get("period_type", "week"),
                        "period_start": period_start.date() if period_start else None,
                        "period_end": period_end.date() if period_end else None,
                        "value": numeric_value,
                        "raw_value": raw_value,
                        "geo_code": province_name,
                        "tags": tags,
                        "unit": metric_template.get("unit"),
                        "dedup_key": dedup_key
                    }
                    
                    observations.append(observation)
        
        print(f"     └─ 解析完成，共生成 {len(observations)} 条观测值", flush=True)
        return observations
