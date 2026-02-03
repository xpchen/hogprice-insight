"""Dispatcher - Sheet分派器，根据ingest_profile的dispatch_rules自动分派parser"""
from typing import Optional, Dict, Any, List
import re
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO

from app.models.ingest_profile import IngestProfile
from app.services.ingestors.profile_loader import get_profile_by_code, get_profile_by_dataset_type, match_sheet_by_dispatch_rules


class Dispatcher:
    """Sheet分派器"""
    
    def __init__(self, db: Session, profile: IngestProfile):
        """
        初始化分派器
        
        Args:
            db: 数据库会话
            profile: IngestProfile对象
        """
        self.db = db
        self.profile = profile
    
    def dispatch_sheet(
        self,
        sheet_name: str,
        worksheet: Optional[Worksheet] = None,
        sheet_columns: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        根据dispatch_rules分派sheet到对应的parser
        
        Args:
            sheet_name: Sheet名称
            worksheet: openpyxl Worksheet对象（可选，用于提取列名）
            sheet_columns: Sheet列名列表（可选，如果提供则直接使用）
        
        Returns:
            分派结果字典：
            {
                "action": "SKIP_META" | "RAW_TABLE_STORE_ONLY" | "PARSE",
                "parser": parser类型（如果action为PARSE）,
                "sheet_config": sheet配置（如果匹配到）
            }
        """
        # 如果没有提供列名，从worksheet提取
        if sheet_columns is None and worksheet is not None:
            sheet_columns = self._extract_column_names(worksheet)
        
        # 优先使用sheet名称直接匹配（最准确）
        sheet_config = self._find_sheet_config_by_name(sheet_name)
        if sheet_config:
            parser = sheet_config.get("parser")
            action = sheet_config.get("action")
            
            if action == "SKIP_META":
                return {
                    "action": "SKIP_META",
                    "parser": None,
                    "sheet_config": None
                }
            elif action == "RAW_TABLE_STORE_ONLY":
                return {
                    "action": "RAW_TABLE_STORE_ONLY",
                    "parser": None,
                    "sheet_config": None
                }
            elif parser and parser not in ['RAW_TABLE_STORE_ONLY', 'SKIP_META']:
                print(f"     └─ 通过sheet名称匹配到配置: parser={parser}", flush=True)
                return {
                    "action": "PARSE",
                    "parser": parser,
                    "sheet_config": sheet_config
                }
        
        # 如果没有通过sheet名称匹配到，使用dispatch_rules匹配
        matched_config = match_sheet_by_dispatch_rules(
            self.profile,
            sheet_name,
            sheet_columns or []
        )
        
        if matched_config:
            # 如果匹配到配置
            action = matched_config.get("action")
            parser = matched_config.get("parser")
            
            if action == "SKIP_META":
                return {
                    "action": "SKIP_META",
                    "parser": None,
                    "sheet_config": None
                }
            elif action == "RAW_TABLE_STORE_ONLY":
                return {
                    "action": "RAW_TABLE_STORE_ONLY",
                    "parser": None,
                    "sheet_config": None
                }
            elif parser:
                # 找到对应的sheet配置（优先使用sheet_name，如果找不到再使用parser类型）
                # 注意：优先使用sheet_name匹配，确保获取到正确的metric_template等配置
                sheet_config = self._find_sheet_config_by_name(sheet_name)
                if not sheet_config:
                    # 如果sheet_name匹配不到，再尝试用parser匹配（但可能配置不完整）
                    sheet_config = self._find_sheet_config_by_parser(parser)
                    if sheet_config:
                        print(f"     └─ ⚠️  通过parser类型匹配到配置（可能不完整）: parser={parser}", flush=True)
                        print(f"        建议：为sheet '{sheet_name}' 添加明确的配置", flush=True)
                
                if sheet_config:
                    print(f"     └─ 通过dispatch_rules匹配到parser: {parser}", flush=True)
                    # 调试：检查配置是否包含metric_template
                    if "metric_template" not in sheet_config:
                        print(f"     └─ ⚠️  sheet_config中没有metric_template！sheet_name={sheet_name}", flush=True)
                    return {
                        "action": "PARSE",
                        "parser": parser,
                        "sheet_config": sheet_config
                    }
        
        # 调试：打印为什么没有匹配到
        print(f"     └─ ⚠️  未找到匹配配置 (sheet_name={sheet_name}, profile.sheets数量={len(self.profile.sheets)})", flush=True)
        if sheet_columns:
            print(f"     └─ 提取的列名（前10个）: {sheet_columns[:10]}", flush=True)
        
        # 默认：只保存raw，不解析
        return {
            "action": "RAW_TABLE_STORE_ONLY",
            "parser": None,
            "sheet_config": None
        }
    
    def _extract_column_names(self, worksheet: Worksheet, max_cols: int = 50) -> List[str]:
        """
        从worksheet提取列名（尝试多个header行）
        
        Args:
            worksheet: openpyxl Worksheet对象
            max_cols: 最大列数
        
        Returns:
            列名列表
        """
        columns = []
        # 尝试读取前5行，找到包含"开始日期"或"结束日期"的行作为header行
        header_row = 1
        for row_idx in range(1, min(6, worksheet.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(worksheet.max_column + 1, max_cols + 1)):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    row_values.append(str(value).strip())
                else:
                    row_values.append("")
            
            # 检查这一行是否包含常见的日期列名
            if any(col in ["开始日期", "结束日期", "日期", "起始日期"] for col in row_values):
                header_row = row_idx
                columns = row_values
                break
        
        # 如果没有找到，使用第一行
        if not columns:
            for col_idx in range(1, min(worksheet.max_column + 1, max_cols + 1)):
                cell = worksheet.cell(row=1, column=col_idx)
                value = cell.value
                if value is not None:
                    columns.append(str(value).strip())
                else:
                    columns.append("")
        
        return columns
    
    def _find_sheet_config_by_parser(self, parser_type: str) -> Optional[Dict[str, Any]]:
        """根据parser类型查找sheet配置"""
        for sheet in self.profile.sheets:
            if sheet.parser == parser_type:
                return sheet.config_json
        return None
    
    def _find_sheet_config_by_name(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """根据sheet名称查找配置"""
        for sheet in self.profile.sheets:
            if sheet.sheet_name == sheet_name:
                return sheet.config_json
        return None


def create_dispatcher(db: Session, profile_code: Optional[str] = None, dataset_type: Optional[str] = None) -> Dispatcher:
    """
    创建Dispatcher实例
    
    Args:
        db: 数据库会话
        profile_code: Profile代码（如YONGYI_DAILY_V1）
        dataset_type: 数据集类型（如YONGYI_DAILY）
    
    Returns:
        Dispatcher实例
    """
    if profile_code:
        profile = get_profile_by_code(db, profile_code)
    elif dataset_type:
        profile = get_profile_by_dataset_type(db, dataset_type)
    else:
        raise ValueError("必须提供profile_code或dataset_type")
    
    if not profile:
        raise ValueError(f"未找到profile: profile_code={profile_code}, dataset_type={dataset_type}")
    
    return Dispatcher(db, profile)
