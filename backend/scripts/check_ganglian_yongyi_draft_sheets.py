"""
检查 2、【生猪产业数据】.xlsx 中的「钢联底稿」和「4.2涌益底稿」表结构，
确认是否存在：淘汰母猪屠宰、能繁母猪饲料环比等 E2 表格1 所需数据。
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import openpyxl

script_dir = Path(__file__).parent.parent
repo_root = script_dir.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = repo_root / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"

# E2 表格1 关心的关键词（用于在表头中高亮）
KEYWORDS = ["淘汰", "屠宰", "能繁", "母猪", "饲料", "环比", "存栏", "宰杀", "涌益", "钢联", "全国", "规模场", "中小散户"]


def cell_matches_keywords(cell):
    if cell is None:
        return False
    s = str(cell).strip()
    if not s:
        return False
    return any(kw in s for kw in KEYWORDS)


def analyze_sheet(wb, sheet_name: str):
    """分析单个 sheet 的表头与相关列"""
    if sheet_name not in wb.sheetnames:
        print(f"\n  [不存在] 未找到 sheet: {sheet_name}")
        return
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    # 前 5 行作为表头/数据示例
    rows = list(ws.iter_rows(max_row=8, values_only=True))
    for row_idx, row in enumerate(rows, 1):
        parts = []
        for col_idx, cell in enumerate(row[:35], 1):
            if cell is None or (isinstance(cell, str) and not cell.strip()):
                continue
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            tag = " [相关]" if cell_matches_keywords(cell) else ""
            parts.append(f"{col_letter}:{cell!r}{tag}")
        if parts:
            print(f"  行{row_idx}: {', '.join(parts[:20])}")
            if len(parts) > 20:
                print(f"        ... 共 {len(parts)} 列")
    # 汇总：含关键词的列
    header_rows = rows[:3] if len(rows) >= 3 else rows
    found = []
    for r_idx, row in enumerate(header_rows):
        for col_idx, cell in enumerate(row[:40], 1):
            if cell_matches_keywords(cell):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                found.append((r_idx + 1, col_letter, cell))
    if found:
        print(f"\n  含关键词的单元格（行, 列, 内容）:")
        for r, c, v in found[:30]:
            print(f"    行{r} {c}列: {v}")
    else:
        print("\n  未在表头前几行中发现「淘汰/屠宰/能繁/母猪/饲料/环比」等关键词。")
    print()


def main():
    if not EXCEL_PATH.exists():
        print(f"文件不存在: {EXCEL_PATH}")
        print("请将 2、【生猪产业数据】.xlsx 放在 docs/生猪/ 目录下后重试。")
        return
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("2、【生猪产业数据】.xlsx 中的全部 Sheet:")
    for i, name in enumerate(wb.sheetnames, 1):
        mark = "  <-- 将检查" if name in ("钢联底稿", "4.2涌益底稿") else ""
        print(f"  {i}. {name}{mark}")
    analyze_sheet(wb, "钢联底稿")
    analyze_sheet(wb, "4.2涌益底稿")

    # 专项：4.2涌益底稿 的列索引（A=0, B=1, ...）便于 API 取数
    if "4.2涌益底稿" in wb.sheetnames:
        ws = wb["4.2涌益底稿"]
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        print("\n4.2涌益底稿 列索引（供 API 用）:")
        for col_idx in range(min(25, len(row1))):
            c1, c2 = (row1[col_idx] if col_idx < len(row1) else None), (row2[col_idx] if col_idx < len(row2) else None)
            if c1 or c2:
                letter = openpyxl.utils.get_column_letter(col_idx + 1)
                print(f"  索引{col_idx} ({letter}列): 行1={c1!r}, 行2={c2!r}")

    # 是否存在「淘汰」「屠宰」列
    print("\n是否存在「淘汰」「屠宰」相关列:")
    for sh_name in ("钢联底稿", "4.2涌益底稿"):
        if sh_name not in wb.sheetnames:
            continue
        ws = wb[sh_name]
        found_cull = []
        for r in range(1, min(6, ws.max_row + 1)):
            for c in range(1, min(50, ws.max_column + 1)):
                cell = ws.cell(row=r, column=c).value
                if cell and ("淘汰" in str(cell) or "屠宰" in str(cell) or "宰杀" in str(cell)):
                    found_cull.append((r, openpyxl.utils.get_column_letter(c), cell))
        print(f"  {sh_name}: {'是 - ' + str(found_cull[:5]) if found_cull else '否'}")

    wb.close()
    print("\n完成。若上述 sheet 中存在「淘汰母猪屠宰」「能繁母猪饲料」等列，可据此调整 multi_source API 的取数列索引。")


if __name__ == "__main__":
    main()
