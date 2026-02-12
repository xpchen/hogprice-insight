# -*- coding: utf-8 -*-
"""Inspect A1供给预测 sheet header structure from Excel."""
import sys
from pathlib import Path
import openpyxl

script_dir = Path(__file__).resolve().parent
backend = script_dir.parent
repo = backend.parent
excel_path = repo / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"

if not excel_path.exists():
    print("File not found:", excel_path)
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)
if "A1供给预测" not in wb.sheetnames:
    print("Sheets:", wb.sheetnames[:25])
    wb.close()
    sys.exit(0)

ws = wb["A1供给预测"]
max_row, max_col = ws.max_row, ws.max_column
print("A1供给预测: max_row=%s max_col=%s" % (max_row, max_col))

# First 3 rows, all columns (sample every few cols if huge)
for ri in range(1, min(4, max_row + 1)):
    row = []
    for ci in range(1, min(max_col + 1, 100)):
        v = ws.cell(ri, ci).value
        row.append(v)
    print("Row %d (first 60 cols): %s" % (ri, row[:60]))

# Find columns containing key phrases (row 1 and 2)
from openpyxl.utils import get_column_letter
keywords = ["积压", "能繁环比", "新生仔猪环比", "定点屠宰", "中大猪", "存栏", "出栏", "体重", "价格", "成本", "指数"]
print("\nColumns with keywords (row 1 and 2):")
for ri in [1, 2]:
    for ci in range(1, min(max_col + 1, 120)):
        v = ws.cell(ri, ci).value
        if v is not None and str(v).strip():
            s = str(v)
            if any(k in s for k in keywords):
                print("  %s row%d: %s" % (get_column_letter(ci), ri, repr(s)[:70]))

wb.close()
