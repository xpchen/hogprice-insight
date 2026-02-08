"""
分析月度-生产指标sheet结构
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import pandas as pd
import openpyxl

script_dir = Path(__file__).parent.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_sheet():
    """分析月度-生产指标sheet"""
    print("=" * 80)
    print("分析月度-生产指标sheet结构")
    print("=" * 80)
    
    # 查找Excel文件
    docs_dir = Path(script_dir.parent) / "docs"
    excel_files = list(docs_dir.glob("*涌益*.xlsx"))
    
    if not excel_files:
        print("\n未找到涌益Excel文件")
        return
    
    excel_file = excel_files[0]
    print(f"\n找到文件: {excel_file.name}")
    
    try:
        # 使用openpyxl读取，保留格式
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # 查找"月度-生产指标"sheet
        sheet_names = wb.sheetnames
        target_sheet_name = None
        
        for name in sheet_names:
            if "生产指标" in name and "月度" in name:
                target_sheet_name = name
                break
        
        if not target_sheet_name:
            print(f"\n未找到'月度-生产指标'sheet")
            print(f"可用sheet列表:")
            for name in sheet_names:
                if "生产指标" in name or "月度" in name:
                    print(f"  - {name}")
            return
        
        print(f"\n找到sheet: {target_sheet_name}")
        
        ws = wb[target_sheet_name]
        
        # 读取前30行数据
        print(f"\nSheet尺寸: {ws.max_row} 行 x {ws.max_column} 列")
        
        print(f"\n前20行数据:")
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value is not None:
                    row_data.append(f"{chr(64+col_idx)}{row_idx}:{value}")
            if row_data:
                print(f"  行{row_idx}: {', '.join(row_data[:10])}")
        
        # 分析列结构
        print(f"\n列结构分析:")
        print(f"  F列（图表1：母猪效能）:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            cell = ws.cell(row_idx, 6)  # F列
            if cell.value:
                print(f"    F{row_idx}: {cell.value}")
        
        print(f"\n  N列（图表2：压栏系数）:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            cell = ws.cell(row_idx, 14)  # N列
            if cell.value:
                print(f"    N{row_idx}: {cell.value}")
        
        # 查找表头行
        print(f"\n查找表头行:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            headers = []
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    headers.append(str(cell.value))
            if headers:
                print(f"  行{row_idx}: {headers[:15]}")
        
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_sheet()
