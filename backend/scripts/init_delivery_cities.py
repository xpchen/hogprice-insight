"""初始化交割城市数据（dim_location + dim_location_alias）

注意：实际城市列表需要从 Excel "交割地市出栏价" Sheet 中提取。
本脚本提供基础框架和常见交割城市数据。
"""
import sys
import os
import io

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.database import SessionLocal
from app.models.dim_location import DimLocation
from app.models.dim_location_alias import DimLocationAlias
from app.models.dim_source import DimSource


# 省份代码映射（基于行政区划代码前2位）
PROVINCE_CODES = {
    "北京": "110000",
    "天津": "120000",
    "河北": "130000",
    "山西": "140000",
    "内蒙古": "150000",
    "辽宁": "210000",
    "吉林": "220000",
    "黑龙江": "230000",
    "上海": "310000",
    "江苏": "320000",
    "浙江": "330000",
    "安徽": "340000",
    "福建": "350000",
    "江西": "360000",
    "山东": "370000",
    "河南": "410000",
    "湖北": "420000",
    "湖南": "430000",
    "广东": "440000",
    "广西": "450000",
    "海南": "460000",
    "重庆": "500000",
    "四川": "510000",
    "贵州": "520000",
    "云南": "530000",
    "西藏": "540000",
    "陕西": "610000",
    "甘肃": "620000",
    "青海": "630000",
    "宁夏": "640000",
    "新疆": "650000",
}

# 常见交割城市列表（需要根据实际 Excel 文件补充）
DELIVERY_CITIES = [
    # 河南
    {"province": "河南", "city": "郑州", "code_suffix": "100"},
    {"province": "河南", "city": "新乡", "code_suffix": "700"},
    {"province": "河南", "city": "开封", "code_suffix": "200"},
    {"province": "河南", "city": "洛阳", "code_suffix": "300"},
    {"province": "河南", "city": "安阳", "code_suffix": "500"},
    {"province": "河南", "city": "商丘", "code_suffix": "400"},
    
    # 山东
    {"province": "山东", "city": "济南", "code_suffix": "100"},
    {"province": "山东", "city": "青岛", "code_suffix": "200"},
    {"province": "山东", "city": "潍坊", "code_suffix": "700"},
    {"province": "山东", "city": "临沂", "code_suffix": "300"},
    {"province": "山东", "city": "烟台", "code_suffix": "600"},
    
    # 江苏
    {"province": "江苏", "city": "南京", "code_suffix": "100"},
    {"province": "江苏", "city": "苏州", "code_suffix": "500"},
    {"province": "江苏", "city": "徐州", "code_suffix": "300"},
    {"province": "江苏", "city": "南通", "code_suffix": "600"},
    
    # 安徽
    {"province": "安徽", "city": "合肥", "code_suffix": "100"},
    {"province": "安徽", "city": "蚌埠", "code_suffix": "300"},
    {"province": "安徽", "city": "阜阳", "code_suffix": "200"},
    
    # 湖北
    {"province": "湖北", "city": "武汉", "code_suffix": "100"},
    {"province": "湖北", "city": "襄阳", "code_suffix": "600"},
    {"province": "湖北", "city": "宜昌", "code_suffix": "500"},
    
    # 湖南
    {"province": "湖南", "city": "长沙", "code_suffix": "100"},
    {"province": "湖南", "city": "株洲", "code_suffix": "200"},
    {"province": "湖南", "city": "衡阳", "code_suffix": "400"},
    
    # 江西
    {"province": "江西", "city": "南昌", "code_suffix": "100"},
    {"province": "江西", "city": "九江", "code_suffix": "400"},
    
    # 四川
    {"province": "四川", "city": "成都", "code_suffix": "100"},
    {"province": "四川", "city": "绵阳", "code_suffix": "700"},
    {"province": "四川", "city": "南充", "code_suffix": "300"},
    
    # 重庆
    {"province": "重庆", "city": "重庆", "code_suffix": "100"},
    
    # 广东
    {"province": "广东", "city": "广州", "code_suffix": "100"},
    {"province": "广东", "city": "深圳", "code_suffix": "300"},
    {"province": "广东", "city": "佛山", "code_suffix": "600"},
    
    # 广西
    {"province": "广西", "city": "南宁", "code_suffix": "100"},
    {"province": "广西", "city": "柳州", "code_suffix": "200"},
    
    # 辽宁
    {"province": "辽宁", "city": "沈阳", "code_suffix": "100"},
    {"province": "辽宁", "city": "大连", "code_suffix": "200"},
    
    # 吉林
    {"province": "吉林", "city": "长春", "code_suffix": "100"},
    
    # 黑龙江
    {"province": "黑龙江", "city": "哈尔滨", "code_suffix": "100"},
    
    # 河北
    {"province": "河北", "city": "石家庄", "code_suffix": "100"},
    {"province": "河北", "city": "唐山", "code_suffix": "200"},
    {"province": "河北", "city": "保定", "code_suffix": "600"},
    
    # 天津
    {"province": "天津", "city": "天津", "code_suffix": "100"},
    
    # 北京
    {"province": "北京", "city": "北京", "code_suffix": "100"},
]


def init_delivery_cities(db):
    """初始化交割城市数据"""
    print("初始化交割城市数据...")
    
    # 检查 YONGYI 数据源是否存在
    source = db.query(DimSource).filter(DimSource.source_code == "YONGYI").first()
    if not source:
        print("  ⚠ 警告：YONGYI 数据源不存在，请先初始化数据源")
        return
    
    created_count = 0
    alias_count = 0
    
    for city_info in DELIVERY_CITIES:
        province_name = city_info["province"]
        city_name = city_info["city"]
        code_suffix = city_info["code_suffix"]
        
        # 获取省份代码
        province_code = PROVINCE_CODES.get(province_name)
        if not province_code:
            print(f"  ⚠ 跳过：未找到省份代码 - {province_name}")
            continue
        
        # 生成城市 location_code（格式：LOC_ + 6位行政区划代码）
        city_location_code = f"LOC_{province_code[:2]}{code_suffix}"
        province_location_code = f"LOC_{province_code}"
        
        # 1. 确保省份存在（如果不存在则创建）
        province_location = db.query(DimLocation).filter(
            DimLocation.location_code == province_location_code
        ).first()
        
        if not province_location:
            province_location = DimLocation(
                location_code=province_location_code,
                level="province",
                parent_code=None,
                name_cn=province_name
            )
            db.add(province_location)
            db.flush()
            print(f"  ✓ 创建省份：{province_name} ({province_location_code})")
        
        # 2. 创建或更新城市
        city_location = db.query(DimLocation).filter(
            DimLocation.location_code == city_location_code
        ).first()
        
        if not city_location:
            city_location = DimLocation(
                location_code=city_location_code,
                level="city",
                parent_code=province_location_code,
                name_cn=city_name
            )
            db.add(city_location)
            db.flush()
            created_count += 1
            print(f"  ✓ 创建城市：{city_name} ({city_location_code})")
        else:
            print(f"  - 城市已存在：{city_name} ({city_location_code})")
        
        # 3. 创建城市别名映射（支持多种变体）
        aliases = [
            city_name,  # 标准名称
            f"{city_name}市",  # 带"市"后缀
        ]
        
        for alias in aliases:
            existing_alias = db.query(DimLocationAlias).filter(
                DimLocationAlias.alias == alias,
                DimLocationAlias.source_code == "YONGYI"
            ).first()
            
            if not existing_alias:
                location_alias = DimLocationAlias(
                    alias=alias,
                    source_code="YONGYI",
                    location_code=city_location_code
                )
                db.add(location_alias)
                alias_count += 1
                print(f"    → 创建别名：{alias} → {city_location_code}")
    
    db.commit()
    print(f"\n完成！创建了 {created_count} 个城市，{alias_count} 个别名映射")


def extract_cities_from_excel(excel_path: str, sheet_name: str = "交割地市出栏价"):
    """
    从 Excel 文件中提取交割城市列表
    
    注意：这是一个辅助函数，需要根据实际 Excel 格式实现。
    建议手动提取城市列表后，更新 DELIVERY_CITIES 列表。
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        
        # 使用 openpyxl 读取复杂格式
        wb = load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' 不存在")
            return []
        
        ws = wb[sheet_name]
        cities = []
        
        # 从第5行（city_header_row）开始读取城市名称
        # 这里需要根据实际格式调整
        for row_idx in range(5, ws.max_row + 1):
            city_cell = ws.cell(row=row_idx, column=2)  # 假设城市在第2列
            if city_cell.value:
                city_name = str(city_cell.value).strip()
                if city_name and city_name not in ["日期", "省份"]:
                    cities.append(city_name)
        
        return list(set(cities))  # 去重
        
    except Exception as e:
        print(f"  ⚠ 提取城市失败：{e}")
        return []


def main():
    """执行初始化"""
    db = SessionLocal()
    
    try:
        print("=" * 60)
        print("初始化交割城市数据")
        print("=" * 60)
        print("\n注意：")
        print("1. 本脚本使用预定义的常见交割城市列表")
        print("2. 实际城市列表需要从 Excel '交割地市出栏价' Sheet 中提取")
        print("3. 提取后可以更新 DELIVERY_CITIES 列表或使用 extract_cities_from_excel() 函数")
        print("\n" + "-" * 60 + "\n")
        
        init_delivery_cities(db)
        
        print("\n" + "=" * 60)
        print("初始化完成！")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n初始化失败: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()


if __name__ == "__main__":
    main()
