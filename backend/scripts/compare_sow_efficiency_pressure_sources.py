# -*- coding: utf-8 -*-
"""
对比 母猪效能、压栏系数 的旧数据源 vs 当前库数据，分析是否一致。

旧数据源：
  - A1供给预测（2、【生猪产业数据】.xlsx）：F列=母猪效能，N列=压栏系数
  - 涌益 月度-生产指标（涌益咨询 周度数据.xlsx）：F列=母猪效能，N列=压栏系数（分析脚本中的说法）

当前 API 数据源：
  - fact_monthly_indicator: 母猪效能=prod_farrowing_count(YONGYI), 压栏系数=prod_healthy_piglets_per_litter(YONGYI)
  - r09 读取的是 涌益 月度-生产指标 sheet，COL_MAP 为 col5=分娩窝数, col7=窝均健仔数（注意：Excel 列5=E, 列6=F, 列7=G, 列8=H, 列14=N）
"""
import sys
from pathlib import Path
from datetime import datetime

backend = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend))

import openpyxl
from app.core.database import SessionLocal
from sqlalchemy import text


def find_excel_paths():
    repo = backend.parent
    industry = repo / "docs" / "0226" / "生猪" / "2、【生猪产业数据】.xlsx"
    if not industry.exists():
        industry = repo / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    yongyi_dir = repo / "docs" / "0226" / "生猪"
    yongyi_files = list(yongyi_dir.glob("*涌益*周度*.xlsx")) if yongyi_dir.exists() else []
    if not yongyi_files:
        yongyi_files = list(repo.glob("docs/**/*涌益*周度*.xlsx"))
    return industry, yongyi_files[0] if yongyi_files else None


def read_a1_f_n(excel_path):
    """读取 A1 供给预测 sheet 的 F列(6)、N列(14) 表头与数据"""
    if not excel_path or not excel_path.exists():
        return None
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "A1供给预测" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["A1供给预测"]
    # 表头：通常前几行
    h1_f, h1_n = ws.cell(1, 6).value, ws.cell(1, 14).value
    h2_f, h2_n = ws.cell(2, 6).value, ws.cell(2, 14).value
    h3_f, h3_n = ws.cell(3, 6).value, ws.cell(3, 14).value
    rows = []
    for ri in range(4, min(ws.max_row + 1, 80)):
        date_val = ws.cell(ri, 1).value
        f_val = ws.cell(ri, 6).value
        n_val = ws.cell(ri, 14).value
        if date_val is None and f_val is None and n_val is None:
            break
        rows.append((date_val, f_val, n_val))
    wb.close()
    return {
        "file": str(excel_path.name),
        "sheet": "A1供给预测",
        "F_header": (h1_f, h2_f, h3_f),
        "N_header": (h1_n, h2_n, h3_n),
        "rows": rows,
    }


def read_yongyi_production_sheet(excel_path):
    """读取涌益 月度-生产指标 sheet：列 5,6,7,8 和 14 的表头与数据（对应 E,F,G,H,N）"""
    if not excel_path or not excel_path.exists():
        return None
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    target = None
    for name in wb.sheetnames:
        if "月度" in name and "生产指标" in name:
            target = name
            break
    if not target:
        wb.close()
        return None
    ws = wb[target]
    # 第2行一般为指标名（r09 注释写的 row 2 = header）
    headers = {}
    for c in [5, 6, 7, 8, 14]:
        h1 = ws.cell(1, c).value
        h2 = ws.cell(2, c).value
        headers[c] = (h1, h2)
    rows = []
    for ri in range(3, min(ws.max_row + 1, 80)):
        date_val = ws.cell(ri, 1).value
        row = [ws.cell(ri, c).value for c in [5, 6, 7, 8, 14]]
        if date_val is None and all(x is None for x in row):
            break
        rows.append((date_val,) + tuple(row))
    wb.close()
    return {
        "file": str(excel_path.name),
        "sheet": target,
        "headers": headers,
        "rows": rows,
    }


def _month_key(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m")
    s = str(v).strip()[:7]
    if len(s) >= 7 and s[4] == "-":
        return s
    return None


def get_db_sow_pressure():
    db = SessionLocal()
    sow = db.execute(text("""
        SELECT month_date, value FROM fact_monthly_indicator
        WHERE indicator_code = 'prod_farrowing_count' AND source = 'YONGYI' AND region_code = 'NATION'
        ORDER BY month_date
    """)).fetchall()
    press = db.execute(text("""
        SELECT month_date, value FROM fact_monthly_indicator
        WHERE indicator_code = 'prod_healthy_piglets_per_litter' AND source = 'YONGYI' AND region_code = 'NATION'
        ORDER BY month_date
    """)).fetchall()
    db.close()
    return (
        {_month_key(r[0]): float(r[1]) for r in sow if r[1] is not None},
        {_month_key(r[0]): float(r[1]) for r in press if r[1] is not None},
    )


def main():
    print("=" * 70)
    print("母猪效能、压栏系数 数据源对比分析")
    print("=" * 70)

    industry_path, yongyi_path = find_excel_paths()

    # ----- A1 表（旧：2、生猪产业数据） -----
    print("\n【1】A1供给预测（2、【生猪产业数据】.xlsx）F列 / N列")
    if industry_path:
        a1 = read_a1_f_n(industry_path)
        if a1:
            print(f"  文件: {a1['file']}")
            print(f"  F列 表头(行1~3): {a1['F_header']}")
            print(f"  N列 表头(行1~3): {a1['N_header']}")
            print(f"  数据行数: {len(a1['rows'])}")
            print("  前 8 行 (日期, F列, N列):")
            for r in a1["rows"][:8]:
                print(f"    {r}")
            a1_by_month = {}
            for date_val, f_val, n_val in a1["rows"]:
                mk = _month_key(date_val)
                if mk:
                    try:
                        a1_by_month[mk] = (float(f_val) if f_val is not None else None, float(n_val) if n_val is not None else None)
                    except (TypeError, ValueError):
                        a1_by_month[mk] = (f_val, n_val)
        else:
            a1_by_month = None
            print("  未找到 A1供给预测 sheet")
    else:
        a1_by_month = None
        print("  未找到 2、【生猪产业数据】.xlsx")

    # ----- 涌益 月度-生产指标（r09 实际读取的 sheet） -----
    print("\n【2】涌益 月度-生产指标 sheet 列 E(5)/F(6)/G(7)/H(8)/N(14)")
    if yongyi_path:
        yy = read_yongyi_production_sheet(yongyi_path)
        if yy:
            print(f"  文件: {yy['file']}, Sheet: {yy['sheet']}")
            for c in [5, 6, 7, 8, 14]:
                let = chr(64 + c) if c <= 26 else f"Col{c}"
                print(f"  列{c}({let}) 表头: {yy['headers'][c]}")
            print("  前 8 行 (日期, col5, col6, col7, col8, col14):")
            for r in yy["rows"][:8]:
                print(f"    {r}")
            # r09 当前写入: col5=prod_farrowing_count, col7=prod_healthy_piglets_per_litter
            yy_col5_col7 = {}
            for r in yy["rows"]:
                mk = _month_key(r[0])
                if mk and len(r) >= 4:
                    try:
                        v5 = float(r[1]) if r[1] is not None else None
                        v7 = float(r[3]) if r[3] is not None else None
                        yy_col5_col7[mk] = (v5, v7)
                    except (TypeError, ValueError):
                        pass
        else:
            yy_col5_col7 = None
            print("  未找到 月度-生产指标 sheet")
    else:
        yy_col5_col7 = None
        print("  未找到涌益周度 Excel")

    # ----- 当前数据库（r09 导入结果） -----
    print("\n【3】当前库 fact_monthly_indicator (YONGYI)")
    db_sow, db_press = get_db_sow_pressure()
    print(f"  母猪效能(prod_farrowing_count): {len(db_sow)} 条, 例如 2024-01={db_sow.get('2024-01')}, 2025-06={db_sow.get('2025-06')}")
    print(f"  压栏系数(prod_healthy_piglets_per_litter): {len(db_press)} 条, 例如 2024-01={db_press.get('2024-01')}, 2025-06={db_press.get('2025-06')}")

    # ----- 对比：A1 F/N vs 库；涌益 col5/col7 vs 库 -----
    print("\n【4】逐月对比（同一月份 A1 F/N、涌益 col5/col7、库 母猪效能/压栏系数）")
    all_months = sorted(set(list(db_sow.keys()) + list(db_press.keys()) + (list(a1_by_month.keys()) if a1_by_month else []) + (list(yy_col5_col7.keys()) if yy_col5_col7 else [])))
    all_months = [m for m in all_months if m and m >= "2023-01"][:24]
    print("  月份         A1_F(母猪效能)  A1_N(压栏系数)  涌益col5  涌益col7   DB母猪效能  DB压栏系数  说明")
    for m in all_months:
        a1_f, a1_n = (a1_by_month.get(m) or (None, None)) if a1_by_month else (None, None)
        yy5, yy7 = (yy_col5_col7.get(m) or (None, None)) if yy_col5_col7 else (None, None)
        ds, dp = db_sow.get(m), db_press.get(m)
        note = ""
        if a1_f is not None and ds is not None and abs((a1_f or 0) - (ds or 0)) > 0.01:
            note += " 母猪效能≠A1_F"
        if a1_n is not None and dp is not None and abs((a1_n or 0) - (dp or 0)) > 0.01:
            note += " 压栏系数≠A1_N"
        if yy5 is not None and ds is not None and abs((yy5 or 0) - (ds or 0)) > 0.01:
            note += " 母猪效能≠涌益col5"
        if yy7 is not None and dp is not None and abs((yy7 or 0) - (dp or 0)) > 0.01:
            note += " 压栏系数≠涌益col7"
        print(f"  {m}  {a1_f!r:>12}  {a1_n!r:>12}  {yy5!r:>8}  {yy7!r:>8}  {ds!r:>10}  {dp!r:>10}  {note}")

    # ----- r09 列号说明 -----
    print("\n【5】r09 当前列映射（_parse_production_metrics）")
    print("  表头(按 r09 注释): 日期, 基础母猪存栏, 环比涨跌, 配种数, 配种数环比, 分娩窝数, 分娩窝数环比, 窝均健仔数, 产房存活率...")
    print("  即: col1=日期 col2=基础母猪存栏 col3=环比 col4=配种数 col5=配种数环比 col6=分娩窝数 col7=分娩窝数环比 col8=窝均健仔数")
    print("  当前 r09 COL_MAP: col5=prod_farrowing_count(分娩窝数), col7=prod_healthy_piglets_per_litter(窝均健仔数)")
    print("  => 若实际表头如上，则 col5 是「配种数环比」、col6 才是「分娩窝数」；col7 是「分娩窝数环比」、col8 才是「窝均健仔数」。")
    print("  => 若数据不一致，应改为 col6=母猪效能(分娩窝数), col8=压栏系数(窝均健仔数)。")


if __name__ == "__main__":
    main()
