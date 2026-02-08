"""
完整分析月度-生产指标sheet结构
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import openpyxl

script_dir = Path(__file__).parent.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_full():
    """完整分析"""
    print("=" * 80)
    print("完整分析月度-生产指标sheet结构")
    print("=" * 80)
    
    docs_dir = Path(script_dir.parent) / "docs"
    excel_files = list(docs_dir.glob("*涌益*.xlsx"))
    
    if not excel_files:
        print("\n未找到涌益Excel文件")
        return
    
    excel_file = excel_files[0]
    print(f"\n文件: {excel_file.name}")
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    sheet_names = wb.sheetnames
    target_sheet_name = None
    
    for name in sheet_names:
        if "生产指标" in name and "月度" in name:
            target_sheet_name = name
            break
    
    if not target_sheet_name:
        print("\n未找到sheet")
        return
    
    print(f"\nSheet: {target_sheet_name}")
    ws = wb[target_sheet_name]
    
    # 读取第1行和第2行，确定列结构
    print(f"\n第1行（省份/分组标题）:")
    row1 = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(1, col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row1.append(f"{col_letter}: {cell.value}")
    for item in row1:
        print(f"  {item}")
    
    print(f"\n第2行（指标标题）:")
    row2 = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(2, col_idx)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row2.append(f"{col_letter}: {cell.value}")
    for item in row2:
        print(f"  {item}")
    
    # 分析F列和N列
    print(f"\nF列（图表1：母猪效能）:")
    print(f"  第1行: {ws.cell(1, 6).value}")
    print(f"  第2行: {ws.cell(2, 6).value}")
    print(f"  数据示例（前5条）:")
    for row_idx in range(3, min(8, ws.max_row + 1)):
        date_val = ws.cell(row_idx, 1).value
        f_val = ws.cell(row_idx, 6).value
        print(f"    {date_val}: {f_val}")
    
    print(f"\nN列（图表2：压栏系数）:")
    print(f"  第1行: {ws.cell(1, 14).value}")
    print(f"  第2行: {ws.cell(2, 14).value}")
    print(f"  数据示例（前5条）:")
    for row_idx in range(3, min(8, ws.max_row + 1)):
        date_val = ws.cell(row_idx, 1).value
        n_val = ws.cell(row_idx, 14).value
        print(f"    {date_val}: {n_val}")
    
    # 分析所有列，找出可能的5个指标列
    print(f"\n所有列分析（找出可能的5个指标）:")
    # 从第2行找出所有不同的指标名称
    indicators = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header1 = ws.cell(1, col_idx).value
        header2 = ws.cell(2, col_idx).value
        if header2:
            key = f"{header1 or ''}|{header2}"
            if key not in indicators:
                indicators[key] = []
            indicators[key].append(col_letter)
    
    print(f"\n指标分组:")
    for key, cols in indicators.items():
        if len(cols) > 0:
            print(f"  {key}: {cols}")

if __name__ == "__main__":
    analyze_full()
