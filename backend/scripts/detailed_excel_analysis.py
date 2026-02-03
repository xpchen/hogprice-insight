"""详细分析Excel文件结构，生成表结构定义"""
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Excel文件列表
EXCEL_FILES = [
    {
        "name": "钢联价格模板",
        "path": project_root / "docs" / "1、价格：钢联自动更新模板.xlsx",
        "source_code": "GANGLIAN",
        "dataset_type": "DAILY"
    },
    {
        "name": "涌益周度数据",
        "path": project_root / "docs" / "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx",
        "source_code": "YONGYI",
        "dataset_type": "WEEKLY"
    },
    {
        "name": "涌益日度数据",
        "path": project_root / "docs" / "2026年2月2日涌益咨询日度数据 - 副本.xlsx",
        "source_code": "YONGYI",
        "dataset_type": "DAILY"
    },
    {
        "name": "DCE期货",
        "path": project_root / "docs" / "lh_ftr.xlsx",
        "source_code": "DCE",
        "dataset_type": "DAILY"
    },
    {
        "name": "DCE期权",
        "path": project_root / "docs" / "lh_opt.xlsx",
        "source_code": "DCE",
        "dataset_type": "DAILY"
    }
]


def analyze_sheet_detailed(workbook, sheet_name, source_code, dataset_type):
    """详细分析单个sheet的结构"""
    result = {
        "sheet_name": sheet_name,
        "source_code": source_code,
        "dataset_type": dataset_type,
        "table_name": None,
        "structure_type": None,
        "header_info": {},
        "columns": [],
        "sample_data": [],
        "notes": []
    }
    
    try:
        ws = workbook[sheet_name]
        result["total_rows"] = ws.max_row
        result["total_cols"] = ws.max_column
        
        # 读取前30行进行详细分析
        data_rows = []
        for row in ws.iter_rows(max_row=30, values_only=True):
            data_rows.append(row)
        
        if not data_rows:
            result["notes"].append("Sheet为空")
            return result
        
        # 分析表头结构
        # 检查是否有合并单元格（通过检查前几行的结构）
        header_candidates = []
        for i in range(min(5, len(data_rows))):
            row = data_rows[i]
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
            if non_empty > 2:
                header_candidates.append({
                    "row_index": i + 1,
                    "non_empty_count": non_empty,
                    "row_data": row[:10]  # 只取前10列
                })
        
        if header_candidates:
            best_header = max(header_candidates, key=lambda x: x["non_empty_count"])
            result["header_info"]["header_row"] = best_header["row_index"]
            result["header_info"]["header_data"] = best_header["row_data"]
        
        # 尝试识别结构类型
        # 1. 检查是否有日期列
        date_cols = []
        for col_idx in range(min(10, result["total_cols"])):
            col_data = [row[col_idx] if col_idx < len(row) else None for row in data_rows[5:15]]
            # 检查是否可能是日期
            date_count = 0
            for val in col_data:
                if val:
                    try:
                        pd.to_datetime(val)
                        date_count += 1
                    except:
                        pass
            if date_count > len(col_data) * 0.5:
                date_cols.append(col_idx)
        
        # 2. 识别结构类型
        if len(date_cols) == 1:
            # 可能是窄表（日期列 + 指标列）
            result["structure_type"] = "NARROW_DATE_ROWS"
            result["notes"].append(f"检测到日期列在第{date_cols[0]+1}列")
        elif len(date_cols) > 1:
            # 可能是宽表（日期作为列）
            result["structure_type"] = "WIDE_DATE_COLS"
            result["notes"].append(f"检测到多个日期列: {[c+1 for c in date_cols]}")
        else:
            # 可能是其他结构
            result["structure_type"] = "UNKNOWN"
            result["notes"].append("未检测到明显的日期列")
        
        # 分析列结构
        if result["header_info"].get("header_row"):
            header_row_idx = result["header_info"]["header_row"] - 1
            if header_row_idx < len(data_rows):
                header_row = data_rows[header_row_idx]
                for col_idx, header_val in enumerate(header_row[:min(20, result["total_cols"])]):
                    if header_val:
                        col_info = {
                            "index": col_idx,
                            "name": str(header_val).strip(),
                            "is_date": col_idx in date_cols,
                            "sample_values": []
                        }
                        
                        # 获取样本值
                        for row in data_rows[header_row_idx+1:header_row_idx+6]:
                            if col_idx < len(row) and row[col_idx] is not None:
                                col_info["sample_values"].append(row[col_idx])
                        
                        result["columns"].append(col_info)
        
        # 保存前10行样本数据
        if len(data_rows) > result["header_info"].get("header_row", 0):
            sample_start = result["header_info"].get("header_row", 0)
            for row in data_rows[sample_start:sample_start+10]:
                result["sample_data"].append(row[:min(15, len(row))])
        
    except Exception as e:
        result["error"] = str(e)
        import traceback
        result["traceback"] = traceback.format_exc()
    
    return result


def generate_table_name(sheet_name, source_code, dataset_type):
    """生成表名"""
    import re
    # 转换为snake_case
    name = re.sub(r'[^\w\s-]', '', sheet_name)
    name = re.sub(r'[\s-]+', '_', name)
    name = name.lower()
    name = re.sub(r'_+', '_', name)
    name = name.strip('_')
    
    table_name = f"{source_code.lower()}_{dataset_type.lower()}_{name}"
    if len(table_name) > 64:
        table_name = table_name[:64]
    return table_name


def analyze_file(file_info, max_sheets=None):
    """分析单个文件"""
    print(f"\n{'='*80}")
    print(f"分析文件: {file_info['name']}")
    print(f"路径: {file_info['path']}")
    print(f"{'='*80}")
    
    if not file_info['path'].exists():
        print(f"❌ 文件不存在: {file_info['path']}")
        return None
    
    try:
        workbook = load_workbook(file_info['path'], data_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"\n找到 {len(sheet_names)} 个sheet")
        if max_sheets:
            sheet_names = sheet_names[:max_sheets]
            print(f"（仅分析前{max_sheets}个）")
        
        results = []
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"\n[{i}/{len(sheet_names)}] 分析: {sheet_name}")
            result = analyze_sheet_detailed(
                workbook,
                sheet_name,
                file_info['source_code'],
                file_info['dataset_type']
            )
            
            # 生成表名
            result["table_name"] = generate_table_name(
                sheet_name,
                file_info['source_code'],
                file_info['dataset_type']
            )
            
            results.append(result)
            
            # 打印摘要
            if "error" in result:
                print(f"  ❌ 错误: {result['error']}")
            else:
                print(f"  ✓ 行数: {result.get('total_rows', 0)}, 列数: {result.get('total_cols', 0)}")
                print(f"  ✓ 结构类型: {result.get('structure_type', 'UNKNOWN')}")
                print(f"  ✓ 表名: {result['table_name']}")
                print(f"  ✓ 列数: {len(result.get('columns', []))}")
        
        workbook.close()
        
        return {
            "file_name": file_info['name'],
            "file_path": str(file_info['path']),
            "source_code": file_info['source_code'],
            "dataset_type": file_info['dataset_type'],
            "total_sheets": len(workbook.sheetnames),
            "analyzed_sheets": len(results),
            "sheets": results
        }
    
    except Exception as e:
        print(f"❌ 分析文件失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """主函数"""
    print("="*80)
    print("Excel文件详细结构分析工具")
    print("="*80)
    
    all_results = []
    
    # 先分析前3个文件（限制每个文件最多分析10个sheet）
    for file_info in EXCEL_FILES[:3]:
        result = analyze_file(file_info, max_sheets=10)
        if result:
            all_results.append(result)
    
    # 保存分析结果
    output_file = project_root / "docs" / "detailed_excel_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n{'='*80}")
    print(f"分析完成！结果已保存到: {output_file}")
    print(f"{'='*80}")
    
    # 生成汇总
    print("\n汇总:")
    for result in all_results:
        print(f"\n{result['file_name']}:")
        print(f"  总sheet数: {result['total_sheets']}")
        print(f"  已分析: {result['analyzed_sheets']}")
        for sheet in result['sheets'][:5]:  # 只显示前5个
            if "error" not in sheet:
                print(f"    - {sheet['sheet_name']}: {sheet.get('structure_type', 'UNKNOWN')} -> {sheet['table_name']}")


if __name__ == "__main__":
    main()
