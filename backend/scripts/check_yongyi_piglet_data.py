"""
检查4.2涌益底稿sheet是否有新生仔猪数据
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import openpyxl

script_dir = Path(__file__).parent.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def check_piglet_data():
    """检查新生仔猪数据"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "4.2涌益底稿" not in wb.sheetnames:
        print("Sheet不存在")
        return
    
    ws = wb["4.2涌益底稿"]
    
    print("=" * 80)
    print("检查4.2涌益底稿sheet结构")
    print("=" * 80)
    
    # 显示前20行，查找新生仔猪相关列
    print("\n前20行数据:")
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        row_data = []
        for col_idx, cell in enumerate(row[:30], 1):
            if cell is not None:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                row_data.append(f"{col_letter}:{cell}")
        if row_data:
            print(f"  行{row_idx}: {', '.join(row_data[:15])}")
    
    # 查找包含"仔猪"或"新生"的列
    print("\n查找包含'仔猪'或'新生'的列:")
    header_row = list(ws.iter_rows(max_row=2, values_only=True))
    if len(header_row) >= 2:
        for col_idx, cell in enumerate(header_row[0][:30], 1):
            if cell and ("仔猪" in str(cell) or "新生" in str(cell)):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  第1行 {col_letter}列: {cell}")
        for col_idx, cell in enumerate(header_row[1][:30], 1):
            if cell and ("仔猪" in str(cell) or "新生" in str(cell)):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  第2行 {col_letter}列: {cell}")
    
    wb.close()

if __name__ == "__main__":
    check_piglet_data()
