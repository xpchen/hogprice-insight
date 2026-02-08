"""详细分析汇总sheet结构"""
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from openpyxl import load_workbook

# 设置输出编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_detailed():
    """详细分析汇总sheet"""
    script_dir = Path(__file__).parent.parent.parent
    excel_file = script_dir / "docs" / "生猪" / "集团企业" / "3.2、集团企业月度数据跟踪.xlsx"
    
    if not excel_file.exists():
        print(f"Excel文件不存在: {excel_file}")
        return
    
    wb = load_workbook(excel_file, data_only=True)
    ws = wb["汇总"]
    
    print("=" * 80)
    print("汇总Sheet详细分析")
    print("=" * 80)
    
    # 读取前3行
    row1_values = [cell.value for cell in ws[1]]
    row2_values = [cell.value for cell in ws[2]]
    
    print("\n第1行（地区分组）:")
    for idx, val in enumerate(row1_values[:25], 1):
        if val:
            print(f"  列{idx}: {val}")
    
    print("\n第2行（指标名称）:")
    for idx, val in enumerate(row2_values[:25], 1):
        if val:
            print(f"  列{idx}: {val}")
    
    # 分析地区分组
    print("\n" + "=" * 80)
    print("地区分组映射:")
    print("=" * 80)
    
    regions_map = {}  # {列索引: 地区名称}
    current_region = None
    region_start_col = None
    
    for idx, val in enumerate(row1_values, 1):
        if val and str(val).strip():
            val_str = str(val).strip()
            if val_str in ["广东", "四川", "贵州", "全国CR20", "全国CR5"]:
                if current_region:
                    regions_map[current_region] = (region_start_col, idx - 1)
                current_region = val_str
                region_start_col = idx
    if current_region:
        regions_map[current_region] = (region_start_col, len(row1_values))
    
    print(f"地区分组: {regions_map}")
    
    # 分析每个地区的指标列
    print("\n" + "=" * 80)
    print("各地区指标列映射:")
    print("=" * 80)
    
    for region, (start_col, end_col) in regions_map.items():
        print(f"\n{region} (列 {start_col}-{end_col}):")
        for col_idx in range(start_col, min(end_col + 1, len(row2_values) + 1)):
            if col_idx <= len(row2_values):
                metric = row2_values[col_idx - 1]
                if metric:
                    print(f"  列{col_idx}: {metric}")
    
    # 分析数据行
    print("\n" + "=" * 80)
    print("数据行示例（前5行）:")
    print("=" * 80)
    
    for row_idx in range(3, min(8, ws.max_row + 1)):
        row = ws[row_idx]
        period_type = row[0].value if len(row) > 0 else None
        date_val = row[1].value if len(row) > 1 else None
        
        print(f"\n第{row_idx}行:")
        print(f"  时间类型: {period_type}")
        print(f"  日期: {date_val}")
        
        # 显示一些数据样本
        sample_data = []
        for col_idx in range(2, min(10, len(row))):
            val = row[col_idx].value
            if val is not None and val != "":
                sample_data.append(f"列{col_idx+1}={val}")
        if sample_data:
            print(f"  数据样本: {', '.join(sample_data[:5])}")

if __name__ == "__main__":
    analyze_detailed()
