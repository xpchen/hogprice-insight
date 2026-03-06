#!/usr/bin/env python3
"""
核查 A6 毛白价差：Excel 解析、DB 数据、缓存。
用于定位「数据更新后图表不显示」的原因。
"""
import os
import re
import sys

BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND)

HEADER_ROW = 2
DATA_START_ROW = 5
_MAO_BAI_RE = re.compile(r"毛白：价差：(\S+?)（日）")
_MAO_BAI_RE_HALF = re.compile(r"毛白：价差：(\S+?)\(日\)")


def main():
    from sqlalchemy import text

    excel_path = os.path.join(
        os.path.dirname(BACKEND), "docs", "0306_生猪", "生猪", "1、价格：钢联自动更新模板.xlsx"
    )
    project_root = os.path.dirname(BACKEND)

    print("=== 1. Excel 毛白价差 sheet ===")
    if not os.path.isfile(excel_path):
        print("  File not found:", excel_path)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        if "毛白价差" not in wb.sheetnames:
            print("  Sheet '毛白价差' not found. Sheets:", wb.sheetnames)
        else:
            ws = wb["毛白价差"]
            rows = list(ws.iter_rows(min_row=1, max_row=DATA_START_ROW + 3, values_only=True))
            headers = list(rows[HEADER_ROW - 1]) if len(rows) >= HEADER_ROW else []
            print("  Row2 (header) first 6 cells:", headers[:6])
            col_maobai = None
            for idx, h in enumerate(headers):
                if not h:
                    continue
                s = str(h)
                if "毛白" in s and "价差" in s and "/" not in s and "移动平均" not in s:
                    if _MAO_BAI_RE.search(s) or _MAO_BAI_RE_HALF.search(s):
                        col_maobai = idx
                        print("  Mao-bai column index:", idx, "header:", s[:60])
                        break
            if col_maobai is None:
                print("  WARNING: No matching mao-bai column found (check full-width  （日） vs half-width (日) )")
            else:
                # 数据从第 DATA_START_ROW 行起（与 r01 一致）
                data_rows = [r for r in rows[DATA_START_ROW - 1:] if len(r) > col_maobai and r[0] is not None]
                def _is_date_like(v):
                    if v is None: return False
                    try: return hasattr(v, "year") or ("-" in str(v) and len(str(v)) >= 8)
                    except: return False
                data_rows = [r for r in data_rows if _is_date_like(r[0])]
                print("  Data rows (from row %d):" % DATA_START_ROW, len(data_rows))
                if data_rows:
                    print("  First date:", data_rows[0][0], "value:", data_rows[0][col_maobai])
                    print("  Last  date:", data_rows[-1][0], "value:", data_rows[-1][col_maobai])
        wb.close()

    print("\n=== 2. DB fact_spread_daily (mao_bai_spread, NATION) ===")
    try:
        from import_tool.db import get_engine
        engine = get_engine()
        with engine.connect() as conn:
            r = conn.execute(text("""
                SELECT COUNT(*), MIN(trade_date), MAX(trade_date), source
                FROM fact_spread_daily
                WHERE spread_type = 'mao_bai_spread' AND region_code = 'NATION'
                GROUP BY source
            """)).fetchall()
            for row in r:
                print(f"  source={row[3]}: count={row[0]}, {row[1]} ~ {row[2]}")
            if not r:
                print("  No rows. API will return empty; chart will show no data.")
    except Exception as e:
        print("  Error:", e)

    print("\n=== 3. Cache (quick_chart_cache) ===")
    try:
        from import_tool.db import get_engine
        engine = get_engine()
        with engine.connect() as conn:
            keys = conn.execute(text("""
                SELECT cache_key FROM quick_chart_cache
                WHERE cache_key LIKE '%live-white-spread%'
                LIMIT 5
            """)).fetchall()
            if keys:
                print("  Cached keys (sample):", [k[0] for k in keys])
                print("  -> If data was updated, clear cache or run refresh_chart_cache so chart gets new data.")
            else:
                print("  No cache entries for live-white-spread.")
    except Exception as e:
        print("  Error:", e)

    print("\n=== 4. Cause & fix ===")
    print("  - If Excel has data but DB empty: run import from docs/0306_生猪/生猪 (file: 1、价格：钢联自动更新模板.xlsx).")
    print("  - If DB has data but chart old: clear chart cache (import_data.py does it) or run: python scripts/refresh_chart_cache.py")
    print("  - If Excel updated same dates: incremental import does NOT overwrite; run BULK import for 钢联 file to refresh 毛白价差.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
