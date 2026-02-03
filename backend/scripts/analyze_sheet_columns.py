"""详细分析sheet的列结构，生成表结构定义"""
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import json
import re

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def analyze_sheet_columns(file_path, sheet_name, max_rows=50):
    """详细分析sheet的列结构"""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        # 读取前几行
        header_rows = []
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(50, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(cell.value)
            header_rows.append(row_data)
        
        # 读取数据样本
        data_rows = []
        data_start = 3  # 假设数据从第3行开始
        for row_idx in range(data_start, min(data_start + max_rows, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(cell.value)
            data_rows.append(row_data)
        
        wb.close()
        
        return {
            "header_rows": header_rows,
            "data_rows": data_rows[:10],  # 只返回前10行数据
            "total_rows": ws.max_row,
            "total_cols": ws.max_column
        }
    except Exception as e:
        return {"error": str(e)}


def main():
    """主函数"""
    files_to_analyze = [
        {
            "name": "钢联价格模板",
            "path": project_root / "docs" / "1、价格：钢联自动更新模板.xlsx",
            "sheets": [
                "分省区猪价",
                "集团企业出栏价",
                "交割库出栏价",
                "区域价差",
                "肥标价差",
                "毛白价差",
                "养殖利润（周度）"
            ]
        },
        {
            "name": "涌益日度数据",
            "path": project_root / "docs" / "2026年2月2日涌益咨询日度数据 - 副本.xlsx",
            "sheets": [
                "出栏价",
                "价格+宰量",
                "散户标肥价差",
                "各省份均价",
                "市场主流标猪肥猪价格",
                "屠宰企业日度屠宰量",
                "市场主流标猪肥猪均价方便作图",
                "交割地市出栏价"
            ]
        }
    ]
    
    results = {}
    
    for file_info in files_to_analyze:
        if not file_info['path'].exists():
            print(f"文件不存在: {file_info['path']}")
            continue
        
        print(f"\n分析文件: {file_info['name']}")
        print(f"路径: {file_info['path']}")
        
        file_results = {}
        for sheet_name in file_info['sheets']:
            print(f"\n分析sheet: {sheet_name}")
            result = analyze_sheet_columns(file_info['path'], sheet_name)
            file_results[sheet_name] = result
            
            if "error" in result:
                print(f"  错误: {result['error']}")
            else:
                print(f"  行数: {result['total_rows']}, 列数: {result['total_cols']}")
                if result['header_rows']:
                    print(f"  表头行数: {len(result['header_rows'])}")
                    # 打印第一行表头
                    if result['header_rows']:
                        first_row = [str(v) if v else "" for v in result['header_rows'][0][:15]]
                        print(f"  第一行表头: {first_row}")
        
        results[file_info['name']] = file_results
    
    # 保存结果
    output_file = project_root / "docs" / "sheet_columns_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n分析完成！结果已保存到: {output_file}")


if __name__ == "__main__":
    main()
