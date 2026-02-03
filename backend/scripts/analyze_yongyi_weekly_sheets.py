"""详细分析涌益周度数据的所有sheet"""
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
import json

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def analyze_sheet_structure(file_path, sheet_name, max_header_rows=10, max_sample_rows=5):
    """分析单个sheet的结构"""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        # 读取表头行
        header_rows = []
        for row_idx in range(1, min(max_header_rows + 1, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(50, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                # 检查合并单元格
                merged_value = None
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = ws[merged_range.min_col][merged_range.min_row]
                        merged_value = top_left.value
                        break
                
                final_value = merged_value if merged_value else value
                if final_value:
                    str_val = str(final_value)
                    if len(str_val) > 30:
                        str_val = str_val[:27] + "..."
                    row_data.append(str_val)
                else:
                    row_data.append("")
            header_rows.append(row_data)
        
        # 读取数据样本
        data_start = max_header_rows + 1
        sample_rows = []
        for row_idx in range(data_start, min(data_start + max_sample_rows, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value:
                    str_val = str(value)
                    if len(str_val) > 20:
                        str_val = str_val[:17] + "..."
                    row_data.append(str_val)
                else:
                    row_data.append("")
            sample_rows.append(row_data)
        
        # 识别表头行（非空单元格最多的行）
        header_row_idx = None
        max_non_empty = 0
        for i, row in enumerate(header_rows[:10], 1):
            non_empty = sum(1 for v in row if v and str(v).strip())
            if non_empty > max_non_empty:
                max_non_empty = non_empty
                header_row_idx = i
        
        # 识别日期列模式
        date_pattern = None
        date_cols = []
        if header_row_idx:
            header_row = header_rows[header_row_idx - 1]
            for col_idx, val in enumerate(header_row[:30], 1):
                if val and isinstance(val, str):
                    # 检查是否是日期格式
                    if any(keyword in val for keyword in ["日期", "日期", "周", "周期", "开始", "结束"]):
                        date_cols.append({"col": col_idx, "value": val})
                    # 检查是否是日期值
                    elif "/" in val or "-" in val:
                        try:
                            # 尝试解析日期
                            date_cols.append({"col": col_idx, "value": val, "is_date_value": True})
                        except:
                            pass
        
        wb.close()
        
        return {
            "sheet_name": sheet_name,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "header_row": header_row_idx,
            "header_rows": header_rows,
            "header_sample": header_rows[header_row_idx - 1][:20] if header_row_idx else [],
            "sample_rows": sample_rows,
            "date_cols": date_cols
        }
    except Exception as e:
        return {
            "sheet_name": sheet_name,
            "error": str(e)
        }


def main():
    """主函数"""
    file_path = project_root / "docs" / "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx"
    
    print("="*80)
    print("涌益周度数据Sheet分析")
    print("="*80)
    print(f"文件: {file_path}")
    
    if not file_path.exists():
        print(f"❌ 文件不存在: {file_path}")
        return
    
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        print(f"\n找到 {len(sheet_names)} 个sheet:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")
        
        print("\n" + "="*80)
        print("开始分析每个sheet...")
        print("="*80)
        
        results = {}
        for sheet_name in sheet_names:
            print(f"\n分析sheet: {sheet_name}")
            result = analyze_sheet_structure(file_path, sheet_name)
            results[sheet_name] = result
            
            if "error" in result:
                print(f"  ❌ 错误: {result['error']}")
            else:
                print(f"  ✓ 行数: {result['total_rows']}, 列数: {result['total_cols']}")
                if result['header_row']:
                    print(f"  ✓ 表头行: {result['header_row']}")
                    print(f"  ✓ 表头样本: {result['header_sample'][:10]}")
                if result['date_cols']:
                    print(f"  ✓ 日期列: {result['date_cols']}")
        
        # 保存结果
        output_file = project_root / "docs" / "yongyi_weekly_sheets_analysis.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2, default=str)
        
        print("\n" + "="*80)
        print(f"分析完成！结果已保存到: {output_file}")
        print("="*80)
        
        # 生成表名建议
        print("\n" + "="*80)
        print("表名建议:")
        print("="*80)
        for sheet_name in sheet_names:
            # 转换为snake_case
            table_name = sheet_name.lower()
            table_name = table_name.replace("周度-", "").replace("周度", "")
            table_name = table_name.replace(" ", "_").replace("-", "_")
            table_name = table_name.replace("（", "_").replace("）", "")
            table_name = table_name.replace("(", "_").replace(")", "")
            table_name = "yongyi_weekly_" + table_name
            print(f"  {sheet_name:40s} -> {table_name}")
        
    except Exception as e:
        print(f"❌ 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
