"""测试涌益周度数据导入功能 - 支持单独测试某个sheet并显示详细调试信息"""
import sys
import os
import io
import json
import hashlib
from pathlib import Path
from datetime import datetime
from sqlalchemy import text, inspect
from openpyxl import load_workbook
import pandas as pd

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 添加backend目录到路径（确保能导入app模块）
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.core.database import SessionLocal
from app.models.sys_user import SysUser
from app.models.dim_source import DimSource
from app.services.ingestors.unified_ingestor import unified_import
from app.services.ingestors.dispatcher import Dispatcher
from app.services.ingestors.parsers import get_parser
from app.services.ingestors.profile_loader import load_profile_from_json, get_profile_by_dataset_type
from app.services.ingestors.validator import ObservationValidator
from app.services.ingestors.error_collector import ErrorCollector
from app.models.import_batch import ImportBatch
from app.models.raw_sheet import RawSheet
from app.services.ingestors.raw_writer import save_raw_file
from app.services.ingestors.observation_upserter import upsert_observations
from app.services.sheet_table_mapper import get_table_name_for_sheet
from app.services.column_mapper import ColumnMapper
from app.services.sheet_table_importer import SheetTableImporter
from sqlalchemy import inspect, text

# 项目根目录
backend_dir = Path(__file__).parent.parent
project_root = backend_dir.parent


def get_or_create_test_user(db):
    """获取或创建测试用户"""
    user = db.query(SysUser).filter(SysUser.username == "admin").first()
    if not user:
        from app.core.security import get_password_hash
        user = SysUser(
            username="test_admin",
            password_hash=get_password_hash("test123456"),
            display_name="测试管理员",
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        print(f"  ✓ 创建测试用户: {user.username} (ID: {user.id})", flush=True)
    else:
        print(f"  ✓ 使用现有用户: {user.username} (ID: {user.id})", flush=True)
    return user


def print_excel_structure(file_path: Path, sheet_name: str, max_rows: int = 10):
    """打印Excel sheet的前几行结构"""
    print(f"\n{'='*80}", flush=True)
    print(f"📊 Excel Sheet结构分析: {sheet_name}", flush=True)
    print(f"{'='*80}", flush=True)
    
    try:
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"  ❌ Sheet '{sheet_name}' 不存在", flush=True)
            print(f"  可用sheets: {wb.sheetnames[:10]}", flush=True)
            return
        
        ws = wb[sheet_name]
        
        # 打印前max_rows行
        print(f"\n前{max_rows}行数据:", flush=True)
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=False), 1):
            row_values = [cell.value if cell.value is not None else "" for cell in row[:20]]  # 只显示前20列
            row_str = " | ".join([str(v)[:15] for v in row_values])
            print(f"  行{row_idx:2d}: {row_str}", flush=True)
        
        # 打印列名（尝试识别header行）
        print(f"\n列名分析（尝试识别header行）:", flush=True)
        for header_row_idx in [1, 2, 3, 4]:
            if header_row_idx <= ws.max_row:
                row_values = [cell.value if cell.value is not None else "" for cell in ws[header_row_idx][:20]]
                row_str = " | ".join([str(v)[:15] for v in row_values])
                print(f"  第{header_row_idx}行: {row_str}", flush=True)
        
        wb.close()
    except Exception as e:
        print(f"  ❌ 读取Excel失败: {e}", flush=True)
        import traceback
        traceback.print_exc()


def test_single_sheet(file_path: Path, sheet_name: str, db, profile, uploader_id: int, auto_import: bool = False):
    """测试单个sheet的解析和入库"""
    print(f"\n{'='*80}", flush=True)
    print(f"🧪 测试Sheet: {sheet_name}", flush=True)
    print(f"{'='*80}", flush=True)
    
    # 1. 显示Excel结构
    print_excel_structure(file_path, sheet_name, max_rows=15)
    
    # 2. 创建Dispatcher并分派
    print(f"\n{'='*80}", flush=True)
    print(f"🔍 Dispatcher分派分析", flush=True)
    print(f"{'='*80}", flush=True)
    
    dispatcher = Dispatcher(db, profile)
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        # 调用dispatcher（传入Worksheet对象，不是DataFrame）
        dispatch_result = dispatcher.dispatch_sheet(sheet_name, worksheet=ws)
        
        print(f"  分派结果:", flush=True)
        print(f"    - Parser: {dispatch_result.get('parser')}", flush=True)
        sheet_config = dispatch_result.get('sheet_config', {})
        print(f"    - Sheet配置: {json.dumps(sheet_config, ensure_ascii=False, indent=2)}", flush=True)
        print(f"    - 原因: {dispatch_result.get('reason', 'N/A')}", flush=True)
        
        # 检查是否有table_config
        table_config = sheet_config.get("table_config")
        if not table_config:
            print(f"\n  ⚠️  警告: Sheet配置中没有table_config，将使用旧架构（fact_observation）", flush=True)
            print(f"  💡 检测到配置文件可能已更新，是否自动重新加载配置？(y/n): ", end="", flush=True)
            reload = input().strip().lower()
            if reload == 'y':
                print(f"  正在重新加载配置...", flush=True)
                try:
                    profile = load_profile_from_json(db, profile_path)
                    db.refresh(profile)
                    # 重新dispatch
                    dispatcher = Dispatcher(db, profile)
                    dispatch_result = dispatcher.dispatch_sheet(sheet_name, worksheet=ws)
                    sheet_config = dispatch_result.get('sheet_config', {})
                    table_config = sheet_config.get("table_config")
                    if table_config:
                        print(f"  ✓ 重新加载成功，现在使用新架构（独立表）", flush=True)
                        print(f"    目标表: {table_config.get('table_name', '未知')}", flush=True)
                    else:
                        print(f"  ⚠️  重新加载后仍没有table_config，将使用旧架构", flush=True)
                except Exception as e:
                    print(f"  ❌ 重新加载失败: {e}", flush=True)
                    print(f"  请手动运行: python scripts/load_ingest_profiles.py", flush=True)
            else:
                print(f"  跳过重新加载，将使用旧架构（fact_observation）", flush=True)
                print(f"  如需使用新架构，请运行: python scripts/load_ingest_profiles.py", flush=True)
        
        if not dispatch_result.get('parser'):
            print(f"\n  ⚠️  未匹配到parser，原因:", flush=True)
            print(f"    {dispatch_result.get('reason', '未知原因')}", flush=True)
            wb.close()
            return
        
        # 3. 使用parser解析
        parser_name = dispatch_result['parser']
        # sheet_config 已经在上面获取了
        
        print(f"\n{'='*80}", flush=True)
        print(f"🔧 Parser解析: {parser_name}", flush=True)
        print(f"{'='*80}", flush=True)
        
        print(f"  Sheet配置:", flush=True)
        print(f"    {json.dumps(sheet_config, ensure_ascii=False, indent=4)}", flush=True)
        
        parser = get_parser(parser_name)
        if not parser:
            print(f"  ❌ Parser '{parser_name}' 不存在", flush=True)
            wb.close()
            return
        
        # 准备解析参数
        profile_defaults = profile.defaults_json or {}
        source_code = profile.source_code
        
        print(f"\n  开始解析...", flush=True)
        observations = parser.parse(
            sheet_data=ws,
            sheet_config=sheet_config,
            profile_defaults=profile_defaults,
            source_code=source_code,
            batch_id=0  # 临时使用0，入库时会创建真实batch
        )
        
        print(f"\n  解析结果:", flush=True)
        print(f"    - 观测值数量: {len(observations)}", flush=True)
        
        if len(observations) == 0:
            print(f"\n  ⚠️  解析出0条数据，无法继续", flush=True)
            print(f"    可能的原因:", flush=True)
            print(f"    1. header_row配置不正确", flush=True)
            print(f"    2. 日期列名不匹配（需要'开始日期'和'结束日期'）", flush=True)
            print(f"    3. 数据行格式不符合预期", flush=True)
            print(f"    4. 省份列识别失败", flush=True)
            wb.close()
            return
        
        # 显示前3条示例
        print(f"\n  前3条观测值示例:", flush=True)
        for i, obs in enumerate(observations[:3], 1):
            print(f"    [{i}] metric_key={obs.get('metric_key')}, geo_code={obs.get('geo_code')}, indicator={obs.get('tags', {}).get('indicator')}", flush=True)
            print(f"        {json.dumps(obs, ensure_ascii=False, indent=6, default=str)}", flush=True)
        
        # 4. 显示目标表信息
        table_config = sheet_config.get("table_config")
        table_name = None
        if table_config:
            # table_name在sheet_config顶层，不在table_config里面
            table_name = sheet_config.get("table_name") or get_table_name_for_sheet(
                sheet_name=sheet_name,
                source_code=profile.source_code,
                dataset_type=profile.dataset_type
            )
            print(f"\n  目标表: {table_name}", flush=True)
            print(f"  架构: 新架构（独立表）", flush=True)
        else:
            print(f"\n  目标表: fact_observation", flush=True)
            print(f"  架构: 旧架构（统一表）", flush=True)
        
        # 5. 询问是否入库（如果auto_import为False）
        if not auto_import:
            print(f"\n{'='*80}", flush=True)
            print(f"💾 入库确认", flush=True)
            print(f"{'='*80}", flush=True)
            
            confirm = input(f"是否将 {len(observations)} 条解析数据导入到数据库？(y/n): ").strip().lower()
            
            if confirm != 'y' and confirm != '是':
                print(f"  已取消入库", flush=True)
                wb.close()
                return
        else:
            print(f"\n{'='*80}", flush=True)
            print(f"💾 自动入库模式", flush=True)
            print(f"{'='*80}", flush=True)
            print(f"  将自动导入 {len(observations)} 条解析数据到数据库", flush=True)
        
        # 6. 执行入库
        print(f"\n{'='*80}", flush=True)
        print(f"🚀 开始入库", flush=True)
        print(f"{'='*80}", flush=True)
        
        # 6.1 创建import_batch
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        file_hash = hashlib.sha256(file_content).hexdigest()
        batch = ImportBatch(
            filename=file_path.name,
            file_hash=file_hash,
            uploader_id=uploader_id,
            source_code=profile.source_code or profile.dataset_type,
            status="processing",
            total_rows=0,
            success_rows=0,
            failed_rows=0
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)
        
        print(f"  ✓ 创建批次: ID={batch.id}", flush=True)
        
        # 6.2 验证数据（需要batch_id）
        print(f"\n{'='*80}", flush=True)
        print(f"✅ 数据验证", flush=True)
        print(f"{'='*80}", flush=True)
        
        error_collector = ErrorCollector(db, batch.id)
        validator = ObservationValidator(error_collector)
        skip_metric_check = bool(table_config)
        
        valid_observations = validator.validate_batch(
            observations,
            sheet_name,
            skip_metric_key_check=skip_metric_check
        )
        
        print(f"  验证结果:", flush=True)
        print(f"    - 总观测值: {len(observations)} 条", flush=True)
        print(f"    - 有效观测值: {len(valid_observations)} 条", flush=True)
        print(f"    - 无效观测值: {len(observations) - len(valid_observations)} 条", flush=True)
        
        if len(valid_observations) == 0:
            print(f"\n  ⚠️  没有有效观测值，无法入库", flush=True)
            if error_collector.errors:
                print(f"  前5个错误:", flush=True)
                for error in error_collector.errors[:5]:
                    print(f"    - {error}", flush=True)
            batch.status = "failed"
            db.commit()
            wb.close()
            return
        
        # 6.2 保存raw_file和raw_sheet
        raw_file = save_raw_file(
            db=db,
            batch_id=batch.id,
            filename=file_path.name,
            file_content=file_content
        )
        print(f"  ✓ 保存raw_file: ID={raw_file.id}", flush=True)
        
        # 6.3 保存raw_sheet（简化版，只保存当前sheet）
        raw_sheet = RawSheet(
            raw_file_id=raw_file.id,
            sheet_name=sheet_name,
            parse_status="parsed",
            parser_type=parser_name,
            observation_count=len(valid_observations)
        )
        db.add(raw_sheet)
        db.commit()
        db.refresh(raw_sheet)
        print(f"  ✓ 保存raw_sheet: ID={raw_sheet.id}", flush=True)
        
        # 6.4 更新observations的batch_id
        for obs in valid_observations:
            obs['batch_id'] = batch.id
        
        # 6.5 根据架构选择入库方式
        if table_config and valid_observations:
            # 新架构：导入到独立表
            print(f"  → 使用新架构导入到独立表...", flush=True)
            
            column_mapping = table_config.get("column_mapping", {})
            unique_key = table_config.get("unique_key", [])
            
            # 转换数据
            print(f"  → 开始转换数据（{len(valid_observations)} 条观测值 -> 表记录）...", flush=True)
            mapper = ColumnMapper()
            records = mapper.map_observations_to_table_records(
                observations=valid_observations,
                column_mapping=column_mapping,
                table_name=table_name,
                batch_id=batch.id,
                sheet_config=sheet_config
            )
            
            print(f"  ✓ 转换后记录数: {len(records)} 条", flush=True)
            
            # 显示第一条记录示例（用于调试）
            if len(records) > 0:
                print(f"\n  第一条记录示例:", flush=True)
                print(f"    {json.dumps(records[0], ensure_ascii=False, indent=4, default=str)}", flush=True)
            
            # 导入到表
            print(f"  → 检查表是否存在...", flush=True)
            inspector = inspect(db.bind)
            if table_name not in inspector.get_table_names():
                print(f"  ❌ 表 {table_name} 不存在！", flush=True)
                print(f"  💡 请先运行迁移脚本创建表:", flush=True)
                print(f"     cd backend", flush=True)
                print(f"     alembic upgrade head", flush=True)
                wb.close()
                return
            
            print(f"  ✓ 表 {table_name} 存在", flush=True)
            print(f"  → 开始导入 {len(records)} 条记录...", flush=True)
            
            importer = SheetTableImporter(db)
            import_result = importer.import_to_table(
                table_name=table_name,
                records=records,
                unique_key=unique_key
            )
            
            inserted = import_result.get("inserted", 0)
            updated = import_result.get("updated", 0)
            errors = import_result.get("errors", 0)
            
            print(f"  ✓ 导入完成:", flush=True)
            print(f"    - 插入: {inserted} 条", flush=True)
            print(f"    - 更新: {updated} 条", flush=True)
            print(f"    - 错误: {errors} 条", flush=True)
            
            # ========== 验证：同时导入到fact_observation ==========
            print(f"\n  → 验证：同时导入到fact_observation...", flush=True)
            try:
                obs_result = upsert_observations(db, valid_observations, batch_id=batch.id, sheet_name=sheet_name)
                obs_inserted = obs_result.get("inserted", 0)
                obs_updated = obs_result.get("updated", 0)
                obs_errors = obs_result.get("errors", 0)
                print(f"  ✓ fact_observation导入完成:", flush=True)
                print(f"    - 插入: {obs_inserted} 条", flush=True)
                print(f"    - 更新: {obs_updated} 条", flush=True)
                print(f"    - 错误: {obs_errors} 条", flush=True)
                
                # 验证metric_key是否正确设置
                if obs_inserted > 0 or obs_updated > 0:
                    from app.models.dim_metric import DimMetric
                    from sqlalchemy import func
                    # 检查是否有metric设置了metric_key
                    metrics_with_key = db.query(DimMetric).filter(
                        func.json_unquote(
                            func.json_extract(DimMetric.parse_json, '$.metric_key')
                        ).isnot(None)
                    ).count()
                    print(f"  ✓ 已设置metric_key的metric数量: {metrics_with_key}", flush=True)
                    
                    # 检查当前sheet相关的metric
                    sheet_metrics = db.query(DimMetric).filter(
                        DimMetric.sheet_name == sheet_name
                    ).all()
                    if sheet_metrics:
                        print(f"  ✓ 当前sheet的metric数量: {len(sheet_metrics)}", flush=True)
                        for metric in sheet_metrics[:3]:  # 只显示前3个
                            metric_key = None
                            if metric.parse_json:
                                metric_key = metric.parse_json.get("metric_key")
                            print(f"    - {metric.metric_name}: metric_key={metric_key}", flush=True)
            except Exception as e:
                print(f"  ⚠️  导入到fact_observation失败: {e}", flush=True)
                import traceback
                traceback.print_exc()
            
            # 验证导入结果
            if inserted == 0 and updated == 0 and len(records) > 0:
                print(f"\n  ⚠️  警告: 没有数据被插入或更新", flush=True)
                print(f"  可能的原因:", flush=True)
                print(f"    1. 表结构不匹配", flush=True)
                print(f"    2. 唯一键冲突但更新失败", flush=True)
                print(f"    3. 数据转换有问题", flush=True)
                if len(records) > 0:
                    print(f"\n  第一条记录示例:", flush=True)
                    print(f"    {json.dumps(records[0], ensure_ascii=False, indent=4, default=str)}", flush=True)
        else:
            # 旧架构：导入到fact_observation
            print(f"  → 使用旧架构导入到fact_observation...", flush=True)
            
            result = upsert_observations(db, valid_observations, batch_id=batch.id, sheet_name=sheet_name)
            
            inserted = result.get("inserted", 0)
            updated = result.get("updated", 0)
            errors = result.get("errors", 0)
            
            print(f"  ✓ 导入完成:", flush=True)
            print(f"    - 插入: {inserted} 条", flush=True)
            print(f"    - 更新: {updated} 条", flush=True)
            print(f"    - 错误: {errors} 条", flush=True)
        
        # 6.6 更新batch状态
        batch.status = "completed"
        db.commit()
        
        print(f"\n  ✅ 入库完成！批次ID: {batch.id}", flush=True)
        
        wb.close()
        
    except Exception as e:
        print(f"  ❌ 测试失败: {e}", flush=True)
        import traceback
        traceback.print_exc()
        db.rollback()


def test_full_import(file_path: Path, db, uploader_id: int):
    """测试完整导入"""
    print(f"\n{'='*80}", flush=True)
    print(f"🚀 完整导入测试", flush=True)
    print(f"{'='*80}", flush=True)
    
    with open(file_path, 'rb') as f:
        file_content = f.read()
    
    print(f"  文件: {file_path.name}", flush=True)
    print(f"  大小: {len(file_content) / 1024 / 1024:.2f} MB", flush=True)
    
    result = unified_import(
        db=db,
        file_content=file_content,
        filename=file_path.name,
        uploader_id=uploader_id,
        dataset_type="YONGYI_WEEKLY",
        source_code="YONGYI"
    )
    
    print(f"\n  导入结果:", flush=True)
    print(f"    - 成功: {result.get('success')}", flush=True)
    print(f"    - 批次ID: {result.get('batch_id')}", flush=True)
    print(f"    - 插入数: {result.get('inserted', 0)}", flush=True)
    print(f"    - 更新数: {result.get('updated', 0)}", flush=True)
    print(f"    - 错误数: {len(result.get('errors', []))}", flush=True)
    
    if result.get('errors'):
        print(f"\n  前5个错误:", flush=True)
        for error in result.get('errors', [])[:5]:
            print(f"    - {error}", flush=True)


def main():
    """主函数"""
    import sys
    
    print(f"\n{'='*80}", flush=True)
    print(f"涌益周度数据导入测试脚本", flush=True)
    print(f"{'='*80}", flush=True)
    
    # 文件路径
    file_path = project_root / "docs" / "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx"
    profile_path = project_root / "docs" / "ingest_profile_yongyi_weekly_v1.json"
    
    if not file_path.exists():
        print(f"❌ 文件不存在: {file_path}", flush=True)
        return
    
    if not profile_path.exists():
        print(f"❌ 配置文件不存在: {profile_path}", flush=True)
        return
    
    # 数据库连接
    db = SessionLocal()
    try:
        # 获取或创建用户
        user = get_or_create_test_user(db)
        
        # 加载profile
        print(f"\n🔧 加载导入配置...", flush=True)
        profile = get_profile_by_dataset_type(db, "YONGYI_WEEKLY")
        if not profile:
            print(f"  ⚠️  数据库中未找到profile，从JSON文件加载...", flush=True)
            profile = load_profile_from_json(db, profile_path)
            if not profile:
                print(f"  ❌ 加载profile失败", flush=True)
                return
        else:
            print(f"  ✓ Profile已存在: {profile.profile_code}", flush=True)
            # 检查是否需要重新加载配置（检查是否有sheet缺少table_config）
            need_reload = False
            if profile.sheets:
                sheets_without_table_config = [
                    s.sheet_name for s in profile.sheets 
                    if s.config_json and not s.config_json.get("table_config") and s.parser not in ['RAW_TABLE_STORE_ONLY', 'SKIP_META']
                ]
                if sheets_without_table_config:
                    print(f"  ⚠️  发现 {len(sheets_without_table_config)} 个sheet缺少table_config配置", flush=True)
                    print(f"  💡 提示: 配置文件可能已更新，建议重新加载配置", flush=True)
                    print(f"     运行: python scripts/load_ingest_profiles.py", flush=True)
        
        print(f"  ✓ Profile已加载: {profile.profile_code} ({len(profile.sheets) if profile.sheets else 0} sheets)", flush=True)
        
        # 检查命令行参数
        if len(sys.argv) > 1:
            # 命令行模式：直接测试指定的sheet
            sheet_name_arg = sys.argv[1]
            # 检查是否有自动入库参数（可能在同一个参数中，用空格分隔）
            auto_import = False
            if len(sys.argv) > 2:
                auto_flag = sys.argv[2].lower()
                if auto_flag in ['-y', '--yes', '--auto', 'y', '是']:
                    auto_import = True
            elif ' ' in sheet_name_arg:
                # 检查是否在sheet名称参数中包含了自动入库标志
                parts = sheet_name_arg.split(' ', 1)
                sheet_name_arg = parts[0]
                if len(parts) > 1 and parts[1].lower() in ['-y', '--yes', '--auto', 'y', '是']:
                    auto_import = True
            
            # 尝试从Excel文件中查找匹配的sheet名称（处理编码问题）
            wb_temp = load_workbook(file_path, data_only=True)
            sheet_name = None
            # 精确匹配
            if sheet_name_arg in wb_temp.sheetnames:
                sheet_name = sheet_name_arg
            else:
                # 模糊匹配（处理编码问题）
                for actual_sheet_name in wb_temp.sheetnames:
                    if sheet_name_arg in actual_sheet_name or actual_sheet_name in sheet_name_arg:
                        sheet_name = actual_sheet_name
                        break
                if not sheet_name:
                    print(f"  ❌ 未找到匹配的sheet: '{sheet_name_arg}'", flush=True)
                    print(f"  可用sheets（前10个）:", flush=True)
                    for i, name in enumerate(wb_temp.sheetnames[:10], 1):
                        print(f"    {i:2d}. {name}", flush=True)
                    wb_temp.close()
                    return
            wb_temp.close()
            
            print(f"\n📋 命令行模式: 测试sheet '{sheet_name}'", flush=True)
            if auto_import:
                print(f"  ✓ 自动入库模式已启用", flush=True)
            test_single_sheet(file_path, sheet_name, db, profile, user.id, auto_import=auto_import)
        else:
            # 交互式选择测试模式
            print(f"\n{'='*80}", flush=True)
            print(f"请选择测试模式:", flush=True)
            print(f"  1. 测试单个sheet（交互式）", flush=True)
            print(f"  2. 测试完整导入", flush=True)
            print(f"  3. 列出所有sheets", flush=True)
            print(f"{'='*80}", flush=True)
            
            choice = input("请输入选项 (1/2/3): ").strip()
            
            if choice == "1":
                # 列出所有sheets
                wb = load_workbook(file_path, data_only=True)
                print(f"\n可用sheets:", flush=True)
                for i, sheet_name in enumerate(wb.sheetnames, 1):
                    print(f"  {i:2d}. {sheet_name}", flush=True)
                wb.close()
                
                sheet_input = input("\n请输入sheet名称或序号: ").strip()
                
                # 尝试按序号查找
                try:
                    sheet_idx = int(sheet_input) - 1
                    wb = load_workbook(file_path, data_only=True)
                    if 0 <= sheet_idx < len(wb.sheetnames):
                        sheet_name = wb.sheetnames[sheet_idx]
                    else:
                        sheet_name = sheet_input
                    wb.close()
                except ValueError:
                    sheet_name = sheet_input
                
                test_single_sheet(file_path, sheet_name, db, profile, user.id, auto_import=False)
                
            elif choice == "2":
                test_full_import(file_path, db, user.id)
                
            elif choice == "3":
                wb = load_workbook(file_path, data_only=True)
                print(f"\n所有sheets ({len(wb.sheetnames)}个):", flush=True)
                for i, sheet_name in enumerate(wb.sheetnames, 1):
                    print(f"  {i:2d}. {sheet_name}", flush=True)
                wb.close()
            else:
                print(f"❌ 无效选项", flush=True)
    
    finally:
        db.close()


if __name__ == "__main__":
    main()
