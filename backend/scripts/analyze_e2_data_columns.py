"""
详细分析E2. 多渠道汇总所需的数据列
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import openpyxl
from datetime import datetime

script_dir = Path(__file__).parent.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_columns():
    """分析所需的数据列"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("E2. 多渠道汇总数据列分析")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 1. NYB数据 - 能繁母猪存栏环比、新生仔猪存栏环比、生猪存栏环比
    print("\n1. NYB sheet - 能繁母猪存栏环比、新生仔猪存栏环比、生猪存栏环比")
    if "NYB" in wb.sheetnames:
        ws = wb["NYB"]
        print("  表头行（第2行）:")
        header_row = list(ws.iter_rows(max_row=2, values_only=True))[1]
        for col_idx, cell in enumerate(header_row[:30], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"    {col_letter}列: {cell}")
        
        print("\n  数据示例（前5行）:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=7, values_only=True), 3):
            date_val = row[1] if len(row) > 1 else None  # B列是日期
            print(f"    行{row_idx}: 日期={date_val}")
            # C列：能繁环比-全国, D列：能繁环比-规模场, E列：能繁环比-CR26, F列：能繁环比-小散户
            # G列：新生仔猪环比-全国, H列：新生仔猪环比-规模场, I列：新生仔猪环比-小散户
            # Q列：存栏环比
            if len(row) > 2:
                print(f"      C列(能繁-全国): {row[2]}, D列(能繁-规模场): {row[3]}, E列(能繁-CR26): {row[4]}, F列(能繁-小散户): {row[5]}")
            if len(row) > 6:
                print(f"      G列(新生仔猪-全国): {row[6]}, H列(新生仔猪-规模场): {row[7]}, I列(新生仔猪-小散户): {row[8]}")
            if len(row) > 16:
                print(f"      Q列(存栏环比): {row[16]}")
    
    # 2. 协会猪料 - 能繁母猪饲料环比、仔猪饲料环比、育肥猪饲料环比
    print("\n\n2. 02.协会猪料 sheet - 能繁母猪饲料环比、仔猪饲料环比、育肥猪饲料环比")
    if "02.协会猪料" in wb.sheetnames:
        ws = wb["02.协会猪料"]
        print("  表头行（第1行）:")
        header_row = list(ws.iter_rows(max_row=1, values_only=True))[0]
        for col_idx, cell in enumerate(header_row[:30], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"    {col_letter}列: {cell}")
        
        print("\n  数据示例（前5行）:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
            date_val = row[1] if len(row) > 1 else None  # B列是日期
            print(f"    行{row_idx}: 日期={date_val}")
            # H列：母猪料产量, I列：母猪料环比
            # M列：仔猪料产量, N列：仔猪料环比
            # R列：育肥料产量, S列：育肥料环比
            if len(row) > 7:
                print(f"      H列(母猪料产量): {row[7]}, I列(母猪料环比): {row[8]}")
            if len(row) > 12:
                print(f"      M列(仔猪料产量): {row[12]}, N列(仔猪料环比): {row[13]}")
            if len(row) > 17:
                print(f"      R列(育肥料产量): {row[17]}, S列(育肥料环比): {row[18]}")
    
    # 3. 涌益数据 - 淘汰母猪屠宰环比、能繁母猪存栏环比、能繁母猪饲料环比、新生仔猪存栏环比、仔猪饲料环比、生猪存栏环比、育肥猪饲料环比
    print("\n\n3. 涌益 sheet - 涌益相关数据")
    if "涌益" in wb.sheetnames:
        ws = wb["涌益"]
        print("  表头行（第2行）:")
        header_row = list(ws.iter_rows(max_row=2, values_only=True))[1]
        for col_idx, cell in enumerate(header_row[:30], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"    {col_letter}列: {cell}")
    
    # 4. 钢联数据 - 能繁母猪存栏环比、新生仔猪存栏环比、生猪存栏环比、能繁母猪饲料环比、仔猪饲料环比、育肥猪饲料环比
    print("\n\n4. 4.1.钢联数据 sheet - 钢联相关数据")
    if "4.1.钢联数据" in wb.sheetnames:
        ws = wb["4.1.钢联数据"]
        print("  表头行（第1行）:")
        header_row = list(ws.iter_rows(max_row=1, values_only=True))[0]
        for col_idx, cell in enumerate(header_row[:30], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"    {col_letter}列: {cell}")
        
        print("\n  数据示例（前5行）:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
            date_val = row[0] if len(row) > 0 else None  # A列是日期
            print(f"    行{row_idx}: 日期={date_val}")
            # L列：能繁存栏环比, O列：新生仔猪数环比, S列：育肥料环比, W列：仔猪饲料环比, AA列：母猪料环比
    
    # 5. 淘汰母猪屠宰 - 需要查找
    print("\n\n5. 查找淘汰母猪屠宰数据")
    if "涌益样本" in wb.sheetnames:
        ws = wb["涌益样本"]
        print("  表头行（第2行）:")
        header_row = list(ws.iter_rows(max_row=2, values_only=True))[1]
        for col_idx, cell in enumerate(header_row[:30], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if "淘汰" in str(cell) or "屠宰" in str(cell):
                    print(f"    {col_letter}列: {cell}")
    
    wb.close()
    
    print("\n\n" + "=" * 80)
    print("数据列映射总结")
    print("=" * 80)
    print("""
表格1：淘汰母猪屠宰环比、能繁母猪存栏环比、能繁母猪饲料环比
  - 淘汰母猪屠宰环比-涌益: 涌益样本 sheet, V列（淘汰母猪屠宰量-涌益）
  - 淘汰母猪屠宰环比-钢联: 需要查找
  - 能繁母猪存栏环比-涌益: 涌益 sheet, E列（能繁母猪）, F列（环比）
  - 能繁母猪存栏环比-钢联-全国: 4.1.钢联数据 sheet, L列（能繁存栏环比）- 中国
  - 能繁母猪存栏环比-钢联-规模场: 4.1.钢联数据 sheet, L列（能繁存栏环比）- 规模场
  - 能繁母猪存栏环比-钢联-中小散户: 4.1.钢联数据 sheet, L列（能繁存栏环比）- 中小散户
  - 能繁母猪存栏环比-NYB: NYB sheet, C列（能繁环比-全国）
  - 能繁母猪饲料环比-涌益: 需要查找
  - 能繁母猪饲料环比-钢联: 4.1.钢联数据 sheet, AA列（母猪料环比）
  - 能繁母猪饲料环比-协会: 02.协会猪料 sheet, I列（母猪料环比）

表格2：新生仔猪存栏环比、仔猪饲料环比
  - 新生仔猪存栏环比-涌益: 涌益 sheet, L列（新生仔猪）, M列（环比）
  - 新生仔猪存栏环比-钢联-全国: 4.1.钢联数据 sheet, O列（新生仔猪数环比）- 中国
  - 新生仔猪存栏环比-钢联-规模场: 4.1.钢联数据 sheet, O列（新生仔猪数环比）- 规模场
  - 新生仔猪存栏环比-钢联-中小散户: 4.1.钢联数据 sheet, O列（新生仔猪数环比）- 中小散户
  - 新生仔猪存栏环比-NYB: NYB sheet, G列（新生仔猪环比-全国）
  - 仔猪饲料环比-涌益: 需要查找
  - 仔猪饲料环比-钢联: 4.1.钢联数据 sheet, W列（仔猪饲料环比）
  - 仔猪饲料环比-协会: 02.协会猪料 sheet, N列（仔猪料环比）

表格3：生猪存栏环比、育肥猪饲料环比
  - 生猪存栏环比-涌益: 涌益 sheet, W列（存栏）, X列（环比）
  - 生猪存栏环比-钢联-全国: 4.1.钢联数据 sheet, B列（存栏环比：中小散户）- 中国
  - 生猪存栏环比-钢联-规模场: 需要查找
  - 生猪存栏环比-钢联-中小散户: 4.1.钢联数据 sheet, B列（存栏环比：中小散户）
  - 生猪存栏环比-NYB: NYB sheet, Q列（存栏环比）
  - 生猪存栏环比-NYB-5月龄: NYB sheet, J列（5月龄及以上大猪环比）
  - 育肥猪饲料环比-涌益: 需要查找
  - 育肥猪饲料环比-钢联: 4.1.钢联数据 sheet, S列（育肥料环比）
  - 育肥猪饲料环比-协会: 02.协会猪料 sheet, S列（育肥料环比）
    """)

if __name__ == "__main__":
    analyze_columns()
