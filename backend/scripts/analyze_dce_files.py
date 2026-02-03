"""分析DCE期货和期权文件结构"""
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
import json

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def analyze_dce_file(file_path, file_name):
    """分析DCE文件结构"""
    print(f"\n{'='*80}")
    print(f"分析文件: {file_name}")
    print(f"路径: {file_path}")
    print(f"{'='*80}")
    
    if not file_path.exists():
        print(f"❌ 文件不存在: {file_path}")
        return None
    
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        
        print(f"\n找到 {len(sheet_names)} 个sheet:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")
        
        results = {}
        for sheet_name in sheet_names:
            print(f"\n分析sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # 读取前20行和前30列
            header_rows = []
            for row_idx in range(1, min(21, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(31, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    if value:
                        str_val = str(value)
                        if len(str_val) > 30:
                            str_val = str_val[:27] + "..."
                        row_data.append(str_val)
                    else:
                        row_data.append("")
                header_rows.append(row_data)
            
            # 识别表头行
            header_row_idx = None
            for i, row in enumerate(header_rows[:10], 1):
                non_empty = sum(1 for v in row if v)
                if non_empty > 3:
                    header_row_idx = i
                    break
            
            result = {
                "sheet_name": sheet_name,
                "total_rows": ws.max_row,
                "total_cols": ws.max_column,
                "header_row": header_row_idx,
                "header_sample": header_rows[header_row_idx-1][:15] if header_row_idx else [],
                "sample_rows": header_rows[:5]
            }
            
            results[sheet_name] = result
            
            print(f"  ✓ 行数: {result['total_rows']}, 列数: {result['total_cols']}")
            if header_row_idx:
                print(f"  ✓ 表头行: {header_row_idx}")
                print(f"  ✓ 表头样本: {result['header_sample'][:10]}")
        
        wb.close()
        
        return {
            "file_name": file_name,
            "file_path": str(file_path),
            "sheets": results
        }
    
    except Exception as e:
        print(f"❌ 分析文件失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """主函数"""
    print("="*80)
    print("DCE文件结构分析工具")
    print("="*80)
    
    files_to_analyze = [
        {
            "name": "DCE期货",
            "path": project_root / "docs" / "lh_ftr.xlsx"
        },
        {
            "name": "DCE期权",
            "path": project_root / "docs" / "lh_opt.xlsx"
        }
    ]
    
    all_results = []
    
    for file_info in files_to_analyze:
        result = analyze_dce_file(file_info['path'], file_info['name'])
        if result:
            all_results.append(result)
    
    # 保存结果
    output_file = project_root / "docs" / "dce_files_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n{'='*80}")
    print(f"分析完成！结果已保存到: {output_file}")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
