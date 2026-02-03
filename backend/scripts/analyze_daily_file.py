"""分析日度数据文件结构"""
import sys
import os
import io
from pathlib import Path

# Set UTF-8 encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pandas as pd
from openpyxl import load_workbook

file_path = Path("d:/Workspace/hogprice-insight/docs/2026年2月2日涌益咨询日度数据.xlsx")

print("="*60)
print("Analyzing Daily Data File")
print("="*60)
print(f"File: {file_path.name}")
print(f"Size: {file_path.stat().st_size / 1024 / 1024:.2f} MB")
print()

wb = load_workbook(file_path, data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print()

for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    row_count = ws.max_row
    col_count = ws.max_column
    
    # 估算JSON大小（稀疏格式，只存非空单元格）
    non_empty_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(100, row_count), values_only=False):
        for cell in row:
            if cell.value is not None:
                non_empty_count += 1
    
    # 估算：每个非空单元格约50字节（稀疏格式）
    estimated_size = non_empty_count * 50 / 1024 / 1024  # MB
    
    print(f"{i}. {sheet_name}")
    print(f"   Rows: {row_count}, Cols: {col_count}")
    print(f"   Estimated JSON size (first 100 rows): {estimated_size:.2f} MB")
    if row_count > 100:
        print(f"   Full size estimate: {estimated_size * (row_count / 100):.2f} MB")
    print()

wb.close()
