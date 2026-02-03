"""快速分析涌益周度数据的所有sheet名称和基本结构"""
import sys
from pathlib import Path
from openpyxl import load_workbook
import json

project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def quick_analyze_sheet(file_path, sheet_name):
    """快速分析sheet的基本结构"""
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        
        # 只读取前5行和前20列
        header_sample = []
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(21, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if value:
                    str_val = str(value)
                    if len(str_val) > 25:
                        str_val = str_val[:22] + "..."
                    row_data.append(str_val)
                else:
                    row_data.append("")
            header_sample.append(row_data)
        
        wb.close()
        
        return {
            "sheet_name": sheet_name,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "header_sample": header_sample
        }
    except Exception as e:
        return {
            "sheet_name": sheet_name,
            "error": str(e)
        }


def main():
    file_path = project_root / "docs" / "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx"
    
    print("快速分析涌益周度数据...")
    print(f"文件: {file_path}")
    
    if not file_path.exists():
        print(f"❌ 文件不存在")
        return
    
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    print(f"\n找到 {len(sheet_names)} 个sheet\n")
    
    # 过滤掉说明性sheet
    data_sheets = []
    skip_keywords = ["样本点", "目录", "数据说明", "历史", "料肉比", "费用", "成本计算附件", "新2022"]
    
    for name in sheet_names:
        if not any(keyword in name for keyword in skip_keywords):
            data_sheets.append(name)
    
    print(f"数据sheet数量: {len(data_sheets)}")
    print("\n数据sheet列表:")
    for i, name in enumerate(data_sheets, 1):
        print(f"  {i:2d}. {name}")
    
    print("\n开始快速分析...")
    results = {}
    
    # 只分析数据sheet
    for sheet_name in data_sheets[:30]:  # 限制前30个避免超时
        print(f"分析: {sheet_name}")
        result = quick_analyze_sheet(file_path, sheet_name)
        results[sheet_name] = result
    
    # 保存结果
    output_file = project_root / "docs" / "yongyi_weekly_quick_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            "all_sheets": sheet_names,
            "data_sheets": data_sheets,
            "analysis": results
        }, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n分析完成！结果已保存到: {output_file}")


if __name__ == "__main__":
    main()
