"""
分析【生猪产业数据】.xlsx文件结构
用于E2. 多渠道汇总数据导入
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import json
import openpyxl

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_excel():
    """分析Excel文件结构"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        print("\n请确认文件路径是否正确")
        return
    
    print("=" * 80)
    print(f"分析文件: {excel_path.name}")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"\nSheet列表 ({len(wb.sheetnames)} 个):")
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")
        
        # 分析每个sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'=' * 80}")
            print(f"Sheet: {sheet_name}")
            print(f"{'=' * 80}")
            
            ws = wb[sheet_name]
            
            # 显示前10行数据
            print(f"\n前10行数据:")
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                row_data = []
                for col_idx, cell in enumerate(row[:30], 1):  # 只显示前30列
                    if cell is not None:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        row_data.append(f"{col_letter}:{cell}")
                if row_data:
                    print(f"  行{row_idx}: {', '.join(row_data[:10])}...")
            
            # 分析列结构
            print(f"\n列结构分析（前30列）:")
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                # 查找可能的表头行（包含"日期"、"月份"等关键词）
                row_str = ' '.join([str(cell) for cell in row[:30] if cell])
                if any(keyword in row_str for keyword in ['日期', '月份', '时间', 'Date', 'Month']):
                    header_row = row
                    print(f"  可能的表头行: 第{row_idx}行")
                    for col_idx, cell in enumerate(header_row[:30], 1):
                        if cell:
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            print(f"    {col_letter}列: {cell}")
                    break
            
            # 统计数据行数
            data_row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row[:10]):
                    data_row_count += 1
            
            print(f"\n数据行数（估算）: {data_row_count}")
        
        wb.close()
        
    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel()
