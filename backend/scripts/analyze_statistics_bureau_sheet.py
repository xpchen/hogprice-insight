"""
分析【生猪产业数据】.xlsx中的"03.统计局季度数据"sheet
用于E4. 统计局数据汇总
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import json
import openpyxl

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_statistics_bureau_sheet():
    """分析统计局季度数据sheet结构"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("分析【03.统计局季度数据】sheet结构")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        if "03.统计局季度数据" not in wb.sheetnames:
            print("未找到'03.统计局季度数据'sheet")
            print(f"\n可用的sheet列表:")
            for idx, sheet_name in enumerate(wb.sheetnames, 1):
                print(f"  {idx}. {sheet_name}")
            wb.close()
            return
        
        ws = wb["03.统计局季度数据"]
        
        print(f"\nSheet: 03.统计局季度数据")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
        # 显示前20行数据
        print(f"\n前20行数据:")
        for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
            row_data = []
            for col_idx, cell in enumerate(row[:30], 1):  # 显示前30列
                if cell is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_str = str(cell)
                    if len(cell_str) > 30:
                        cell_str = cell_str[:30] + "..."
                    row_data.append(f"{col_letter}:{cell_str}")
            if row_data:
                print(f"  行{row_idx}: {', '.join(row_data[:10])}...")
        
        # 分析表头行
        print(f"\n分析表头行:")
        # 查找可能的表头行
        for row_idx in range(1, min(5, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            row_str = ' '.join([str(cell) for cell in row[:30] if cell])
            
            # 检查是否包含季度、能繁、存栏等关键词
            keywords = ['季度', '能繁', '存栏', '出栏', '屠宰', '产量', '进口']
            if any(keyword in row_str for keyword in keywords):
                print(f"\n  可能的表头行: 第{row_idx}行")
                print(f"  前30列内容:")
                for col_idx, cell in enumerate(row[:30], 1):
                    if cell:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        print(f"    {col_letter}列 ({col_idx}): {cell}")
                break
        
        # 分析数据行结构
        print(f"\n分析数据行结构（第3-15行）:")
        for row_idx in range(3, min(16, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            
            # A列（索引0）应该是季度
            quarter = row[0] if len(row) > 0 else None
            
            # B-Y列（索引1-24）是数据列
            data_cols = []
            for col_idx in range(1, min(25, len(row))):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)  # B列是索引1
                value = row[col_idx]
                if value is not None:
                    data_cols.append(f"{col_letter}:{value}")
            
            if quarter:
                print(f"  行{row_idx} - 季度:{quarter}, 数据列数:{len(data_cols)}")
                if data_cols:
                    print(f"    前5列: {', '.join(data_cols[:5])}")
        
        # 统计有效数据行数
        print(f"\n统计有效数据行数:")
        data_row_count = 0
        for row_idx in range(2, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            # 检查A列是否有季度标识
            if len(row) > 0 and row[0] is not None:
                data_row_count += 1
        
        print(f"  有效数据行数: {data_row_count}")
        
        wb.close()
        
    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_statistics_bureau_sheet()
