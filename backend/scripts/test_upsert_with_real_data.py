"""使用实际导入的数据测试upsert_observations"""
import sys
import os
import io
import json
from pathlib import Path
from openpyxl import load_workbook

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 添加backend目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.core.database import SessionLocal
from app.models.import_batch import ImportBatch
from app.models.sys_user import SysUser
from app.services.ingestors.observation_upserter import upsert_observations
from app.services.ingestors.dispatcher import create_dispatcher
from app.services.ingestors.parsers import get_parser
from app.services.ingestors.validator import ObservationValidator
from app.services.ingestors.error_collector import ErrorCollector


def test_with_real_data(file_path: str, sheet_name: str = "周度-体重", limit: int = 100):
    """使用实际数据测试"""
    db = SessionLocal()
    try:
        print("=" * 80)
        print(f"使用实际数据测试 upsert_observations")
        print(f"文件: {file_path}")
        print(f"Sheet: {sheet_name}")
        print(f"限制: {limit} 条")
        print("=" * 80)
        
        # 1. 获取用户
        user = db.query(SysUser).filter(SysUser.username == "admin").first()
        if not user:
            print("  ❌ 未找到admin用户")
            return
        
        print(f"  ✓ 使用用户: {user.username} (ID: {user.id})")
        
        # 2. 加载profile
        print(f"\n  加载导入配置...")
        dispatcher = create_dispatcher(db, dataset_type="YONGYI_WEEKLY")
        profile = dispatcher.profile
        print(f"  ✓ Profile: {profile.profile_code}")
        
        # 3. 解析sheet
        print(f"\n  解析sheet '{sheet_name}'...")
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"  ❌ Sheet '{sheet_name}' 不存在")
            print(f"  可用sheets: {wb.sheetnames[:10]}")
            wb.close()
            return
        
        worksheet = wb[sheet_name]
        
        # 分派parser
        dispatch_result = dispatcher.dispatch_sheet(sheet_name, worksheet=worksheet)
        if dispatch_result["action"] != "PARSE":
            print(f"  ❌ 分派失败: {dispatch_result['action']}")
            wb.close()
            return
        
        parser_type = dispatch_result["parser"]
        sheet_config = dispatch_result["sheet_config"]
        
        print(f"  ✓ Parser: {parser_type}")
        print(f"  ✓ Sheet配置已加载")
        
        # 获取parser
        parser = get_parser(parser_type)
        if not parser:
            print(f"  ❌ 未找到parser: {parser_type}")
            wb.close()
            return
        
        # 解析数据
        observations = parser.parse(
            sheet_data=worksheet,
            sheet_config=sheet_config,
            profile_defaults=profile.defaults_json or {},
            source_code=profile.source_code,
            batch_id=0  # 临时batch_id
        )
        
        wb.close()
        
        print(f"  ✓ 解析完成: {len(observations)} 条观测值")
        
        if len(observations) == 0:
            print(f"  ⚠️  没有解析到数据")
            return
        
        # 限制数量
        if limit and len(observations) > limit:
            observations = observations[:limit]
            print(f"  ✓ 限制为前 {limit} 条")
        
        # 显示前3条
        print(f"\n  前3条观测值:")
        for i, obs in enumerate(observations[:3], 1):
            print(f"    [{i}] metric_key={obs.get('metric_key')}, geo_code={obs.get('geo_code')}, "
                  f"period_end={obs.get('period_end')}, value={obs.get('value')}")
            if obs.get('dedup_key'):
                print(f"        dedup_key={obs.get('dedup_key')[:32]}...")
            else:
                print(f"        ⚠️  dedup_key为空！")
        
        # 4. 验证数据
        print(f"\n  验证数据...")
        batch = ImportBatch(
            filename=Path(file_path).name,
            file_hash="test_hash",
            uploader_id=user.id,
            source_code=profile.source_code,
            status="processing",
            total_rows=0,
            success_rows=0,
            failed_rows=0
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)
        
        error_collector = ErrorCollector(db, batch.id)
        validator = ObservationValidator(error_collector)
        
        # 检查是否有table_config（新架构）
        table_config = sheet_config.get("table_config")
        skip_metric_check = bool(table_config)
        
        valid_observations = validator.validate_batch(
            observations,
            sheet_name,
            skip_metric_key_check=skip_metric_check
        )
        
        print(f"    总观测值: {len(observations)} 条")
        print(f"    有效观测值: {len(valid_observations)} 条")
        print(f"    无效观测值: {len(observations) - len(valid_observations)} 条")
        
        if len(valid_observations) == 0:
            print(f"  ❌ 没有有效观测值")
            if error_collector.errors:
                print(f"  前5个错误:")
                for error in error_collector.errors[:5]:
                    print(f"    - {error}")
            return
        
        # 5. 测试upsert
        print(f"\n  {'='*80}")
        print(f"  开始测试 upsert_observations...")
        print(f"  {'='*80}")
        
        result = upsert_observations(
            db=db,
            observations=valid_observations,
            batch_id=batch.id,
            sheet_name=sheet_name
        )
        
        # 6. 显示结果
        print(f"\n  {'='*80}")
        print(f"  测试结果")
        print(f"  {'='*80}")
        print(f"    插入: {result.get('inserted', 0)} 条")
        print(f"    更新: {result.get('updated', 0)} 条")
        print(f"    错误: {result.get('errors', 0)} 条")
        
        if result.get('errors', 0) > 0:
            print(f"\n  ⚠️  有 {result.get('errors', 0)} 条错误")
            print(f"  请查看上方的详细错误日志")
        else:
            print(f"\n  ✅ 所有记录都成功处理！")
        
        print(f"\n  {'='*80}")
        print(f"  测试完成")
        print(f"  {'='*80}")
        
    except Exception as e:
        db.rollback()
        print(f"\n  ❌ 测试失败: {type(e).__name__}: {str(e)}", flush=True)
        import traceback
        print(f"  详细堆栈:\n{traceback.format_exc()}", flush=True)
    finally:
        db.close()


if __name__ == "__main__":
    # 查找测试文件
    project_root = Path(__file__).parent.parent.parent
    docs_dir = project_root / "docs"
    
    # 查找周度数据文件
    weekly_files = list(docs_dir.glob("*涌益咨询*周度数据*.xlsx"))
    
    if not weekly_files:
        print("❌ 未找到周度数据文件")
        print(f"  请在 {docs_dir} 目录下放置周度数据文件")
        sys.exit(1)
    
    file_path = str(weekly_files[0])
    print(f"使用文件: {file_path}\n")
    
    # 测试所有数据（不限制）
    import sys
    limit = None
    if len(sys.argv) > 1:
        try:
            limit = int(sys.argv[1])
        except:
            pass
    
    test_with_real_data(file_path, sheet_name="周度-体重", limit=limit)
