"""测试钢联（Ganglian）数据导入功能"""
import sys
import os
import io
import json
from pathlib import Path
from datetime import datetime
from sqlalchemy import text, inspect
from openpyxl import load_workbook

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 添加backend目录到路径（确保能导入app模块）
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.core.database import SessionLocal
from app.models.sys_user import SysUser
from app.models.dim_source import DimSource
from app.services.ingestors.unified_ingestor import unified_import
from app.services.sheet_table_mapper import SheetTableMapper
from app.services.ingestors.profile_loader import load_profile_from_json, get_profile_by_dataset_type

# 项目根目录
backend_dir = Path(__file__).parent.parent
project_root = backend_dir.parent


def get_or_create_test_user(db):
    """获取或创建测试用户"""
    user = db.query(SysUser).filter(SysUser.username == "admin").first()
    if not user:
        from app.core.security import get_password_hash
        user = SysUser(
            username="test_admin",
            password_hash=get_password_hash("test123456"),
            display_name="测试管理员",
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        print(f"  ✓ 创建测试用户: {user.username} (ID: {user.id})")
    else:
        print(f"  ✓ 使用现有用户: {user.username} (ID: {user.id})")
    return user


def get_table_record_count(db, table_name: str) -> int:
    """获取表中的记录数"""
    try:
        inspector = inspect(db.bind)
        if table_name not in inspector.get_table_names():
            return -1  # 表不存在
        
        result = db.execute(text(f"SELECT COUNT(*) as cnt FROM `{table_name}`"))
        row = result.fetchone()
        return row[0] if row else 0
    except Exception as e:
        print(f"    ⚠️  查询表 {table_name} 记录数失败: {e}")
        return -1


def get_sheet_table_mapping_from_profile(profile_path: Path) -> dict:
    """从ingest_profile配置文件中读取sheet到表的映射"""
    sheet_mapping = {}
    
    try:
        with open(profile_path, 'r', encoding='utf-8') as f:
            profile = json.load(f)
        
        for sheet_config in profile.get('sheets', []):
            sheet_name = sheet_config.get('sheet_name')
            table_name = sheet_config.get('table_name')
            
            if sheet_name and table_name:
                sheet_mapping[sheet_name] = {
                    "table_name": table_name,
                    "parser": sheet_config.get('parser'),
                    "has_table_config": bool(sheet_config.get('table_config'))
                }
    except Exception as e:
        print(f"    ⚠️  读取配置文件失败: {e}")
    
    return sheet_mapping


def get_sheet_tables_from_excel(file_path: Path, source_code: str, dataset_type: str, 
                                profile_path: Path = None) -> dict:
    """从Excel文件读取sheet名称，并获取对应的表名和记录数"""
    sheet_tables = {}
    
    # 优先从profile配置文件读取映射
    sheet_mapping = {}
    if profile_path and profile_path.exists():
        sheet_mapping = get_sheet_table_mapping_from_profile(profile_path)
    
    try:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        db = SessionLocal()
        try:
            for sheet_name in sheet_names:
                # 优先使用配置文件中的映射
                if sheet_name in sheet_mapping:
                    table_name = sheet_mapping[sheet_name]["table_name"]
                else:
                    # 否则使用mapper生成表名
                    table_name = SheetTableMapper.sheet_name_to_table_name(
                        sheet_name=sheet_name,
                        source_code=source_code,
                        dataset_type=dataset_type
                    )
                
                # 获取表记录数
                count = get_table_record_count(db, table_name)
                sheet_tables[sheet_name] = {
                    "table_name": table_name,
                    "count_before": count,
                    "count_after": count  # 将在导入后更新
                }
        finally:
            db.close()
    except Exception as e:
        print(f"    ⚠️  读取Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
    
    return sheet_tables


def print_sheet_status(sheet_name: str, table_name: str, count_before: int, 
                      count_after: int, status: str = "完成"):
    """打印sheet导入状态"""
    print(f"\n  ┌─ Sheet: {sheet_name}")
    print(f"  │  表名: {table_name}")
    print(f"  │  导入前记录数: {count_before if count_before >= 0 else '表不存在'}")
    print(f"  │  导入后记录数: {count_after if count_after >= 0 else '表不存在'}")
    if count_before >= 0 and count_after >= 0:
        inserted = count_after - count_before
        print(f"  │  新增记录数: {inserted}")
        if inserted > 0:
            status = "✅ 成功"
        elif count_after > 0:
            status = "⚠️  无新增（可能已存在）"
        else:
            status = "⚠️  无数据"
    elif count_after >= 0:
        status = "✅ 成功（新表）"
    else:
        status = "❌ 表不存在"
    print(f"  │  状态: {status}")
    print(f"  └─")


def test_ganglian_import(file_path: Path, profile_path: Path = None):
    """测试导入钢联数据文件"""
    print(f"\n{'='*80}")
    print(f"测试导入钢联数据文件")
    print(f"文件: {file_path.name}")
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}")
    
    if not file_path.exists():
        print(f"❌ 文件不存在: {file_path}")
        return None
    
    # 读取文件
    with open(file_path, 'rb') as f:
        file_content = f.read()
    
    print(f"文件大小: {len(file_content) / 1024 / 1024:.2f} MB")
    
    # 获取数据库连接
    db = SessionLocal()
    try:
        # 检查并创建数据源（如果不存在）
        print(f"\n🔧 检查数据源...")
        source = db.query(DimSource).filter(DimSource.source_code == "GANGLIAN").first()
        if not source:
            print(f"  ⚠️  数据源 'GANGLIAN' 不存在，正在创建...")
            source = DimSource(
                source_code="GANGLIAN",
                source_name="钢联数据",
                update_freq="daily",
                source_type="vendor",
                license_note="钢联数据，需授权使用"
            )
            db.add(source)
            db.commit()
            print(f"  ✓ 成功创建数据源: GANGLIAN")
        else:
            print(f"  ✓ 数据源已存在: {source.source_code} - {source.source_name}")
        
        # 检查并加载profile（如果不存在）
        print(f"\n🔧 检查导入配置...")
        profile = get_profile_by_dataset_type(db, "GANGLIAN_DAILY")
        if not profile:
            print(f"  ⚠️  Profile不存在，尝试加载...")
            if profile_path and profile_path.exists():
                try:
                    profile = load_profile_from_json(db, str(profile_path))
                    print(f"  ✓ 成功加载Profile: {profile.profile_code} ({len(profile.sheets)} sheets)")
                except Exception as e:
                    print(f"  ❌ 加载Profile失败: {e}")
                    import traceback
                    traceback.print_exc()
                    return None
            else:
                # 使用默认配置文件路径
                default_profile = project_root / "docs" / "ingest_profile_ganglian_daily_v1.json"
                if default_profile.exists():
                    try:
                        profile = load_profile_from_json(db, str(default_profile))
                        print(f"  ✓ 成功加载Profile: {profile.profile_code} ({len(profile.sheets)} sheets)")
                    except Exception as e:
                        print(f"  ❌ 加载Profile失败: {e}")
                        import traceback
                        traceback.print_exc()
                        return None
                else:
                    print(f"  ❌ Profile不存在且未找到默认配置文件")
                    print(f"  请提供配置文件路径作为第二个参数")
                    return None
        else:
            print(f"  ✓ Profile已存在: {profile.profile_code} ({len(profile.sheets)} sheets)")
        
        # 获取或创建测试用户
        user = get_or_create_test_user(db)
        
        # 导入前：获取所有sheet对应的表记录数
        print(f"\n📊 导入前状态检查...")
        sheet_tables = get_sheet_tables_from_excel(
            file_path=file_path,
            source_code="GANGLIAN",
            dataset_type="GANGLIAN_DAILY",
            profile_path=profile_path
        )
        
        if not sheet_tables:
            print("  ⚠️  未能读取到sheet信息")
            return None
        
        print(f"  找到 {len(sheet_tables)} 个sheet:")
        for sheet_name, info in sheet_tables.items():
            count = info["count_before"]
            table_name = info["table_name"]
            status = f"{count} 条记录" if count >= 0 else "表不存在"
            print(f"    - {sheet_name} -> {table_name}: {status}")
        
        # 执行导入
        print(f"\n🚀 开始导入...")
        result = unified_import(
            db=db,
            file_content=file_content,
            filename=file_path.name,
            uploader_id=user.id,
            dataset_type="GANGLIAN_DAILY",
            source_code="GANGLIAN"
        )
        
        batch_id = result.get('batch_id')
        if not batch_id:
            print(f"❌ 导入失败：未返回批次ID")
            if result.get('errors'):
                errors = result.get('errors', [])
                if isinstance(errors, list):
                    for error in errors[:5]:
                        print(f"  错误: {error}")
                else:
                    print(f"  错误数量: {errors}")
            return None
        
        print(f"\n✅ 导入完成，批次ID: {batch_id}")
        
        # 导入后：重新查询所有表的记录数
        print(f"\n📊 导入后状态检查...")
        
        # 更新每个sheet对应的表记录数
        for sheet_name in sheet_tables.keys():
            table_name = sheet_tables[sheet_name]["table_name"]
            count_after = get_table_record_count(db, table_name)
            sheet_tables[sheet_name]["count_after"] = count_after
        
        # 打印每个sheet的导入状态
        print(f"\n📋 Sheet导入详情:")
        print(f"{'─'*80}")
        
        for sheet_name, info in sheet_tables.items():
            table_name = info["table_name"]
            count_before = info["count_before"]
            count_after = info["count_after"]
            
            print_sheet_status(
                sheet_name=sheet_name,
                table_name=table_name,
                count_before=count_before,
                count_after=count_after
            )
        
        print(f"{'─'*80}")
        
        # 汇总结果
        print(f"\n📈 导入结果汇总:")
        print(f"  成功: {result.get('success', False)}")
        print(f"  插入: {result.get('inserted', 0)}")
        print(f"  更新: {result.get('updated', 0)}")
        print(f"  总sheet数: {result.get('total_sheets', 0)}")
        print(f"  解析sheet数: {result.get('parsed_sheets', 0)}")
        
        errors = result.get('errors', 0)
        if isinstance(errors, int):
            print(f"  错误数量: {errors}")
        elif isinstance(errors, list) and len(errors) > 0:
            print(f"  错误数量: {len(errors)}")
            print(f"\n⚠️  错误列表（前10个）:")
            for i, error in enumerate(errors[:10], 1):
                error_msg = error if isinstance(error, str) else str(error)
                print(f"    {i}. {error_msg}")
        
        print(f"  批次ID: {batch_id}")
        print(f"  结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 计算总新增记录数
        total_inserted = 0
        for info in sheet_tables.values():
            count_before = info["count_before"]
            count_after = info["count_after"]
            if count_before >= 0 and count_after >= 0:
                total_inserted += max(0, count_after - count_before)
        
        print(f"\n📊 表记录数统计:")
        print(f"  总新增记录数（按表统计）: {total_inserted}")
        
        return result
    
    except Exception as e:
        print(f"❌ 导入失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        db.close()


def main():
    """主函数"""
    print("="*80)
    print("钢联（Ganglian）数据导入测试")
    print("="*80)
    
    # 默认文件路径
    default_file = project_root / "docs" / "1、价格：钢联自动更新模板.xlsx"
    default_profile = project_root / "docs" / "ingest_profile_ganglian_daily_v1.json"
    
    # 检查文件是否存在
    if not default_file.exists():
        print(f"\n❌ 默认文件不存在: {default_file}")
        print(f"\n请提供文件路径作为命令行参数，例如:")
        print(f"  python {Path(__file__).name} <文件路径> [配置文件路径]")
        print(f"\n或者直接运行（使用默认路径）:")
        print(f"  python {Path(__file__).name}")
        return
    
    # 检查配置文件是否存在
    if not default_profile.exists():
        print(f"\n⚠️  配置文件不存在: {default_profile}")
        print(f"  将使用默认的sheet到表名映射")
        default_profile = None
    
    # 执行测试
    result = test_ganglian_import(
        file_path=default_file,
        profile_path=default_profile
    )
    
    if result:
        print(f"\n{'='*80}")
        print("测试完成")
        print(f"{'='*80}")
    else:
        print(f"\n{'='*80}")
        print("测试失败")
        print(f"{'='*80}")


if __name__ == "__main__":
    # 如果提供了命令行参数，使用指定的文件路径
    if len(sys.argv) > 1:
        file_path = Path(sys.argv[1])
        profile_path = None
        if len(sys.argv) > 2:
            profile_path = Path(sys.argv[2])
        
        result = test_ganglian_import(
            file_path=file_path,
            profile_path=profile_path
        )
    else:
        main()
