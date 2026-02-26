# -*- coding: utf-8 -*-
"""
Analyze A1 supply forecast sheet: AG column (slaughter), multi-header structure.
Output: structure analysis for supply-demand curve backend calculation.
"""
import sys
import io
from pathlib import Path

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def col_letter(idx: int) -> str:
    """0-based index to Excel column letter: 0=A, 26=AA, 32=AG"""
    result = ""
    idx += 1
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def analyze_from_db():
    """Analyze from raw_table (database)"""
    from sqlalchemy.orm import Session
    from app.core.database import SessionLocal
    from app.models.raw_sheet import RawSheet
    from app.models.raw_table import RawTable
    from app.models.raw_file import RawFile
    import json

    db = SessionLocal()
    try:
        sheet = db.query(RawSheet).join(RawFile).filter(
            RawSheet.sheet_name == "A1供给预测",
            RawFile.filename.like('%生猪产业数据%')
        ).first()
        if not sheet:
            print("A1 supply forecast sheet not found in DB")
            return

        raw = db.query(RawTable).filter(RawTable.raw_sheet_id == sheet.id).first()
        if not raw:
            print("Raw table not found")
            return

        data = raw.table_json
        if isinstance(data, str):
            data = json.loads(data)

        print("=" * 70)
        print("A1 supply forecast - multi-header analysis (from DB)")
        print("=" * 70)
        print("Total rows:", len(data))

        # AG = index 32
        ag_idx = 32
        # Show headers: rows 0, 1, 2 (maybe 3)
        for ri in range(min(5, len(data))):
            row = data[ri]
            ag_val = row[ag_idx] if len(row) > ag_idx else None
            # Context: cols 28-36 (AC to AI)
            ctx = []
            for j in range(max(0, ag_idx - 4), min(len(row), ag_idx + 5)):
                v = row[j] if j < len(row) else None
                ctx.append(f"{col_letter(j)}:{repr(v)[:30]}")
            print(f"\nRow {ri+1} (header?): AG={repr(ag_val)}")
            print("  Context [AC..AI]:", ctx)

        # Data rows: find where slaughter-like numbers (1509, 832, etc) appear
        data_start = 3
        if len(data) > 3:
            row2 = data[2] if len(data) > 2 else []
            row2_str = " ".join(str(c or "") for c in row2[:80])
            if "环比" in row2_str or "同比" in row2_str:
                data_start = 3
            else:
                data_start = 2

        print(f"\nData start row index: {data_start}")
        print("\nAG column (slaughter) sample - first 25 data rows:")
        for i in range(data_start, min(data_start + 25, len(data))):
            row = data[i]
            if len(row) <= ag_idx:
                continue
            date_val = row[1] if len(row) > 1 else None
            ag_val = row[ag_idx]
            print(f"  row{i+1}: date={date_val}, AG(slaughter)={ag_val}")

        # Find columns with numeric values like 1509, 832 (slaughter data)
        print("\n--- Searching for columns with slaughter-like numbers (800-2500) ---")
        sample_row_idx = data_start + 12 * 4  # around 2021
        for trial_row in range(data_start, min(data_start + 120, len(data))):
            row = data[trial_row]
            for j in range(min(50, len(row))):
                v = row[j]
                try:
                    if v is not None and v != "":
                        n = float(v)
                        if 500 < n < 3000 and n == int(n):
                            print(f"  Row {trial_row+1} col {col_letter(j)}(idx{j}): {n}")
                            break
                except:
                    pass
            else:
                continue
            break

        # Full row sample for 2021-01
        print("\n--- Full row sample (2021-01, cols 25-45) ---")
        for i in range(data_start, min(len(data), data_start + 60)):
            row = data[i]
            d = row[1] if len(row) > 1 else None
            if d and "2021-01" in str(d):
                ctx = [(col_letter(j), row[j] if j < len(row) else None) for j in range(25, min(46, len(row)))]
                print(f"  Row {i+1}: {ctx}")
                break

        # Year range and AG data availability
        from datetime import datetime
        years = set()
        ag_years = set()
        for i in range(data_start, len(data)):
            row = data[i]
            if len(row) < 2:
                continue
            d = row[1]
            ag_val = row[ag_idx] if len(row) > ag_idx else None
            if d:
                try:
                    if isinstance(d, str) and len(d) >= 4:
                        y = int(d[:4])
                        years.add(y)
                        if ag_val is not None and ag_val != "":
                            try:
                                float(ag_val)
                                ag_years.add(y)
                            except:
                                pass
                    elif hasattr(d, 'year'):
                        years.add(d.year)
                        if ag_val is not None and ag_val != "":
                            try:
                                float(ag_val)
                                ag_years.add(d.year)
                            except:
                                pass
                except:
                    pass
        print("\nYears in data:", sorted(years) if years else "none")
        print("Years with AG(slaughter) data:", sorted(ag_years) if ag_years else "none")

    finally:
        db.close()


def analyze_from_excel():
    """Analyze from Excel file directly"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed, skip Excel analysis")
        return

    excel_path = Path(script_dir).parent / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    if not excel_path.exists():
        print("Excel not found:", excel_path)
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "A1供给预测" not in wb.sheetnames:
        print("A1 sheet not found")
        wb.close()
        return

    ws = wb["A1供给预测"]
    print("\n" + "=" * 70)
    print("A1 supply forecast - from Excel (AG=col 33)")
    print("=" * 70)

    # AG = column 33 (1-based)
    ag_col = 33
    for ri in range(1, min(6, ws.max_row + 1)):
        v = ws.cell(ri, ag_col).value
        print(f"Row {ri}: {get_column_letter(ag_col)} = {repr(v)}")

    # Context: AC to AI (cols 29-35)
    print("\nHeader context (AC-AI, rows 1-3):")
    for ri in range(1, 4):
        vals = [ws.cell(ri, c).value for c in range(29, 36)]
        letters = [get_column_letter(c) for c in range(29, 36)]
        print(f"  Row {ri}: {list(zip(letters, vals))}")

    wb.close()


if __name__ == "__main__":
    analyze_from_db()
    analyze_from_excel()
