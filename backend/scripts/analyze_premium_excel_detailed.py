"""
详细分析升贴水Excel文件结构
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import pandas as pd
import openpyxl
from datetime import datetime

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_detailed():
    """详细分析Excel文件"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "4.1、生猪期货升贴水数据（盘面结算价）.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("详细分析升贴水Excel文件")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        print(f"\nSheet名称: {ws.title}")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
        # 分析表头
        print("\n表头分析（第1-3行）:")
        header_row = []
        for col_idx in range(1, ws.max_column + 1):
            cell1 = ws.cell(row=1, column=col_idx)
            cell2 = ws.cell(row=2, column=col_idx)
            cell3 = ws.cell(row=3, column=col_idx)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell1.value or cell2.value or cell3.value:
                print(f"  {col_letter}列: 指标={cell1.value}, 单位={cell2.value}, 来源={cell3.value}")
                header_row.append({
                    'col': col_idx,
                    'col_letter': col_letter,
                    'indicator': cell1.value,
                    'unit': cell2.value,
                    'source': cell3.value
                })
        
        # 分析数据范围
        print("\n数据范围分析:")
        print(f"  数据从第4行开始，共{ws.max_row - 3}行数据")
        
        # 查看前10行数据
        print("\n前10行数据示例:")
        for row_idx in range(4, min(14, ws.max_row + 1)):
            date_cell = ws.cell(row=row_idx, column=1)
            print(f"\n  行{row_idx}:")
            print(f"    日期: {date_cell.value}")
            for header in header_row[1:]:  # 跳过日期列
                value_cell = ws.cell(row=row_idx, column=header['col'])
                if value_cell.value is not None:
                    print(f"    {header['indicator']}: {value_cell.value}")
        
        # 查看最后10行数据
        print("\n最后10行数据示例:")
        for row_idx in range(max(4, ws.max_row - 9), ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=1)
            print(f"\n  行{row_idx}:")
            print(f"    日期: {date_cell.value}")
            for header in header_row[1:3]:  # 只显示前2个指标
                value_cell = ws.cell(row=row_idx, column=header['col'])
                if value_cell.value is not None:
                    print(f"    {header['indicator']}: {value_cell.value}")
        
        # 分析日期范围
        print("\n日期范围分析:")
        first_date = ws.cell(row=4, column=1).value
        last_date = ws.cell(row=ws.max_row, column=1).value
        print(f"  最早日期: {first_date}")
        print(f"  最晚日期: {last_date}")
        
        # 统计每个合约的数据量
        print("\n各合约数据统计:")
        contract_cols = {
            '01合约': 'B',
            '03合约': 'C',
            '05合约': 'D',
            '07合约': 'E',
            '09合约': 'F',
            '11合约': 'G'
        }
        
        for contract_name, col_letter in contract_cols.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            non_null_count = 0
            for row_idx in range(4, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    non_null_count += 1
            print(f"  {contract_name} ({col_letter}列): {non_null_count}条数据")
    
    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_detailed()
