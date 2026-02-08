"""分析集团企业出栏跟踪Excel文件结构"""
# -*- coding: utf-8 -*-
import zipfile
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
import json

# 设置输出编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def extract_and_analyze():
    """解压并分析Excel文件"""
    script_dir = Path(__file__).parent.parent.parent
    zip_path = script_dir / "docs" / "生猪" / "数据库模板02：集团企业.zip"
    extract_dir = script_dir / "docs" / "生猪" / "集团企业"
    
    # 解压文件
    if zip_path.exists():
        print(f"找到zip文件: {zip_path}")
        extract_dir.mkdir(exist_ok=True)
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        print(f"已解压到: {extract_dir}")
    else:
        print(f"未找到zip文件: {zip_path}")
        return
    
    # 查找Excel文件
    excel_files = list(extract_dir.rglob("*.xlsx"))
    if not excel_files:
        print("未找到Excel文件")
        return
    
    print(f"\n找到 {len(excel_files)} 个Excel文件:")
    for f in excel_files:
        print(f"  - {f.name}")
    
    # 分析每个Excel文件
    for excel_file in excel_files:
        print(f"\n{'='*60}")
        print(f"分析文件: {excel_file.name}")
        print(f"{'='*60}")
        
        try:
            wb = load_workbook(excel_file, data_only=True)
            print(f"Sheet列表: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                print(f"\n--- Sheet: {sheet_name} ---")
                ws = wb[sheet_name]
                
                # 读取前20行分析结构
                print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
                print("\n前20行数据:")
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                    row_data = [str(cell)[:30] if cell is not None else "" for cell in row[:15]]
                    print(f"  Row {i:2d}: {row_data}")
                
                # 查找关键sheet
                if "CR5" in sheet_name or "cr5" in sheet_name.lower():
                    print(f"\n*** 这是CR5日度sheet ***")
                if "西南" in sheet_name or "southwest" in sheet_name.lower():
                    print(f"\n*** 这是西南汇总sheet ***")
                    
        except Exception as e:
            print(f"分析文件失败: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    extract_and_analyze()
