"""测试基于sheet的独立表导入功能"""
import sys
import os
import io
from pathlib import Path
from datetime import datetime
from sqlalchemy import text, inspect

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 添加backend目录到路径（确保能导入app模块）
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.core.database import SessionLocal
from app.models.sys_user import SysUser
from app.services.ingestors.unified_ingestor import unified_import
from app.services.sheet_table_mapper import SheetTableMapper
from openpyxl import load_workbook

# 项目根目录
backend_dir = Path(__file__).parent.parent
project_root = backend_dir.parent


def get_or_create_test_user(db):
    """获取或创建测试用户"""
    user = db.query(SysUser).filter(SysUser.username == "admin").first()
    if not user:
        from app.core.security import get_password_hash
        user = SysUser(
            username="test_admin",
            password_hash=get_password_hash("test123456"),
            display_name="测试管理员",
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        print(f"  ✓ 创建测试用户: {user.username} (ID: {user.id})")
    else:
        print(f"  ✓ 使用现有用户: {user.username} (ID: {user.id})")
    return user


def get_table_record_count(db, table_name: str) -> int:
    """获取表中的记录数"""
    try:
        inspector = inspect(db.bind)
        if table_name not in inspector.get_table_names():
            return -1  # 表不存在
        
        result = db.execute(text(f"SELECT COUNT(*) as cnt FROM `{table_name}`"))
        row = result.fetchone()
        return row[0] if row else 0
    except Exception as e:
        print(f"    ⚠️  查询表 {table_name} 记录数失败: {e}")
        return -1


def get_sheet_tables_from_excel(file_path: Path, source_code: str, dataset_type: str) -> dict:
    """从Excel文件读取sheet名称，并获取对应的表名和记录数"""
    sheet_tables = {}
    
    try:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        db = SessionLocal()
        try:
            for sheet_name in sheet_names:
                # 对于DCE期货/期权，使用专门的表
                if dataset_type == "LH_FTR":
                    table_name = "fact_futures_daily"
                elif dataset_type == "LH_OPT":
                    table_name = "fact_options_daily"
                else:
                    table_name = SheetTableMapper.sheet_name_to_table_name(
                        sheet_name=sheet_name,
                        source_code=source_code,
                        dataset_type=dataset_type
                    )
                
                # 获取表记录数
                count = get_table_record_count(db, table_name)
                sheet_tables[sheet_name] = {
                    "table_name": table_name,
                    "count_before": count,
                    "count_after": count  # 将在导入后更新
                }
        finally:
            db.close()
    except Exception as e:
        print(f"    ⚠️  读取Excel文件失败: {e}")
    
    return sheet_tables


def log_sheet_import_status(db, sheet_name: str, table_name: str, batch_id: int, 
                           count_before: int, count_after: int, status: str = "完成"):
    """记录sheet导入状态"""
    print(f"\n  ┌─ Sheet: {sheet_name}")
    print(f"  │  表名: {table_name}")
    print(f"  │  导入前记录数: {count_before if count_before >= 0 else '表不存在'}")
    print(f"  │  导入后记录数: {count_after if count_after >= 0 else '表不存在'}")
    if count_before >= 0 and count_after >= 0:
        inserted = count_after - count_before
        print(f"  │  新增记录数: {inserted}")
    print(f"  │  状态: {status}")
    print(f"  └─")


def test_import_file(file_path: Path, dataset_type: str, source_code: str = None):
    """测试导入单个文件"""
    print(f"\n{'='*80}")
    print(f"测试导入文件: {file_path.name}")
    print(f"数据集类型: {dataset_type}")
    print(f"数据源: {source_code}")
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}")
    
    if not file_path.exists():
        print(f"❌ 文件不存在: {file_path}")
        return None
    
    # 读取文件
    with open(file_path, 'rb') as f:
        file_content = f.read()
    
    print(f"文件大小: {len(file_content) / 1024 / 1024:.2f} MB")
    
    # 获取数据库连接
    db = SessionLocal()
    try:
        # 获取或创建测试用户
        user = get_or_create_test_user(db)
        
        # 导入前：获取目标表的当前记录数
        print(f"\n📊 导入前状态检查...")
        
        # 根据dataset_type确定目标表
        if dataset_type == "LH_FTR":
            target_table = "fact_futures_daily"
        elif dataset_type == "LH_OPT":
            target_table = "fact_options_daily"
        else:
            target_table = None  # 将使用sheet表
        
        count_before = {}
        sheet_tables_before = {}
        
        if target_table:
            count = get_table_record_count(db, target_table)
            count_before[target_table] = count
            print(f"  {target_table}: {count if count >= 0 else '表不存在'} 条记录")
        else:
            # 对于使用sheet独立表的类型，先读取Excel获取sheet信息
            print(f"  读取Excel文件sheet信息...")
            sheet_tables_before = get_sheet_tables_from_excel(file_path, source_code, dataset_type)
            if sheet_tables_before:
                print(f"  找到 {len(sheet_tables_before)} 个sheet")
                for sheet_name, info in list(sheet_tables_before.items())[:5]:  # 只显示前5个
                    print(f"    - {sheet_name} -> {info['table_name']}: {info['count_before'] if info['count_before'] >= 0 else '表不存在'} 条记录")
                if len(sheet_tables_before) > 5:
                    print(f"    ... 还有 {len(sheet_tables_before) - 5} 个sheet")
            else:
                print(f"  ⚠️  无法读取sheet信息")
        
        # 执行导入
        print(f"\n🚀 开始导入...")
        
        # 对于LH_FTR和LH_OPT，使用专门的导入器
        if dataset_type == "LH_FTR":
            from app.services.ingestors.futures_ingestor import import_lh_ftr
            from app.models.import_batch import ImportBatch
            import hashlib
            
            # 创建batch
            file_hash = hashlib.sha256(file_content).hexdigest()
            batch = ImportBatch(
                filename=file_path.name,
                file_hash=file_hash,
                uploader_id=user.id,
                status="processing",
                source_code=source_code or dataset_type
            )
            db.add(batch)
            db.flush()
            batch_id = batch.id
            
            print(f"  ✓ 创建批次: {batch_id}")
            
            # 导入数据
            result = import_lh_ftr(db, file_content, batch_id)
            
            # 更新batch状态
            batch.status = "success" if result.get("success", False) else "failed"
            batch.inserted_count = result.get("inserted", 0)
            batch.updated_count = result.get("updated", 0)
            db.commit()
            
            # 包装结果格式（errors 可能是 int 或 list）
            _err = result.get("errors")
            _err_cnt = len(_err) if isinstance(_err, (list, tuple)) else (_err or 0)
            result = {
                "batch_id": batch_id,
                "success": result.get("success", False),
                "inserted": result.get("inserted", 0),
                "updated": result.get("updated", 0),
                "error_count": _err_cnt,
                "errors": _err if isinstance(_err, (list, tuple)) else []
            }
            
        elif dataset_type == "LH_OPT":
            from app.services.ingestors.options_ingestor import import_lh_opt
            from app.models.import_batch import ImportBatch
            import hashlib
            
            # 创建batch
            file_hash = hashlib.sha256(file_content).hexdigest()
            batch = ImportBatch(
                filename=file_path.name,
                file_hash=file_hash,
                uploader_id=user.id,
                status="processing",
                source_code=source_code or dataset_type
            )
            db.add(batch)
            db.flush()
            batch_id = batch.id
            
            print(f"  ✓ 创建批次: {batch_id}")
            
            # 导入数据
            result = import_lh_opt(db, file_content, batch_id)
            
            # 更新batch状态
            batch.status = "success" if result.get("success", False) else "failed"
            batch.inserted_count = result.get("inserted", 0)
            batch.updated_count = result.get("updated", 0)
            db.commit()
            
            # 包装结果格式（errors 可能是 int 或 list）
            _err = result.get("errors")
            _err_cnt = len(_err) if isinstance(_err, (list, tuple)) else (_err or 0)
            result = {
                "batch_id": batch_id,
                "success": result.get("success", False),
                "inserted": result.get("inserted", 0),
                "updated": result.get("updated", 0),
                "error_count": _err_cnt,
                "errors": _err if isinstance(_err, (list, tuple)) else []
            }
        else:
            # 使用unified_import
            result = unified_import(
                db=db,
                file_content=file_content,
                filename=file_path.name,
                uploader_id=user.id,
                dataset_type=dataset_type,
                source_code=source_code
            )
        
        batch_id = result.get('batch_id')
        if not batch_id:
            print(f"❌ 导入失败：未返回批次ID")
            return None
        
        print(f"\n✅ 导入完成，批次ID: {batch_id}")
        
        # 导入后：获取目标表的记录数
        print(f"\n📊 导入后状态检查...")
        
        if target_table:
            # 对于LH_FTR/LH_OPT，直接查询目标表
            count_after = get_table_record_count(db, target_table)
            count_before_val = count_before.get(target_table, 0)
            
            print(f"\n📋 导入详情:")
            print(f"{'─'*80}")
            log_sheet_import_status(
                db=db,
                sheet_name="日历史行情",
                table_name=target_table,
                batch_id=batch_id,
                count_before=count_before_val,
                count_after=count_after,
                status="✅ 成功" if count_after > count_before_val else "⚠️  无新增数据"
            )
            print(f"{'─'*80}")
        else:
            # 对于其他类型，使用导入前获取的sheet信息
            if sheet_tables_before:
                # 记录每个sheet的导入状态
                print(f"\n📋 Sheet导入详情:")
                print(f"{'─'*80}")
                
                for sheet_name, table_info in sheet_tables_before.items():
                    table_name = table_info["table_name"]
                    count_before_val = table_info["count_before"]
                    
                    # 重新查询导入后的记录数
                    count_after = get_table_record_count(db, table_name)
                    
                    # 计算新增记录数
                    if count_before_val >= 0 and count_after >= 0:
                        inserted = count_after - count_before_val
                        status = "✅ 成功" if inserted > 0 or count_after > 0 else "⚠️  无数据"
                    elif count_after >= 0:
                        inserted = count_after
                        status = "✅ 成功（新表）"
                    else:
                        inserted = 0
                        status = "❌ 表不存在"
                    
                    log_sheet_import_status(
                        db=db,
                        sheet_name=sheet_name,
                        table_name=table_name,
                        batch_id=batch_id,
                        count_before=count_before_val,
                        count_after=count_after,
                        status=status
                    )
                
                print(f"{'─'*80}")
            else:
                print(f"\n⚠️  无法获取sheet导入详情（导入前未读取sheet信息）")
        
        # 汇总结果
        print(f"\n📈 导入结果汇总:")
        print(f"  成功: {result.get('success', False)}")
        print(f"  插入: {result.get('inserted', 0)}")
        print(f"  更新: {result.get('updated', 0)}")
        print(f"  错误: {result.get('error_count', 0)}")
        print(f"  批次ID: {batch_id}")
        print(f"  结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        errors = result.get('errors', [])
        error_count = result.get('error_count', 0)
        
        # unified_import返回的errors可能是整数（错误数量）或列表
        if error_count > 0 or (isinstance(errors, int) and errors > 0):
            error_num = error_count if error_count > 0 else (errors if isinstance(errors, int) else 0)
            print(f"\n⚠️  错误数量: {error_num}")
            
            # 如果是列表，显示错误详情
            if isinstance(errors, list) and len(errors) > 0:
                print(f"⚠️  错误列表（前10个）:")
                for i, error in enumerate(errors[:10], 1):
                    if isinstance(error, dict):
                        error_msg = error.get('reason', str(error))
                    else:
                        error_msg = str(error)
                    print(f"  {i}. {error_msg}")
        
        return result
    
    except Exception as e:
        print(f"❌ 导入失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        db.close()


def main():
    """主函数"""
    print("="*80)
    print("基于Sheet的独立表架构导入测试")
    print("="*80)
    
    # 测试文件列表
    test_files = [
        {
            "path": project_root / "docs" / "lh_ftr.xlsx",
            "dataset_type": "LH_FTR",
            "source_code": "DCE",
            "name": "DCE 生猪期货"
        },
        {
            "path": project_root / "docs" / "2026年2月2日涌益咨询日度数据.xlsx",
            "dataset_type": "YONGYI_DAILY",
            "source_code": "YONGYI",
            "name": "涌益日度数据"
        }
    ]
    
    # 检查文件是否存在
    print(f"\n📁 检查测试文件...")
    for file_info in test_files:
        if file_info["path"].exists():
            print(f"  ✓ {file_info['path'].name}")
        else:
            print(f"  ❌ {file_info['path'].name} - 文件不存在")
            print(f"     路径: {file_info['path']}")
    
    results = []
    
    for file_info in test_files:
        if not file_info["path"].exists():
            print(f"\n⚠️  跳过文件: {file_info['path'].name} (文件不存在)")
            continue
            
        result = test_import_file(
            file_path=file_info["path"],
            dataset_type=file_info["dataset_type"],
            source_code=file_info["source_code"]
        )
        if result:
            results.append({
                "file": file_info["path"].name,
                "result": result
            })
    
    # 汇总
    print(f"\n{'='*80}")
    print("测试汇总")
    print(f"{'='*80}")
    
    if results:
        total_inserted = sum(r["result"].get("inserted", 0) for r in results)
        total_updated = sum(r["result"].get("updated", 0) for r in results)
        total_errors = sum(r["result"].get("error_count", 0) for r in results)
        
        print(f"总插入: {total_inserted}")
        print(f"总更新: {total_updated}")
        print(f"总错误: {total_errors}")
        
        for r in results:
            print(f"\n{r['file']}:")
            print(f"  插入: {r['result'].get('inserted', 0)}")
            print(f"  更新: {r['result'].get('updated', 0)}")
            print(f"  错误: {r['result'].get('error_count', 0)}")
    else:
        print("⚠️  没有成功导入的文件")


if __name__ == "__main__":
    main()
