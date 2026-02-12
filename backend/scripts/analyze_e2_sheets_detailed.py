"""
详细分析E2需要的sheet结构
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import openpyxl

script_dir = Path(__file__).parent.parent
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_sheets():
    """分析sheet结构"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    sheets_to_analyze = ["NYB", "4.1.钢联数据", "4.2涌益底稿", "02.协会猪料"]
    
    for sheet_name in sheets_to_analyze:
        if sheet_name not in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"Sheet '{sheet_name}' 不存在")
            continue
        
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*80}")
        
        ws = wb[sheet_name]
        
        # 显示前15行数据，找出表头和数据行
        print(f"\n前15行数据:")
        for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
            row_data = []
            for col_idx, cell in enumerate(row[:30], 1):
                if cell is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    row_data.append(f"{col_letter}:{cell}")
            if row_data:
                print(f"  行{row_idx}: {', '.join(row_data[:15])}")
        
        # 统计总行数
        total_rows = ws.max_row
        print(f"\n总行数: {total_rows}")
        
        # 查找有数据的行
        print(f"\n查找有数据的行（前20行）:")
        data_count = 0
        for row_idx, row in enumerate(ws.iter_rows(max_row=min(100, total_rows), values_only=True), 1):
            # 检查是否有非空值
            has_data = False
            for cell in row[:30]:
                if cell is not None and cell != "":
                    if isinstance(cell, (int, float)):
                        has_data = True
                        break
                    elif isinstance(cell, str) and cell.strip() and not cell.startswith("日期") and not cell.startswith("月度"):
                        has_data = True
                        break
            
            if has_data:
                data_count += 1
                date_val = row[0] if len(row) > 0 else None
                if data_count <= 5:
                    print(f"  行{row_idx}: 日期={date_val}, 前5列={row[:5]}")
        
        print(f"  找到 {data_count} 行有数据")
    
    wb.close()

if __name__ == "__main__":
    analyze_sheets()
