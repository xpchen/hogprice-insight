#!/usr/bin/env python3
"""
核查区域价差 sheet 表头：为何广东-重庆等未入库。
读取 docs/0306_生猪/生猪/1、价格：钢联自动更新模板.xlsx 的「区域价差」sheet，
逐列检查正则是否匹配、省份是否识别。
"""
import os
import re
import sys

# backend 为根
BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BACKEND)

# 与 r01 一致
HEADER_ROW = 2
_SPREAD_REGION_RE = re.compile(
    r"出栏均价：(\S+?)（日）\s*-\s*.*?出栏均价：(\S+?)（日）"
)
# 兼容半角括号
_SPREAD_REGION_RE_HALF = re.compile(
    r"出栏均价：(\S+?)\(日\)\s*-\s*.*?出栏均价：(\S+?)\(日\)"
)

from import_tool.utils import province_to_code


def _region_lookup(name: str):
    if not name:
        return None
    n = name.strip().replace("省", "").replace("市", "").replace("自治区", "").replace("特别行政区", "")
    return province_to_code(n) or (None if name.strip() != "中国" else "NATION")


def main():
    excel_path = os.path.join(
        os.path.dirname(BACKEND), "docs", "0306_生猪", "生猪", "1、价格：钢联自动更新模板.xlsx"
    )
    out_lines = []
    def log(msg):
        out_lines.append(msg)
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("utf-8", errors="replace").decode("utf-8"))

    if not os.path.isfile(excel_path):
        log(f"File not found: {excel_path}")
        return 1

    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if "区域价差" not in wb.sheetnames:
        log("Sheet '区域价差' not in workbook. Sheets: " + str(wb.sheetnames))
        wb.close()
        return 1

    ws = wb["区域价差"]
    rows_1_3 = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    log("Region spread sheet: first 3 rows (first 8 cells each):")
    for i, row in enumerate(rows_1_3, 1):
        cells = list(row)[:8] if row else []
        log(f"  Row{i}: {cells}")

    headers = list(rows_1_3[HEADER_ROW - 1]) if len(rows_1_3) >= HEADER_ROW else []
    log(f"\nUsing row {HEADER_ROW} as header. Column check:")
    for idx, h in enumerate(headers):
        if idx == 0 or h is None:
            continue
        s = str(h).strip()
        if not s:
            continue
        m = _SPREAD_REGION_RE.search(s)
        if not m:
            m = _SPREAD_REGION_RE_HALF.search(s)
        if m:
            high_name, low_name = m.group(1), m.group(2)
            high_code = _region_lookup(high_name)
            low_code = _region_lookup(low_name)
            if high_code and low_code:
                st = f"region_spread_{high_code}_{low_code}"
                log(f"  Col{idx}: OK -> {high_name}-{low_name} -> {st}")
            else:
                log(f"  Col{idx}: regex OK but province unknown: high={high_name!r}({high_code}), low={low_name!r}({low_code})")
        else:
            preview = repr(s[:80])
            log(f"  Col{idx}: regex NO MATCH. Preview: {preview}")
    wb.close()

    # write to file for safe utf-8
    out_path = os.path.join(os.path.dirname(BACKEND), "docs", "region_spread_header_check.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))
    log(f"\nOutput also written to: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
