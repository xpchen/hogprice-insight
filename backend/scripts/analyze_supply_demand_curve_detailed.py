"""
详细分析【供需曲线】sheet的数据结构
确定如何提取定点屠宰系数和猪价系数
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import json
import openpyxl
from datetime import datetime

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_detailed():
    """详细分析供需曲线sheet"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("详细分析【供需曲线】sheet")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["供需曲线"]
        
        # 分析定点屠宰部分
        print("\n" + "=" * 80)
        print("定点屠宰部分（第2-14行）")
        print("=" * 80)
        
        # 读取第2行表头
        header_row_2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        print("\n第2行表头:")
        for col_idx, cell in enumerate(header_row_2[:20], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  {col_letter}列 ({col_idx}): {cell}")
        
        # 读取定点屠宰数据（第3-14行）
        print("\n定点屠宰数据（第3-14行，1-12月）:")
        slaughter_data = []
        for row_idx in range(3, 15):  # 第3-14行
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            month = row[1] if len(row) > 1 else None  # B列是月份
            mean_value = row[9] if len(row) > 9 else None  # J列是均值
            ratio_2019 = row[13] if len(row) > 13 else None  # N列是2019年比率
            ratio_2020 = row[14] if len(row) > 14 else None  # O列是2020年比率
            
            # 各年份数据
            years_data = {}
            year_labels = ['2019年', '2020年', '2021年', '2022年', '2023年', '2024年', '2025年']
            for i, year_label in enumerate(year_labels):
                col_idx = 2 + i  # C列开始
                if len(row) > col_idx:
                    years_data[year_label] = row[col_idx]
            
            slaughter_data.append({
                'month': month,
                'mean': mean_value,
                'ratio_2019': ratio_2019,
                'ratio_2020': ratio_2020,
                'years': years_data
            })
            
            print(f"  行{row_idx} - {month}: 均值={mean_value}, 2019比率={ratio_2019}, 2020比率={ratio_2020}")
        
        # 分析猪价部分
        print("\n" + "=" * 80)
        print("猪价部分（第16行开始）")
        print("=" * 80)
        
        # 读取第16行表头
        header_row_16 = list(ws.iter_rows(min_row=16, max_row=16, values_only=True))[0]
        print("\n第16行表头:")
        for col_idx, cell in enumerate(header_row_16[:20], 1):
            if cell:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  {col_letter}列 ({col_idx}): {cell}")
        
        # 读取猪价数据（第17行开始，到第28行，1-12月）
        print("\n猪价数据（第17-28行，1-12月）:")
        price_data = []
        for row_idx in range(17, 29):  # 第17-28行
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            month = row[1] if len(row) > 1 else None  # B列是月份
            mean_value = row[9] if len(row) > 9 else None  # J列是均值
            ratio_2019 = row[13] if len(row) > 13 else None  # N列是2019年比率
            ratio_2020 = row[14] if len(row) > 14 else None  # O列是2020年比率
            
            # 各年份数据
            years_data = {}
            year_labels = ['2019年', '2020年', '2021年', '2022年', '2023年', '2024年', '2025年']
            for i, year_label in enumerate(year_labels):
                col_idx = 2 + i  # C列开始
                if len(row) > col_idx:
                    years_data[year_label] = row[col_idx]
            
            price_data.append({
                'month': month,
                'mean': mean_value,
                'ratio_2019': ratio_2019,
                'ratio_2020': ratio_2020,
                'years': years_data
            })
            
            print(f"  行{row_idx} - {month}: 均值={mean_value}, 2019比率={ratio_2019}, 2020比率={ratio_2020}")
        
        # 分析数据结构
        print("\n" + "=" * 80)
        print("数据结构分析")
        print("=" * 80)
        print("\n定点屠宰系数计算方式:")
        print("  - 比率列（N列、O列等）应该是：当月值 / 均值")
        print("  - 例如：2019年1月比率 = 2019年1月屠宰量 / 均值")
        
        print("\n猪价系数计算方式:")
        print("  - 比率列（N列、O列等）应该是：当月值 / 均值")
        print("  - 例如：2019年1月比率 = 2019年1月猪价 / 均值")
        
        # 检查是否有时间序列数据（按年月）
        print("\n" + "=" * 80)
        print("检查是否有时间序列格式的数据")
        print("=" * 80)
        
        # 查看第30行之后是否有其他格式的数据
        print("\n第30-50行数据:")
        for row_idx in range(30, min(51, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            row_data = []
            for col_idx, cell in enumerate(row[:10], 1):
                if cell is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_str = str(cell)
                    if len(cell_str) > 20:
                        cell_str = cell_str[:20] + "..."
                    row_data.append(f"{col_letter}:{cell_str}")
            if row_data:
                print(f"  行{row_idx}: {', '.join(row_data)}")
        
        wb.close()
        
    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_detailed()
