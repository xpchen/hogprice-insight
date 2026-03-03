# -*- coding: utf-8 -*-
"""
列出 A1供给预测 sheet 所有列的表头（第1、2、3行），便于核对哪些列已/未导入 fact_monthly_indicator。
已导入（r02）：B=日期, F=母猪效能, N=压栏系数, AC=生猪存栏, AE/AF=生猪存栏环比/同比, AG/AH/AI=定点屠宰/环比/同比。
运行方式：在 backend 目录执行 python scripts/list_a1_sheet_columns.py
需存在 docs/0226/生猪/ 或 docs/生猪/ 下的 2、【生猪产业数据】.xlsx
"""
import sys
from pathlib import Path

script_dir = Path(__file__).resolve().parent
backend = script_dir.parent
repo = backend.parent

for sub in ("0226/生猪", "生猪"):
    excel_path = repo / "docs" / sub / "2、【生猪产业数据】.xlsx"
    if excel_path.exists():
        break
else:
    print("未找到 2、【生猪产业数据】.xlsx，请放在 docs/0226/生猪/ 或 docs/生猪/ 下")
    sys.exit(1)

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(excel_path, data_only=True)
if "A1供给预测" not in wb.sheetnames:
    print("Sheet 'A1供给预测' 不存在")
    sys.exit(0)

ws = wb["A1供给预测"]
max_row = ws.max_row
max_col = ws.max_column
print("A1供给预测: max_row=%s max_col=%s\n" % (max_row, max_col))

# 已导入列（0-based）: 1=B日期, 5=F, 13=N, 28=AC, 30=AE, 31=AF, 32=AG, 33=AH, 34=AI
imported = {1: "日期(B)", 5: "母猪效能(F)", 13: "压栏系数(N)", 28: "生猪存栏(AC)", 30: "环比(AE)", 31: "同比(AF)", 32: "定点屠宰(AG)", 33: "环比(AH)", 34: "同比(AI)"}

print("列号(Excel) | 0-based | 第1行表头 | 第2行表头 | 第3行表头 | 是否已导入")
print("-" * 100)
for ci in range(min(max_col, 60)):  # 最多显示到 60 列
    col_letter = get_column_letter(ci + 1)
    r1 = ws.cell(1, ci + 1).value
    r2 = ws.cell(2, ci + 1).value
    r3 = ws.cell(3, ci + 1).value
    r1 = (r1 if r1 is not None else "")
    r2 = (r2 if r2 is not None else "")
    r3 = (r3 if r3 is not None else "")
    status = imported.get(ci, "")
    print("%3s %8s   %-20s %-20s %-20s %s" % (col_letter, ci, str(r1)[:20], str(r2)[:20], str(r3)[:20], status or "未导入"))
wb.close()
print("\n说明：若某列表头为 能繁存栏/新生仔猪/5月大猪/残差率/生猪出栏/累计出栏/累同 等且未出现在「已导入」中，可在 r02 _read_a1_supply_forecast 中增加对应列解析。")
