"""
从 docs/生猪/涌益生猪项目数据库.zip 解压并导入包含「月度-淘汰母猪屠宰厂宰杀量」的 xlsx。
若 zip 内无 xlsx，则尝试从 docs/生猪/涌益生猪项目数据库/ 目录下查找并导入。
"""
# -*- coding: utf-8 -*-
import sys
import io
import zipfile
import hashlib
from pathlib import Path

script_dir = Path(__file__).parent.parent
repo = script_dir.parent
sys.path.insert(0, str(script_dir))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

TARGET_SHEET = "月度-淘汰母猪屠宰厂宰杀量"
ZIP_PATH = repo / "docs" / "生猪" / "涌益生猪项目数据库.zip"
OUT_DIR = repo / "docs" / "生猪" / "涌益生猪项目数据库"


def find_xlsx_with_sheet(path_or_zip):
    """在目录或 zip 内查找包含目标 sheet 的 xlsx，返回 (path_or_none, file_content_bytes_or_none)。"""
    try:
        import openpyxl
    except ImportError:
        return None, None
    if path_or_zip.is_file() and path_or_zip.suffix.lower() == ".zip":
        with zipfile.ZipFile(path_or_zip, "r") as zf:
            for name in zf.namelist():
                if not name.lower().endswith(".xlsx"):
                    continue
                try:
                    with zf.open(name) as f:
                        content = f.read()
                    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
                    if TARGET_SHEET in wb.sheetnames:
                        wb.close()
                        return name, content
                    wb.close()
                except Exception:
                    continue
        return None, None
    if path_or_zip.is_dir():
        for p in path_or_zip.glob("*.xlsx"):
            try:
                wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
                if TARGET_SHEET in wb.sheetnames:
                    wb.close()
                    with open(p, "rb") as f:
                        return str(p.name), f.read()
                wb.close()
            except Exception:
                continue
    return None, None


def main():
    from app.core.database import SessionLocal
    from app.models.import_batch import ImportBatch
    from app.models.raw_file import RawFile
    from app.models.raw_sheet import RawSheet
    from app.models.raw_table import RawTable
    from app.services.ingestors.raw_writer import save_raw_file, save_all_sheets_from_workbook

    # 1) 若已解压目录存在，先从目录找
    xlsx_name, content = None, None
    if OUT_DIR.exists():
        xlsx_name, content = find_xlsx_with_sheet(OUT_DIR)
    if not content and ZIP_PATH.exists():
        print("从 zip 解压并查找...")
        xlsx_name, content = find_xlsx_with_sheet(ZIP_PATH)
    if not content:
        print("未找到包含「月度-淘汰母猪屠宰厂宰杀量」的 xlsx。")
        print("请将 2026.1.16-2026.1.22涌益咨询 周度数据.xlsx 放入 docs/生猪/涌益生猪项目数据库/ 或确保 zip 内含有该文件。")
        sys.exit(1)

    # 使用 zip 内文件名或“涌益咨询 周度数据.xlsx”作为存储名
    filename = Path(xlsx_name).name if xlsx_name else "涌益咨询 周度数据.xlsx"
    if "2026" in filename or "2026" in xlsx_name:
        filename = "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx"

    print(f"找到: {xlsx_name or filename}")
    print("导入到 raw 层...")

    db = SessionLocal()
    try:
        file_hash = hashlib.sha256(content).hexdigest()
        existing = db.query(RawFile).filter(RawFile.file_hash == file_hash).first()
        if existing:
            print(f"该文件已存在（相同内容），raw_file.id = {existing.id}, 文件名 = {existing.filename}")
            db.close()
            return

        batch = ImportBatch(
            filename=filename,
            file_hash=file_hash,
            uploader_id=1,
            status="completed",
            source_code="YONGYI_WEEKLY",
            total_rows=0,
            success_rows=0,
            failed_rows=0,
        )
        db.add(batch)
        db.flush()
        raw_file = save_raw_file(db=db, batch_id=batch.id, filename=filename, file_content=content)
        raw_sheets = save_all_sheets_from_workbook(
            db=db,
            raw_file_id=raw_file.id,
            workbook_or_path=io.BytesIO(content),
            sparse=False,
            max_rows=2000,
        )
        db.commit()
        print(f"已导入 {len(raw_sheets)} 个 sheet。")
        for s in raw_sheets:
            if s.sheet_name == TARGET_SHEET:
                print(f"  ✓ {s.sheet_name}")
    except Exception as e:
        db.rollback()
        print(f"导入失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
