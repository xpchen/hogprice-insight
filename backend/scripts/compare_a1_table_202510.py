# -*- coding: utf-8 -*-
"""
比对 A1供给预测 2025-10-01 行：Excel 该行各列 vs 规模场表格使用的 fact_monthly_indicator 数据。
用于排查「表格与 A1 表不一致」的原因（列索引错位、数据源混用、缺列等）。
运行：在 backend 目录执行 python scripts/compare_a1_table_202510.py
"""
import sys
from pathlib import Path
from datetime import date

script_dir = Path(__file__).resolve().parent
backend = script_dir.parent
repo = backend.parent
sys.path.insert(0, str(backend))

# 列字母
def col_letter(i: int) -> str:
    i += 1
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def main():
    # 1) 从 Excel 读 A1 表 2025-10-01 行
    for sub in ("0226/生猪", "生猪"):
        excel_path = repo / "docs" / sub / "2、【生猪产业数据】.xlsx"
        if excel_path.exists():
            break
    else:
        print("未找到 2、【生猪产业数据】.xlsx")
        return

    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "A1供给预测" not in wb.sheetnames:
        print("无 A1供给预测 sheet")
        wb.close()
        return

    ws = wb["A1供给预测"]
    target = date(2025, 10, 1)
    from import_tool.utils import parse_month

    excel_row = None
    excel_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=35, values_only=True), start=4):
        row = list(row)
        if len(row) < 2:
            continue
        md = parse_month(row[1] if len(row) > 1 else None)
        if md and md.year == target.year and md.month == target.month:
            excel_row = row
            excel_row_idx = row_idx
            break
    wb.close()

    if not excel_row:
        print("A1 表中未找到 2025-10 行（日期在 B 列，从第 4 行起）")
    else:
        print("=" * 80)
        print("A1供给预测 Excel 行（2025-10-01 对应行，Excel 行号 %d）" % excel_row_idx)
        print("=" * 80)
        # 当前 r02 使用的列索引 -> 表头含义
        a1_col_map = [
            (1, "B", "日期"),
            (5, "F", "母猪效能"),
            (9, "J", "5月大猪"),
            (10, "K", "5月大猪环比"),
            (11, "L", "5月大猪同比"),
            (12, "M", "残差率"),
            (13, "N", "压栏系数"),
            (14, "O", "生猪出栏"),
            (15, "P", "生猪出栏环比"),
            (16, "Q", "生猪出栏同比"),
            (17, "R", "累计出栏"),
            (18, "S", "累同"),
            (28, "AC", "定点屠宰"),
            (29, "AD", "定点屠宰环比"),
            (30, "AE", "定点屠宰同比"),
        ]
        for idx, letter, name in a1_col_map:
            v = excel_row[idx] if idx < len(excel_row) else None
            print("  %s (idx %2d) %-16s = %s" % (letter, idx, name, v))
        # 能繁/新生仔猪若在 A1 中，通常在 C,D,E 和 G,H,I
        print("\n  其他常用列（若表头与此一致则需在 r02 中增加导入）：")
        for idx in (2, 3, 4, 6, 7, 8):
            v = excel_row[idx] if idx < len(excel_row) else None
            print("  %s (idx %2d)                  = %s" % (col_letter(idx), idx, v))

    # 2) 从 DB 读 2025-10-01 的 fact_monthly_indicator（表格用到的指标）
    from sqlalchemy import text
    from import_tool.db import get_engine

    engine = get_engine()
    # 表格列与 (indicator_code, source, sub_category, value_type) 对应
    table_columns = [
        ("能繁存栏", "breeding_sow_inventory", None, None, "abs"),
        ("能繁环比", "breeding_sow_inventory", "NYB", "nation", "mom_pct"),
        ("能繁同比", "breeding_sow_inventory", None, None, "pct"),
        ("母猪效能", "prod_farrowing_count", "A1", "", "abs"),
        ("新生仔猪", "piglet_inventory", None, None, "abs"),
        ("新生仔猪环比", "piglet_inventory", "NYB", "nation", "mom_pct"),
        ("新生仔猪同比", "piglet_inventory", None, None, "pct"),
        ("5月大猪", "medium_large_hog", "A1", "", "abs"),
        ("5月大猪环比", "medium_large_hog", "NYB/A1", None, "mom_pct"),
        ("5月大猪同比", "medium_large_hog", "A1", "", "yoy_pct"),
        ("残差率", "supply_residual_rate", "A1", "", "pct"),
        ("生猪出栏", "hog_turnover", "A1", "", "abs"),
        ("生猪出栏环比", "hog_turnover", "NYB/A1", None, "mom_pct"),
        ("生猪出栏同比", "hog_turnover", "A1", "", "yoy_pct"),
        ("累计出栏", "hog_turnover", "A1", "cumulative", "abs"),
        ("累同", "hog_turnover", "A1", "cumulative", "yoy_pct"),
        ("定点屠宰", "designated_slaughter", "A1/STATISTICS_BUREAU", "", "abs"),
        ("定点屠宰环比", "designated_slaughter", "A1", "", "mom_pct"),
        ("定点屠宰同比", "designated_slaughter", "A1", "", "yoy_pct"),
    ]

    print("\n" + "=" * 80)
    print("fact_monthly_indicator 中 2025-10-01 与表格相关的所有记录（source 含 A1/NYB/STATISTICS_BUREAU）")
    print("=" * 80)

    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT indicator_code, source, sub_category, value_type, value
                FROM fact_monthly_indicator
                WHERE month_date = :d
                  AND region_code = 'NATION'
                  AND source IN ('A1','NYB','STATISTICS_BUREAU')
                  AND value IS NOT NULL
                ORDER BY indicator_code, source, sub_category, value_type
            """),
            {"d": target},
        ).fetchall()

    db_by_key = {}
    for r in rows:
        key = (r[0], r[1], r[2] or "", r[3])
        db_by_key[key] = r[4]

    for label, code, source, sub_cat, vtype in table_columns:
        if source and "/" in source:
            sources = [s.strip() for s in source.split("/")]
        elif source:
            sources = [source]
        else:
            sources = []
        vals = []
        for src in sources:
            if "nation" in str(sub_cat).lower() or (sub_cat and sub_cat != "cumulative"):
                sc = sub_cat
            else:
                sc = sub_cat or ""
            k = (code, src, sc, vtype)
            if k in db_by_key:
                vals.append("%s=%.4g" % (src, db_by_key[k]))
        if not vals and code:
            # 单 source 或任意
            for k, v in db_by_key.items():
                if k[0] == code:
                    vals.append("%s(%s)=%.4g" % (k[1], k[2] or "-", v))
        print("  %-16s %s" % (label, ", ".join(vals) if vals else "(无)"))

    # 3) 逐项对比：A1 列 vs 表格应显示的值（仅对来自 A1 的列）
    if excel_row:
        print("\n" + "=" * 80)
        print("比对：A1 该行数值 vs 库中 A1 同指标（不一致则可能列索引错或未入库）")
        print("=" * 80)
        a1_expect = [
            (5, "母猪效能", "prod_farrowing_count", "", "abs"),
            (9, "5月大猪", "medium_large_hog", "", "abs"),
            (10, "5月大猪环比", "medium_large_hog", "", "mom_pct"),
            (11, "5月大猪同比", "medium_large_hog", "", "yoy_pct"),
            (12, "残差率", "supply_residual_rate", "", "pct"),
            (13, "压栏系数", "prod_healthy_piglets_per_litter", "", "abs"),
            (14, "生猪出栏", "hog_turnover", "", "abs"),
            (15, "生猪出栏环比", "hog_turnover", "", "mom_pct"),
            (16, "生猪出栏同比", "hog_turnover", "", "yoy_pct"),
            (17, "累计出栏", "hog_turnover", "cumulative", "abs"),
            (18, "累同", "hog_turnover", "cumulative", "yoy_pct"),
            (28, "定点屠宰", "designated_slaughter", "", "abs"),
            (29, "定点屠宰环比", "designated_slaughter", "", "mom_pct"),
            (30, "定点屠宰同比", "designated_slaughter", "", "yoy_pct"),
        ]
        for idx, name, code, sub, vtype in a1_expect:
            excel_val = excel_row[idx] if idx < len(excel_row) else None
            try:
                excel_f = float(excel_val) if excel_val is not None else None
            except (TypeError, ValueError):
                excel_f = None
            key = (code, "A1", sub, vtype)
            db_val = db_by_key.get(key)
            db_f = float(db_val) if db_val is not None else None
            match = "OK" if (excel_f is not None and db_f is not None and abs(excel_f - db_f) < 1e-6) else ("OK" if excel_f is None and db_f is None else "不一致")
            if match != "OK":
                print("  [%s] %s  Excel=%s  库=%s  -> %s" % (col_letter(idx), name, excel_val, db_val, match))


if __name__ == "__main__":
    main()
