"""分析所有Excel文件的结构，生成表结构定义"""
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import json

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Excel文件列表
EXCEL_FILES = [
    {
        "name": "钢联价格模板",
        "path": project_root / "docs" / "1、价格：钢联自动更新模板.xlsx",
        "source_code": "GANGLIAN",
        "dataset_type": "DAILY"
    },
    {
        "name": "涌益周度数据",
        "path": project_root / "docs" / "2026.1.16-2026.1.22涌益咨询 周度数据.xlsx",
        "source_code": "YONGYI",
        "dataset_type": "WEEKLY"
    },
    {
        "name": "涌益日度数据",
        "path": project_root / "docs" / "2026年2月2日涌益咨询日度数据 - 副本.xlsx",
        "source_code": "YONGYI",
        "dataset_type": "DAILY"
    },
    {
        "name": "DCE期货",
        "path": project_root / "docs" / "lh_ftr.xlsx",
        "source_code": "DCE",
        "dataset_type": "DAILY"
    },
    {
        "name": "DCE期权",
        "path": project_root / "docs" / "lh_opt.xlsx",
        "source_code": "DCE",
        "dataset_type": "DAILY"
    }
]


def analyze_sheet_structure(workbook, sheet_name, source_code, dataset_type):
    """分析单个sheet的结构"""
    try:
        ws = workbook[sheet_name]
        
        # 读取前20行进行分析
        data_rows = []
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            data_rows.append(row)
        
        # 转换为DataFrame（尝试找到表头行）
        df = None
        header_row = None
        
        # 尝试识别表头（通常在前5行）
        for i in range(min(5, len(data_rows))):
            try:
                test_df = pd.DataFrame(data_rows[i:], columns=data_rows[i])
                # 检查是否有合理的列名（非空且不全是数字）
                if test_df.columns.notna().sum() > 0:
                    non_numeric_cols = sum(1 for col in test_df.columns if col and not str(col).replace('.', '').isdigit())
                    if non_numeric_cols > 2:  # 至少3个非数字列
                        df = test_df
                        header_row = i + 1
                        break
            except:
                continue
        
        if df is None:
            # 如果无法识别表头，使用第一行
            df = pd.DataFrame(data_rows[1:], columns=data_rows[0] if data_rows else None)
            header_row = 1
        
        # 分析列结构
        columns_info = []
        for col in df.columns:
            col_data = df[col].dropna()
            col_info = {
                "name": str(col) if col else f"Column_{df.columns.get_loc(col)}",
                "sample_values": col_data.head(5).values.tolist() if len(col_data) > 0 else [],
                "data_type": "unknown",
                "has_numeric": False,
                "has_date": False,
                "unique_count": col_data.nunique() if len(col_data) > 0 else 0
            }
            
            # 判断数据类型
            if len(col_data) > 0:
                # 检查是否是日期
                try:
                    pd.to_datetime(col_data.head(1).iloc[0])
                    col_info["has_date"] = True
                    col_info["data_type"] = "date"
                except:
                    pass
                
                # 检查是否是数字
                try:
                    numeric_values = pd.to_numeric(col_data, errors='coerce')
                    if numeric_values.notna().sum() > len(col_data) * 0.5:  # 至少50%是数字
                        col_info["has_numeric"] = True
                        col_info["data_type"] = "numeric"
                except:
                    pass
            
            columns_info.append(col_info)
        
        return {
            "sheet_name": sheet_name,
            "source_code": source_code,
            "dataset_type": dataset_type,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "header_row": header_row,
            "columns": columns_info,
            "sample_data": [dict(row) for row in df.head(10).to_dict('records')] if len(df) > 0 else []
        }
    except Exception as e:
        return {
            "sheet_name": sheet_name,
            "error": str(e)
        }


def analyze_excel_file(file_info):
    """分析单个Excel文件"""
    print(f"\n{'='*80}")
    print(f"分析文件: {file_info['name']}")
    print(f"路径: {file_info['path']}")
    print(f"{'='*80}")
    
    if not file_info['path'].exists():
        print(f"❌ 文件不存在: {file_info['path']}")
        return None
    
    try:
        workbook = load_workbook(file_info['path'], data_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"\n找到 {len(sheet_names)} 个sheet:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")
        
        analysis_results = []
        for sheet_name in sheet_names:
            print(f"\n分析sheet: {sheet_name}")
            result = analyze_sheet_structure(
                workbook,
                sheet_name,
                file_info['source_code'],
                file_info['dataset_type']
            )
            analysis_results.append(result)
            
            if "error" in result:
                print(f"  ❌ 错误: {result['error']}")
            else:
                print(f"  ✓ 行数: {result['total_rows']}, 列数: {result['total_cols']}")
                print(f"  ✓ 表头行: {result['header_row']}")
                print(f"  ✓ 列数: {len(result['columns'])}")
        
        workbook.close()
        
        return {
            "file_name": file_info['name'],
            "file_path": str(file_info['path']),
            "source_code": file_info['source_code'],
            "dataset_type": file_info['dataset_type'],
            "sheets": analysis_results
        }
    
    except Exception as e:
        print(f"❌ 分析文件失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """主函数"""
    print("="*80)
    print("Excel文件结构分析工具")
    print("="*80)
    
    all_results = []
    
    for file_info in EXCEL_FILES:
        result = analyze_excel_file(file_info)
        if result:
            all_results.append(result)
    
    # 保存分析结果
    output_file = project_root / "docs" / "excel_structure_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n{'='*80}")
    print(f"分析完成！结果已保存到: {output_file}")
    print(f"{'='*80}")
    
    # 生成汇总报告
    print("\n汇总报告:")
    print("-" * 80)
    for result in all_results:
        print(f"\n{result['file_name']} ({result['source_code']} {result['dataset_type']}):")
        print(f"  Sheet数量: {len(result['sheets'])}")
        for sheet in result['sheets']:
            if "error" not in sheet:
                print(f"    - {sheet['sheet_name']}: {sheet['total_rows']}行 x {sheet['total_cols']}列")


if __name__ == "__main__":
    main()
