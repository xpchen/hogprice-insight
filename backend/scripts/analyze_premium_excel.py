"""
分析升贴水Excel文件结构
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import pandas as pd
import openpyxl

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_excel():
    """分析Excel文件结构"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "4.1、生猪期货升贴水数据（盘面结算价）.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("分析升贴水Excel文件结构")
    print("=" * 80)
    
    try:
        # 使用openpyxl读取，保留格式
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        print(f"\n所有sheet名称: {wb.sheetnames}")
        
        # 分析每个sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"Sheet: {sheet_name}")
            print(f"{'='*80}")
            
            ws = wb[sheet_name]
            
            # 显示前20行数据
            print(f"\n前20行数据:")
            for row_idx in range(1, min(21, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(30, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        row_data.append(f"{col_letter}{row_idx}={cell.value}")
                
                if row_data:
                    print(f"  行{row_idx}: {', '.join(row_data[:10])}")  # 只显示前10列
        
        # 尝试用pandas读取第一个sheet，查看数据结构
        print(f"\n\n使用pandas读取第一个sheet:")
        try:
            df = pd.read_excel(excel_path, sheet_name=0, header=None, engine='openpyxl')
            print(f"  形状: {df.shape}")
            print(f"\n前10行数据:")
            print(df.head(10).to_string())
        except Exception as e:
            print(f"  pandas读取失败: {e}")
    
    except Exception as e:
        print(f"读取失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel()
