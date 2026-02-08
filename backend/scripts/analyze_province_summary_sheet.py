"""分析重点省份旬度汇总sheet结构"""
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from openpyxl import load_workbook

# 设置输出编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_sheet():
    """分析重点省份旬度汇总sheet"""
    script_dir = Path(__file__).parent.parent.parent
    excel_file = script_dir / "docs" / "生猪" / "集团企业" / "3.1、集团企业出栏跟踪【分省区】.xlsx"
    
    if not excel_file.exists():
        print(f"Excel文件不存在: {excel_file}")
        return
    
    wb = load_workbook(excel_file, data_only=True)
    
    # 查找"汇总"或"重点省区汇总"sheet
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "汇总" in sheet_name:
            target_sheet = sheet_name
            break
    
    if not target_sheet:
        print("未找到汇总sheet")
        return
    
    print(f"分析Sheet: {target_sheet}")
    print("=" * 80)
    
    ws = wb[target_sheet]
    
    # 读取前30行分析结构
    print(f"\n行数: {ws.max_row}, 列数: {ws.max_column}")
    print("\n前30行数据:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        row_data = []
        for j, cell in enumerate(row[:20], 1):  # 只显示前20列
            if cell is None:
                row_data.append("")
            elif isinstance(cell, (int, float)):
                row_data.append(str(cell))
            else:
                cell_str = str(cell)[:20]
                row_data.append(cell_str)
        print(f"  Row {i:2d}: {row_data}")
    
    # 分析结构
    print("\n" + "=" * 80)
    print("结构分析:")
    print("=" * 80)
    
    # 第1行：地区分组
    row1 = [str(cell) if cell else "" for cell in ws[1]]
    print(f"\n第1行（地区分组）: {row1[:15]}")
    
    # 第2行：指标名称
    row2 = [str(cell) if cell else "" for cell in ws[2]]
    print(f"\n第2行（指标名称）: {row2[:15]}")
    
    # 分析地区分组
    regions = {}
    current_region = None
    region_start_col = None
    for idx, cell in enumerate(row1, 1):
        if cell and cell.strip():
            cell_str = cell.strip()
            # 检查是否是地区名称
            if cell_str in ["广东", "四川", "贵州", "全国CR20", "全国CR5", "西南", "华南", "东北"]:
                if current_region:
                    regions[current_region] = (region_start_col, idx - 1)
                current_region = cell_str
                region_start_col = idx
    if current_region:
        regions[current_region] = (region_start_col, len(row1))
    
    print(f"\n地区分组: {regions}")
    
    # 分析每个地区的指标
    print("\n各地区指标:")
    for region, (start_col, end_col) in regions.items():
        print(f"\n{region} (列 {start_col}-{end_col}):")
        region_metrics = []
        for col_idx in range(start_col, min(end_col + 1, len(row2) + 1)):
            if col_idx <= len(row2):
                metric = row2[col_idx - 1] if row2[col_idx - 1] else ""
                if metric.strip():
                    region_metrics.append(f"  列{col_idx}: {metric}")
        print("\n".join(region_metrics))
    
    # 分析数据行（第3行开始）
    print("\n" + "=" * 80)
    print("数据行分析（前10行）:")
    print("=" * 80)
    for i in range(3, min(13, ws.max_row + 1)):
        row = ws[i]
        date_cell = row[0].value if len(row) > 0 else None
        date_cell2 = row[1].value if len(row) > 1 else None
        print(f"\n第{i}行:")
        print(f"  第1列: {date_cell}")
        print(f"  第2列: {date_cell2}")
        # 显示前几个数据列
        data_sample = []
        for j in range(2, min(8, len(row))):
            val = row[j].value
            if val is not None:
                data_sample.append(f"列{j+1}={val}")
        if data_sample:
            print(f"  数据样本: {', '.join(data_sample)}")

if __name__ == "__main__":
    analyze_sheet()
