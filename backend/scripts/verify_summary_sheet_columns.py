"""
核对 3.2、集团企业月度数据跟踪.xlsx 的「汇总」sheet 与 r04 列映射是否一致。
用法：在项目根目录执行
  python -m backend.scripts.verify_summary_sheet_columns [path_to_3.2.xlsx]
不传路径时默认 docs/0306_生猪/生猪/3.2、集团企业月度数据跟踪.xlsx
"""
import sys
from pathlib import Path

# 期望的列布局（与 r04 _read_summary 一致）
# A=1 旬度, B=2 日期, C-G 广东(3-7), H-L 四川(8-12), M-Q 贵州(13-17)
EXPECTED_COL_MAP = {
    1: "旬度",
    2: "日期",
    3: "广东-出栏计划",
    4: "广东-实际出栏量",
    5: "广东-计划完成率",
    6: "广东-均重",
    7: "广东-均价",
    8: "四川-出栏计划",
    9: "四川-实际出栏量",
    10: "四川-计划完成率",
    11: "四川-均重",
    12: "四川-计划均重",
    13: "贵州-计划出栏量",
    14: "贵州-实际出栏量",
    15: "贵州-计划达成率",
    16: "贵州-实际均重",
    17: "贵州-计划均重",
}


def main():
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent
    default_path = repo_root / "docs" / "0306_生猪" / "生猪" / "3.2、集团企业月度数据跟踪.xlsx"

    path = sys.argv[1] if len(sys.argv) > 1 else default_path
    path = Path(path)
    if not path.exists():
        print(f"文件不存在: {path}")
        print("请传入正确的 3.2 集团企业月度数据跟踪.xlsx 路径")
        return 1

    try:
        import openpyxl
    except ImportError:
        print("请安装 openpyxl: pip install openpyxl")
        return 1

    wb = openpyxl.load_workbook(path, data_only=True)
    if "汇总" not in wb.sheetnames:
        print("该 Excel 中未找到「汇总」sheet，当前 sheets:", wb.sheetnames)
        return 1

    ws = wb["汇总"]
    print("=" * 60)
    print("汇总 sheet 结构核对")
    print("=" * 60)

    # 第 1、2 行：表头（可能合并）
    print("\n【第 1 行】预期为省份等分组:")
    row1 = [ws.cell(row=1, column=c).value for c in range(1, 18)]
    for c, v in enumerate(row1, 1):
        print(f"  列{c}({chr(64+c)}): {v!r}")

    print("\n【第 2 行】预期为指标名（出栏计划、实际出栏量等）:")
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 18)]
    for c, v in enumerate(row2, 1):
        exp = EXPECTED_COL_MAP.get(c, "")
        print(f"  列{c}({chr(64+c)}): {v!r}  <- 期望: {exp}")

    # 数据行 3.. 最后
    print("\n【数据行 3～末尾】旬度 | 日期 | 各列值（与上表头对应）:")
    for r in range(3, min(20, ws.max_row + 1)):
        period = ws.cell(row=r, column=1).value
        dt = ws.cell(row=r, column=2).value
        vals = [ws.cell(row=r, column=c).value for c in range(3, 18)]
        print(f"  行{r}: 旬度={period!r} 日期={dt!r}")
        print(f"        C～Q(3-17): {vals}")

    wb.close()
    print("\n若实际列顺序与 EXPECTED_COL_MAP 不一致，需修改 r04 _read_summary 的 COL_MAP。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
