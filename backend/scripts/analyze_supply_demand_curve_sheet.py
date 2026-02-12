"""
分析【生猪产业数据】.xlsx中的"供需曲线"sheet
用于E3. 供需曲线 - 长周期猪价推演
"""
# -*- coding: utf-8 -*-
import sys
import io
from pathlib import Path
import json
import openpyxl
from datetime import datetime

script_dir = Path(__file__).parent.parent
sys.path.insert(0, str(script_dir))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_supply_demand_curve_sheet():
    """分析供需曲线sheet结构"""
    excel_path = Path(script_dir.parent) / "docs" / "生猪" / "2、【生猪产业数据】.xlsx"
    
    if not excel_path.exists():
        print(f"文件不存在: {excel_path}")
        return
    
    print("=" * 80)
    print("分析【供需曲线】sheet结构")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        if "供需曲线" not in wb.sheetnames:
            print("未找到'供需曲线'sheet")
            wb.close()
            return
        
        ws = wb["供需曲线"]
        
        print(f"\nSheet: 供需曲线")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        
        # 显示前20行数据
        print(f"\n前20行数据:")
        for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
            row_data = []
            for col_idx, cell in enumerate(row[:15], 1):  # 显示前15列
                if cell is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_str = str(cell)
                    if len(cell_str) > 30:
                        cell_str = cell_str[:30] + "..."
                    row_data.append(f"{col_letter}:{cell_str}")
            if row_data:
                print(f"  行{row_idx}: {', '.join(row_data)}")
        
        # 查找表头行
        print(f"\n查找表头行:")
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            row_str = ' '.join([str(cell) for cell in row[:15] if cell])
            if any(keyword in row_str for keyword in ['日期', '月份', '时间', 'Date', 'Month', '定点屠宰', '猪价', '系数']):
                print(f"  可能的表头行: 第{row_idx}行")
                for col_idx, cell in enumerate(row[:15], 1):
                    if cell:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        print(f"    {col_letter}列: {cell}")
        
        # 分析数据列结构
        print(f"\n分析数据列结构（前30列）:")
        # 尝试找到日期列和数据列
        header_row_idx = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            row_str = ' '.join([str(cell) for cell in row[:15] if cell])
            if any(keyword in row_str for keyword in ['日期', '月份', '定点屠宰', '猪价']):
                header_row_idx = row_idx
                break
        
        if header_row_idx:
            print(f"  找到表头行: 第{header_row_idx}行")
            header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
            
            # 查找关键列
            date_col_idx = None
            slaughter_col_idx = None
            price_col_idx = None
            
            for col_idx, cell in enumerate(header_row[:30], 1):
                if cell:
                    cell_str = str(cell).strip()
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    if any(keyword in cell_str for keyword in ['日期', '月份', '时间', 'Date', 'Month']):
                        date_col_idx = col_idx
                        print(f"    日期列: {col_letter}列 (索引{col_idx})")
                    elif any(keyword in cell_str for keyword in ['定点屠宰', '屠宰系数', '屠宰']):
                        slaughter_col_idx = col_idx
                        print(f"    定点屠宰列: {col_letter}列 (索引{col_idx})")
                    elif any(keyword in cell_str for keyword in ['猪价', '价格系数', '价格']):
                        price_col_idx = col_idx
                        print(f"    猪价列: {col_letter}列 (索引{col_idx})")
            
            # 显示数据样本（表头行之后的前10行）
            print(f"\n数据样本（表头行之后的前10行）:")
            for data_row_idx in range(header_row_idx + 1, min(header_row_idx + 11, ws.max_row + 1)):
                row = list(ws.iter_rows(min_row=data_row_idx, max_row=data_row_idx, values_only=True))[0]
                row_data = []
                
                if date_col_idx and len(row) >= date_col_idx:
                    date_val = row[date_col_idx - 1]
                    row_data.append(f"日期:{date_val}")
                
                if slaughter_col_idx and len(row) >= slaughter_col_idx:
                    slaughter_val = row[slaughter_col_idx - 1]
                    row_data.append(f"屠宰:{slaughter_val}")
                
                if price_col_idx and len(row) >= price_col_idx:
                    price_val = row[price_col_idx - 1]
                    row_data.append(f"猪价:{price_val}")
                
                if row_data:
                    print(f"  行{data_row_idx}: {', '.join(row_data)}")
        
        # 统计有效数据行数
        print(f"\n统计有效数据行数:")
        data_row_count = 0
        for row_idx in range(1, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            # 检查是否有非空单元格（至少前5列中有数据）
            if any(cell is not None and str(cell).strip() != '' for cell in row[:5]):
                data_row_count += 1
        
        print(f"  有效数据行数: {data_row_count}")
        
        wb.close()
        
    except Exception as e:
        print(f"分析失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_supply_demand_curve_sheet()
