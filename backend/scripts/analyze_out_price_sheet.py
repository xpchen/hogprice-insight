"""详细分析出栏价sheet的结构，特别是日期跨列和子列"""
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
import json

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# 设置UTF-8编码输出（Windows兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def analyze_out_price_sheet():
    """分析出栏价sheet的详细结构"""
    file_path = project_root / "docs" / "2026年2月2日涌益咨询日度数据 - 副本.xlsx"
    sheet_name = "出栏价"
    
    print(f"分析文件: {file_path}")
    print(f"Sheet: {sheet_name}")
    print("="*80)
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # 读取前10行和前50列
    print("\n前10行 x 前50列的结构:")
    print("-"*80)
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(51, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value:
                # 截断过长的值
                str_val = str(value)
                if len(str_val) > 20:
                    str_val = str_val[:17] + "..."
                row_data.append(str_val)
            else:
                row_data.append("")
        
        print(f"\n第{row_idx}行:")
        for col_idx, val in enumerate(row_data[:30], 1):  # 只显示前30列
            if val:
                print(f"  列{col_idx}: {val}")
    
    # 分析日期跨列结构
    print("\n" + "="*80)
    print("分析日期跨列结构:")
    print("-"*80)
    
    # 读取第1行和第2行
    date_row = []
    subheader_row = []
    
    for col_idx in range(1, min(100, ws.max_column + 1)):
        date_cell = ws.cell(1, col_idx)
        subheader_cell = ws.cell(2, col_idx)
        
        date_val = date_cell.value
        subheader_val = subheader_cell.value
        
        # 检查合并单元格
        merged_date = None
        merged_subheader = None
        
        for merged_range in ws.merged_cells.ranges:
            if date_cell.coordinate in merged_range:
                top_left = ws[merged_range.min_col][merged_range.min_row]
                merged_date = top_left.value
            if subheader_cell.coordinate in merged_range:
                top_left = ws[merged_range.min_col][merged_range.min_row]
                merged_subheader = top_left.value
        
        date_row.append(merged_date if merged_date else date_val)
        subheader_row.append(merged_subheader if merged_subheader else subheader_val)
    
    # 识别日期组
    print("\n日期组识别（前100列）:")
    current_date = None
    current_subheaders = []
    col_start = None
    
    for col_idx in range(min(100, len(date_row))):
        date_val = date_row[col_idx]
        subheader_val = subheader_row[col_idx]
        
        if date_val and str(date_val).strip():
            # 新日期组开始
            if current_date and col_start:
                print(f"\n日期组: {current_date}")
                print(f"  列范围: {col_start} - {col_idx-1}")
                print(f"  子列: {current_subheaders}")
            
            current_date = date_val
            current_subheaders = []
            col_start = col_idx + 1
        
        if subheader_val and str(subheader_val).strip():
            current_subheaders.append(str(subheader_val).strip())
    
    # 最后一个日期组
    if current_date and col_start:
        print(f"\n日期组: {current_date}")
        print(f"  列范围: {col_start} - {len(date_row)}")
        print(f"  子列: {current_subheaders}")
    
    # 检查是否有"去年同期"、"同比"、"明日预计"等列
    print("\n" + "="*80)
    print("查找特殊列（去年同期、同比、明日预计）:")
    print("-"*80)
    
    special_keywords = ["去年同期", "同比", "明日预计", "较昨日", "涨跌"]
    
    for row_idx in range(1, min(5, ws.max_row + 1)):
        for col_idx in range(1, min(100, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value:
                str_val = str(value)
                for keyword in special_keywords:
                    if keyword in str_val:
                        print(f"  找到 '{keyword}' 在第{row_idx}行第{col_idx}列: {str_val}")
    
    wb.close()


if __name__ == "__main__":
    analyze_out_price_sheet()
