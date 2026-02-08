"""分析汇总sheet结构（旬度数据）"""
# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from openpyxl import load_workbook

# 设置输出编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_summary_sheet():
    """分析汇总sheet（包含旬度数据）"""
    script_dir = Path(__file__).parent.parent.parent
    
    # 检查多个可能的文件
    excel_files = [
        script_dir / "docs" / "生猪" / "集团企业" / "3.1、集团企业出栏跟踪【分省区】.xlsx",
        script_dir / "docs" / "生猪" / "集团企业" / "3.2、集团企业月度数据跟踪.xlsx",
    ]
    
    for excel_file in excel_files:
        if not excel_file.exists():
            continue
        
        print(f"\n{'='*80}")
        print(f"分析文件: {excel_file.name}")
        print(f"{'='*80}")
        
        wb = load_workbook(excel_file, data_only=True)
        print(f"Sheet列表: {wb.sheetnames}")
        
        # 查找包含"汇总"的sheet
        for sheet_name in wb.sheetnames:
            if "汇总" in sheet_name and sheet_name != "重点省区汇总":
                print(f"\n{'='*80}")
                print(f"分析Sheet: {sheet_name}")
                print(f"{'='*80}")
                
                ws = wb[sheet_name]
                print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
                
                # 读取前20行
                print("\n前20行数据:")
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                    row_data = []
                    for j, cell in enumerate(row[:25], 1):  # 显示前25列
                        if cell is None:
                            row_data.append("")
                        elif isinstance(cell, (int, float)):
                            row_data.append(str(cell))
                        else:
                            cell_str = str(cell)[:25]
                            row_data.append(cell_str)
                    print(f"  Row {i:2d}: {row_data}")
                
                # 分析结构
                print("\n结构分析:")
                print("-" * 80)
                
                # 读取前3行分析
                row1 = [str(cell).strip() if cell else "" for cell in ws[1]]
                row2 = [str(cell).strip() if cell else "" for cell in ws[2]]
                row3 = [str(cell).strip() if cell else "" for cell in ws[3]] if ws.max_row >= 3 else []
                
                print(f"\n第1行（地区分组）: {row1[:20]}")
                print(f"\n第2行（指标名称）: {row2[:20]}")
                if row3:
                    print(f"\n第3行（时间/数据）: {row3[:20]}")
                
                # 分析地区分组
                print("\n地区分组分析:")
                regions_found = []
                for idx, cell in enumerate(row1, 1):
                    if cell and cell.strip():
                        cell_str = cell.strip()
                        if cell_str in ["广东", "四川", "贵州", "全国CR20", "全国CR5", "西南", "华南", "东北"]:
                            regions_found.append((idx, cell_str))
                print(f"找到的地区: {regions_found}")
                
                # 分析指标
                print("\n指标分析:")
                metrics_found = []
                for idx, cell in enumerate(row2, 1):
                    if cell and cell.strip():
                        cell_str = cell.strip()
                        if any(keyword in cell_str for keyword in ["计划", "出栏", "完成", "达成", "均重", "均价"]):
                            metrics_found.append((idx, cell_str))
                print(f"找到的指标: {metrics_found[:15]}")

if __name__ == "__main__":
    analyze_summary_sheet()
