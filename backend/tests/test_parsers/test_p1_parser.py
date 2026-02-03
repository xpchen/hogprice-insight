"""P1解析器单元测试"""
import pytest
from datetime import date
from openpyxl import Workbook
from app.services.ingestors.parsers.p1_narrow_date_rows import P1NarrowDateRowsParser


def test_p1_parse_narrow_table():
    """测试P1解析器处理窄表"""
    parser = P1NarrowDateRowsParser()
    
    # 创建测试workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "价格+宰量"
    
    # 设置表头
    ws['A1'] = "日期"
    ws['B1'] = "全国均价"
    ws['C1'] = "日屠宰量合计1"
    
    # 设置数据
    ws['A2'] = date(2026, 2, 1)
    ws['B2'] = 15.5
    ws['C2'] = 1000
    
    ws['A3'] = date(2026, 2, 2)
    ws['B3'] = 15.8
    ws['C3'] = 1100
    
    # 配置
    sheet_config = {
        "sheet_name": "价格+宰量",
        "header": {"header_row": 1},
        "date_col": "日期",
        "metrics": [
            {
                "metric_key": "YY_D_PRICE_NATION_AVG",
                "metric_name": "全国均价",
                "unit": "元/公斤",
                "raw_header": "全国均价",
                "tags": {"scope": "nation"}
            }
        ],
        "geo": {"type": "NATION"}
    }
    
    profile_defaults = {
        "period_type": "day",
        "freq": "D"
    }
    
    # 解析
    observations = parser.parse(
        sheet_data=ws,
        sheet_config=sheet_config,
        profile_defaults=profile_defaults,
        source_code="YONGYI",
        batch_id=1
    )
    
    # 验证
    assert len(observations) == 2
    assert observations[0]["metric_key"] == "YY_D_PRICE_NATION_AVG"
    assert observations[0]["value"] == 15.5
    assert observations[0]["obs_date"] == date(2026, 2, 1)
