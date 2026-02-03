"""合并单元格处理测试"""
import pytest
from openpyxl import Workbook
from app.utils.merged_cell_handler import forward_fill_merged_cells, get_merged_cell_value


def test_forward_fill_merged_cells():
    """测试合并单元格值向下填充"""
    wb = Workbook()
    ws = wb.active
    
    # 设置合并单元格
    ws['A1'] = "日期"
    ws.merge_cells('B1:D1')
    ws['B1'] = "2026-02-01"
    
    # 填充
    forward_fill_merged_cells(ws)
    
    # 验证
    assert ws['B1'].value == "2026-02-01"
    assert ws['C1'].value == "2026-02-01"
    assert ws['D1'].value == "2026-02-01"
