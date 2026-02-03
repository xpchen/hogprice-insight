"""统一导入工作流集成测试"""
import pytest
from datetime import date
from sqlalchemy.orm import Session
from app.core.database import SessionLocal
from app.services.ingestors.unified_ingestor import unified_import
from app.models.import_batch import ImportBatch
from app.models.raw_file import RawFile
from app.models.raw_sheet import RawSheet
from openpyxl import Workbook
from io import BytesIO


@pytest.fixture
def db():
    """数据库会话fixture"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def test_unified_import_workflow(db: Session):
    """测试统一导入工作流"""
    # 创建测试workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "价格+宰量"
    ws['A1'] = "日期"
    ws['B1'] = "全国均价"
    ws['A2'] = date(2026, 2, 1)
    ws['B2'] = 15.5
    
    # 保存到BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_content = buffer.read()
    
    # 执行导入
    result = unified_import(
        db=db,
        file_content=file_content,
        filename="test.xlsx",
        uploader_id=1,
        dataset_type="YONGYI_DAILY",
        source_code="YONGYI"
    )
    
    # 验证结果
    assert result.get("success") is True
    assert result.get("batch_id") is not None
    
    # 验证raw_file和raw_sheet已创建
    batch_id = result.get("batch_id")
    raw_file = db.query(RawFile).filter(RawFile.batch_id == batch_id).first()
    assert raw_file is not None
    
    raw_sheets = db.query(RawSheet).filter(RawSheet.raw_file_id == raw_file.id).all()
    assert len(raw_sheets) > 0
