"""
Reader 03 - 集团企业出栏跟踪【分省区】
Source: 3.1、集团企业出栏跟踪【分省区】.xlsx (16 sheets)
Target: fact_enterprise_daily
Schema: (trade_date, company_code, region_code, metric_type, value, unit, batch_id)

Sheet layout summary
====================
CR5日度:          Row1=headers(col2..10), Row2+=data; col1=date(serial), col2=实际出栏, col4=全国均价, col5..9=company volumes
重点省区汇总:     Row1=region groups, Row2=province headers, Row3+=data; col2=date
西南汇总:         Row1=region groups, Row2=metric headers, Row3+=data; col2=date
四川/贵州/广西:   Summary sheets with planned/actual per company per day
四川日度/贵州日度/广西日度/广东日度/江西日度/湖南日度: col2=date, col3=unit, col4=indicator, col5..=companies
陕西日度:         col1=date, col2=indicator, col3..=companies
东北日度:         Row1=province groups, Row2=headers, Row3+=data; col1=date, col2=indicator, col3..=province subtotals + companies
华南合计:         Row1=province groups, Row2=company headers, Row3+=data; col2=date
MS猪价:          Row1=region groups, Row2=provinces, Row5=indicator names, Row9+=data; col1=date, col2..=prices per province
"""
import logging
from typing import Optional
import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, clean_value, province_to_code

logger = logging.getLogger(__name__)

# ── company name -> code mapping ──
COMPANY_NAME_MAP = {
    "牧原": "MUYUAN", "温氏": "WENS", "双胞胎": "SHUANGBAOTAI",
    "新希望": "XINWUFENG",  # Note: 新希望六和
    "德康": "DEKANG", "正邦": "ZHENGBANG", "巨星": "JUXING",
    "正大": "ZHENGDA", "驰阳": "CHIYANG", "日泉": "RIQUAN",
    "齐全": "QIQUAN", "力源": "LIYUAN", "铁骑力士": "TIEQI",
    "兴新鑫": "XINGXINXIN", "海大": "HAIDA", "富之源": "FUZHIYUAN",
    "傲农": "AONONG", "大北农": "DABEINONG", "罗牛山": "LUONIUSHAN",
    "铁骑": "TIEQI", "扬翔": "YANGXIANG", "雄桂": "XIONGGUI",
    "园丰": "YUANFENG", "唐人神": "TANGRENSHEN", "五祥": "WUXIANG",
    "京基智农": "JINGJI", "京基": "JINGJI", "神农": "SHENNONG",
    "新天地": "XINTIANDI", "天邦": "TIANBANG", "富凤": "FUFENG",
    "广垦": "GUANGKEN", "天农": "TIANNONG", "金新农": "JINXINNONG",
    "万科": "VANKE", "宝金": "BAOJIN", "大广": "DAGUANG",
    "大家": "DAJIA", "东方希望": "DFXW", "中粮": "COFCO",
    "石羊": "SHIYANG", "正能": "ZHENGNENG", "天康": "TIANKANG",
    "新五丰": "XINWUFENG_SZ",  # 新五丰 is different from 新希望
    "禾丰": "HEFENG", "大伟嘉": "DAWEIJIA", "杨翔": "YANGXIANG",
    "利源": "LIYUAN_HN",  # 湖南利源
    "新五丰": "XINWUFENG_SZ",
    "食出宝金": "BAOJIN",
    "东瑞": "DONGRUI",
}

# CR5日度 sheet: column -> company
CR5_COL_COMPANY = {
    5: "MUYUAN", 6: "WENS", 7: "SHUANGBAOTAI",
    8: "XINWUFENG", 9: "DEKANG",
}

# 重点省区汇总: column -> province
SUMMARY_COL_PROVINCE = {
    3: "SICHUAN", 4: "GUIZHOU", 5: "GUANGXI", 6: "SOUTHWEST",
    7: "JIANGXI", 8: "GUANGDONG", 9: "HUNAN",
    10: "HEILONGJIANG", 11: "JILIN", 12: "NEIMENGGU", 13: "LIAONING", 14: "NORTHEAST",
    15: "SHAANXI",
}

# MS猪价: column -> province (row2 has names)
MS_COL_PROVINCE = {
    2: "HEILONGJIANG", 3: "JILIN", 4: "LIAONING", 5: "NEIMENGGU",
    6: "SHANXI", 7: "HEBEI", 8: "SHANDONG", 9: "ANHUI",
    10: "JIANGSU", 11: "ZHEJIANG", 12: "HENAN", 13: "HUBEI",
    14: "HUNAN", 15: "JIANGXI", 16: "SICHUAN", 17: "GUIZHOU",
    18: "GUANGXI", 19: "YUNNAN",
}

# Province daily sheets: sheet name -> region_code
PROVINCE_DAILY_SHEETS = {
    "四川日度": "SICHUAN",
    "贵州日度": "GUIZHOU",
    "广西日度": "GUANGXI",
    "广东日度": "GUANGDONG",
    "江西日度": "JIANGXI",
    "湖南日度": "HUNAN",
}


def _company_code(name: str) -> Optional[str]:
    """Map Chinese company name to company_code."""
    if not name or not isinstance(name, str):
        return None
    name = name.strip().replace("\xa0", "")
    return COMPANY_NAME_MAP.get(name)


class EnterpriseProvinceReader(BaseSheetReader):
    """Reader for 3.1 集团企业出栏跟踪【分省区】"""

    FILE_PATTERN = "集团企业出栏跟踪"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        records = []
        wb = openpyxl.load_workbook(filepath, data_only=True)
        try:
            records += self._read_cr5(wb)
            records += self._read_province_summary(wb)
            records += self._read_southwest_summary(wb)
            records += self._read_province_detail_sheets(wb)
            records += self._read_province_daily_sheets(wb)
            records += self._read_shaanxi_daily(wb)
            records += self._read_northeast_daily(wb)
            records += self._read_south_summary(wb)
            records += self._read_ms_price(wb)
        finally:
            wb.close()

        logger.info(f"[r03] Total records: {len(records)}")
        return {"fact_enterprise_daily": records}

    # ────────────────────────────────────────────
    # CR5日度
    # ────────────────────────────────────────────
    def _read_cr5(self, wb) -> list[dict]:
        """
        CR5日度: col1=date(serial), col2=实际出栏(CR5 total), col3=月度计划,
        col4=全国均价, col5=牧原, col6=温氏, col7=双胞胎, col8=新希望, col9=德康,
        col10=牧原均重
        """
        if "CR5日度" not in wb.sheetnames:
            return []
        ws = wb["CR5日度"]
        records = []
        for row in range(2, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=1).value)
                if not dt:
                    continue

                # CR5 total output (col 2)
                total_val = clean_value(ws.cell(row=row, column=2).value)
                if total_val is not None:
                    records.append({
                        "trade_date": dt, "company_code": "CR5",
                        "region_code": "NATION", "metric_type": "output_cumulative",
                        "value": total_val, "unit": "万头", "batch_id": self.batch_id,
                    })

                # CR5 月度计划 (col 3)，与旧版图表「计划量」一致
                plan_val = clean_value(ws.cell(row=row, column=3).value)
                if plan_val is not None:
                    records.append({
                        "trade_date": dt, "company_code": "CR5",
                        "region_code": "NATION", "metric_type": "planned_volume",
                        "value": plan_val, "unit": "万头", "batch_id": self.batch_id,
                    })

                # National avg price (col 4)
                price_val = clean_value(ws.cell(row=row, column=4).value)
                if price_val is not None:
                    records.append({
                        "trade_date": dt, "company_code": "CR5",
                        "region_code": "NATION", "metric_type": "avg_price",
                        "value": price_val, "unit": "元/公斤", "batch_id": self.batch_id,
                    })

                # Individual companies (col 5-9)
                for col, code in CR5_COL_COMPANY.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": code,
                            "region_code": "NATION", "metric_type": "output_cumulative",
                            "value": val, "unit": "万头", "batch_id": self.batch_id,
                        })

                # Muyuan avg weight (col 10)
                wt = clean_value(ws.cell(row=row, column=10).value)
                if wt is not None:
                    records.append({
                        "trade_date": dt, "company_code": "MUYUAN",
                        "region_code": "NATION", "metric_type": "avg_weight",
                        "value": wt, "unit": "公斤", "batch_id": self.batch_id,
                    })
            except Exception as e:
                logger.warning(f"[r03] CR5日度 row {row} error: {e}")
        logger.info(f"[r03] CR5日度: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 重点省区汇总
    # ────────────────────────────────────────────
    def _read_province_summary(self, wb) -> list[dict]:
        """
        重点省区汇总: Row1=region groups, Row2=province headers
        Row3+: col1=计划量(ignored), col2=date, col3..15=province output values
        """
        if "重点省区汇总" not in wb.sheetnames:
            return []
        ws = wb["重点省区汇总"]
        records = []
        for row in range(3, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=2).value)
                if not dt:
                    continue
                for col, region in SUMMARY_COL_PROVINCE.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": "SAMPLE",
                            "region_code": region, "metric_type": "planned_volume",
                            "value": val, "unit": "头", "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] 重点省区汇总 row {row} error: {e}")
        logger.info(f"[r03] 重点省区汇总: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 西南汇总
    # ────────────────────────────────────────────
    def _read_southwest_summary(self, wb) -> list[dict]:
        """
        西南汇总: Row1=region groups(四川/广西/贵州/西南样本企业)
        Row2=metric headers, Row3+=data; col2=date
        四川: col3=挂牌, col4=实际成交, col5=成交率, col6=计划日均, col7=最新日均, col8=价格MS
        广西: col9=挂牌, col10=实际成交, col11=成交率, col12=价格MS
        贵州: col13=出栏, col14=价格MS
        西南样本: col15=量, col16=重, col17=量MA7, col18=重MA7, col19=价
        """
        if "西南汇总" not in wb.sheetnames:
            return []
        ws = wb["西南汇总"]
        records = []

        METRIC_MAP = {
            3: ("SICHUAN", "listed_volume", "头"),
            4: ("SICHUAN", "actual_volume", "头"),   # D列 实际成交 → 日度出栏
            5: ("SICHUAN", "deal_rate", "%"),       # E列 成交率
            6: ("SICHUAN", "planned_daily", "头"),  # F列 计划日均 → 计划出栏
            8: ("SICHUAN", "ms_price", "元/公斤"),
            9: ("GUANGXI", "listed_volume", "头"),
            10: ("GUANGXI", "actual_volume", "头"),
            11: ("GUANGXI", "deal_rate", "%"),
            12: ("GUANGXI", "ms_price", "元/公斤"),
            13: ("GUIZHOU", "output", "头"),
            14: ("GUIZHOU", "ms_price", "元/公斤"),
            15: ("SOUTHWEST", "sample_volume", "头"),
            16: ("SOUTHWEST", "sample_weight", "公斤"),
            19: ("SOUTHWEST", "sample_price", "元/公斤"),
        }

        for row in range(3, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=2).value)
                if not dt:
                    continue
                for col, (region, metric, unit) in METRIC_MAP.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        # 成交率：Excel 多为小数 0.85=85%，与 P8 一致转为 0-100 存库
                        if metric == "deal_rate" and unit == "%":
                            v = float(val)
                            if 0 < v <= 1.5:
                                val = round(v * 100, 2)
                        records.append({
                            "trade_date": dt, "company_code": "SAMPLE",
                            "region_code": region, "metric_type": metric,
                            "value": val, "unit": unit, "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] 西南汇总 row {row} error: {e}")
        logger.info(f"[r03] 西南汇总: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 四川 / 贵州 / 广西 (province detail, non-daily)
    # ────────────────────────────────────────────
    def _read_province_detail_sheets(self, wb) -> list[dict]:
        """
        四川: Row4=company headers (col2=date); actual_sales cols3..10, planned cols12..19
        贵州: Row1=headers; col2=date(planned), col3..10=companies(planned); col11=date(price), col12..19=price ranges
        广西: Row1=headers; col1=date, col2..9=planned by company, col17..25=actual by company
        """
        records = []

        # ── 四川 (non-daily) ──
        if "四川" in wb.sheetnames:
            ws = wb["四川"]
            # Row 4 has company headers for actual sales (cols 3-10) and planned (cols 12-19)
            actual_companies = {}
            planned_companies = {}
            for c in range(3, 11):
                name = ws.cell(row=4, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    actual_companies[c] = code
            for c in range(12, 20):
                name = ws.cell(row=4, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    planned_companies[c] = code

            for row in range(5, ws.max_row + 1):
                try:
                    dt = parse_date(ws.cell(row=row, column=2).value)
                    if not dt:
                        continue
                    for col, code in actual_companies.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": code,
                                "region_code": "SICHUAN", "metric_type": "actual_sales",
                                "value": val, "unit": "头", "batch_id": self.batch_id,
                            })
                    # Planned from col 11 date
                    dt_plan = parse_date(ws.cell(row=row, column=11).value)
                    if dt_plan:
                        for col, code in planned_companies.items():
                            val = clean_value(ws.cell(row=row, column=col).value)
                            if val is not None:
                                records.append({
                                    "trade_date": dt_plan, "company_code": code,
                                    "region_code": "SICHUAN", "metric_type": "planned_volume",
                                    "value": val, "unit": "头", "batch_id": self.batch_id,
                                })
                except Exception as e:
                    logger.warning(f"[r03] 四川 row {row} error: {e}")

        # ── 贵州 (non-daily) ──
        if "贵州" in wb.sheetnames:
            ws = wb["贵州"]
            # Row 1: col2=date header(planned), companies col3..10; col11=date(price), col12..19=prices
            plan_companies = {}
            for c in range(3, 11):
                name = ws.cell(row=1, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    plan_companies[c] = code

            for row in range(2, ws.max_row + 1):
                try:
                    dt = parse_date(ws.cell(row=row, column=2).value)
                    if not dt:
                        continue
                    for col, code in plan_companies.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": code,
                                "region_code": "GUIZHOU", "metric_type": "planned_volume",
                                "value": val, "unit": "头", "batch_id": self.batch_id,
                            })
                except Exception as e:
                    logger.warning(f"[r03] 贵州 row {row} error: {e}")

        # ── 广西 (non-daily) ──
        if "广西" in wb.sheetnames:
            ws = wb["广西"]
            # Row 1: col1=date, col2..9=planned companies, col17=实际销售量, col18..25=actual companies
            plan_companies = {}
            actual_companies = {}
            for c in range(2, 10):
                name = ws.cell(row=1, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    plan_companies[c] = code
            for c in range(18, 26):
                name = ws.cell(row=1, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    actual_companies[c] = code

            for row in range(2, ws.max_row + 1):
                try:
                    dt = parse_date(ws.cell(row=row, column=1).value)
                    if not dt:
                        continue
                    for col, code in plan_companies.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": code,
                                "region_code": "GUANGXI", "metric_type": "planned_volume",
                                "value": val, "unit": "头", "batch_id": self.batch_id,
                            })
                    for col, code in actual_companies.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": code,
                                "region_code": "GUANGXI", "metric_type": "actual_sales",
                                "value": val, "unit": "头", "batch_id": self.batch_id,
                            })
                except Exception as e:
                    logger.warning(f"[r03] 广西 row {row} error: {e}")

        logger.info(f"[r03] Province detail (四川/贵州/广西): {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # Province daily sheets (四川日度/贵州日度/广西日度/广东日度/江西日度/湖南日度)
    # ────────────────────────────────────────────
    def _read_province_daily_sheets(self, wb) -> list[dict]:
        """
        Common format: col2=date, col3=unit, col4=indicator, col5..N=companies
        Row1 = company headers; data starts at row 2.
        Indicators: 计划订购量, 实际销售量, 成交率, 体重范围, 价格范围, etc.
        We extract: 计划订购量 and 实际销售量 (numeric head counts).
        """
        records = []
        for sheet_name, region_code in PROVINCE_DAILY_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # Build company map from row 1 (col5+)
            company_map = {}
            for c in range(5, ws.max_column + 1):
                name = ws.cell(row=1, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    company_map[c] = code

            count = 0
            for row in range(2, ws.max_row + 1):
                try:
                    dt = parse_date(ws.cell(row=row, column=2).value)
                    if not dt:
                        continue
                    indicator_raw = ws.cell(row=row, column=4).value
                    if not indicator_raw:
                        continue
                    indicator = str(indicator_raw).strip().replace("\xa0", "")

                    # Map indicator to metric_type
                    if indicator == "计划订购量":
                        metric = "planned_volume"
                        unit = "头"
                    elif indicator == "实际销售量":
                        metric = "actual_sales"
                        unit = "头"
                    elif indicator == "成交率":
                        metric = "deal_rate"
                        unit = "%"
                    elif indicator == "二育销量":
                        metric = "secondary_fattening_volume"
                        unit = "头"
                    else:
                        # Skip non-numeric indicators (体重范围, 价格范围, etc.)
                        continue

                    for col, code in company_map.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": code,
                                "region_code": region_code, "metric_type": metric,
                                "value": val, "unit": unit, "batch_id": self.batch_id,
                            })
                            count += 1
                except Exception as e:
                    logger.warning(f"[r03] {sheet_name} row {row} error: {e}")
            logger.info(f"[r03] {sheet_name}: {count} records")
        return records

    # ────────────────────────────────────────────
    # 陕西日度
    # ────────────────────────────────────────────
    def _read_shaanxi_daily(self, wb) -> list[dict]:
        """
        陕西日度: different layout - col1=date, col2=indicator(计划量/体重范围/成交均价),
        col3..12=companies, col13=陕西(total)
        Row1=company headers, Row2+=data (groups of 3 rows per date).
        """
        if "陕西日度" not in wb.sheetnames:
            return []
        ws = wb["陕西日度"]
        records = []

        # Row 1: company headers starting col3
        company_map = {}
        for c in range(3, ws.max_column + 1):
            name = ws.cell(row=1, column=c).value
            if name and str(name).strip() == "陕西":
                company_map[c] = "SAMPLE_SHAANXI"  # total for Shaanxi
            else:
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    company_map[c] = code

        for row in range(2, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=1).value)
                if not dt:
                    continue
                indicator_raw = ws.cell(row=row, column=2).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip()

                if indicator == "计划量":
                    metric, unit = "planned_volume", "头"
                elif indicator in ("成交量", "实际成交"):
                    metric, unit = "actual_sales", "头"
                elif indicator == "成交均价":
                    # Price values may contain text like "11.6-11.95 / 12.1-12.15",
                    # only extract pure numeric values
                    metric, unit = "avg_price", "元/公斤"
                else:
                    continue

                for col, code in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": code,
                            "region_code": "SHAANXI", "metric_type": metric,
                            "value": val, "unit": unit, "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] 陕西日度 row {row} error: {e}")
        logger.info(f"[r03] 陕西日度: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 东北日度
    # ────────────────────────────────────────────
    def _read_northeast_daily(self, wb) -> list[dict]:
        """
        东北日度: Row1=province groups (col7=黑龙江, col11=吉林, col18=内蒙古, col24=辽宁)
        Row2: col3..6=province subtotals, col7..=companies per province
        Row3+: col1=date, col2=indicator, col3..=values

        Province company ranges (from row1 groups):
          黑龙江: col7..10; 吉林: col11..17; 内蒙古: col18..23; 辽宁: col24..33
        """
        if "东北日度" not in wb.sheetnames:
            return []
        ws = wb["东北日度"]
        records = []

        # Province subtotals (col3..6) map from row 2
        subtotal_map = {3: "HEILONGJIANG", 4: "JILIN", 5: "NEIMENGGU", 6: "LIAONING"}

        # Company groups by province from row 1 grouping and row 2 names
        province_groups = {
            "HEILONGJIANG": (7, 10),
            "JILIN": (11, 17),
            "NEIMENGGU": (18, 23),
            "LIAONING": (24, 33),
        }

        # Build company map from row 2
        company_map = {}  # col -> (company_code, province_code)
        for province, (start_col, end_col) in province_groups.items():
            for c in range(start_col, min(end_col + 1, ws.max_column + 1)):
                name = ws.cell(row=2, column=c).value
                code = _company_code(str(name).replace("\xa0", "")) if name else None
                if code:
                    company_map[c] = (code, province)

        for row in range(3, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=1).value)
                if not dt:
                    continue
                indicator_raw = ws.cell(row=row, column=2).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip()

                if indicator == "计划量":
                    metric, unit = "planned_volume", "头"
                elif indicator in ("成交量", "实际成交"):
                    metric, unit = "actual_sales", "头"
                elif indicator.startswith("成交价格"):
                    metric, unit = "avg_price", "元/公斤"
                else:
                    continue  # skip 体重段 etc.

                # Province subtotals (only for volume metrics)
                if metric in ("planned_volume", "actual_sales"):
                    for col, region in subtotal_map.items():
                        val = clean_value(ws.cell(row=row, column=col).value)
                        if val is not None:
                            records.append({
                                "trade_date": dt, "company_code": "SAMPLE",
                                "region_code": region, "metric_type": metric,
                                "value": val, "unit": unit, "batch_id": self.batch_id,
                            })

                # Per-company values
                for col, (code, province) in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": code,
                            "region_code": province, "metric_type": metric,
                            "value": val, "unit": unit, "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] 东北日度 row {row} error: {e}")
        logger.info(f"[r03] 东北日度: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 华南合计
    # ────────────────────────────────────────────
    def _read_south_summary(self, wb) -> list[dict]:
        """
        华南合计: Row1=province groups (col3=江西, col11=广东, col24=湖南)
        Row2=company headers + 合计 columns
        Row3+: col2=date, then data per company.
        We read subtotals per province (合计 columns).
        江西合计=col10, 广东合计=col23, 湖南合计=col30
        """
        if "华南合计" not in wb.sheetnames:
            return []
        ws = wb["华南合计"]
        records = []

        # Province subtotals from 合计 columns
        subtotal_cols = {}
        # Also read per-company data
        company_map = {}  # col -> (company_code, province_code)

        # Detect province boundaries from row 1
        province_ranges = {}
        current_province = None
        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v and isinstance(v, str) and v.strip():
                current_province = province_to_code(v.strip())
            if current_province:
                header = ws.cell(row=2, column=c).value
                if header:
                    h = str(header).strip()
                    if h in ("合计", "合计01", "合计02"):
                        subtotal_cols[c] = current_province
                    else:
                        code = _company_code(h)
                        if code:
                            company_map[c] = (code, current_province)

        for row in range(3, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=2).value)
                if not dt:
                    continue
                # Subtotals
                for col, region in subtotal_cols.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": "SAMPLE",
                            "region_code": region, "metric_type": "planned_volume",
                            "value": val, "unit": "头", "batch_id": self.batch_id,
                        })
                # Per-company
                for col, (code, region) in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": code,
                            "region_code": region, "metric_type": "planned_volume",
                            "value": val, "unit": "头", "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] 华南合计 row {row} error: {e}")
        logger.info(f"[r03] 华南合计: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # MS猪价
    # ────────────────────────────────────────────
    def _read_ms_price(self, wb) -> list[dict]:
        """
        MS猪价: Row1=region groups, Row2=province names, Row5=indicator names,
        Row6=data source, Row7=description, Row8=update time.
        Data rows start at row 9; col1=date, col2..19=price per province.
        Col20 exists in some rows (extra province like 云南).
        """
        if "MS猪价" not in wb.sheetnames:
            return []
        ws = wb["MS猪价"]
        records = []
        for row in range(9, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=1).value)
                if not dt:
                    continue
                for col, region in MS_COL_PROVINCE.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "trade_date": dt, "company_code": "MS",
                            "region_code": region, "metric_type": "ms_avg_price",
                            "value": val, "unit": "元/公斤", "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r03] MS猪价 row {row} error: {e}")
        logger.info(f"[r03] MS猪价: {len(records)} records")
        return records
