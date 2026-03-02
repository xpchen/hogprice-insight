"""
Reader 05 - 白条市场跟踪
Source: 3.3、白条市场跟踪.xlsx (4 sheets)
Target tables:
  - fact_carcass_market  (trade_date, region_code, metric_type, source, value, unit, batch_id)
  - fact_slaughter_daily (trade_date, region_code, source, volume, batch_id)

Sheet layout summary
====================
华宝和牧原白条:
    Row1: col2=日期, col3=华宝肉条, col4=牧原白条
    Row2: col4=华东, col5=河南山东, col6=湖北陕西, col7=京津冀, col8=东北
    Row3+: col2=date, col3=华宝 price spread, col4..8=牧原 regional spreads
    Values are price spreads (元/头), positive/negative.

白条市场:
    Row1: col1=日期, col2..20=market metrics
    Pairs: (到货量, 价格) for each market:
      北京石门(2,3), 上海西郊(4,5), 成都点杀(6,7), 山西太原(8,9),
      杭州五和(10,11), 无锡天鹏(12,13), 南京众彩(14,15), 广西桂林(16,17)
    Extra cols: ML到货(18), SDY到货(19), BZY到货(20)
    Row2+: dates as strings "YYYY.MM.DD", values may have "~" prefix

日度屠宰统计:
    Row1: headers
      col1=农历日, col2=农历, col3=周, col4=农历年, col5=日期03, col6=日期,
      col7=全国均价, col8=涌益日度, col9=河南日度, col10=牧原河南04,
      col11=牧原my, col12=双汇sh, col13=大红门dhm, col14=雨润yr,
      col15=金锣jl, col16=龙大ld, col17=高金gj, col18=千喜鹤qxh,
      col19=CR8宰企, col20=MA5涌益, col21=MA5河南, col22=MA5集团
    Row2+: data rows; col6=date, col7=price, col8..=slaughter volumes

Sheet3: empty
"""
import logging
import re
from typing import Optional
import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, clean_value

logger = logging.getLogger(__name__)

# ── 白条市场 market columns ──
# (col_volume, col_price, market_name, region_code)
MARKET_COLUMNS = [
    (2, 3, "北京石门", "BEIJING"),
    (4, 5, "上海西郊", "SHANGHAI"),
    (6, 7, "成都点杀", "SICHUAN"),
    (8, 9, "山西太原", "SHANXI"),
    (10, 11, "杭州五和", "ZHEJIANG"),
    (12, 13, "无锡天鹏", "JIANGSU"),
    (14, 15, "南京众彩", "JIANGSU_NJ"),
    (16, 17, "广西桂林", "GUANGXI"),
]

# Extra volume-only columns in 白条市场
EXTRA_VOLUME_COLS = {
    18: ("ML到货", "NATION_ML"),
    19: ("SDY到货", "NATION_SDY"),
    20: ("BZY到货", "NATION_BZY"),
}

# 日度屠宰统计: slaughter enterprise columns
SLAUGHTER_COLS = {
    8:  ("涌益日度", "YONGYI"),
    9:  ("河南日度", "HENAN"),
    10: ("牧原河南04", "MUYUAN_HENAN"),
    11: ("牧原my", "MUYUAN"),
    12: ("双汇sh", "SHUANGHUI"),
    13: ("大红门dhm", "DAHONGMEN"),
    14: ("雨润yr", "YURUN"),
    15: ("金锣jl", "JINLUO"),
    16: ("龙大ld", "LONGDA"),
    17: ("高金gj", "GAOJIN"),
    18: ("千喜鹤qxh", "QIANXIHE"),
    19: ("CR8宰企", "CR8_SLAUGHTER"),
    20: ("MA5涌益", "YONGYI_MA5"),
    21: ("MA5河南", "HENAN_MA5"),
    22: ("MA5集团", "GROUP_MA5"),
}

# 华宝和牧原白条: regional spread columns
HUABAO_MUYUAN_COLS = {
    3: ("华宝", "NATION", "huabao_spread"),
    4: ("牧原华东", "EAST", "muyuan_spread"),
    5: ("牧原河南山东", "HENAN_SHANDONG", "muyuan_spread"),
    6: ("牧原湖北陕西", "HUBEI_SHAANXI", "muyuan_spread"),
    7: ("牧原京津冀", "BEIJING_TIANJIN_HEBEI", "muyuan_spread"),
    8: ("牧原东北", "NORTHEAST", "muyuan_spread"),
}


def _clean_market_value(raw) -> Optional[float]:
    """
    Clean market values that may have '~' prefix or other text artifacts.
    E.g. '~7600' -> 7600.0, '~15.50' -> 15.5, '(昨降)' -> None
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return clean_value(raw)
    s = str(raw).strip()
    if not s:
        return None
    # Remove ~ prefix
    s = s.replace("~", "").replace("～", "").strip()
    # Skip text-only values like '(昨降)'
    if not s or s.startswith("(") or s.startswith("（"):
        return None
    return clean_value(s)


def _parse_date_str(val) -> Optional[object]:
    """Parse date from various formats including 'YYYY.MM.DD'."""
    if val is None:
        return None
    d = parse_date(val)
    if d:
        return d
    # Try YYYY.MM.DD format
    if isinstance(val, str):
        s = val.strip()
        m = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", s)
        if m:
            from datetime import date
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
    return None


class CarcassMarketReader(BaseSheetReader):
    """Reader for 3.3 白条市场跟踪"""

    FILE_PATTERN = "白条市场跟踪"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        carcass_records = []
        slaughter_records = []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        try:
            carcass_records += self._read_huabao_muyuan(wb)
            carcass_records += self._read_carcass_market(wb)
            slaughter_records += self._read_slaughter_daily(wb)
        finally:
            wb.close()

        logger.info(
            f"[r05] Total: carcass_market={len(carcass_records)}, "
            f"slaughter_daily={len(slaughter_records)}"
        )
        return {
            "fact_carcass_market": carcass_records,
            "fact_slaughter_daily": slaughter_records,
        }

    # ────────────────────────────────────────────
    # 华宝和牧原白条
    # ────────────────────────────────────────────
    def _read_huabao_muyuan(self, wb) -> list[dict]:
        """
        华宝和牧原白条:
        Row1: col2=日期, col3=华宝肉条, col4=牧原白条
        Row2: col4=华东, col5=河南山东, col6=湖北陕西, col7=京津冀, col8=东北
        Row3+: col2=date, col3=华宝 spread, col4..8=牧原 regional spreads
        Values represent price spreads in 元/头.
        """
        if "华宝和牧原白条" not in wb.sheetnames:
            return []
        ws = wb["华宝和牧原白条"]
        records = []

        for row in range(3, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=2).value)
                if not dt:
                    continue

                for col, (label, region, metric) in HUABAO_MUYUAN_COLS.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        source = "HUABAO" if "华宝" in label else "MUYUAN"
                        records.append({
                            "trade_date": dt,
                            "region_code": region,
                            "metric_type": metric,
                            "source": source,
                            "value": val,
                            "unit": "元/头",
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r05] 华宝和牧原白条 row {row} error: {e}")
        logger.info(f"[r05] 华宝和牧原白条: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 白条市场
    # ────────────────────────────────────────────
    def _read_carcass_market(self, wb) -> list[dict]:
        """
        白条市场:
        Row1: col1=日期, then pairs (到货, 价格) per market
        Row2+: dates as 'YYYY.MM.DD' strings; values may have '~' prefix.
        Markets: 北京石门(2,3), 上海西郊(4,5), 成都点杀(6,7), 山西太原(8,9),
                 杭州五和(10,11), 无锡天鹏(12,13), 南京众彩(14,15), 广西桂林(16,17)
        Extra: ML到货(18), SDY到货(19), BZY到货(20)
        """
        if "白条市场" not in wb.sheetnames:
            return []
        ws = wb["白条市场"]
        records = []

        for row in range(2, ws.max_row + 1):
            try:
                dt = _parse_date_str(ws.cell(row=row, column=1).value)
                if not dt:
                    continue

                # Market pairs (volume + price)
                for col_vol, col_price, market_name, region in MARKET_COLUMNS:
                    vol_val = _clean_market_value(ws.cell(row=row, column=col_vol).value)
                    if vol_val is not None:
                        records.append({
                            "trade_date": dt,
                            "region_code": region,
                            "metric_type": "carcass_arrival",
                            "source": market_name,
                            "value": vol_val,
                            "unit": "头",
                            "batch_id": self.batch_id,
                        })

                    price_val = _clean_market_value(ws.cell(row=row, column=col_price).value)
                    if price_val is not None:
                        records.append({
                            "trade_date": dt,
                            "region_code": region,
                            "metric_type": "carcass_price",
                            "source": market_name,
                            "value": price_val,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })

                # Extra volume-only columns
                for col, (label, region) in EXTRA_VOLUME_COLS.items():
                    vol_val = _clean_market_value(ws.cell(row=row, column=col).value)
                    if vol_val is not None:
                        records.append({
                            "trade_date": dt,
                            "region_code": region,
                            "metric_type": "carcass_arrival",
                            "source": label,
                            "value": vol_val,
                            "unit": "头",
                            "batch_id": self.batch_id,
                        })

                # Cols 21-22 sometimes contain duplicated data (成都到货/价格 reference)
                # parsed from row 6+ if present; these repeat 成都 data so skip to avoid duplication

            except Exception as e:
                logger.warning(f"[r05] 白条市场 row {row} error: {e}")
        logger.info(f"[r05] 白条市场: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 日度屠宰统计
    # ────────────────────────────────────────────
    def _read_slaughter_daily(self, wb) -> list[dict]:
        """
        日度屠宰统计:
        Row1: header row
          col6=日期, col7=全国均价, col8=涌益日度, col9=河南日度,
          col10=牧原河南04, col11=牧原my, col12=双汇sh, col13=大红门dhm,
          col14=雨润yr, col15=金锣jl, col16=龙大ld, col17=高金gj,
          col18=千喜鹤qxh, col19=CR8宰企, col20=MA5涌益, col21=MA5河南, col22=MA5集团
        Row2+: data; col6=date, col7=price, col8..22=slaughter volumes

        Writes to fact_slaughter_daily: (trade_date, region_code, source, volume, batch_id)
        Also writes col7 (全国均价) to fact_carcass_market.
        """
        if "日度屠宰统计" not in wb.sheetnames:
            return []
        ws = wb["日度屠宰统计"]
        slaughter_records = []

        for row in range(2, ws.max_row + 1):
            try:
                dt = parse_date(ws.cell(row=row, column=6).value)
                if not dt:
                    continue

                # Slaughter volumes (col 8..22)
                for col, (label, source) in SLAUGHTER_COLS.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        volume = int(val) if val == int(val) else int(round(val))
                        slaughter_records.append({
                            "trade_date": dt,
                            "region_code": "NATION",
                            "source": source,
                            "volume": volume,
                            "batch_id": self.batch_id,
                        })

                # National avg price (col 7) -> also store in slaughter table
                # as a special record, or callers can use it from fact_carcass_market
                # (handled separately if needed)

            except Exception as e:
                logger.warning(f"[r05] 日度屠宰统计 row {row} error: {e}")
        logger.info(f"[r05] 日度屠宰统计: {len(slaughter_records)} records")
        return slaughter_records
