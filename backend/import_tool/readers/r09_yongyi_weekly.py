"""
r09_yongyi_weekly.py  --  涌益咨询 周度数据.xlsx (65 sheets, ~2.3 MB)

Target tables
─────────────
  fact_weekly_indicator   (week_end, week_start, region_code, indicator_code, source, value, unit, batch_id)
  fact_monthly_indicator  (month_date, region_code, indicator_code, sub_category, source, value, value_type, unit, batch_id)
"""
from __future__ import annotations

import logging
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, parse_month, clean_value, province_to_code

logger = logging.getLogger(__name__)

# ────────────────────────────────────────────────────────────────────────
# Constants
# ────────────────────────────────────────────────────────────────────────
SOURCE = "YONGYI"


def _week_start_from_end(week_end: date) -> date:
    """Given a week-ending date (typically Thursday or Friday), return the Monday."""
    return week_end - timedelta(days=6)


def _safe_region(name: Optional[str]) -> Optional[str]:
    """Convert province / area name to region_code.  Handles area names too."""
    if not name:
        return None
    name = str(name).strip().replace("\u3000", "").replace(" ", "")
    # Direct area-level mapping
    area_map = {
        "全国": "NATION", "全国1": "NATION", "全国2": "NATION",
        "全  国1": "NATION", "全  国2": "NATION",
        "东北": "NORTHEAST", "华北": "NORTH", "华东": "EAST",
        "华中": "CENTRAL", "华南": "SOUTH", "西南": "SOUTHWEST", "西北": "NORTHWEST",
    }
    code = area_map.get(name)
    if code:
        return code
    return province_to_code(name)


# ╔════════════════════════════════════════════════════════════════════════╗
# ║  READER CLASS                                                         ║
# ╚════════════════════════════════════════════════════════════════════════╝

class YongyiWeeklyReader(BaseSheetReader):
    FILE_PATTERN = "涌益咨询 周度数据"

    # ── public entry point ──────────────────────────────────────────────
    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        weekly_records: list[dict] = []
        monthly_records: list[dict] = []

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                w, m = self._dispatch(sheet_name, ws)
                weekly_records.extend(w)
                monthly_records.extend(m)
            except Exception:
                logger.exception("Sheet [%s] failed", sheet_name)

        wb.close()
        logger.info(
            "YongyiWeekly: %d weekly, %d monthly records",
            len(weekly_records), len(monthly_records),
        )
        return {
            "fact_weekly_indicator": weekly_records,
            "fact_monthly_indicator": monthly_records,
        }

    # ── dispatcher ──────────────────────────────────────────────────────
    def _dispatch(self, name: str, ws) -> tuple[list[dict], list[dict]]:
        """Route sheet name to the appropriate parser.  Returns (weekly, monthly)."""
        empty: list[dict] = []

        # ── Weekly: Pattern A  (start_date, end_date, province columns) ──
        if name == "周度-商品猪出栏价":
            return self._parse_weekly_regional(ws, "hog_price_out", "元/公斤", header_row=2, data_start=3), empty
        if name == "周度-体重":
            return self._parse_weight_multi(ws), empty
        if name == "周度-50公斤二元母猪价格":
            return self._parse_weekly_regional(ws, "sow_price_50kg", "元/头", header_row=2, data_start=3), empty
        if name == "周度-规模场15公斤仔猪出栏价":
            return self._parse_weekly_regional(ws, "piglet_price_15kg", "元/头", header_row=2, data_start=3), empty
        if name == "周度-淘汰母猪价格":
            return self._parse_weekly_regional(ws, "cull_sow_price", "元/斤", header_row=2, data_start=3), empty
        if name == "周度-宰后结算价":
            return self._parse_weekly_regional(ws, "carcass_price", "元/公斤", header_row=2, data_start=3), empty
        if name == "周度-冻品库存":
            return self._parse_weekly_regional(ws, "frozen_rate", "", header_row=2, data_start=3), empty
        if name == "周度-猪肉价（前三等级白条均价）":
            return self._parse_weekly_regional(ws, "pork_carcass_price", "元/公斤", header_row=2, data_start=3), empty

        # ── Weekly: 鲜销率 (starts data at a later row, irregular header) ──
        if name == "鲜销率":
            return self._parse_fresh_sale_rate(ws), empty

        # ── Weekly: Pattern B  (with '指标' sub-category column) ──
        if name == "周度-屠宰厂宰前活猪重":
            return self._parse_slaughter_weight(ws), empty
        if name == "周度-各体重段价差":
            return self._parse_weight_spread(ws), empty
        if name == "周度-体重拆分":
            return self._parse_weight_split(ws), empty
        if name == "周度-猪肉产品价格":
            return self._parse_pork_product_price(ws), empty

        # ── Weekly: slaughter volume (new-style, aggregated) ──
        if name == "周度-屠宰新2022.10.28":
            return self._parse_slaughter_new(ws), empty

        # ── Weekly: daily slaughter by province ──
        if name == "周度-屠宰企业日度屠宰量":
            return self._parse_daily_slaughter(ws), empty

        # ── Weekly: profit (new format with scale breakdowns) ──
        if name == "周度-养殖利润最新":
            return self._parse_profit_latest(ws), empty

        # ── Weekly: feed price ──
        if name == "育肥全价料价格":
            return self._parse_feed_price(ws), empty

        # ── Weekly: frozen inventory multi-sample (by area) ──
        if name == "周度-冻品库存多样本":
            return self._parse_frozen_multi(ws), empty

        # ── Weekly: gross-net spread (毛白价差) ──
        if name == "周度-毛白价差":
            return self._parse_mao_bai_spread(ws), empty

        # ── Monthly: new-sample inventory (2020+) ──
        if name == "月度-能繁母猪存栏（2020年2月新增）":
            return empty, self._parse_monthly_inventory_new(ws, "breeding_sow_inventory", "头")
        if name == "月度-小猪存栏（2020年5月新增）":
            return empty, self._parse_monthly_inventory_new(ws, "piglet_inventory", "头")
        if name == "月度-大猪存栏（2020年5月新增）":
            return empty, self._parse_monthly_inventory_new(ws, "hog_inventory", "头")

        # ── Monthly: old-sample inventory ──
        if name == "月度-能繁母猪存栏量":
            return empty, self._parse_monthly_inventory_old(ws, "breeding_sow_inventory_old", "头")
        if name == "月度-商品猪出栏量":
            return empty, self._parse_monthly_inventory_old(ws, "hog_output_volume", "头", include_mom_col=True)
        if name == "月度-小猪（50公斤以下）存栏":
            return empty, self._parse_monthly_inventory_old(ws, "piglet_inventory_old", "头")

        # ── Monthly: feed sales (猪料销量) ──
        if name == "月度-猪料销量":
            return empty, self._parse_feed_sales(ws)

        # ── Monthly: cull sow slaughter (淘汰母猪屠宰厂宰杀量) ──
        if name == "月度-淘汰母猪屠宰厂宰杀量":
            return empty, self._parse_cull_slaughter(ws)

        # ── Monthly: slaughter utilization rate ──
        if name == "月度-屠宰企业开工率":
            return empty, self._parse_slaughter_utilization(ws)

        # ── Monthly: gender ratio ──
        if name == "月度-屠宰厂公母比例":
            return empty, self._parse_gender_ratio(ws)

        # ── Monthly: production metrics ──
        if name == "月度-生产指标（2021.5.7新增）" or name == "月度-生产指标":
            return empty, self._parse_production_metrics(ws)
        if name == "月度-生产指标2":
            return empty, self._parse_production_metrics2(ws)

        # 销售计划：月度计划出栏量 — 本月计划销售 → fact_monthly_indicator (yongyi_planned_output)
        if name == "月度计划出栏量":
            return empty, self._parse_monthly_planned_output(ws)

        # Skip all other sheets (reference / summary / not needed)
        return empty, empty

    # ════════════════════════════════════════════════════════════════════
    # WEEKLY PARSERS
    # ════════════════════════════════════════════════════════════════════

    # ── Pattern A: generic weekly regional ──────────────────────────────
    def _parse_weekly_regional(
        self, ws, indicator_code: str, unit: str,
        header_row: int = 2, data_start: int = 3,
    ) -> list[dict]:
        """
        Pattern A:  col-A = start_date, col-B = end_date,
        then province names across columns from header_row.
        """
        records: list[dict] = []
        # Read header to get province names
        header = self._row_values(ws, header_row, max_col=30)
        col_offset = 2  # first two cols are start/end dates

        province_map: dict[int, str] = {}  # col_index -> region_code
        for ci in range(col_offset, len(header)):
            name = header[ci]
            if name is None:
                continue
            rc = _safe_region(str(name))
            if rc:
                province_map[ci] = rc

        for row in ws.iter_rows(min_row=data_start, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1]) if len(row) > 1 else None
                if end_dt is None and start_dt is None:
                    continue
                week_end = end_dt or start_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": indicator_code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row in %s", indicator_code, exc_info=True)
        return records

    # ── 鲜销率 (fresh sale rate, irregular header) ──────────────────────
    def _parse_fresh_sale_rate(self, ws) -> list[dict]:
        """
        Row 2-3 are region headers (row 3 is province names).
        Data rows may start with text or dates.
        Only parse rows where col-A is a date.
        """
        header = self._row_values(ws, 3, max_col=25)
        province_map: dict[int, str] = {}
        for ci in range(2, len(header)):
            name = header[ci]
            if name is None:
                continue
            rc = _safe_region(str(name))
            if rc:
                province_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=4, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1]) if len(row) > 1 else None
                if start_dt is None:
                    continue
                week_end = end_dt or start_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": "fresh_sale_rate",
                        "source": SOURCE,
                        "value": val,
                        "unit": "",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row fresh_sale_rate", exc_info=True)
        return records

    # ── 屠宰厂宰前活猪重 (slaughter weight, has '较上周' columns) ───────
    def _parse_slaughter_weight(self, ws) -> list[dict]:
        """
        Row 4 = header: 开始日期, 结束日期, 指标, 河南, 较上周, 湖南, 较上周, ...
        Data starts row 5.  We skip the '较上周' columns.
        """
        header = self._row_values(ws, 4, max_col=25)
        province_map: dict[int, str] = {}
        for ci in range(3, len(header)):
            name = header[ci]
            if name is None:
                continue
            name_str = str(name).strip()
            if name_str in ("较上周", "较上期"):
                continue
            rc = _safe_region(name_str)
            if rc:
                province_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=5, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if end_dt is None:
                    continue
                week_end = end_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": "weight_slaughter",
                        "source": SOURCE,
                        "value": val,
                        "unit": "公斤",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row weight_slaughter", exc_info=True)
        return records

    # ── 体重 (multi-row: 均重, 90kg以下, 150kg以上) ─────────────────────
    def _parse_weight_multi(self, ws) -> list[dict]:
        """
        Row 2 = header: 开始日期, 结束日期, 指标, 河南, 湖南, ...
        Data from row 3.  Multiple rows per date with different 指标:
          均重     -> weight_avg (公斤)
          90kg以下  -> weight_pct_under90 (ratio)
          150kg以上 -> weight_pct_over150 (ratio)
        """
        header = self._row_values(ws, 2, max_col=25)
        province_map: dict[int, str] = {}
        for ci in range(3, len(header)):
            name = header[ci]
            if name is None:
                continue
            rc = _safe_region(str(name))
            if rc:
                province_map[ci] = rc

        INDICATOR_MAP = {
            "均重": ("weight_avg", "公斤"),
            "90kg以下": ("weight_pct_under90", ""),
            "150kg以上": ("weight_pct_over150", ""),
        }

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if end_dt is None:
                    continue
                week_end = end_dt
                week_start = start_dt or _week_start_from_end(week_end)

                label = str(row[2]).strip() if row[2] else ""
                mapping = INDICATOR_MAP.get(label)
                if mapping is None:
                    continue
                code, unit = mapping

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row weight_multi", exc_info=True)
        return records

    # ── 体重拆分 (national weight avg + group/scatter split) ────────────
    def _parse_weight_split(self, ws) -> list[dict]:
        """
        周度-体重拆分 sheet layout:
        Row 1: title
        Row 2: header groups: "体重拆分(单位:公斤)", "全国", ..., "出栏权重"
        Row 3: sub-header: "结束日期", "全国均重", "集团", "散户", "集团"(权重), "散户"(权重)
        Data from row 4.
        """
        records: list[dict] = []
        for row in ws.iter_rows(min_row=4, max_col=6, values_only=True):
            try:
                row = list(row)
                end_dt = parse_date(row[0])
                if end_dt is None:
                    continue
                week_end = end_dt
                week_start = _week_start_from_end(week_end)

                # col 1 = 全国均重
                val_nation = clean_value(row[1])
                if val_nation is not None:
                    records.append({
                        "week_end": week_end, "week_start": week_start,
                        "region_code": "NATION", "indicator_code": "weight_avg",
                        "source": SOURCE, "value": val_nation, "unit": "公斤",
                        "batch_id": self.batch_id,
                    })
                # col 2 = 集团
                val_group = clean_value(row[2])
                if val_group is not None:
                    records.append({
                        "week_end": week_end, "week_start": week_start,
                        "region_code": "NATION", "indicator_code": "weight_group",
                        "source": SOURCE, "value": val_group, "unit": "公斤",
                        "batch_id": self.batch_id,
                    })
                # col 3 = 散户
                val_scatter = clean_value(row[3])
                if val_scatter is not None:
                    records.append({
                        "week_end": week_end, "week_start": week_start,
                        "region_code": "NATION", "indicator_code": "weight_scatter",
                        "source": SOURCE, "value": val_scatter, "unit": "公斤",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row weight_split", exc_info=True)
        return records

    # ── 各体重段价差 (weight-segment spread, multi-row per date) ────────
    def _parse_weight_spread(self, ws) -> list[dict]:
        """
        Row 2 = header: 开始日期, 结束日期, 指标, 辽宁, 河北, ...
        Data from row 3.  Multiple rows share same date pair with different 指标 values.
        """
        header = self._row_values(ws, 2, max_col=20)
        province_map: dict[int, str] = {}
        for ci in range(3, len(header)):
            name = header[ci]
            if name is None:
                continue
            name_str = str(name).strip()
            if name_str == "平均":
                province_map[ci] = "NATION"
                continue
            rc = _safe_region(name_str)
            if rc:
                province_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if end_dt is None:
                    continue
                indicator_label = str(row[2]).strip() if row[2] else ""
                if not indicator_label:
                    continue
                # Map indicator label
                code = f"weight_spread_{indicator_label}"
                week_end = end_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/斤",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row weight_spread", exc_info=True)
        return records

    # ── 猪肉产品价格 (pork product price, with 指标类型 column) ─────────
    def _parse_pork_product_price(self, ws) -> list[dict]:
        """
        Row 2 = header: 开始日期, 结束日期, 指标类型, 辽宁, 黑龙江, ...
        Data from row 3.  Multiple rows per date with different 指标类型.
        """
        header = self._row_values(ws, 2, max_col=25)
        province_map: dict[int, str] = {}
        for ci in range(3, len(header)):
            name = header[ci]
            if name is None:
                continue
            rc = _safe_region(str(name))
            if rc:
                province_map[ci] = rc

        INDICATOR_MAP = {
            "白条价前三级均价": "pork_carcass_top3",
            "2号肉鲜品价": "pork_no2_fresh",
            "2号肉冻品价": "pork_no2_frozen",
            "2号肉鲜-冻": "pork_no2_fresh_frozen_spread",
            "肋排鲜品价": "rib_fresh",
            "肋排冻品价": "rib_frozen",
        }

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=len(header), values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if start_dt is None:
                    continue
                week_end = end_dt or start_dt
                week_start = start_dt

                label = str(row[2]).strip() if row[2] else ""
                code = INDICATOR_MAP.get(label, f"pork_product_{label}")

                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": rc,
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row pork_product", exc_info=True)
        return records

    # ── 屠宰新2022.10.28 (aggregated slaughter volume) ─────────────────
    def _parse_slaughter_new(self, ws) -> list[dict]:
        """
        Row 2 = header: 开始日期, 结束日期, 规模屠宰厂全国前10家, 环比, 各省规模屠宰厂100家, 环比, ...
        Data from row 3.
        """
        records: list[dict] = []
        COL_MAP = {
            2: ("slaughter_top10", "头"),
            4: ("slaughter_top100", "头"),
            6: ("slaughter_county", "头"),
        }
        for row in ws.iter_rows(min_row=3, max_col=10, values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if end_dt is None:
                    continue
                week_end = end_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, (code, unit) in COL_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row slaughter_new", exc_info=True)
        return records

    # ── 屠宰企业日度屠宰量 (daily slaughter by province) ────────────────
    def _parse_daily_slaughter(self, ws) -> list[dict]:
        """
        Row 1 = header: 省份, 辽宁, 吉林, ... , 合计（单位：头）, ...
        Data from row 2, col A = date.
        We store each daily record as a 'weekly' indicator with week_end = week_start = the date.
        """
        header = self._row_values(ws, 1, max_col=25)
        province_map: dict[int, str] = {}
        total_col: int | None = None
        for ci in range(1, len(header)):
            name = header[ci]
            if name is None:
                continue
            name_str = str(name).strip()
            if "合计" in name_str:
                total_col = ci
                continue
            rc = _safe_region(name_str)
            if rc:
                province_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_col=len(header), values_only=True):
            try:
                row = list(row)
                dt = parse_date(row[0])
                if dt is None:
                    continue
                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": dt,
                        "week_start": dt,
                        "region_code": rc,
                        "indicator_code": "slaughter_volume_daily",
                        "source": SOURCE,
                        "value": val,
                        "unit": "头",
                        "batch_id": self.batch_id,
                    })
                # Also store the national total
                if total_col is not None and total_col < len(row):
                    val = clean_value(row[total_col])
                    if val is not None:
                        records.append({
                            "week_end": dt,
                            "week_start": dt,
                            "region_code": "NATION",
                            "indicator_code": "slaughter_volume_daily",
                            "source": SOURCE,
                            "value": val,
                            "unit": "头",
                            "batch_id": self.batch_id,
                        })
            except Exception:
                logger.debug("skip row daily_slaughter", exc_info=True)
        return records

    # ── 养殖利润最新 (breeding profit by scale) ─────────────────────────
    def _parse_profit_latest(self, ws) -> list[dict]:
        """
        Row 3 = header: 开始日期, 结束日期, 项目, 母猪50头以下, 50-200, ...
        Data from row 4.  col[2] = '利润'/'成本'.
        We map scale names to indicator_code suffixes.
        """
        SCALE_MAP = {
            3: ("profit_breeding_lt50", "元/头"),
            4: ("profit_breeding_50_200", "元/头"),
            5: ("profit_breeding_200_500", "元/头"),
            6: ("profit_breeding_500_2000", "元/头"),
            7: ("profit_breeding_2000_5000", "元/头"),
            8: ("profit_breeding_5000_10000", "元/头"),
            9: ("profit_piglet_purchase", "元/头"),
            10: ("profit_contract_farming", "元/头"),
        }
        records: list[dict] = []
        for row in ws.iter_rows(min_row=4, max_col=15, values_only=True):
            try:
                row = list(row)
                start_dt = parse_date(row[0])
                end_dt = parse_date(row[1])
                if end_dt is None:
                    continue
                label = str(row[2]).strip() if row[2] else ""
                if label != "利润":
                    continue  # Only import profit rows
                week_end = end_dt
                week_start = start_dt or _week_start_from_end(week_end)

                for ci, (code, unit) in SCALE_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": week_start,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row profit_latest", exc_info=True)
        return records

    # ── 育肥全价料价格 (feed price by province) ─────────────────────────
    def _parse_feed_price(self, ws) -> list[dict]:
        """
        Row 3 = area header: 日期, 东北, 华北, ..., 全国
        Row 4 = province header: None, 辽宁, 山东, 河南, 江苏, 湖南, 广东, 广西, 四川, 均价, 均价
        Data from row 5.  Col A = date.
        """
        header_area = self._row_values(ws, 3, max_col=15)
        header_prov = self._row_values(ws, 4, max_col=15)

        province_map: dict[int, str] = {}
        for ci in range(1, len(header_prov)):
            name = header_prov[ci]
            if name is None:
                # fallback to area header
                name = header_area[ci] if ci < len(header_area) else None
            if name is None:
                continue
            name_str = str(name).strip()
            if name_str == "均价":
                # Determine which one: column 9 = factory price, column 10 = farm price
                if ci == 9:
                    province_map[ci] = "NATION"
                # skip col 10 (farm gate price) to avoid duplicating
                continue
            rc = _safe_region(name_str)
            if rc:
                province_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=5, max_col=len(header_prov), values_only=True):
            try:
                row = list(row)
                dt = parse_date(row[0])
                if dt is None:
                    continue
                for ci, rc in province_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": dt,
                        "week_start": dt,
                        "region_code": rc,
                        "indicator_code": "feed_price_complete",
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/吨",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row feed_price", exc_info=True)
        return records

    # ── 冻品库存多样本 (frozen inventory by area) ───────────────────────
    def _parse_frozen_multi(self, ws) -> list[dict]:
        """
        Row 2 = header: 日期, 东北, 华北, 华东, 华中, 西南, 华南, 全国
        Data from row 3.  Single date column.
        """
        header = self._row_values(ws, 2, max_col=12)
        area_map: dict[int, str] = {}
        for ci in range(1, len(header)):
            name = header[ci]
            if name is None:
                continue
            rc = _safe_region(str(name))
            if rc:
                area_map[ci] = rc

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=len(header), values_only=True):
            try:
                row = list(row)
                dt = parse_date(row[0])
                if dt is None:
                    continue
                for ci, rc in area_map.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": dt,
                        "week_start": dt,
                        "region_code": rc,
                        "indicator_code": "frozen_inventory_multi",
                        "source": SOURCE,
                        "value": val,
                        "unit": "",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row frozen_multi", exc_info=True)
        return records

    # ── 毛白价差 (live-to-carcass spread, national) ─────────────────────
    def _parse_mao_bai_spread(self, ws) -> list[dict]:
        """
        Row 1 = header: 毛白价差, 前三级别白条价, 生猪出栏价, 价差
        Data from row 2.  Col A = date.
        """
        records: list[dict] = []
        COL_MAP = {
            1: ("pork_carcass_avg", "元/公斤"),
            2: ("hog_price_out_nation", "元/公斤"),
            3: ("mao_bai_spread", "元/公斤"),
        }
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            try:
                row = list(row)
                dt = parse_date(row[0])
                if dt is None:
                    continue
                for ci, (code, unit) in COL_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "week_end": dt,
                        "week_start": dt,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row mao_bai", exc_info=True)
        return records

    # ════════════════════════════════════════════════════════════════════
    # MONTHLY PARSERS
    # ════════════════════════════════════════════════════════════════════

    # ── New-sample monthly inventory (2020+) ────────────────────────────
    def _parse_monthly_inventory_new(self, ws, indicator_code: str, unit: str) -> list[dict]:
        """
        Sheets like 月度-能繁母猪存栏（2020年2月新增）, 月度-小猪存栏, 月度-大猪存栏.
        Auto-detects header row by scanning for row containing '日期' and '全国'.
        Header pattern: 日期, 全国, 环比, [同比], 日期, 华北, 环比, ...
        Data starts the row after the header.

        We extract absolute values for each region.
        """
        # Auto-detect header row (scan first 6 rows for one containing '日期' and '全国')
        header_row = 4  # default
        for ri in range(1, 7):
            candidate = self._row_values(ws, ri, max_col=30)
            strs = [str(c).strip() if c else "" for c in candidate]
            if "日期" in strs and "全国" in strs:
                header_row = ri
                break

        header = self._row_values(ws, header_row, max_col=30)

        # Detect region columns: find positions of region names in header
        region_cols: list[tuple[int, str]] = []  # (value_col_index, region_code)

        # The first group: col 0=日期, col 1=全国, col 2=环比, col 3=同比
        # Then repeating groups: 日期, <region>, 环比
        # For 大猪存栏 the first group has no 同比: 日期, 全国, 环比, 日期, 华北, 环比, ...
        i = 0
        while i < len(header):
            h = str(header[i]).strip() if header[i] else ""
            if h == "日期" or h == "备注":
                # Next column should be a region name
                if i + 1 < len(header):
                    rname = str(header[i + 1]).strip() if header[i + 1] else ""
                    rc = _safe_region(rname)
                    if rc:
                        region_cols.append((i + 1, rc))
                i += 1
            else:
                i += 1

        records: list[dict] = []
        data_start = header_row + 1
        for row in ws.iter_rows(min_row=data_start, max_col=30, values_only=True):
            try:
                row = list(row)
                # First column should be the date
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                for val_col, rc in region_cols:
                    if val_col >= len(row):
                        continue
                    val = clean_value(row[val_col])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": rc,
                        "indicator_code": indicator_code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row monthly_new %s", indicator_code, exc_info=True)
        return records

    # ── Old-sample monthly inventory ────────────────────────────────────
    def _parse_monthly_inventory_old(
        self, ws, indicator_code: str, unit: str, include_mom_col: bool = False
    ) -> list[dict]:
        """
        Sheets like 月度-能繁母猪存栏量, 月度-商品猪出栏量, 月度-小猪（50公斤以下）存栏.
        Header row 2 pattern: 日期, 全国, 环比, 同比, 较非瘟前, 日期, 华北, 环比, 同比, 较非瘟前, ...
        Each group has 5 fields.  Data starts at row 3.
        When include_mom_col=True, also read the column after value (环比) and write as value_type='mom_pct'.
        """
        header = self._row_values(ws, 2, max_col=40)

        # Find region columns: (value_col, region_code); 环比 is at value_col+1
        region_cols: list[tuple[int, str]] = []
        i = 0
        while i < len(header):
            h = str(header[i]).strip() if header[i] else ""
            if h == "日期":
                if i + 1 < len(header):
                    rname = str(header[i + 1]).strip() if header[i + 1] else ""
                    rc = _safe_region(rname)
                    if rc:
                        region_cols.append((i + 1, rc))
                i += 1
            else:
                i += 1

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=40, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                for val_col, rc in region_cols:
                    if val_col >= len(row):
                        continue
                    val = clean_value(row[val_col])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": rc,
                        "indicator_code": indicator_code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
                    if include_mom_col and val_col + 1 < len(row):
                        mom_val = clean_value(row[val_col + 1])
                        if mom_val is not None:
                            records.append({
                                "month_date": month_dt,
                                "region_code": rc,
                                "indicator_code": indicator_code,
                                "sub_category": "",
                                "source": SOURCE,
                                "value": mom_val,
                                "value_type": "mom_pct",
                                "unit": "%",
                                "batch_id": self.batch_id,
                            })
            except Exception:
                logger.debug("skip row monthly_old %s", indicator_code, exc_info=True)
        return records

    # ── 月度计划出栏量（销售计划 D3 — 计划+实际同表同单位万头）────────────────
    def _parse_monthly_planned_output(self, ws) -> list[dict]:
        """
        Sheet 月度计划出栏量：Row1 有「上月样本企业合计销售」「本月计划销售」「万头」「本月计划较上月实际销售」。
        - 本月计划销售 → yongyi_planned_output（当月计划，万头）
        - 上月样本企业合计销售：当行月份为 M 时表示 M-1 月实际，写入 month_date=M-1 的 yongyi_actual_sample_sales（万头），
          这样 2026-02 行的「上月」= 2026-01 实际，与 2026-01 行的「本月计划」同口径，销售计划页 实际/计划/达成率 与旧版一致。
        """
        row1 = self._row_values(ws, 1, max_col=25)
        plan_col: Optional[int] = None
        actual_prev_col: Optional[int] = None  # 上月样本企业合计销售
        for idx, cell in enumerate(row1):
            if not cell:
                continue
            s = str(cell).strip()
            if s == "本月计划销售":
                plan_col = idx
            elif s == "上月样本企业合计销售":
                actual_prev_col = idx
        if plan_col is None:
            logger.warning("月度计划出栏量: 未找到「本月计划销售」列，跳过")
            return []

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=max(plan_col + 1, actual_prev_col + 1 if actual_prev_col is not None else 0, 18), values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                if plan_col < len(row):
                    val = clean_value(row[plan_col])
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "region_code": "NATION",
                            "indicator_code": "yongyi_planned_output",
                            "sub_category": "",
                            "source": SOURCE,
                            "value": val,
                            "value_type": "abs",
                            "unit": "万头",
                            "batch_id": self.batch_id,
                        })
                # 上月样本企业合计销售 → 上一月的实际（万头），与旧版销售计划口径一致
                if actual_prev_col is not None and actual_prev_col < len(row):
                    prev_val = clean_value(row[actual_prev_col])
                    if prev_val is not None:
                        prev_month = month_dt.month - 1
                        prev_year = month_dt.year
                        if prev_month <= 0:
                            prev_month += 12
                            prev_year -= 1
                        prev_month_dt = date(prev_year, prev_month, 1)
                        records.append({
                            "month_date": prev_month_dt,
                            "region_code": "NATION",
                            "indicator_code": "yongyi_actual_sample_sales",
                            "sub_category": "",
                            "source": SOURCE,
                            "value": prev_val,
                            "value_type": "abs",
                            "unit": "万头",
                            "batch_id": self.batch_id,
                        })
            except Exception:
                logger.debug("skip row 月度计划出栏量", exc_info=True)
        logger.info("月度计划出栏量: %d 条 (yongyi_planned_output + yongyi_actual_sample_sales)", len(records))
        return records

    # ── 猪料销量 (feed sales, national, with sub-categories) ────────────
    def _parse_feed_sales(self, ws) -> list[dict]:
        """
        Row 1 header: 月份, 猪料销量（环比）, 后备母猪料, 其他母猪料, 教保料, 育肥料, ...
        Data from row 2.  Col A = date (Excel serial number for month).
        All values are MoM ratios.
        """
        SUB_MAP = {
            1: ("feed_sales_total", "环比"),
            2: ("feed_sales_reserve_sow", "环比"),
            3: ("feed_sales_other_sow", "环比"),
            4: ("feed_sales_nursery", "环比"),
            5: ("feed_sales_fattening", "环比"),
        }
        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                for ci, (code, vtype) in SUB_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "mom",
                        "unit": "",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row feed_sales", exc_info=True)
        return records

    # ── 淘汰母猪屠宰厂宰杀量 (cull sow slaughter by province) ─────────
    def _parse_cull_slaughter(self, ws) -> list[dict]:
        """
        Row 1: 淘汰母猪宰杀量, 省份, 河南, None, 山东, None, 辽宁, None, 四川, None, 广西, None, 合计, ...
        Row 2: 日期, 类别, 5家, 月度环比, 6家, 月度环比, 3家, 月度环比, ...
        Data from row 3.
        We extract absolute slaughter values (odd columns 2,4,6,8,10) and total (col 12).
        """
        header1 = self._row_values(ws, 1, max_col=20)
        province_cols: list[tuple[int, str]] = []
        for ci in [2, 4, 6, 8, 10]:
            if ci < len(header1):
                name = header1[ci]
                if name:
                    rc = _safe_region(str(name))
                    if rc:
                        province_cols.append((ci, rc))
        # Total column
        total_col = 12

        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=15, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                label = str(row[1]).strip() if row[1] else ""
                if label != "宰杀量":
                    continue

                for ci, rc in province_cols:
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": rc,
                        "indicator_code": "cull_slaughter",
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": "头",
                        "batch_id": self.batch_id,
                    })
                if total_col < len(row):
                    val = clean_value(row[total_col])
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "region_code": "NATION",
                            "indicator_code": "cull_slaughter",
                            "sub_category": "",
                            "source": SOURCE,
                            "value": val,
                            "value_type": "abs",
                            "unit": "头",
                            "batch_id": self.batch_id,
                        })
            except Exception:
                logger.debug("skip row cull_slaughter", exc_info=True)
        return records

    # ── 屠宰企业开工率 (slaughter utilization rate, national) ───────────
    def _parse_slaughter_utilization(self, ws) -> list[dict]:
        """
        Row 1: title, Row 2+: date, value (single column)
        """
        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                val = clean_value(row[1])
                if val is None:
                    continue
                records.append({
                    "month_date": month_dt,
                    "region_code": "NATION",
                    "indicator_code": "slaughter_utilization",
                    "sub_category": "",
                    "source": SOURCE,
                    "value": val,
                    "value_type": "abs",
                    "unit": "",
                    "batch_id": self.batch_id,
                })
            except Exception:
                logger.debug("skip row slaughter_util", exc_info=True)
        return records

    # ── 屠宰厂公母比例 (gender ratio) ──────────────────────────────────
    def _parse_gender_ratio(self, ws) -> list[dict]:
        """
        Rows: 月份, 头数, 公猪占比
        """
        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                # Volume
                vol = clean_value(row[1])
                if vol is not None:
                    records.append({
                        "month_date": month_dt,
                        "region_code": "NATION",
                        "indicator_code": "slaughter_volume_monthly",
                        "sub_category": "",
                        "source": SOURCE,
                        "value": vol,
                        "value_type": "abs",
                        "unit": "头",
                        "batch_id": self.batch_id,
                    })
                # Male ratio
                ratio = clean_value(row[2])
                if ratio is not None:
                    records.append({
                        "month_date": month_dt,
                        "region_code": "NATION",
                        "indicator_code": "male_ratio",
                        "sub_category": "",
                        "source": SOURCE,
                        "value": ratio,
                        "value_type": "abs",
                        "unit": "",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row gender_ratio", exc_info=True)
        return records

    # ── 生产指标（2021.5.7新增） ────────────────────────────────────────
    def _parse_production_metrics(self, ws) -> list[dict]:
        """
        Row 2 = header: 日期, 基础母猪存栏, 环比涨跌, 配种数, 配种数环比,
                         分娩窝数, 分娩窝数环比, 窝均健仔数, 产房存活率, ...
        列对应: col1=日期 col2=基础母猪 col3=环比 col4=配种数 col5=配种数环比
               col6=分娩窝数(母猪效能) col7=分娩窝数环比 col8=窝均健仔数(压栏系数) col9=产房存活率
        Data from row 3.  National-level aggregated metrics.
        """
        COL_MAP = {
            1: ("prod_base_sow_inventory", "头"),
            3: ("prod_mating_count", "窝"),
            6: ("prod_farrowing_count", "窝"),   # 母猪效能：分娩窝数（原误用 col5 为配种数环比）
            8: ("prod_healthy_piglets_per_litter", "头"),  # 压栏系数：窝均健仔数（原误用 col7 为分娩窝数环比）
            9: ("prod_farrowing_survival_rate", ""),  # 产房存活率（原误用 col8）
        }
        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=20, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                for ci, (code, unit) in COL_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row prod_metrics", exc_info=True)
        return records

    # ── 生产指标2 ──────────────────────────────────────────────────────
    def _parse_production_metrics2(self, ws) -> list[dict]:
        """
        Row 2 = header: 日期, 基础母猪存栏, 后备母猪数, 配种数, 分娩母猪窝数,
                         窝均健仔数, 产房存活率, 配种分娩率, 断奶成活率, 育肥出栏成活率
        Data from row 3.
        """
        COL_MAP = {
            1: ("prod2_base_sow_inventory", "头"),
            3: ("prod2_mating_count", "窝"),
            4: ("prod2_farrowing_count", "窝"),
            5: ("prod2_healthy_piglets_per_litter", "头"),
            6: ("prod2_farrowing_survival_rate", ""),
            7: ("prod2_mating_farrowing_rate", ""),
            8: ("prod2_weaning_survival_rate", ""),
            9: ("prod2_fattening_survival_rate", ""),
        }
        records: list[dict] = []
        for row in ws.iter_rows(min_row=3, max_col=12, values_only=True):
            try:
                row = list(row)
                month_dt = parse_month(row[0])
                if month_dt is None:
                    continue
                for ci, (code, unit) in COL_MAP.items():
                    if ci >= len(row):
                        continue
                    val = clean_value(row[ci])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_dt,
                        "region_code": "NATION",
                        "indicator_code": code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug("skip row prod_metrics2", exc_info=True)
        return records

    # ════════════════════════════════════════════════════════════════════
    # HELPERS
    # ════════════════════════════════════════════════════════════════════

    @staticmethod
    def _row_values(ws, row_num: int, max_col: int = 30) -> list:
        """Read a single row from a worksheet (1-indexed)."""
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, max_col=max_col, values_only=True):
            return list(row)
        return []
