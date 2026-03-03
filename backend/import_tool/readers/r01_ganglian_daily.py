"""
Reader: 钢联自动更新模板.xlsx  (14 sheets)
数据源: 钢联（Ganglian / Mysteel）日度/周度/月度综合数据
产出表:
  - fact_price_daily          ← 分省区猪价, 交割库出栏价
  - fact_enterprise_daily     ← 集团企业出栏价
  - fact_spread_daily         ← 区域价差, 肥标价差, 毛白价差
  - fact_weekly_indicator     ← 养殖利润（周度）, 冻品库容率, 宰后均重
  - fact_monthly_indicator    ← 月度数据, 月度出栏, 猪肉进口, 淘汰母猪屠宰
  - fact_futures_basis        ← 仓单数据
"""
import logging
import re
from typing import Optional

import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, parse_month, clean_value, province_to_code

logger = logging.getLogger(__name__)

# ── 常量 ──────────────────────────────────────────────

SOURCE = "GANGLIAN"
DATA_START_ROW = 5          # 所有 sheet 数据从第 5 行开始（row1=标题, 2=指标, 3=单位, 4=更新时间）
HEADER_ROW = 2              # 指标名称行

# 交割库城市 → region_code（交割库所在城市并非省级，需要手动映射）
CITY_TO_REGION = {
    "亳州": "ANHUI", "阜阳": "ANHUI", "宿州": "ANHUI",
    "广州": "GUANGDONG",
    "贵港": "GUANGXI", "南宁": "GUANGXI",
    "张家口": "HEBEI",
    "南阳": "HENAN", "平顶山": "HENAN", "濮阳": "HENAN",
    "商丘": "HENAN", "驻马店": "HENAN",
    "荆州": "HUBEI", "武汉": "HUBEI", "襄阳": "HUBEI",
    "常德": "HUNAN", "衡阳": "HUNAN", "株洲": "HUNAN",
    "松原": "JILIN",
    "苏州": "JIANGSU", "宿迁": "JIANGSU", "徐州": "JIANGSU",
    "盐城": "JIANGSU",
    "南昌": "JIANGXI",
    "赤峰": "NEIMENGGU", "呼和浩特": "NEIMENGGU", "乌兰察布": "NEIMENGGU",
}

# 集团企业名 → company_code
ENTERPRISE_TO_CODE = {
    "辽宁大北农": "DABEINONG_LN",
    "内蒙古牧原": "MUYUAN_NMG",
    "黑龙江大北农": "DABEINONG_HLJ",
    "中粮家佳康（吉林）": "COFCO_JL",
    "山东新希望": "NEWHOPE_SD",
    "河北新好农牧": "XINHAO_HB",
    "河南牧原": "MUYUAN_HN",
    "湖北牧原": "MUYUAN_HUB",
    "陕西温氏": "WENS_SAX",
    "江苏温氏": "WENS_JS",
    "安徽牧原": "MUYUAN_AH",
    "江西温氏": "WENS_JX",
    "湖南唐人神": "TANGRENSHEN_HUN",
    "广东温氏": "WENS_GD",
    "四川德康": "DEKANG_SC",
    "贵州富之源": "FUZHIYUAN_GZ",
    "云南神农": "SHENNONG_YN",
}

# 区域价差 header → (region_high, region_low) + spread_type
# 格式: "商品猪：出栏均价：广东（日） - 商品猪：出栏均价：广西（日）"
_SPREAD_REGION_RE = re.compile(
    r"出栏均价：(\S+?)（日）\s*-\s*.*?出栏均价：(\S+?)（日）"
)

# 肥标价差 header → region_code
# 格式: "生猪标肥：价差：中国（日）" or "生猪标肥：价差：四川（日）"
_FAT_SPREAD_RE = re.compile(r"标肥：价差：(\S+?)（日）")

# 毛白价差 header → metric key
# 格式: "毛白：价差：中国（日）", "毛白：价差：中国（日） / 商品猪：出栏均价：中国（日）"
_MAO_BAI_RE = re.compile(r"毛白：价差：(\S+?)（日）")

# 养殖利润周度 indicator_code 映射
WEEKLY_PROFIT_INDICATORS = {
    "二元母猪：每头重50kg：样本养殖企业：均价：中国（周）": ("gilt_price", "元/头"),
    "仔猪：每头重7kg：规模化养殖场：出栏均价：中国（周）": ("piglet_price", "元/头"),
    "淘汰母猪：样本养殖企业：均价：中国（周）": ("cull_sow_price", "元/千克"),
    "淘汰母猪：样本养殖企业：均价：中国（周） / 标猪：市场价：中国（日）": ("cull_sow_std_ratio", None),
    "标猪：市场价：中国（日）(变频周平均值)": ("std_pig_price_wavg", "元/千克"),
    "白条肉：前三级：均价：中国（日）(变频周平均值)": ("carcass_top3_price_wavg", "元/千克"),
    "鲜猪肉：1号：市场均价：中国（周）": ("fresh_pork_no1_price", "元/千克"),
    "冻肉：2号：市场价：中国（日）(变频周平均值)": ("frozen_no2_price_wavg", "元/千克"),
    "冻肉：4号：市场价：中国（日）(变频周平均值)": ("frozen_no4_price_wavg", "元/千克"),
    "猪粮比：中国（周）": ("pig_grain_ratio", None),
    "生猪饲料：比价：中国（周）": ("pig_feed_ratio", None),
    "猪：屠宰利润（周）": ("slaughter_profit", "元/头"),
    "猪：外购利润（周）": ("external_purchase_profit", "元/头"),
}

# 宰后均重 header → region_code
# 格式: "白条肉：屠宰企业：宰后均重：中国（周）"
_WEIGHT_RE = re.compile(r"宰后均重：(\S+?)（周）")

# 月度数据 indicator_code 映射
MONTHLY_DATA_INDICATORS = {
    "猪：出栏数：中国（月）": ("hog_slaughter_count", "万头", "abs"),
    "猪：计划出栏量：中国（月）": ("hog_planned_output", "万头", "abs"),
    "母猪饲料：销量环比：中国（月）": ("sow_feed_sales_mom", "%", "mom"),
    "仔猪饲料：销量环比：中国（月）": ("piglet_feed_sales_mom", "%", "mom"),
    "育肥猪饲料：销量环比：中国（月）": ("fattening_feed_sales_mom", "%", "mom"),
    "生猪饲料：销量环比：中国（月）": ("pig_feed_sales_mom", "%", "mom"),
    "商品猪：综合养殖场：存栏数：中国（月）": ("hog_inventory_total", "万头", "abs"),
    "商品猪：规模化养殖场：存栏数：中国（月）": ("hog_inventory_large", "万头", "abs"),
    "商品猪：中小散：存栏数：中国（月）": ("hog_inventory_small", "万头", "abs"),
    "仔猪：出生数环比：中国（月）": ("piglet_birth_mom", "%", "mom"),
    "仔猪：销量环比：中国（月）": ("piglet_sales_mom", "%", "mom"),
    "能繁殖母猪：综合养殖场：存栏数：中国（月）": ("breeding_sow_inventory_total", "万头", "abs"),
    "能繁殖母猪：规模化养殖场：存栏数：中国（月）": ("breeding_sow_inventory_large", "万头", "abs"),
    "能繁殖母猪：中小散：存栏数：中国（月）": ("breeding_sow_inventory_small", "万头", "abs"),
}

# 月度出栏
MONTHLY_OUTPUT_INDICATORS = {
    "商品猪：综合养殖场：出栏数：中国（月）": ("hog_output_total", "万头"),
    "商品猪：规模化养殖场：出栏数：中国（月）": ("hog_output_large", "万头"),
    "商品猪：中小散：出栏数：中国（月）": ("hog_output_small", "万头"),
}

# 猪肉进口 - 国家名提取
_IMPORT_COUNTRY_RE = re.compile(r"进口数量(?:合计)?：(\S+?)→中国（月）")

# 仓单数据 - warehouse name extraction
_WAREHOUSE_RE = re.compile(r"注册仓单(?:：(.+?))?（日）")


# ── Helper ─────────────────────────────────────────────

def _read_headers(ws) -> list[Optional[str]]:
    """读取第 HEADER_ROW 行的 header（指标名称）"""
    headers = []
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        headers = list(row)
    return headers


def _region_lookup(name: str) -> Optional[str]:
    """省份/地区名 → region_code，扩展处理 '中国' = '全国'"""
    if not name:
        return None
    rc = province_to_code(name)
    if rc:
        return rc
    # 钢联 header 用 "中国" 而非 "全国"
    if name.strip() == "中国":
        return "NATION"
    return None


def _extract_province_from_header(header: str) -> Optional[str]:
    """从 header 如 '商品猪：出栏均价：黑龙江（日）' 提取省份名并返回 region_code"""
    # 匹配最后一个 ： 到 （ 之间的文本（省份名不含 ： ）
    m = re.search(r"：([^：]+?)（", header)
    if m:
        return _region_lookup(m.group(1))
    return None


def _extract_city_from_header(header: str) -> Optional[str]:
    """从 header 如 '猪：每头重110-125kg：出栏价：亳州（日）' 提取城市名"""
    m = re.search(r"出栏价：(\S+?)（日）", header)
    if m:
        return m.group(1)
    return None


def _extract_enterprise_from_header(header: str) -> tuple[Optional[str], Optional[str]]:
    """
    从 header 提取企业名和省份
    例: '外三元猪：每头重110-125kg：出栏价：辽宁：辽宁大北农（日）'
    返回: ('辽宁大北农', 'LIAONING')
    """
    # 格式: ...出栏价：省份：企业名（日）
    m = re.search(r"出栏价：(\S+?)：(\S+?)（日）", header)
    if m:
        province_name = m.group(1)
        enterprise_name = m.group(2)
        region_code = _region_lookup(province_name)
        return enterprise_name, region_code
    return None, None


# ── Reader 类 ──────────────────────────────────────────

class GanglianDailyReader(BaseSheetReader):
    """钢联自动更新模板 Excel 阅读器"""

    FILE_PATTERN = "钢联自动更新模板"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        """读取 Excel 文件，返回 {table_name: [records]}"""
        logger.info(f"[GanglianDailyReader] 开始读取: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

        results: dict[str, list[dict]] = {
            "fact_price_daily": [],
            "fact_enterprise_daily": [],
            "fact_spread_daily": [],
            "fact_weekly_indicator": [],
            "fact_monthly_indicator": [],
            "fact_futures_basis": [],
        }

        # 各 sheet 处理器映射
        handlers = {
            "分省区猪价": self._read_province_price,
            "集团企业出栏价": self._read_enterprise_price,
            "交割库出栏价": self._read_delivery_price,
            "区域价差": self._read_region_spread,
            "肥标价差": self._read_fat_std_spread,
            "毛白价差": self._read_mao_bai_spread,
            "养殖利润（周度）": self._read_weekly_profit,
            "冻品库容率": self._read_frozen_storage,
            "宰后均重": self._read_post_slaughter_weight,
            "月度数据": self._read_monthly_data,
            "月度出栏": self._read_monthly_output,
            "猪肉进口": self._read_pork_import,
            "仓单数据": self._read_warehouse_receipt,
            "淘汰母猪屠宰": self._read_cull_sow_slaughter,
        }

        for sheet_name, handler in handlers.items():
            if sheet_name not in wb.sheetnames:
                if sheet_name == "仓单数据":
                    logger.warning("  缺少 sheet: 仓单数据 → 前端 C4 仓单数据页将无数据，请确认钢联模板含该 sheet")
                else:
                    logger.warning(f"  缺少 sheet: {sheet_name}")
                continue
            ws = wb[sheet_name]
            try:
                handler(ws, results)
                logger.info(f"  ✓ {sheet_name} 处理完成")
            except Exception:
                logger.exception(f"  ✗ {sheet_name} 处理异常")

        wb.close()

        # 汇总日志
        for tbl, recs in results.items():
            logger.info(f"  {tbl}: {len(recs)} records")

        return results

    # ────────────────────────────────────────────────────
    # Sheet 1: 分省区猪价 → fact_price_daily
    # ────────────────────────────────────────────────────
    def _read_province_price(self, ws, results: dict):
        """
        行: 日期 | 全国价格 | 黑龙江 | 吉林 | ...
        每列一个省份，header 格式: 商品猪：出栏均价：黑龙江（日）
        """
        headers = _read_headers(ws)

        # 解析列 → region_code 映射
        col_map: list[tuple[int, str]] = []  # (col_idx, region_code)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            rc = _extract_province_from_header(str(h))
            if rc:
                col_map.append((idx, rc))
            else:
                logger.debug(f"  分省区猪价: 无法识别 header[{idx}]: {h}")

        records = results["fact_price_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, region_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "price_type": "hog_avg_price",
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  分省区猪价: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 2: 集团企业出栏价 → fact_enterprise_daily
    # ────────────────────────────────────────────────────
    def _read_enterprise_price(self, ws, results: dict):
        """
        行: 日期 | 企业1价格 | 企业2价格 | ...
        header 格式: 外三元猪：每头重110-125kg：出栏价：辽宁：辽宁大北农（日）
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, str]] = []  # (col_idx, company_code, region_code)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            enterprise_name, region_code = _extract_enterprise_from_header(str(h))
            if enterprise_name and enterprise_name in ENTERPRISE_TO_CODE:
                company_code = ENTERPRISE_TO_CODE[enterprise_name]
                col_map.append((idx, company_code, region_code or "NATION"))
            else:
                logger.debug(f"  集团企业出栏价: 无法识别 header[{idx}]: {h}")

        records = results["fact_enterprise_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, company_code, region_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "company_code": company_code,
                        "region_code": region_code,
                        "metric_type": "output_price",
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  集团企业出栏价: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 3: 交割库出栏价 → fact_price_daily
    # ────────────────────────────────────────────────────
    def _read_delivery_price(self, ws, results: dict):
        """
        行: 日期 | 城市1价格 | 城市2价格 | ...
        header 格式: 猪：每头重110-125kg：出栏价：亳州（日）
        映射到省级 region_code, price_type = delivery_price_{city}
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, str]] = []  # (col_idx, region_code, city_name)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            city = _extract_city_from_header(str(h))
            if city and city in CITY_TO_REGION:
                col_map.append((idx, CITY_TO_REGION[city], city))
            else:
                logger.debug(f"  交割库出栏价: 无法识别城市 header[{idx}]: {h}")

        records = results["fact_price_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, region_code, city_name in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "price_type": f"delivery_{city_name}",
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  交割库出栏价: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 4: 区域价差 → fact_spread_daily
    # ────────────────────────────────────────────────────
    def _read_region_spread(self, ws, results: dict):
        """
        header 格式: 商品猪：出栏均价：广东（日） - 商品猪：出栏均价：广西（日）
        spread_type = region_spread_{high}_{low}
        region_code 取 high 端省份
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, str]] = []  # (col_idx, region_code, spread_type)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            m = _SPREAD_REGION_RE.search(str(h))
            if m:
                high_name = m.group(1)
                low_name = m.group(2)
                high_code = _region_lookup(high_name)
                low_code = _region_lookup(low_name)
                if high_code and low_code:
                    spread_type = f"region_spread_{high_code}_{low_code}"
                    col_map.append((idx, high_code, spread_type))
                else:
                    logger.debug(f"  区域价差: 省份未识别 {high_name}/{low_name}")
            else:
                logger.debug(f"  区域价差: 无法解析 header[{idx}]: {h}")

        records = results["fact_spread_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, region_code, spread_type in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "spread_type": spread_type,
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  区域价差: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 5: 肥标价差 → fact_spread_daily
    # ────────────────────────────────────────────────────
    def _read_fat_std_spread(self, ws, results: dict):
        """
        header 格式: 生猪标肥：价差：中国（日）
        spread_type = fat_std_spread
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str]] = []  # (col_idx, region_code)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            m = _FAT_SPREAD_RE.search(str(h))
            if m:
                region_name = m.group(1)
                rc = _region_lookup(region_name)
                if rc:
                    col_map.append((idx, rc))
                else:
                    logger.debug(f"  肥标价差: 省份未识别: {region_name}")
            else:
                logger.debug(f"  肥标价差: 无法解析 header[{idx}]: {h}")

        records = results["fact_spread_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, region_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "spread_type": "fat_std_spread",
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  肥标价差: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 6: 毛白价差 → fact_spread_daily
    # ────────────────────────────────────────────────────
    def _read_mao_bai_spread(self, ws, results: dict):
        """
        header 列:
          col1: 毛白：价差：中国（日）               → spread value
          col2: 毛白：价差：中国（日） / 出栏均价     → ratio (skip)
          col3: 商品猪：出栏均价：中国（日）          → (duplicated in 分省区猪价, skip)
          col4-6: 移动平均 (skip)
        只取 col1 的毛白价差
        """
        headers = _read_headers(ws)

        # 找到 "毛白：价差" 且不含 "/" 和 "移动平均" 的列
        col_map: list[tuple[int, str]] = []
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            h_str = str(h)
            if "毛白：价差" in h_str and "/" not in h_str and "移动平均" not in h_str:
                m = _MAO_BAI_RE.search(h_str)
                if m:
                    region_name = m.group(1)
                    rc = _region_lookup(region_name)
                    if rc:
                        col_map.append((idx, rc))

        records = results["fact_spread_daily"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, region_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "spread_type": "mao_bai_spread",
                        "source": SOURCE,
                        "value": val,
                        "unit": "元/千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  毛白价差: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 7: 养殖利润（周度） → fact_weekly_indicator
    # ────────────────────────────────────────────────────
    def _read_weekly_profit(self, ws, results: dict):
        """
        行: 周末日期 | 指标1 | 指标2 | ...
        每列对应一个全国级别周度指标
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, Optional[str]]] = []  # (col_idx, indicator_code, unit)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            h_str = str(h).strip()
            if h_str in WEEKLY_PROFIT_INDICATORS:
                code, unit = WEEKLY_PROFIT_INDICATORS[h_str]
                col_map.append((idx, code, unit))
            else:
                logger.debug(f"  养殖利润: 未映射的指标 header[{idx}]: {h_str}")

        records = results["fact_weekly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                week_end = parse_date(row[0])
                if not week_end:
                    continue
                for col_idx, indicator_code, unit in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": None,
                        "region_code": "NATION",
                        "indicator_code": indicator_code,
                        "source": SOURCE,
                        "value": val,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  养殖利润: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 8: 冻品库容率 → fact_weekly_indicator
    # ────────────────────────────────────────────────────
    def _read_frozen_storage(self, ws, results: dict):
        """
        单列数据: 冷冻猪肉：重点屠宰企业：库容率：中国（周）
        indicator_code = frozen_storage_rate
        """
        records = results["fact_weekly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                week_end = parse_date(row[0])
                if not week_end:
                    continue
                val = clean_value(row[1]) if len(row) > 1 else None
                if val is None:
                    continue
                records.append({
                    "week_end": week_end,
                    "week_start": None,
                    "region_code": "NATION",
                    "indicator_code": "frozen_storage_rate",
                    "source": SOURCE,
                    "value": val,
                    "unit": "%",
                    "batch_id": self.batch_id,
                })
            except Exception:
                logger.debug(f"  冻品库容率: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 9: 宰后均重 → fact_weekly_indicator
    # ────────────────────────────────────────────────────
    def _read_post_slaughter_weight(self, ws, results: dict):
        """
        行: 周末日期 | 全国均重 | 安徽 | 重庆 | ...
        header 格式: 白条肉：屠宰企业：宰后均重：中国（周）
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str]] = []
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            m = _WEIGHT_RE.search(str(h))
            if m:
                region_name = m.group(1)
                rc = _region_lookup(region_name)
                if rc:
                    col_map.append((idx, rc))
                else:
                    logger.debug(f"  宰后均重: 省份未识别: {region_name}")

        records = results["fact_weekly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                week_end = parse_date(row[0])
                if not week_end:
                    continue
                for col_idx, region_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "week_end": week_end,
                        "week_start": None,
                        "region_code": region_code,
                        "indicator_code": "post_slaughter_weight",
                        "source": SOURCE,
                        "value": val,
                        "unit": "千克",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  宰后均重: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 10: 月度数据 → fact_monthly_indicator
    # ────────────────────────────────────────────────────
    def _read_monthly_data(self, ws, results: dict):
        """
        行: 月末日期 | 指标1 | 指标2 | ...
        每列一个全国级别月度指标
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, str, str]] = []  # (col_idx, indicator_code, unit, value_type)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            h_str = str(h).strip()
            if h_str in MONTHLY_DATA_INDICATORS:
                code, unit, vtype = MONTHLY_DATA_INDICATORS[h_str]
                col_map.append((idx, code, unit, vtype))
            else:
                logger.debug(f"  月度数据: 未映射的指标 header[{idx}]: {h_str}")

        records = results["fact_monthly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                month_date = parse_month(row[0])
                if not month_date:
                    continue
                for col_idx, indicator_code, unit, value_type in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_date,
                        "region_code": "NATION",
                        "indicator_code": indicator_code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": value_type,
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  月度数据: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 11: 月度出栏 → fact_monthly_indicator
    # ────────────────────────────────────────────────────
    def _read_monthly_output(self, ws, results: dict):
        """
        行: 月末日期 | 综合出栏 | 规模化出栏 | 中小散出栏
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str, str]] = []
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            h_str = str(h).strip()
            if h_str in MONTHLY_OUTPUT_INDICATORS:
                code, unit = MONTHLY_OUTPUT_INDICATORS[h_str]
                col_map.append((idx, code, unit))
            else:
                logger.debug(f"  月度出栏: 未映射的指标 header[{idx}]: {h_str}")

        records = results["fact_monthly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                month_date = parse_month(row[0])
                if not month_date:
                    continue
                for col_idx, indicator_code, unit in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_date,
                        "region_code": "NATION",
                        "indicator_code": indicator_code,
                        "sub_category": "",
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": unit,
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  月度出栏: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 12: 猪肉进口 → fact_monthly_indicator
    # ────────────────────────────────────────────────────
    def _read_pork_import(self, ws, results: dict):
        """
        header 格式: 猪肉及猪杂碎：进口数量合计：全球→中国（月）
                     猪肉及猪杂碎：进口数量：巴西→中国（月）
        sub_category = 来源国 (全球/巴西/丹麦/...)
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str]] = []  # (col_idx, country_name)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            m = _IMPORT_COUNTRY_RE.search(str(h))
            if m:
                country = m.group(1)
                col_map.append((idx, country))
            else:
                logger.debug(f"  猪肉进口: 无法解析来源 header[{idx}]: {h}")

        records = results["fact_monthly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                month_date = parse_month(row[0])
                if not month_date:
                    continue
                for col_idx, country in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "month_date": month_date,
                        "region_code": "NATION",
                        "indicator_code": "pork_import",
                        "sub_category": country,
                        "source": SOURCE,
                        "value": val,
                        "value_type": "abs",
                        "unit": "吨",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  猪肉进口: 行解析失败 row[0]={row[0]}", exc_info=True)

    # ────────────────────────────────────────────────────
    # Sheet 13: 仓单数据 → fact_futures_basis
    # ────────────────────────────────────────────────────
    def _read_warehouse_receipt(self, ws, results: dict):
        """
        header 格式:
          DCE：猪：注册仓单（日）                     → 总仓单
          DCE：猪：注册仓单：德康农牧库（日）          → 集团汇总
          DCE：猪：注册仓单：常熟德康（德康农牧）库（日） → 子库
        indicator_code = warehouse_receipt / warehouse_receipt_{name}
        """
        headers = _read_headers(ws)

        col_map: list[tuple[int, str]] = []  # (col_idx, indicator_code)
        for idx, h in enumerate(headers):
            if idx == 0 or not h:
                continue
            h_str = str(h).strip()
            m = _WAREHOUSE_RE.search(h_str)
            if m:
                warehouse_name = m.group(1)
                if warehouse_name:
                    # 清理仓库名：去除（xxx）中的母公司标注
                    clean_name = re.sub(r"（\S+?）", "", warehouse_name).strip()
                    indicator_code = f"warehouse_receipt_{clean_name}"
                else:
                    indicator_code = "warehouse_receipt_total"
                col_map.append((idx, indicator_code))
            else:
                logger.debug(f"  仓单数据: 无法解析 header[{idx}]: {h_str}")

        records = results["fact_futures_basis"]
        start_count = len(records)
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                trade_date = parse_date(row[0])
                if not trade_date:
                    continue
                for col_idx, indicator_code in col_map:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    records.append({
                        "trade_date": trade_date,
                        "contract_code": None,
                        "region_code": "NATION",
                        "indicator_code": indicator_code,
                        "value": val,
                        "unit": "手",
                        "batch_id": self.batch_id,
                    })
            except Exception:
                logger.debug(f"  仓单数据: 行解析失败 row[0]={row[0]}", exc_info=True)
        if len(col_map) == 0:
            logger.warning("  仓单数据: 表头未解析到「注册仓单（日）」格式，C4 仓单页将无数据；需如 DCE：猪：注册仓单（日） 或 注册仓单：库名（日）")
        elif len(records) == start_count:
            logger.warning("  仓单数据: 未解析到任何行数据，请检查日期列与数据列格式")

    # ────────────────────────────────────────────────────
    # Sheet 14: 淘汰母猪屠宰 → fact_monthly_indicator
    # ────────────────────────────────────────────────────
    def _read_cull_sow_slaughter(self, ws, results: dict):
        """
        单列数据: 淘汰母猪：屠宰量：中国（月）
        indicator_code = cull_sow_slaughter
        """
        records = results["fact_monthly_indicator"]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            try:
                month_date = parse_month(row[0])
                if not month_date:
                    continue
                val = clean_value(row[1]) if len(row) > 1 else None
                if val is None:
                    continue
                records.append({
                    "month_date": month_date,
                    "region_code": "NATION",
                    "indicator_code": "cull_sow_slaughter",
                    "sub_category": "",
                    "source": SOURCE,
                    "value": val,
                    "value_type": "abs",
                    "unit": "头",
                    "batch_id": self.batch_id,
                })
            except Exception:
                logger.debug(f"  淘汰母猪屠宰: 行解析失败 row[0]={row[0]}", exc_info=True)
