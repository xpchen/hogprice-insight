"""
Reader: 涌益咨询日度数据.xlsx  (Yongyi Daily Data)

Excel 文件包含 8 个 sheet，解析到以下目标表：
  - fact_price_daily      日度价格
  - fact_spread_daily     标肥价差
  - fact_slaughter_daily  日度屠宰量
"""

import logging
from datetime import date, datetime
from typing import Optional

import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import clean_value, parse_date, province_to_code

logger = logging.getLogger(__name__)

# ── 全国 / 特殊名称映射 ──
_SPECIAL_REGION = {
    "全国均价": "NATION",
    "全国": "NATION",
    "合计（单位：头）": None,   # 跳过合计行
    "合计2（单位：头）": None,
    "合计": None,
    "备注": None,
}

# ── Sheet 7 城市 → region_code 映射 ──
# 交割地市以城市名为列头，需映射到所属省份 region_code
_CITY_TO_PROVINCE_CODE = {
    "广州市": "GUANGDONG",
    "常德市": "HUNAN", "株洲市": "HUNAN",
    "贵港市": "GUANGXI", "南宁市": "GUANGXI", "崇左市": "GUANGXI",
    "南昌市": "JIANGXI",
    "荆门市": "HUBEI", "襄阳市": "HUBEI", "武汉市": "HUBEI", "荆州市": "HUBEI",
    "澄城县": "SHAANXI", "渭南市": "SHAANXI",
    "张家口": "HEBEI",
    "南京市": "JIANGSU", "宿迁市": "JIANGSU", "苏州市": "JIANGSU", "盐城市": "JIANGSU",
    "德州市": "SHANDONG",
    "江安": "SICHUAN", "泸州": "SICHUAN", "达州": "SICHUAN", "绵阳": "SICHUAN",
    "松原": "JILIN",
    "贞丰": "GUIZHOU",
    "陆良": "YUNNAN",
}


def _resolve_region(name: str) -> Optional[str]:
    """省份名称 / 特殊名称 → region_code，返回 None 表示跳过该行"""
    if not name:
        return None
    name = str(name).strip()
    if name in _SPECIAL_REGION:
        return _SPECIAL_REGION[name]
    code = province_to_code(name)
    if code:
        return code
    # 尝试去掉可能的 "省" 后缀
    for suffix in ("省", "市"):
        stripped = name.replace(suffix, "")
        code = province_to_code(stripped)
        if code:
            return code
    return None


class YongyiDailyReader(BaseSheetReader):
    """涌益咨询日度数据 reader"""

    FILE_PATTERN = "涌益咨询日度数据"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        """读取 Excel 文件，返回 {table_name: [records]}"""
        logger.info("Opening workbook: %s", filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        price_records: list[dict] = []
        spread_records: list[dict] = []
        slaughter_records: list[dict] = []

        try:
            # ── Sheet 0: 出栏价 ──
            self._read_sheet0_chulanjia(wb, price_records)
            # ── Sheet 1: 价格+宰量 ──
            self._read_sheet1_price_slaughter(wb, price_records, slaughter_records)
            # ── Sheet 2: 散户标肥价差 ──
            self._read_sheet2_spread(wb, price_records, spread_records)
            # ── Sheet 3: 各省份均价 ──
            self._read_sheet3_province_avg(wb, price_records)
            # ── Sheet 4: 市场主流标猪肥猪价格 ──
            self._read_sheet4_mainstream(wb, price_records)
            # ── Sheet 5: 屠宰企业日度屠宰量 ──
            self._read_sheet5_slaughter(wb, slaughter_records)
            # ── Sheet 6: 市场主流标猪肥猪均价方便作图 ──
            self._read_sheet6_charting(wb, price_records)
            # ── Sheet 7: 交割地市出栏价 ──
            self._read_sheet7_delivery_city(wb, price_records)
        finally:
            wb.close()

        logger.info(
            "Yongyi daily totals: prices=%d, spreads=%d, slaughter=%d",
            len(price_records), len(spread_records), len(slaughter_records),
        )
        return {
            "fact_price_daily": price_records,
            "fact_spread_daily": spread_records,
            "fact_slaughter_daily": slaughter_records,
        }

    # ──────────────────────────────────────────────
    # Sheet 0: 出栏价
    # 横向展开：Row1=日期（分布在不同列），Row2=子表头（规模场/小散户/均价/涨跌）
    # Row3+ 每行一个省份，col 0=省份名
    # 第一个日期块 3 列（无涨跌），后续块 4 列
    # ──────────────────────────────────────────────
    def _read_sheet0_chulanjia(self, wb, price_records: list[dict]):
        sheet_name = "出栏价"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return

        row1 = list(rows[0])  # dates
        # row2 = list(rows[1])  # sub-headers (not needed, positions are fixed)

        # Build date_columns: list of (date, col_index_of_均价)
        # Scan row1 for datetime values; each indicates start of a date block
        date_blocks: list[tuple[date, int]] = []
        for col_idx, val in enumerate(row1):
            d = parse_date(val)
            if d is not None:
                date_blocks.append((d, col_idx))

        # For each date block, 均价 is at offset +2 from the date column
        # (规模场=+0, 小散户=+1, 均价=+2)
        for data_row in rows[2:]:
            province_name = data_row[0] if len(data_row) > 0 else None
            if not province_name:
                continue
            region_code = _resolve_region(str(province_name))
            if region_code is None:
                continue

            for trade_date, block_start in date_blocks:
                try:
                    # 均价 is always at block_start + 2
                    avg_col = block_start + 2
                    if avg_col >= len(data_row):
                        continue
                    avg_val = clean_value(data_row[avg_col])
                    if avg_val is None:
                        continue

                    price_records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "price_type": "出栏均价",
                        "source": "YONGYI",
                        "value": avg_val,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })

                    # 规模场价格 at block_start + 0
                    scale_val = clean_value(data_row[block_start])
                    if scale_val is not None:
                        price_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "price_type": "规模场出栏价",
                            "source": "YONGYI",
                            "value": scale_val,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })

                    # 小散户价格 at block_start + 1
                    small_val = clean_value(data_row[block_start + 1])
                    if small_val is not None:
                        price_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "price_type": "散户出栏价",
                            "source": "YONGYI",
                            "value": small_val,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })
                except Exception:
                    logger.debug(
                        "Sheet '%s' row parse error: province=%s, date=%s",
                        sheet_name, province_name, trade_date, exc_info=True,
                    )

        logger.info("Sheet '%s': extracted %d price records so far", sheet_name, len(price_records))

    # ──────────────────────────────────────────────
    # Sheet 1: 价格+宰量
    # 纵向：Col0=日期, Col1=全国均价, Col2=日屠宰量合计1
    # Row1 是表头，Row2+ 是数据
    # ──────────────────────────────────────────────
    def _read_sheet1_price_slaughter(self, wb, price_records: list[dict], slaughter_records: list[dict]):
        sheet_name = "价格+宰量"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        count_price = 0
        count_slaughter = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=4, values_only=True)):
            try:
                trade_date = parse_date(row[0])
                if trade_date is None:
                    continue

                # 全国均价
                avg_price = clean_value(row[1])
                if avg_price is not None:
                    price_records.append({
                        "trade_date": trade_date,
                        "region_code": "NATION",
                        "price_type": "全国均价",
                        "source": "YONGYI",
                        "value": avg_price,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })
                    count_price += 1

                # 日屠宰量合计
                slaughter_vol = clean_value(row[2])
                if slaughter_vol is not None:
                    slaughter_records.append({
                        "trade_date": trade_date,
                        "region_code": "NATION",
                        "source": "YONGYI",
                        "volume": slaughter_vol,
                        "batch_id": self.batch_id,
                    })
                    count_slaughter += 1
            except Exception:
                logger.debug("Sheet '%s' row %d parse error", sheet_name, i + 2, exc_info=True)

        logger.info("Sheet '%s': prices=%d, slaughter=%d", sheet_name, count_price, count_slaughter)

    # ──────────────────────────────────────────────
    # Sheet 2: 散户标肥价差
    # 横向展开：Row1=日期, Row2=子表头
    # 子列：市场散户标重猪, 150公斤左右较标猪, 175公斤左右较标猪 [, 今日采购二育情绪较昨日]
    # 第一个日期块 3 列（无情绪），后续 4 列
    # ──────────────────────────────────────────────
    def _read_sheet2_spread(self, wb, price_records: list[dict], spread_records: list[dict]):
        sheet_name = "散户标肥价差"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return

        row1 = list(rows[0])  # dates

        # Build date blocks from row 1
        date_blocks: list[tuple[date, int]] = []
        for col_idx, val in enumerate(row1):
            d = parse_date(val)
            if d is not None:
                date_blocks.append((d, col_idx))

        count_price = 0
        count_spread = 0
        for data_row in rows[2:]:
            province_name = data_row[0] if len(data_row) > 0 else None
            if not province_name:
                continue
            region_code = _resolve_region(str(province_name))
            if region_code is None:
                continue

            for trade_date, block_start in date_blocks:
                try:
                    # 市场散户标重猪（标猪价格）at block_start + 0
                    std_price = clean_value(data_row[block_start])
                    if std_price is not None:
                        price_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "price_type": "散户标猪价",
                            "source": "YONGYI",
                            "value": std_price,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })
                        count_price += 1

                    # 150公斤左右较标猪（价差）at block_start + 1
                    spread_150 = clean_value(data_row[block_start + 1])
                    if spread_150 is not None:
                        spread_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "spread_type": "150kg较标猪价差",
                            "source": "YONGYI",
                            "value": spread_150,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })
                        count_spread += 1

                    # 175公斤左右较标猪（价差）at block_start + 2
                    spread_175 = clean_value(data_row[block_start + 2])
                    if spread_175 is not None:
                        spread_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "spread_type": "175kg较标猪价差",
                            "source": "YONGYI",
                            "value": spread_175,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })
                        count_spread += 1
                except Exception:
                    logger.debug(
                        "Sheet '%s' row parse error: province=%s, date=%s",
                        sheet_name, province_name, trade_date, exc_info=True,
                    )

        logger.info("Sheet '%s': prices=%d, spreads=%d", sheet_name, count_price, count_spread)

    # ──────────────────────────────────────────────
    # Sheet 3: 各省份均价
    # 纵向：Row1=表头（日期, 河南, 湖南, ...全国均价）
    # Row2+ 每行一天的各省份均价
    # ──────────────────────────────────────────────
    def _read_sheet3_province_avg(self, wb, price_records: list[dict]):
        sheet_name = "各省份均价"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows_iter = ws.iter_rows(values_only=True)
        header_row = list(next(rows_iter))  # Row 1: ['日期', '河南', '湖南', ...]

        # Build province mapping: col_index -> region_code
        province_cols: list[tuple[int, str]] = []
        for col_idx in range(1, len(header_row)):
            name = header_row[col_idx]
            if name is None:
                continue
            region_code = _resolve_region(str(name))
            if region_code:
                province_cols.append((col_idx, region_code))

        count = 0
        for i, row in enumerate(rows_iter):
            try:
                trade_date = parse_date(row[0])
                if trade_date is None:
                    continue

                for col_idx, region_code in province_cols:
                    if col_idx >= len(row):
                        continue
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    price_records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "price_type": "省份均价",
                        "source": "YONGYI",
                        "value": val,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })
                    count += 1
            except Exception:
                logger.debug("Sheet '%s' row %d parse error", sheet_name, i + 2, exc_info=True)

        logger.info("Sheet '%s': %d price records", sheet_name, count)

    # ──────────────────────────────────────────────
    # Sheet 4: 市场主流标猪肥猪价格
    # 横向展开：Row1=日期, Row2=子表头
    # 每日期块 5 列：标猪均价, 标猪体重段, 90-100kg均价, 130-140kg均价, 150kg左右均价
    # ──────────────────────────────────────────────
    def _read_sheet4_mainstream(self, wb, price_records: list[dict]):
        sheet_name = "市场主流标猪肥猪价格"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return

        row1 = list(rows[0])  # dates

        # Build date blocks
        date_blocks: list[tuple[date, int]] = []
        for col_idx, val in enumerate(row1):
            d = parse_date(val)
            if d is not None:
                date_blocks.append((d, col_idx))

        # Sub-column offsets (from date column):
        # +0: 标猪均价, +1: 标猪体重段(skip), +2: 90-100kg均价, +3: 130-140kg均价, +4: 150kg左右均价
        PRICE_TYPES = [
            (0, "标猪均价"),
            (2, "90-100kg均价"),
            (3, "130-140kg均价"),
            (4, "150kg左右均价"),
        ]

        count = 0
        for data_row in rows[2:]:
            province_name = data_row[0] if len(data_row) > 0 else None
            if not province_name:
                continue
            pname = str(province_name).strip()
            if pname in ("备注",):
                continue
            region_code = _resolve_region(pname)
            if region_code is None:
                continue

            for trade_date, block_start in date_blocks:
                for offset, price_type in PRICE_TYPES:
                    try:
                        col = block_start + offset
                        if col >= len(data_row):
                            continue
                        val = clean_value(data_row[col])
                        if val is None:
                            continue
                        price_records.append({
                            "trade_date": trade_date,
                            "region_code": region_code,
                            "price_type": price_type,
                            "source": "YONGYI",
                            "value": val,
                            "unit": "元/公斤",
                            "batch_id": self.batch_id,
                        })
                        count += 1
                    except Exception:
                        logger.debug(
                            "Sheet '%s' parse error: province=%s, date=%s, type=%s",
                            sheet_name, province_name, trade_date, price_type,
                            exc_info=True,
                        )

        logger.info("Sheet '%s': %d price records", sheet_name, count)

    # ──────────────────────────────────────────────
    # Sheet 5: 屠宰企业日度屠宰量
    # Row1: ['省份', date1, date2, ...] — 日期作为列头
    # Row2+: 每行一个省份，各列为对应日期的屠宰量
    # ──────────────────────────────────────────────
    def _read_sheet5_slaughter(self, wb, slaughter_records: list[dict]):
        sheet_name = "屠宰企业日度屠宰量"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows_iter = ws.iter_rows(values_only=True)
        header_row = list(next(rows_iter))  # Row 1: ['省份', date1, date2, ...]

        # Build date mapping: col_index -> date
        date_cols: list[tuple[int, date]] = []
        for col_idx in range(1, len(header_row)):
            d = parse_date(header_row[col_idx])
            if d is not None:
                date_cols.append((col_idx, d))

        count = 0
        for i, row in enumerate(rows_iter):
            try:
                province_name = row[0] if len(row) > 0 else None
                if not province_name:
                    continue
                pname = str(province_name).strip()
                region_code = _resolve_region(pname)
                if region_code is None:
                    # 跳过合计行等
                    continue

                for col_idx, trade_date in date_cols:
                    if col_idx >= len(row):
                        continue
                    vol = clean_value(row[col_idx])
                    if vol is None:
                        continue
                    slaughter_records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "source": "YONGYI",
                        "volume": vol,
                        "batch_id": self.batch_id,
                    })
                    count += 1
            except Exception:
                logger.debug("Sheet '%s' row %d parse error", sheet_name, i + 2, exc_info=True)

        logger.info("Sheet '%s': %d slaughter records", sheet_name, count)

    # ──────────────────────────────────────────────
    # Sheet 6: 市场主流标猪肥猪均价方便作图
    # 纵向：Col0=日期, Col1=全国均价, Col2=90-100kg均价,
    #       Col3=130-140kg均价, Col4=150-170kg均价
    # ──────────────────────────────────────────────
    def _read_sheet6_charting(self, wb, price_records: list[dict]):
        sheet_name = "市场主流标猪肥猪均价方便作图"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        COL_TYPES = [
            (1, "主流标猪全国均价"),
            (2, "主流90-100kg均价"),
            (3, "主流130-140kg均价"),
            (4, "主流150-170kg均价"),
        ]

        count = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=5, values_only=True)):
            try:
                trade_date = parse_date(row[0])
                if trade_date is None:
                    continue

                for col_idx, price_type in COL_TYPES:
                    val = clean_value(row[col_idx])
                    if val is None:
                        continue
                    price_records.append({
                        "trade_date": trade_date,
                        "region_code": "NATION",
                        "price_type": price_type,
                        "source": "YONGYI",
                        "value": val,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })
                    count += 1
            except Exception:
                logger.debug("Sheet '%s' row %d parse error", sheet_name, i + 2, exc_info=True)

        logger.info("Sheet '%s': %d price records", sheet_name, count)

    # ──────────────────────────────────────────────
    # Sheet 7: 交割地市出栏价
    # Row1: 省份分组 header（广东省, 湖南省, ...）
    # Row2: 升贴水（LH2505及以后）
    # Row3: 升贴水（LH2409-LH2503）
    # Row4: 交易均重
    # Row5: 城市名称 header（日期, 广州市, 常德市, ...）
    # Row6+: 每行一天的各城市出栏价
    # ──────────────────────────────────────────────
    def _read_sheet7_delivery_city(self, wb, price_records: list[dict]):
        sheet_name = "交割地市出栏价"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sheet_name)
            return
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            return

        # Row 5 (index 4): city names
        city_row = list(rows[4])

        # Build city mapping: col_index -> (city_name, region_code)
        city_cols: list[tuple[int, str, str]] = []
        for col_idx in range(1, len(city_row)):
            city_name = city_row[col_idx]
            if city_name is None:
                continue
            city_name = str(city_name).strip()
            region_code = _CITY_TO_PROVINCE_CODE.get(city_name)
            if region_code:
                city_cols.append((col_idx, city_name, region_code))
            else:
                logger.debug("Sheet '%s': unknown city '%s' at col %d, skipping", sheet_name, city_name, col_idx)

        count = 0
        for i, data_row in enumerate(rows[5:]):
            try:
                trade_date = parse_date(data_row[0])
                if trade_date is None:
                    continue

                for col_idx, city_name, region_code in city_cols:
                    if col_idx >= len(data_row):
                        continue
                    val = clean_value(data_row[col_idx])
                    if val is None:
                        continue
                    price_records.append({
                        "trade_date": trade_date,
                        "region_code": region_code,
                        "price_type": f"交割地市出栏价_{city_name}",
                        "source": "YONGYI",
                        "value": val,
                        "unit": "元/公斤",
                        "batch_id": self.batch_id,
                    })
                    count += 1
            except Exception:
                logger.debug("Sheet '%s' data row %d parse error", sheet_name, i + 6, exc_info=True)

        logger.info("Sheet '%s': %d price records", sheet_name, count)
