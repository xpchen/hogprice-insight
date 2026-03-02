"""
Reader 04 - 集团企业月度数据跟踪
Source: 3.2、集团企业月度数据跟踪.xlsx (5 sheets)
Target: fact_enterprise_monthly
Schema: (month_date, company_code, region_code, indicator, value, unit, batch_id)

Sheet layout summary
====================
汇总:          Row1=region groups (广东/四川/贵州/全国CR20/全国CR5)
               Row2=metric headers per region
               Row3+: col1=period(上旬/中旬/月度), col2=date, col3..=values

四川:          Row1=年度/月度/旬度/指标/company1..companyN/合计
               Row2+: year, month, period, indicator, values per company

广东:          Row1=年度/月度/(col3 optional)/企业名称/company1..companyN/合计
               Row2+: year, month, indicator, values per company
               Row8+: with col3=period(月度/20日) for sub-monthly tracking

贵州:          Row1=年度/月度/公司名称/company1..companyN/合计
               Row2+: year, month, indicator, values per company

集团企业全国:  Row1=企业/company headers col3..25 (CR20 col24, CR5 col25)
               Row2+: col1=month_date, col2=indicator, col3..=values per company
"""
import logging
import re
from datetime import date
from typing import Optional
import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, parse_month, clean_value

logger = logging.getLogger(__name__)

# ── Company name -> code mapping ──
COMPANY_NAME_MAP = {
    "牧原": "MUYUAN", "温氏": "WENS", "双胞胎": "SHUANGBAOTAI",
    "新希望": "XINWUFENG", "德康": "DEKANG", "正邦": "ZHENGBANG",
    "巨星": "JUXING", "正大": "ZHENGDA", "驰阳": "CHIYANG",
    "日泉": "RIQUAN", "齐全": "QIQUAN", "兴新鑫": "XINGXINXIN",
    "傲农": "AONONG", "大北农": "DABEINONG", "海大": "HAIDA",
    "富之源": "FUZHIYUAN", "罗牛山": "LUONIUSHAN", "铁骑": "TIEQI",
    "铁骑力士": "TIEQI", "天邦": "TIANBANG", "神农": "SHENNONG",
    "唐人神": "TANGRENSHEN", "新五丰": "XINWUFENG_SZ",
    "立华": "LIHUA", "京基智农": "JINGJI", "京基": "JINGJI",
    "扬翔": "YANGXIANG", "利源": "LIYUAN", "天康": "TIANKANG",
    "中粮": "COFCO", "天农": "TIANNONG", "广垦": "GUANGKEN",
    "金新农": "JINXINNONG", "万科": "VANKE", "大家": "DAJIA",
    "大广": "DAGUANG", "力源": "LIYUAN", "食出宝金": "BAOJIN",
    "宝金": "BAOJIN", "东瑞": "DONGRUI", "东方希望": "DFXW",
    "石羊": "SHIYANG", "正能": "ZHENGNENG",
    "CR20": "CR20", "CR5": "CR5", "合计": "TOTAL",
}

# Month string to integer
MONTH_MAP = {
    "1月": 1, "2月": 2, "3月": 3, "4月": 4, "5月": 5, "6月": 6,
    "7月": 7, "8月": 8, "9月": 9, "10月": 10, "11月": 11, "12月": 12,
}


def _company_code(name: str) -> Optional[str]:
    """Map Chinese company name to company_code."""
    if not name or not isinstance(name, str):
        return None
    name = name.strip().replace("\xa0", "")
    return COMPANY_NAME_MAP.get(name)


def _parse_year_month(year_str, month_str) -> Optional[date]:
    """Parse year/month strings like '2025年' / '12月' into a date (first of month)."""
    if not year_str or not month_str:
        return None
    try:
        year_s = str(year_str).strip().replace("年", "")
        year = int(year_s)
    except (ValueError, TypeError):
        return None

    month_s = str(month_str).strip()
    month = MONTH_MAP.get(month_s)
    if not month:
        # Try extracting number
        m = re.search(r"(\d+)", month_s)
        if m:
            month = int(m.group(1))
    if not month or month < 1 or month > 12:
        return None
    return date(year, month, 1)


class EnterpriseMonthlyReader(BaseSheetReader):
    """Reader for 3.2 集团企业月度数据跟踪"""

    FILE_PATTERN = "集团企业月度数据跟踪"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        records = []
        wb = openpyxl.load_workbook(filepath, data_only=True)
        try:
            records += self._read_summary(wb)
            records += self._read_sichuan(wb)
            records += self._read_guangdong(wb)
            records += self._read_guizhou(wb)
            records += self._read_national(wb)
        finally:
            wb.close()

        logger.info(f"[r04] Total records: {len(records)}")
        return {"fact_enterprise_monthly": records}

    # ────────────────────────────────────────────
    # 汇总
    # ────────────────────────────────────────────
    def _read_summary(self, wb) -> list[dict]:
        """
        汇总: Row1=region groups; Row2=metric headers per region
        Row3+: col1=period(上旬/中旬/月度), col2=date, col3..=values

        Region / metric columns:
          广东 (col3..7): 出栏计划(3), 实际出栏量(4), 计划完成率(5), 均重(6), 销售均价(7)
          四川 (col8..12): 出栏计划(8), 实际出栏量(9), 计划完成率(10), 均重(11), 计划均重(12)
          贵州 (col13..17): 计划出栏量(13), 实际出栏量(14), 计划达成率(15), 实际均重(16), 计划均重(17)
          全国CR20 (col18..20): 计划出栏量(18), 实际出栏量(19), 计划完成率(20)
          全国CR5 (col21..22): 计划出栏量(21), 实际出栏量(22)
        """
        if "汇总" not in wb.sheetnames:
            return []
        ws = wb["汇总"]
        records = []

        COL_MAP = {
            3:  ("GUANGDONG", "TOTAL", "planned_output", "万头"),
            4:  ("GUANGDONG", "TOTAL", "actual_output", "万头"),
            5:  ("GUANGDONG", "TOTAL", "plan_completion_rate", "%"),
            6:  ("GUANGDONG", "TOTAL", "avg_weight", "公斤"),
            7:  ("GUANGDONG", "TOTAL", "avg_price", "元/公斤"),
            8:  ("SICHUAN", "TOTAL", "planned_output", "万头"),
            9:  ("SICHUAN", "TOTAL", "actual_output", "万头"),
            10: ("SICHUAN", "TOTAL", "plan_completion_rate", "%"),
            11: ("SICHUAN", "TOTAL", "avg_weight", "公斤"),
            12: ("SICHUAN", "TOTAL", "planned_avg_weight", "公斤"),
            13: ("GUIZHOU", "TOTAL", "planned_output", "万头"),
            14: ("GUIZHOU", "TOTAL", "actual_output", "万头"),
            15: ("GUIZHOU", "TOTAL", "plan_completion_rate", "%"),
            16: ("GUIZHOU", "TOTAL", "avg_weight", "公斤"),
            17: ("GUIZHOU", "TOTAL", "planned_avg_weight", "公斤"),
            18: ("NATION", "CR20", "planned_output", "万头"),
            19: ("NATION", "CR20", "actual_output", "万头"),
            20: ("NATION", "CR20", "plan_completion_rate", "%"),
            21: ("NATION", "CR5", "planned_output", "万头"),
            22: ("NATION", "CR5", "actual_output", "万头"),
        }

        for row in range(3, ws.max_row + 1):
            try:
                period = ws.cell(row=row, column=1).value
                if not period:
                    continue
                period_str = str(period).strip()

                dt = parse_date(ws.cell(row=row, column=2).value)
                if not dt:
                    continue
                month_dt = dt.replace(day=1)

                # Build indicator suffix from period
                if period_str == "月度":
                    period_tag = "monthly"
                elif period_str == "上旬":
                    period_tag = "first_10d"
                elif period_str == "中旬":
                    period_tag = "mid_10d"
                else:
                    period_tag = period_str

                for col, (region, company, metric, unit) in COL_MAP.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        indicator = f"{metric}_{period_tag}"
                        records.append({
                            "month_date": month_dt,
                            "company_code": company,
                            "region_code": region,
                            "indicator": indicator,
                            "value": val,
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r04] 汇总 row {row} error: {e}")
        logger.info(f"[r04] 汇总: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 四川
    # ────────────────────────────────────────────
    def _read_sichuan(self, wb) -> list[dict]:
        """
        四川: Row1=年度/月度/旬度/指标/company headers(col5..15)/合计(col16)
        Row2+: year, month, period, indicator, values
        Indicators: 实际出栏量, 计划出栏量, 实际均重, 计划均重, 完成率
        """
        if "四川" not in wb.sheetnames:
            return []
        ws = wb["四川"]
        records = []

        # Build company map from row 1 (col5..16)
        company_map = {}
        for c in range(5, ws.max_column + 1):
            name = ws.cell(row=1, column=c).value
            code = _company_code(str(name).replace("\xa0", "")) if name else None
            if code:
                company_map[c] = code

        for row in range(2, ws.max_row + 1):
            try:
                year_str = ws.cell(row=row, column=1).value
                month_str = ws.cell(row=row, column=2).value
                period_str = ws.cell(row=row, column=3).value
                indicator_raw = ws.cell(row=row, column=4).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip().replace("\xa0", "")

                month_dt = _parse_year_month(year_str, month_str)
                if not month_dt:
                    continue

                # Determine unit from indicator
                unit = "万头"
                if "均重" in indicator:
                    unit = "公斤"
                elif "完成率" in indicator or "率" in indicator:
                    unit = "%"

                # Build full indicator name with period
                period = str(period_str).strip() if period_str else "月度"
                full_indicator = f"{indicator}_{period}"

                for col, code in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "company_code": code,
                            "region_code": "SICHUAN",
                            "indicator": full_indicator,
                            "value": val,
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r04] 四川 row {row} error: {e}")
        logger.info(f"[r04] 四川: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 广东
    # ────────────────────────────────────────────
    def _read_guangdong(self, wb) -> list[dict]:
        """
        广东: Row1 = 年度(1)/月度(2)/(col3 optional)/企业名称(4)/companies(5..27)/合计(28)
        Row2-7: year, month, indicator, values (no col3 period)
        Row8+: year, month, period(月度/20日), indicator, values

        Indicators: 实际出栏量, 实际均重, 计划出栏量, 能繁母猪量, 正品均价, 计划达成率, 完成度, 出栏均重
        """
        if "广东" not in wb.sheetnames:
            return []
        ws = wb["广东"]
        records = []

        # Build company map from row 1 (col5..28)
        company_map = {}
        for c in range(5, ws.max_column + 1):
            name = ws.cell(row=1, column=c).value
            code = _company_code(str(name).replace("\xa0", "")) if name else None
            if code:
                company_map[c] = code

        for row in range(2, ws.max_row + 1):
            try:
                year_str = ws.cell(row=row, column=1).value
                month_str = ws.cell(row=row, column=2).value
                # col3 may be period or empty; col4 is always the indicator label
                col3_val = ws.cell(row=row, column=3).value
                indicator_raw = ws.cell(row=row, column=4).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip().replace("\xa0", "")

                month_dt = _parse_year_month(year_str, month_str)
                if not month_dt:
                    continue

                # Determine period from col3
                period = str(col3_val).strip() if col3_val else "月度"

                # Determine unit
                unit = "万头"
                if "均重" in indicator:
                    unit = "公斤"
                elif "均价" in indicator or "价" in indicator:
                    unit = "元/公斤"
                elif "率" in indicator or "完成度" in indicator:
                    unit = "%"
                elif "母猪" in indicator:
                    unit = "万头"

                full_indicator = f"{indicator}_{period}"

                for col, code in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "company_code": code,
                            "region_code": "GUANGDONG",
                            "indicator": full_indicator,
                            "value": val,
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r04] 广东 row {row} error: {e}")
        logger.info(f"[r04] 广东: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 贵州
    # ────────────────────────────────────────────
    def _read_guizhou(self, wb) -> list[dict]:
        """
        贵州: Row1 = 年度(1)/月度(2)/公司名称(3)/companies(4..18)/合计(19)
        Row2+: year, month, indicator, values per company
        Indicators: 实际出栏量, 实际均重, 计划出栏量, 计划均重, 能繁母猪量
        """
        if "贵州" not in wb.sheetnames:
            return []
        ws = wb["贵州"]
        records = []

        # Build company map from row 1 (col4..19)
        company_map = {}
        for c in range(4, ws.max_column + 1):
            name = ws.cell(row=1, column=c).value
            code = _company_code(str(name).replace("\xa0", "")) if name else None
            if code:
                company_map[c] = code

        for row in range(2, ws.max_row + 1):
            try:
                year_str = ws.cell(row=row, column=1).value
                month_str = ws.cell(row=row, column=2).value
                indicator_raw = ws.cell(row=row, column=3).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip().replace("\xa0", "")

                month_dt = _parse_year_month(year_str, month_str)
                if not month_dt:
                    continue

                # Determine unit
                unit = "万头"
                if "均重" in indicator:
                    unit = "公斤"
                elif "均价" in indicator:
                    unit = "元/公斤"
                elif "率" in indicator:
                    unit = "%"
                elif "母猪" in indicator:
                    unit = "万头"

                for col, code in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "company_code": code,
                            "region_code": "GUIZHOU",
                            "indicator": indicator,
                            "value": val,
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r04] 贵州 row {row} error: {e}")
        logger.info(f"[r04] 贵州: {len(records)} records")
        return records

    # ────────────────────────────────────────────
    # 集团企业全国
    # ────────────────────────────────────────────
    def _read_national(self, wb) -> list[dict]:
        """
        集团企业全国: Row1 = 企业(1)/(col2 empty)/companies col3..25
        Company headers include: 牧原, 温氏, 双胞胎, 新希望, 德康, 天邦, 神农,
          海大, 唐人神, 新五丰, 立华, 京基智农, 铁骑力士, 扬翔, 利源, 天康,
          傲农, 巨星, 中粮, 正邦, 天农, CR20(col24), CR5(col25)
        Row2+: col1=month_date (or serial number), col2=indicator, col3..=values

        Indicators: 实际出栏量, 计划出栏量, 能繁母猪存栏
        """
        if "集团企业全国" not in wb.sheetnames:
            return []
        ws = wb["集团企业全国"]
        records = []

        # Build company map from row 1 (col3..25)
        company_map = {}
        for c in range(3, ws.max_column + 1):
            name = ws.cell(row=1, column=c).value
            code = _company_code(str(name).replace("\xa0", "")) if name else None
            if code:
                company_map[c] = code

        for row in range(2, ws.max_row + 1):
            try:
                raw_date = ws.cell(row=row, column=1).value
                month_dt = parse_month(raw_date)
                if not month_dt:
                    continue

                indicator_raw = ws.cell(row=row, column=2).value
                if not indicator_raw:
                    continue
                indicator = str(indicator_raw).strip().replace("\xa0", "")

                # Determine unit
                unit = "万头"
                if "均重" in indicator:
                    unit = "公斤"
                elif "均价" in indicator:
                    unit = "元/公斤"

                for col, code in company_map.items():
                    val = clean_value(ws.cell(row=row, column=col).value)
                    if val is not None:
                        records.append({
                            "month_date": month_dt,
                            "company_code": code,
                            "region_code": "NATION",
                            "indicator": indicator,
                            "value": val,
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception as e:
                logger.warning(f"[r04] 集团企业全国 row {row} error: {e}")
        logger.info(f"[r04] 集团企业全国: {len(records)} records")
        return records
