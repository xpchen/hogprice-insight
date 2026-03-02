"""
R02 - 生猪产业数据 Reader
数据源: 2、【生猪产业数据】.xlsx (26 sheets)

输出表:
  - fact_monthly_indicator : 月度产业指标(环比/绝对值)
  - fact_quarterly_stats   : 统计局季度数据

关键 sheet 及列映射
─────────────────────────────────────────────────────────────
NYB (row0=大类表头, row1=子表头, data从row2起, 日期在B列=col1)
  能繁环比: C(2)=全国, D(3)=规模场, E(4)=CR26, F(5)=小散户
  新生仔猪环比: G(6)=全国, H(7)=规模场, I(8)=小散户
  中大猪环比: J(9)=中大猪, K(10)=全国环比
  存栏环比: Q(16)=全国, R(17)=规模场, S(18)=小散户

02.协会猪料 (row0/1=表头, data从row2起, 日期在B列=col1)
  母猪料: H(7)=产量, I(8)=环比
  仔猪料: M(12)=产量, N(13)=环比
  育肥料: R(17)=产量, S(18)=环比

4.2涌益底稿 (row0/1=表头, data从row2起, 日期在A列=col0)
  B(1)=大猪存栏环比(已是%), G(6)=能繁环比(已是%)
  饲料环比: P(15)=猪料, Q(16)=后备母猪料, R(17)=其他母猪料, S(18)=教保料, T(19)=育肥料, U(20)=小猪料

涌益样本 (row0/1/2=表头, data从row3起, 日期在A列=col0)
  淘汰母猪屠宰量: V(21)=屠宰量(底稿原值,需算环比)

4.1.钢联数据 (row0/1=表头, data从row2起, 日期在A列=col0)
  存栏环比(中小散): B(1)=中国, C(2)=东北, D(3)=华北, E(4)=西南
  能繁存栏环比: L(11)=中国, M(12)=东北, N(13)=西南
  新生仔猪环比(中小散): O(14)=中国, P(15)=东北, Q(16)=西南, R(17)=华中
  育肥料环比: S(18)=中国, T(19)=东北, U(20)=华北, V(21)=西南
  仔猪饲料环比: W(22)=中国, X(23)=东北, Y(24)=华北, Z(25)=西南
  母猪料环比: AA(26)=中国, AB(27)=东北, AC(28)=华北, AD(29)=西南

钢联底稿 (row0-3=表头/说明, row4=单位, data从row7起, 日期在A列=col0)
  存栏(中小散绝对值): B(1)=中国, C(2)=东北, D(3)=华北, E(4)=西南
  能繁(中小散绝对值): L(11)=中国, M(12)=东北, N(13)=西南
  新生仔猪(中小散绝对值): O(14)=中国

03.统计局季度数据 (row0/1=表头, data从row2起, 日期在B列=col1)
  能繁母猪: C(2)=存栏量, D(3)=环比, E(4)=同比
  生猪存栏: F(5)=存栏量, G(6)=商品猪, H(7)=环比, I(8)=同比
  生猪出栏: J(9)=出栏量, K(10)=比例, L(11)=环比, M(12)=同比, N(13)=累计, O(14)=同比
  猪肉产量: U(20)=产量, V(21)=同比

分省区存栏 (row0/1/2=表头, data从row3起, 日期在A列=col0)
  山东: B(1)=生猪环比, E(4)=能繁存栏, F(5)=能繁环比
  广西: H(7)=生猪环比, J(9)=能繁环比
  广东: L(11)=生猪环比, N(13)=能繁环比
  川渝: P(15)=生猪指数, Q(16)=生猪环比, S(18)=能繁指数, T(19)=能繁环比

饲料数据汇总 (row0/1=表头, data从row2起, 日期在B列=col1)
  协会口径(col2-24): 母猪料产量(10)/环比(11), 仔猪料产量(15)/环比(16), 育肥料产量(20)/环比(21), 猪料合计产量(25)/环比(26)
  涌益口径(col28-45): 母猪料产量(28)/环比(29), 仔猪料产量(33)/环比(34), 育肥料产量(38)/环比(39)

猪肉进口 (row0/1/2=表头, data从row3起, 日期在B列=col1)
  猪肉进口量: C(2)=进口量(万吨), D(3)=同比
"""

import logging
from collections import defaultdict
from datetime import date
from typing import Optional

import openpyxl

from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_month, clean_value, compute_mom_pct, province_to_code

logger = logging.getLogger(__name__)


class IndustryDataReader(BaseSheetReader):
    """Reader for 2、【生猪产业数据】.xlsx"""

    FILE_PATTERN = "生猪产业数据"

    # ── helpers ──────────────────────────────────────────────

    @staticmethod
    def _cell(row: tuple, idx: int):
        """Safe cell accessor from openpyxl row tuple."""
        if row is None or idx < 0 or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _cv(row: tuple, idx: int) -> Optional[float]:
        """Clean numeric value from cell."""
        return clean_value(IndustryDataReader._cell(row, idx))

    @staticmethod
    def _make_record(
        month_date,
        indicator_code: str,
        sub_category: str,
        source: str,
        value: float,
        value_type: str = "mom_pct",
        unit: str = "%",
        region_code: str = "NATION",
        batch_id: int = 0,
    ) -> dict:
        return {
            "month_date": month_date,
            "region_code": region_code,
            "indicator_code": indicator_code,
            "sub_category": sub_category,
            "source": source,
            "value": round(value, 6),
            "value_type": value_type,
            "unit": unit,
            "batch_id": batch_id,
        }

    # ── main entry point ─────────────────────────────────────

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        logger.info("Reading industry data: %s", filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        monthly: list[dict] = []
        quarterly: list[dict] = []

        try:
            monthly += self._read_nyb(wb)
        except Exception:
            logger.exception("Error reading NYB sheet")

        try:
            monthly += self._read_association_feed(wb)
        except Exception:
            logger.exception("Error reading 02.协会猪料 sheet")

        try:
            monthly += self._read_yongyi_draft(wb)
        except Exception:
            logger.exception("Error reading 4.2涌益底稿 sheet")

        try:
            monthly += self._read_yongyi_sample(wb)
        except Exception:
            logger.exception("Error reading 涌益样本 sheet")

        try:
            monthly += self._read_ganglian_data(wb)
        except Exception:
            logger.exception("Error reading 4.1.钢联数据 sheet")

        try:
            monthly += self._read_ganglian_draft(wb)
        except Exception:
            logger.exception("Error reading 钢联底稿 sheet")

        try:
            quarterly += self._read_statistics_bureau(wb)
        except Exception:
            logger.exception("Error reading 03.统计局季度数据 sheet")

        try:
            monthly += self._read_provincial_inventory(wb)
        except Exception:
            logger.exception("Error reading 分省区存栏 sheet")

        try:
            monthly += self._read_feed_summary(wb)
        except Exception:
            logger.exception("Error reading 饲料数据汇总 sheet")

        try:
            monthly += self._read_pork_import(wb)
        except Exception:
            logger.exception("Error reading 猪肉进口 sheet")

        try:
            monthly += self._read_designated_slaughter(wb)
        except Exception:
            logger.exception("Error reading 定点屠宰 sheet")

        wb.close()

        logger.info(
            "Industry data extracted: %d monthly records, %d quarterly records",
            len(monthly), len(quarterly),
        )
        return {
            "fact_monthly_indicator": monthly,
            "fact_quarterly_stats": quarterly,
        }

    # ── Sheet: NYB ───────────────────────────────────────────

    def _read_nyb(self, wb) -> list[dict]:
        """NYB sheet: 能繁/仔猪/存栏环比 (已是环比百分比, 直接使用)
        Row 0: 大类表头 (能繁环比 | 新生仔猪环比 | 中大猪环比 | 存栏环比 ...)
        Row 1: 子表头 (月度 | 全国 | 规模场 | CR26 | 小散户 ...)
        Data from row 2+, date in col B(1).
        """
        if "NYB" not in wb.sheetnames:
            logger.warning("Sheet 'NYB' not found")
            return []

        ws = wb["NYB"]
        records = []

        # Column mappings: (col_idx, indicator_code, sub_category)
        col_map = [
            # 能繁环比
            (2, "breeding_sow_inventory", "nation"),
            (3, "breeding_sow_inventory", "scale"),
            (5, "breeding_sow_inventory", "small"),
            # 新生仔猪环比
            (6, "piglet_inventory", "nation"),
            (7, "piglet_inventory", "scale"),
            (8, "piglet_inventory", "small"),
            # 中大猪环比
            (10, "medium_large_hog", "nation"),
            # 存栏环比
            (16, "hog_inventory", "nation"),
            (17, "hog_inventory", "scale"),
            (18, "hog_inventory", "small"),
            # 出栏环比
            (19, "hog_turnover", "nation"),
            (20, "hog_turnover", "scale"),
            (21, "hog_turnover", "small"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                md = parse_month(self._cell(row, 1))
                if not md:
                    continue

                for col_idx, indicator, sub_cat in col_map:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat,
                            source="NYB",
                            value=v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("NYB row parse error", exc_info=True)

        logger.info("NYB: extracted %d records", len(records))
        return records

    # ── Sheet: 02.协会猪料 ───────────────────────────────────

    def _read_association_feed(self, wb) -> list[dict]:
        """协会猪料 sheet: 母猪料/仔猪料/育肥料 产量和环比
        Row 0/1: headers. Data from row 2+, date in col B(1).
        母猪料: 产量=H(7), 环比=I(8)
        仔猪料: 产量=M(12), 环比=N(13)
        育肥料: 产量=R(17), 环比=S(18)
        """
        sheet_name = "02.协会猪料"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        feed_cols = [
            # (prod_col, mom_col, indicator_code)
            (7, 8, "breeding_sow_feed"),
            (12, 13, "piglet_feed"),
            (17, 18, "fattening_feed"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                md = parse_month(self._cell(row, 1))
                if not md:
                    continue

                for prod_col, mom_col, indicator in feed_cols:
                    # Absolute production value
                    prod_v = self._cv(row, prod_col)
                    if prod_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category="production",
                            source="ASSOCIATION",
                            value=prod_v,
                            value_type="abs",
                            unit="万吨",
                            batch_id=self.batch_id,
                        ))

                    # MoM percentage
                    mom_v = self._cv(row, mom_col)
                    if mom_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category="",
                            source="ASSOCIATION",
                            value=mom_v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("协会猪料 row parse error", exc_info=True)

        logger.info("协会猪料: extracted %d records", len(records))
        return records

    # ── Sheet: 4.2涌益底稿 ───────────────────────────────────

    def _read_yongyi_draft(self, wb) -> list[dict]:
        """涌益底稿: 能繁/大猪存栏环比(已是环比%), 饲料环比(直接用)
        Row 0/1: headers. Data from row 2+, date in col A(0).
        B(1)=大猪存栏环比(%), G(6)=能繁母猪环比(%)
        Note: Despite the column header saying "大猪存栏" / "能繁母猪", the actual
        values are month-on-month percentage changes (e.g. 0.4, 2.36, -2),
        NOT absolute inventory numbers. We store them directly as mom_pct.
        P(15)=猪料环比, Q(16)=后备母猪料, R(17)=其他母猪料, S(18)=教保料, T(19)=育肥料, U(20)=小猪料
        """
        sheet_name = "4.2涌益底稿"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Direct MoM columns: (col_idx, indicator_code, sub_category)
        direct_mom_cols = [
            # Inventory MoM (already percentage values)
            (6, "breeding_sow_inventory", ""),
            (1, "hog_inventory", ""),
            # Feed MoM (already percentage values)
            (15, "pig_feed_sales", "total"),
            (16, "breeding_sow_feed", "reserve"),
            (17, "breeding_sow_feed", "other"),
            (18, "piglet_feed", "nursery"),
            (19, "fattening_feed", ""),
            (20, "piglet_feed", "small"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                md = parse_month(self._cell(row, 0))
                if not md:
                    continue

                for col_idx, indicator, sub_cat in direct_mom_cols:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat,
                            source="YONGYI",
                            value=v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("涌益底稿 row parse error", exc_info=True)

        logger.info("涌益底稿: extracted %d records", len(records))
        return records

    # ── Sheet: 涌益样本 ─────────────────────────────────────

    def _read_yongyi_sample(self, wb) -> list[dict]:
        """涌益样本: 淘汰母猪屠宰量(V=col21, 底稿原值->算环比), 存栏结构等
        Row 0/1/2: headers. Data from row 3+, date in col A(0).
        存栏量: J(9)=总存栏, K(10)=大猪, L(11)=中猪, M(12)=小猪
        环比: N(13)=存栏, O(14)=大猪, P(15)=中猪, Q(16)=小猪
        淘汰母猪: V(21)=屠宰量(绝对值)
        """
        sheet_name = "涌益样本"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Collect raw slaughter values for MoM
        cull_raw: dict[str, float] = {}

        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                md = parse_month(self._cell(row, 0))
                if not md:
                    continue
                month_key = md.strftime("%Y-%m")

                # 淘汰母猪屠宰量 (absolute value, col 21)
                cull_v = self._cv(row, 21)
                if cull_v is not None:
                    cull_raw[month_key] = cull_v
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="cull_slaughter",
                        sub_category="abs_value",
                        source="YONGYI",
                        value=cull_v,
                        value_type="abs",
                        unit="头",
                        batch_id=self.batch_id,
                    ))

                # Inventory MoM from sample (col 13 = total inventory MoM)
                inv_mom = self._cv(row, 13)
                if inv_mom is not None:
                    # The values appear to be decimal ratios (e.g. -0.02 = -2%)
                    if 0 < abs(inv_mom) <= 1.5:
                        inv_mom = inv_mom * 100
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="hog_inventory",
                        sub_category="sample_total",
                        source="YONGYI",
                        value=round(inv_mom, 2),
                        value_type="mom_pct",
                        unit="%",
                        batch_id=self.batch_id,
                    ))

            except Exception:
                logger.debug("涌益样本 row parse error", exc_info=True)

        # Compute cull slaughter MoM
        cull_mom = compute_mom_pct(cull_raw)
        for mk, mom_v in cull_mom.items():
            md = parse_month(mk + "-01")
            if md is None:
                continue
            records.append(self._make_record(
                month_date=md,
                indicator_code="cull_slaughter",
                sub_category="",
                source="YONGYI",
                value=mom_v,
                value_type="mom_pct",
                unit="%",
                batch_id=self.batch_id,
            ))

        logger.info("涌益样本: extracted %d records", len(records))
        return records

    # ── Sheet: 4.1.钢联数据 ─────────────────────────────────

    def _read_ganglian_data(self, wb) -> list[dict]:
        """4.1.钢联数据: 存栏/能繁/仔猪/饲料 环比 (已是环比百分比)
        Row 0/1: headers. Data from row 2+, date in col A(0).
        """
        sheet_name = "4.1.钢联数据"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Column mappings: (col_idx, indicator_code, sub_category)
        col_map = [
            # 存栏环比(中小散)
            (1, "hog_inventory", "small_nation"),
            (2, "hog_inventory", "small_northeast"),
            (3, "hog_inventory", "small_north"),
            (4, "hog_inventory", "small_southwest"),
            # 能繁存栏环比
            (11, "breeding_sow_inventory", "nation"),
            (12, "breeding_sow_inventory", "northeast"),
            (13, "breeding_sow_inventory", "southwest"),
            # 新生仔猪数环比(中小散)
            (14, "piglet_inventory", "nation"),
            (15, "piglet_inventory", "northeast"),
            (16, "piglet_inventory", "southwest"),
            (17, "piglet_inventory", "central"),
            # 育肥料环比
            (18, "fattening_feed", "nation"),
            (19, "fattening_feed", "northeast"),
            (20, "fattening_feed", "north"),
            (21, "fattening_feed", "southwest"),
            # 仔猪饲料环比
            (22, "piglet_feed", "nation"),
            (23, "piglet_feed", "northeast"),
            (24, "piglet_feed", "north"),
            (25, "piglet_feed", "southwest"),
            # 母猪料环比
            (26, "breeding_sow_feed", "nation"),
            (27, "breeding_sow_feed", "northeast"),
            (28, "breeding_sow_feed", "north"),
            (29, "breeding_sow_feed", "southwest"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                md = parse_month(self._cell(row, 0))
                if not md:
                    continue

                for col_idx, indicator, sub_cat in col_map:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat,
                            source="GANGLIAN",
                            value=v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("4.1.钢联数据 row parse error", exc_info=True)

        logger.info("4.1.钢联数据: extracted %d records", len(records))
        return records

    # ── Sheet: 钢联底稿 ─────────────────────────────────────

    def _read_ganglian_draft(self, wb) -> list[dict]:
        """钢联底稿: 存栏/能繁/仔猪 绝对值 (底稿原值->算环比)
        Row 0-6: headers/metadata. Data from row 7+, date in col A(0).
        存栏(中小散): B(1)=中国, C(2)=东北, D(3)=华北, E(4)=西南
        能繁(中小散): L(11)=中国, M(12)=东北, N(13)=西南
        新生仔猪: O(14)=中国
        """
        sheet_name = "钢联底稿"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Collect raw absolute values for MoM calculation
        raw_data: dict[str, dict[str, float]] = defaultdict(dict)
        # Key structure: raw_data[month_key][(indicator, sub_cat)] = value

        abs_cols = [
            # (col_idx, indicator_code, sub_category)
            (1, "hog_inventory", "small_nation"),
            (2, "hog_inventory", "small_northeast"),
            (3, "hog_inventory", "small_north"),
            (4, "hog_inventory", "small_southwest"),
            (11, "breeding_sow_inventory", "small_nation"),
            (12, "breeding_sow_inventory", "small_northeast"),
            (13, "breeding_sow_inventory", "small_southwest"),
            (14, "piglet_inventory", "small_nation"),
        ]

        for row in ws.iter_rows(min_row=8, values_only=True):
            try:
                md = parse_month(self._cell(row, 0))
                if not md:
                    continue
                month_key = md.strftime("%Y-%m")

                for col_idx, indicator, sub_cat in abs_cols:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        raw_data[month_key][(indicator, sub_cat)] = v
                        # Store absolute value
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat + "_abs",
                            source="GANGLIAN",
                            value=v,
                            value_type="abs",
                            unit="万头",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("钢联底稿 row parse error", exc_info=True)

        # Compute MoM for each indicator/sub_category series
        series_keys = set()
        for mk, kv in raw_data.items():
            series_keys.update(kv.keys())

        for series_key in series_keys:
            indicator, sub_cat = series_key
            by_month = {}
            for mk, kv in raw_data.items():
                if series_key in kv:
                    by_month[mk] = kv[series_key]

            mom_dict = compute_mom_pct(by_month)
            for mk, mom_v in mom_dict.items():
                md = parse_month(mk + "-01")
                if md is None:
                    continue
                records.append(self._make_record(
                    month_date=md,
                    indicator_code=indicator,
                    sub_category=sub_cat,
                    source="GANGLIAN",
                    value=mom_v,
                    value_type="mom_pct",
                    unit="%",
                    batch_id=self.batch_id,
                ))

        logger.info("钢联底稿: extracted %d records", len(records))
        return records

    # ── Sheet: 03.统计局季度数据 ─────────────────────────────

    def _read_statistics_bureau(self, wb) -> list[dict]:
        """统计局季度数据 -> fact_quarterly_stats
        Row 0/1: headers. Data from row 2+, date in col B(1).
        C(2)=能繁存栏量, D(3)=能繁环比, E(4)=能繁同比
        F(5)=生猪存栏量, H(7)=生猪环比, I(8)=生猪同比
        J(9)=出栏量, L(11)=出栏环比, M(12)=出栏同比, N(13)=累计出栏
        U(20)=猪肉产量, V(21)=猪肉产量同比
        """
        sheet_name = "03.统计局季度数据"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        col_map = [
            # (col_idx, indicator_code, unit)
            (2, "breeding_sow_inventory", "万头"),
            (3, "breeding_sow_inventory_qoq", "%"),
            (4, "breeding_sow_inventory_yoy", "%"),
            (5, "hog_inventory", "万头"),
            (6, "commercial_hog_inventory", "万头"),
            (7, "hog_inventory_qoq", "%"),
            (8, "hog_inventory_yoy", "%"),
            (9, "hog_turnover", "万头"),
            (11, "hog_turnover_qoq", "%"),
            (12, "hog_turnover_yoy", "%"),
            (13, "hog_turnover_cumulative", "万头"),
            (14, "hog_turnover_cumulative_yoy", "%"),
            (15, "designated_slaughter", "万头"),
            (16, "designated_slaughter_yoy", "%"),
            (20, "pork_production", "万吨"),
            (21, "pork_production_yoy", "%"),
            (25, "pork_import", "万吨"),
            (26, "pork_import_yoy", "%"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                raw_date = self._cell(row, 1)
                qd = parse_month(raw_date)
                if not qd:
                    continue

                for col_idx, indicator, unit in col_map:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        records.append({
                            "quarter_date": qd,
                            "indicator_code": indicator,
                            "region_code": "NATION",
                            "value": round(v, 4),
                            "unit": unit,
                            "batch_id": self.batch_id,
                        })
            except Exception:
                logger.debug("统计局季度数据 row parse error", exc_info=True)

        logger.info("统计局季度数据: extracted %d records", len(records))
        return records

    # ── Sheet: 分省区存栏 ───────────────────────────────────

    def _read_provincial_inventory(self, wb) -> list[dict]:
        """分省区存栏: 各省份存栏环比
        Row 0/1/2: headers. Data from row 3+, date in col A(0).
        山东: B(1)=生猪环比, F(5)=能繁环比
        广西: H(7)=生猪环比, J(9)=能繁环比
        广东: L(11)=生猪环比, N(13)=能繁环比
        川渝: Q(16)=生猪环比, T(19)=能繁环比
        """
        sheet_name = "分省区存栏"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        province_cols = [
            # (hog_col, breeding_col, province_name)
            (1, 5, "山东"),
            (7, 9, "广西"),
            (11, 13, "广东"),
            (16, 19, "川渝"),
        ]

        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                md = parse_month(self._cell(row, 0))
                if not md:
                    continue

                for hog_col, breeding_col, prov_name in province_cols:
                    region = province_to_code(prov_name)
                    if prov_name == "川渝":
                        region = "SICHUAN"  # 川渝 mapped to Sichuan

                    if not region:
                        continue

                    hog_v = self._cv(row, hog_col)
                    if hog_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code="hog_inventory",
                            sub_category="provincial",
                            source="ASSOCIATION",
                            value=hog_v,
                            value_type="mom_pct",
                            unit="%",
                            region_code=region,
                            batch_id=self.batch_id,
                        ))

                    breeding_v = self._cv(row, breeding_col)
                    if breeding_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code="breeding_sow_inventory",
                            sub_category="provincial",
                            source="ASSOCIATION",
                            value=breeding_v,
                            value_type="mom_pct",
                            unit="%",
                            region_code=region,
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("分省区存栏 row parse error", exc_info=True)

        logger.info("分省区存栏: extracted %d records", len(records))
        return records

    # ── Sheet: 饲料数据汇总 ─────────────────────────────────

    def _read_feed_summary(self, wb) -> list[dict]:
        """饲料数据汇总: 协会口径+涌益口径 饲料产量/环比
        Row 0/1: headers. Data from row 2+, date in col B(1).
        协会口径:
          母猪料: 产量=K(10), 环比=L(11)
          仔猪料: 产量=P(15), 环比=Q(16)
          育肥料: 产量=U(20), 环比=V(21)
          猪料合计: 产量=Z(25), 环比=AA(26)
        涌益口径:
          母猪料: 产量=AC(28), 环比=AD(29)
          仔猪料: 产量=AH(33), 环比=AI(34)
          育肥料: 产量=AM(38), 环比=AN(39)
        """
        sheet_name = "饲料数据汇总"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Association feed columns
        assoc_cols = [
            # (prod_col, mom_col, indicator, sub_cat)
            (10, 11, "breeding_sow_feed", "production"),
            (15, 16, "piglet_feed", "production"),
            (20, 21, "fattening_feed", "production"),
            (25, 26, "feed_sales", "total"),
        ]

        # Yongyi feed columns
        yongyi_cols = [
            (28, 29, "breeding_sow_feed", "production"),
            (33, 34, "piglet_feed", "production"),
            (38, 39, "fattening_feed", "production"),
        ]

        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                md = parse_month(self._cell(row, 1))
                if not md:
                    continue

                # Association feed
                for prod_col, mom_col, indicator, sub_cat in assoc_cols:
                    prod_v = self._cv(row, prod_col)
                    if prod_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat,
                            source="ASSOCIATION",
                            value=prod_v,
                            value_type="abs",
                            unit="万吨",
                            batch_id=self.batch_id,
                        ))

                    mom_v = self._cv(row, mom_col)
                    if mom_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat + "_mom",
                            source="ASSOCIATION",
                            value=mom_v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))

                # Yongyi feed
                for prod_col, mom_col, indicator, sub_cat in yongyi_cols:
                    prod_v = self._cv(row, prod_col)
                    if prod_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat,
                            source="YONGYI",
                            value=prod_v,
                            value_type="abs",
                            unit="万吨",
                            batch_id=self.batch_id,
                        ))

                    mom_v = self._cv(row, mom_col)
                    if mom_v is not None:
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code=indicator,
                            sub_category=sub_cat + "_mom",
                            source="YONGYI",
                            value=mom_v,
                            value_type="mom_pct",
                            unit="%",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("饲料数据汇总 row parse error", exc_info=True)

        logger.info("饲料数据汇总: extracted %d records", len(records))
        return records

    # ── Sheet: 猪肉进口 ─────────────────────────────────────

    def _read_pork_import(self, wb) -> list[dict]:
        """猪肉进口: 月度进口量/同比
        Row 0/1/2: headers. Data from row 3+, date in col B(1).
        C(2)=进口量(万吨), D(3)=同比, F(5)=累计, G(6)=累计同比
        H(7)=猪肉+杂碎进口量
        """
        sheet_name = "猪肉进口"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                md = parse_month(self._cell(row, 1))
                if not md:
                    continue

                # Pork import volume
                import_v = self._cv(row, 2)
                if import_v is not None:
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="pork_import",
                        sub_category="pork_only",
                        source="CUSTOMS",
                        value=import_v,
                        value_type="abs",
                        unit="万吨",
                        batch_id=self.batch_id,
                    ))

                # YoY
                yoy_v = self._cv(row, 3)
                if yoy_v is not None:
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="pork_import",
                        sub_category="pork_only_yoy",
                        source="CUSTOMS",
                        value=yoy_v,
                        value_type="yoy_pct",
                        unit="%",
                        batch_id=self.batch_id,
                    ))

                # Cumulative
                cum_v = self._cv(row, 5)
                if cum_v is not None:
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="pork_import",
                        sub_category="pork_cumulative",
                        source="CUSTOMS",
                        value=cum_v,
                        value_type="abs",
                        unit="万吨",
                        batch_id=self.batch_id,
                    ))

                # Pork + offal import
                offal_v = self._cv(row, 7)
                if offal_v is not None:
                    records.append(self._make_record(
                        month_date=md,
                        indicator_code="pork_import",
                        sub_category="pork_and_offal",
                        source="CUSTOMS",
                        value=offal_v,
                        value_type="abs",
                        unit="万吨",
                        batch_id=self.batch_id,
                    ))
            except Exception:
                logger.debug("猪肉进口 row parse error", exc_info=True)

        logger.info("猪肉进口: extracted %d records", len(records))
        return records

    # ── Sheet: 定点屠宰 ─────────────────────────────────────

    def _read_designated_slaughter(self, wb) -> list[dict]:
        """定点屠宰: 横向年份布局, 每行=一个月份
        Row 0: headers (屠宰量, 2009年, 2010年, ...)
        Row 1-12: 1月~12月, 各年屠宰量
        年份列从C(2)开始: 2009=col2, 2010=col3, ..., 2025=col18
        """
        sheet_name = "定点屠宰"
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found", sheet_name)
            return []

        ws = wb[sheet_name]
        records = []

        # Read header row to get year columns
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = list(row)
            break

        if not header_row:
            return records

        # Parse year from header (e.g. "2009年" -> 2009)
        year_cols: list[tuple[int, int]] = []  # (col_idx, year)
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            s = str(cell).strip().replace("年", "")
            try:
                year = int(s)
                if 2000 <= year <= 2030:
                    year_cols.append((col_idx, year))
            except ValueError:
                continue

        # Data rows: row 1-12 correspond to months 1-12
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=13, values_only=True)):
            month_num = row_idx + 1  # 1-based month
            try:
                for col_idx, year in year_cols:
                    v = self._cv(row, col_idx)
                    if v is not None:
                        md = date(year, month_num, 1)
                        records.append(self._make_record(
                            month_date=md,
                            indicator_code="designated_slaughter",
                            sub_category="",
                            source="STATISTICS_BUREAU",
                            value=v,
                            value_type="abs",
                            unit="万头",
                            batch_id=self.batch_id,
                        ))
            except Exception:
                logger.debug("定点屠宰 row parse error", exc_info=True)

        logger.info("定点屠宰: extracted %d records", len(records))
        return records
