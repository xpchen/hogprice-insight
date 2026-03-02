"""
R07 - 生猪基差和月间价差研究
Source: 4、生猪基差和月间价差研究.xlsx (10 sheets)
Target: fact_futures_basis

Sheets 概要:
  主力合约基差       - 主力合约结算价 / 全国均价 / 升贴水比率
  导出数据           - 现货价格(元/kg) / 全国均价(元/吨)
  现货盘面和升贴水底稿 - 各合约结算价 & 各合约升贴水
  03合约             - 03合约分年结算价 / 升贴水 / 03-05价差
  05合约             - 05合约升贴水 / 分年结算价
  07合约             - 07合约升贴水 / 分年结算价
  09合约             - 09合约分年结算价 / 升贴水
  Sheet3 (11合约)    - 11合约升贴水 / 分年结算价
  09合约季节性       - 09合约季节性升贴水（MM-DD × 年份）
  11合约季节性       - 11合约季节性升贴水（MM-DD × 年份）
"""
import logging
from openpyxl import load_workbook
from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, clean_value

logger = logging.getLogger(__name__)


class FuturesBasisReader(BaseSheetReader):
    """读取生猪基差和月间价差 → fact_futures_basis"""

    FILE_PATTERN = "4、生猪基差和月间价差研究"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        logger.info("开始读取基差/价差文件: %s", filepath)
        wb = load_workbook(filepath, data_only=True, read_only=True)

        records: list[dict] = []

        records += self._read_main_contract_basis(wb)
        records += self._read_export_data(wb)
        records += self._read_spot_premium_draft(wb)
        records += self._read_contract_03(wb)
        records += self._read_contract_05(wb)
        records += self._read_contract_07(wb)
        records += self._read_contract_09(wb)
        records += self._read_contract_11(wb)
        records += self._read_seasonal_09(wb)
        records += self._read_seasonal_11(wb)

        wb.close()
        logger.info("基差/价差读取完成: 共 %d 条记录", len(records))
        return {"fact_futures_basis": records}

    # ──────────────────────────────────────────────
    # Sheet 1: 主力合约基差
    # Row 1: 日期(A) | 主力合约(B) | 全国均价(C) | 盘面升贴水比率(D)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_main_contract_basis(self, wb) -> list[dict]:
        sheet_name = "主力合约基差"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[0].value if len(row) > 0 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 主力合约结算价
                settle = clean_value(row[1].value if len(row) > 1 else None)
                if settle is not None:
                    results.append(self._rec(trade_date, "LH_MAIN", "NATION",
                                             "settle_main_contract", settle, "元/吨"))

                # 全国现货均价
                spot = clean_value(row[2].value if len(row) > 2 else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_MAIN", "NATION",
                                             "spot_national_avg", spot, "元/吨"))

                # 升贴水比率
                premium = clean_value(row[3].value if len(row) > 3 else None)
                if premium is not None:
                    results.append(self._rec(trade_date, "LH_MAIN", "NATION",
                                             "premium_rate_main", premium, "ratio"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 2: 导出数据
    # Row 1: 日期(A) | 生猪价格(B, 元/kg) | 全国均价(C, 元/吨)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_export_data(self, wb) -> list[dict]:
        sheet_name = "导出数据"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[0].value if len(row) > 0 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 生猪价格 (元/kg)
                price_kg = clean_value(row[1].value if len(row) > 1 else None)
                if price_kg is not None:
                    results.append(self._rec(trade_date, None, "NATION",
                                             "spot_price_kg", price_kg, "元/公斤"))

                # 全国均价 (元/吨)
                price_ton = clean_value(row[2].value if len(row) > 2 else None)
                if price_ton is not None:
                    results.append(self._rec(trade_date, None, "NATION",
                                             "spot_price_ton", price_ton, "元/吨"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 3: 现货盘面和升贴水底稿
    # Row 1: _ | 日期(B) | 09合约(C) | 11合约(D) | 01合约(E) | 03合约(F) | 05合约(G) | 07合约(H) |
    #         _ | 主力合约(J) | 全国均价(K) | 09升贴水(L) | 11升贴水(M) | 03升贴水(N) | 05升贴水(O) | 07升贴水(P)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_spot_premium_draft(self, wb) -> list[dict]:
        sheet_name = "现货盘面和升贴水底稿"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        # 合约结算价列: col_idx(0-based) -> (contract_code, indicator_code)
        settle_cols = {
            2: ("LH_M09", "settle_09"),
            3: ("LH_M11", "settle_11"),
            4: ("LH_M01", "settle_01"),
            5: ("LH_M03", "settle_03"),
            6: ("LH_M05", "settle_05"),
            7: ("LH_M07", "settle_07"),
        }
        # 主力合约 / 全国均价
        main_col = 9       # J (0-based)
        spot_col = 10       # K
        # 升贴水列
        premium_cols = {
            11: ("LH_M09", "premium_09"),
            12: ("LH_M11", "premium_11"),
            13: ("LH_M03", "premium_03"),
            14: ("LH_M05", "premium_05"),
            15: ("LH_M07", "premium_07"),
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 各合约结算价
                for ci, (cc, ic) in settle_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION", ic, val, "元/吨"))

                # 主力合约结算价
                main_val = clean_value(row[main_col].value if len(row) > main_col else None)
                if main_val is not None:
                    results.append(self._rec(trade_date, "LH_MAIN", "NATION",
                                             "settle_main_draft", main_val, "元/吨"))

                # 全国现货均价
                spot_val = clean_value(row[spot_col].value if len(row) > spot_col else None)
                if spot_val is not None:
                    results.append(self._rec(trade_date, None, "NATION",
                                             "spot_avg_draft", spot_val, "元/吨"))

                # 各合约升贴水
                for ci, (cc, ic) in premium_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION", ic, val, "ratio"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 4: 03合约
    # Row 1: _ | 日期(B) | 2203(C) | 2303(D) | 2403(E) | 2503(F) | 2603(G) | 全国均价(H) | 03合约升贴水(I)
    #         _ | _ | 2203(K) | 2303(L) | 2403(M) | 2503(N) | 2603(O)  <- 03-05价差
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_contract_03(self, wb) -> list[dict]:
        sheet_name = "03合约"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        # 年份合约结算价列 (0-based col idx -> header contract code)
        settle_year_cols = {2: "LH2203", 3: "LH2303", 4: "LH2403", 5: "LH2503", 6: "LH2603"}
        premium_col = 8  # col I (0-based)

        # 03-05 spread columns (0-based): same yearly codes
        spread_year_cols = {10: "LH2203", 11: "LH2303", 12: "LH2403", 13: "LH2503", 14: "LH2603"}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 各年份03合约结算价
                for ci, cc in settle_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "settle_03", val, "元/吨"))

                # 全国均价
                spot = clean_value(row[7].value if len(row) > 7 else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_M03", "NATION",
                                             "spot_avg_03", spot, "元/吨"))

                # 03合约升贴水（综合）
                prem = clean_value(row[premium_col].value if len(row) > premium_col else None)
                if prem is not None:
                    results.append(self._rec(trade_date, "LH_M03", "NATION",
                                             "premium_03", prem, "ratio"))

                # 03-05价差（分年份）
                for ci, cc in spread_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "spread_03_05", val, "ratio"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 5: 05合约
    # Row 1: _ | 日期(B) | 05合约升贴水(C) | 全国均价(D) | 2205(E) | 2305(F) | 2405(G) | 2505(H) | 2605(I)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_contract_05(self, wb) -> list[dict]:
        sheet_name = "05合约"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        settle_year_cols = {4: "LH2205", 5: "LH2305", 6: "LH2405", 7: "LH2505", 8: "LH2605"}
        premium_col = 2  # col C (0-based)
        spot_col = 3     # col D

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 05合约升贴水
                prem = clean_value(row[premium_col].value if len(row) > premium_col else None)
                if prem is not None:
                    results.append(self._rec(trade_date, "LH_M05", "NATION",
                                             "premium_05", prem, "ratio"))

                # 全国均价
                spot = clean_value(row[spot_col].value if len(row) > spot_col else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_M05", "NATION",
                                             "spot_avg_05", spot, "元/吨"))

                # 各年份05合约结算价
                for ci, cc in settle_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "settle_05", val, "元/吨"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 6: 07合约
    # Row 1: _ | 日期(B) | 全国均价(C) | 07合约升贴水(D) | 2207(E) | 2307(F) | 2407(G) | 2507(H) | 2607(I)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_contract_07(self, wb) -> list[dict]:
        sheet_name = "07合约"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        settle_year_cols = {4: "LH2207", 5: "LH2307", 6: "LH2407", 7: "LH2507", 8: "LH2607"}
        spot_col = 2     # col C (0-based)
        premium_col = 3  # col D

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 全国均价
                spot = clean_value(row[spot_col].value if len(row) > spot_col else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_M07", "NATION",
                                             "spot_avg_07", spot, "元/吨"))

                # 07合约升贴水
                prem = clean_value(row[premium_col].value if len(row) > premium_col else None)
                if prem is not None:
                    results.append(self._rec(trade_date, "LH_M07", "NATION",
                                             "premium_07", prem, "ratio"))

                # 各年份07合约结算价
                for ci, cc in settle_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "settle_07", val, "元/吨"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 7: 09合约
    # Row 1: _ | _ | 全国均价(C) | 2109(D) | 2209(E) | 2309(F) | 2409(G) | 2509(H) | 2609(I) | 09合约升贴水(J)
    # Row 2+: data (date in col B, 0-based idx 1)
    # ──────────────────────────────────────────────
    def _read_contract_09(self, wb) -> list[dict]:
        sheet_name = "09合约"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        settle_year_cols = {3: "LH2109", 4: "LH2209", 5: "LH2309", 6: "LH2409", 7: "LH2509", 8: "LH2609"}
        spot_col = 2     # col C (0-based)
        premium_col = 9  # col J

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 全国均价
                spot = clean_value(row[spot_col].value if len(row) > spot_col else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_M09", "NATION",
                                             "spot_avg_09", spot, "元/吨"))

                # 各年份09合约结算价
                for ci, cc in settle_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "settle_09", val, "元/吨"))

                # 09合约升贴水
                prem = clean_value(row[premium_col].value if len(row) > premium_col else None)
                if prem is not None:
                    results.append(self._rec(trade_date, "LH_M09", "NATION",
                                             "premium_09", prem, "ratio"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 8: Sheet3 (11合约)
    # Row 1: _ | 日期(B) | 全国均价(C) | 11合约升贴水(D) | 2111(E) | 2211(F) | 2311(G) | 2411(H) | 2411(I) | 2611(J)
    # Note: col H and I both say 2411 in header; col I is likely 2511 (typo in source)
    # Row 2+: data
    # ──────────────────────────────────────────────
    def _read_contract_11(self, wb) -> list[dict]:
        sheet_name = "Sheet3"
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s (11合约)", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        # Note: col H(idx7) and col I(idx8) both have header "2411"; col I is likely 2511
        settle_year_cols = {4: "LH2111", 5: "LH2211", 6: "LH2311", 7: "LH2411", 8: "LH2511", 9: "LH2611"}
        spot_col = 2     # col C (0-based)
        premium_col = 3  # col D

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                trade_date = parse_date(row[1].value if len(row) > 1 else None)
                if trade_date is None:
                    skipped += 1
                    continue

                # 全国均价
                spot = clean_value(row[spot_col].value if len(row) > spot_col else None)
                if spot is not None:
                    results.append(self._rec(trade_date, "LH_M11", "NATION",
                                             "spot_avg_11", spot, "元/吨"))

                # 11合约升贴水
                prem = clean_value(row[premium_col].value if len(row) > premium_col else None)
                if prem is not None:
                    results.append(self._rec(trade_date, "LH_M11", "NATION",
                                             "premium_11", prem, "ratio"))

                # 各年份11合约结算价
                for ci, cc in settle_year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is not None:
                        results.append(self._rec(trade_date, cc, "NATION",
                                                 "settle_11", val, "元/吨"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1
        logger.info("[Sheet3/11合约] %d 条, 跳过 %d", len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Sheet 9: 09合约季节性
    # Row 1: _ | _ | 2109(C) | 2209(D) | 2309(E) | 2409(F) | 2509(G) | 2609(H)
    # Row 2+: _ | MM-DD(B) | premium values per year
    # Note: No full date — only MM-DD label. We store with contract_code per year.
    # ──────────────────────────────────────────────
    def _read_seasonal_09(self, wb) -> list[dict]:
        return self._read_seasonal_sheet(
            wb,
            sheet_name="09合约季节性",
            year_cols={2: "LH2109", 3: "LH2209", 4: "LH2309", 5: "LH2409", 6: "LH2509", 7: "LH2609"},
            indicator_code="seasonal_premium_09",
        )

    # ──────────────────────────────────────────────
    # Sheet 10: 11合约季节性
    # Row 1: _ | _ | 2111(C) | 2211(D) | 2311(E) | 2411(F) | 2511(G) | 2611(H)
    # Row 2+: _ | MM-DD(B) | premium values per year
    # ──────────────────────────────────────────────
    def _read_seasonal_11(self, wb) -> list[dict]:
        return self._read_seasonal_sheet(
            wb,
            sheet_name="11合约季节性",
            year_cols={2: "LH2111", 3: "LH2211", 4: "LH2311", 5: "LH2411", 6: "LH2511", 7: "LH2611"},
            indicator_code="seasonal_premium_11",
        )

    def _read_seasonal_sheet(self, wb, sheet_name: str, year_cols: dict,
                             indicator_code: str) -> list[dict]:
        """读取季节性 sheet（MM-DD × 年份矩阵）。
        季节性数据没有完整日期，利用合约代码中的年份 + MM-DD 构造 trade_date。
        例如 LH2309 + "10-08" → 2022-10-08（交易年份 = 交割年份 - 1）
        """
        if sheet_name not in wb.sheetnames:
            logger.warning("缺少 sheet: %s", sheet_name)
            return []
        ws = wb[sheet_name]
        results = []
        skipped = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                # Col B (0-based idx 1) = "MM-DD" string like "10-08"
                mm_dd_raw = row[1].value if len(row) > 1 else None
                if mm_dd_raw is None:
                    skipped += 1
                    continue
                mm_dd = str(mm_dd_raw).strip()
                if not mm_dd or "-" not in mm_dd:
                    skipped += 1
                    continue

                parts = mm_dd.split("-")
                if len(parts) != 2:
                    skipped += 1
                    continue
                try:
                    month_num = int(parts[0])
                    day_num = int(parts[1])
                except ValueError:
                    skipped += 1
                    continue

                for ci, cc in year_cols.items():
                    val = clean_value(row[ci].value if len(row) > ci else None)
                    if val is None:
                        continue

                    # 从合约代码推导年份: LH2309 → delivery year 2023, month 09
                    # 季节性数据的交易日期通常在交割年份的前一年
                    year_code = cc.replace("LH", "")  # e.g. "2309"
                    delivery_year = 2000 + int(year_code[:2])
                    delivery_month = int(year_code[2:])
                    # 如果 MM > delivery_month，交易发生在前一年
                    if month_num >= delivery_month:
                        trade_year = delivery_year - 1
                    else:
                        trade_year = delivery_year

                    try:
                        from datetime import date
                        trade_date = date(trade_year, month_num, day_num)
                    except ValueError:
                        skipped += 1
                        continue

                    results.append(self._rec(trade_date, cc, "NATION", indicator_code, val, "ratio"))

            except Exception as e:
                logger.warning("[%s] Row %d 异常: %s", sheet_name, row_idx, e)
                skipped += 1

        logger.info("[%s] %d 条, 跳过 %d", sheet_name, len(results), skipped)
        return results

    # ──────────────────────────────────────────────
    # Helper
    # ──────────────────────────────────────────────
    def _rec(self, trade_date, contract_code, region_code, indicator_code, value, unit) -> dict:
        """构建一条 fact_futures_basis 记录"""
        return {
            "trade_date": trade_date,
            "contract_code": contract_code,
            "region_code": region_code,
            "indicator_code": indicator_code,
            "value": value,
            "unit": unit,
            "batch_id": self.batch_id,
        }
