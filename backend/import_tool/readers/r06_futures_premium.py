"""
R06 - 生猪期货升贴水数据（盘面结算价）
Source: 4.1、生猪期货升贴水数据（盘面结算价）.xlsx
Target: fact_futures_daily

Sheet 结构 (期货结算价(1月交割连续)_生猪):
  Row 1: 指标名称 | 期货结算价(1月交割连续):生猪 | ... | 期货持仓量:生猪
  Row 2: 单位行
  Row 3: 来源行
  Row 4+: 数据行（日期降序）
  Col A = 交易日期
  Col B = 1月交割连续结算价 (LH_M01)
  Col C = 3月交割连续结算价 (LH_M03)
  Col D = 5月交割连续结算价 (LH_M05)
  Col E = 7月交割连续结算价 (LH_M07)
  Col F = 9月交割连续结算价 (LH_M09)
  Col G = 11月交割连续结算价 (LH_M11)
  Col H = 连续合约结算价 (LH_CONT)
  Col I = 期货持仓量
"""
import logging
from openpyxl import load_workbook
from import_tool.base_reader import BaseSheetReader
from import_tool.utils import parse_date, clean_value

logger = logging.getLogger(__name__)

# 列号 → 连续合约代码
COL_CONTRACT_MAP = {
    2: "LH_M01",   # 1月交割连续
    3: "LH_M03",   # 3月交割连续
    4: "LH_M05",   # 5月交割连续
    5: "LH_M07",   # 7月交割连续
    6: "LH_M09",   # 9月交割连续
    7: "LH_M11",   # 11月交割连续
    8: "LH_CONT",  # 连续
}

SHEET_NAME = "期货结算价(1月交割连续)_生猪"
DATA_START_ROW = 4  # 跳过表头、单位、来源行
OI_COL = 9          # 持仓量列


class FuturesPremiumReader(BaseSheetReader):
    """读取期货结算价（盘面结算价）数据 → fact_futures_daily"""

    FILE_PATTERN = "4.1、生猪期货升贴水数据（盘面结算价）"

    def read_file(self, filepath: str) -> dict[str, list[dict]]:
        logger.info("开始读取期货结算价文件: %s", filepath)
        wb = load_workbook(filepath, data_only=True, read_only=True)

        try:
            ws = wb[SHEET_NAME]
        except KeyError:
            logger.error("找不到 sheet: %s，可用: %s", SHEET_NAME, wb.sheetnames)
            wb.close()
            return {"fact_futures_daily": []}

        records: list[dict] = []
        skipped = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=False), start=DATA_START_ROW):
            try:
                # Col A (1) = 日期
                raw_date = row[0].value if len(row) > 0 else None
                trade_date = parse_date(raw_date)
                if trade_date is None:
                    skipped += 1
                    continue

                # 读取持仓量（所有合约共享同一个持仓量值）
                raw_oi = row[OI_COL - 1].value if len(row) >= OI_COL else None
                oi_val = clean_value(raw_oi)
                open_interest = int(oi_val) if oi_val is not None else None

                # 每列生成一条 settle 记录
                for col_idx, contract_code in COL_CONTRACT_MAP.items():
                    raw_val = row[col_idx - 1].value if len(row) >= col_idx else None
                    settle = clean_value(raw_val)
                    if settle is None:
                        continue

                    rec = {
                        "contract_code": contract_code,
                        "trade_date": trade_date,
                        "open": None,
                        "high": None,
                        "low": None,
                        "close": None,
                        "settle": settle,
                        "pre_settle": None,
                        "chg": None,
                        "volume": None,
                        "open_interest": open_interest if contract_code == "LH_CONT" else None,
                        "oi_chg": None,
                        "turnover": None,
                        "batch_id": self.batch_id,
                    }
                    records.append(rec)

            except Exception as e:
                logger.warning("Row %d 解析异常: %s", row_idx, e)
                skipped += 1
                continue

        wb.close()
        logger.info(
            "期货结算价读取完成: %d 条记录, %d 行跳过",
            len(records), skipped,
        )
        return {"fact_futures_daily": records}
